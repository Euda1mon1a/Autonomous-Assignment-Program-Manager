"""
Tests for the /imports endpoint for parsing Excel files.
"""

import io
import pytest
from openpyxl import Workbook


def create_test_xlsx(rows: list[list]) -> bytes:
    """Create a test xlsx file in memory."""
    wb = Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


class TestParseXlsxEndpoint:
    """Tests for POST /imports/parse-xlsx endpoint."""

    def test_parse_simple_xlsx(self, client):
        """Test parsing a simple Excel file."""
        xlsx_data = create_test_xlsx(
            [
                ["name", "type", "email"],
                ["John Doe", "resident", "john@example.com"],
                ["Jane Smith", "faculty", "jane@example.com"],
            ]
        )

        response = client.post(
            "/api/v1/imports/parse-xlsx",
            files={
                "file": (
                    "test.xlsx",
                    xlsx_data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        assert response.status_code == 200
        data = response.json()
        assert data["success"] is True
        assert data["total_rows"] == 2
        assert data["columns"] == ["name", "type", "email"]
        assert len(data["rows"]) == 2
        assert data["rows"][0]["name"] == "John Doe"
        assert data["rows"][1]["name"] == "Jane Smith"

    def test_parse_xlsx_with_empty_rows(self, client):
        """Test parsing Excel file skips empty rows by default."""
        xlsx_data = create_test_xlsx(
            [
                ["name", "value"],
                ["Row1", "A"],
                [None, None],  # Empty row
                ["Row2", "B"],
            ]
        )

        response = client.post(
            "/api/v1/imports/parse-xlsx",
            files={
                "file": (
                    "test.xlsx",
                    xlsx_data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            data={"skip_empty_rows": "true"},
        )

        assert response.status_code == 200
        data = response.json()
        assert data["total_rows"] == 2  # Empty row skipped
        assert data["rows"][0]["name"] == "Row1"
        assert data["rows"][1]["name"] == "Row2"

    def test_parse_xlsx_invalid_extension(self, client):
        """Test that invalid file extensions are rejected."""
        response = client.post(
            "/api/v1/imports/parse-xlsx",
            files={"file": ("test.txt", b"not an excel file", "text/plain")},
        )

        assert response.status_code == 400
        data = response.json()
        assert "detail" in data
        assert data["detail"]["error_code"] == "INVALID_EXTENSION"

    def test_parse_xlsx_corrupt_file(self, client):
        """Test that corrupt files are handled gracefully."""
        response = client.post(
            "/api/v1/imports/parse-xlsx",
            files={
                "file": (
                    "test.xlsx",
                    b"not valid xlsx content",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        assert response.status_code == 400
        data = response.json()
        assert "detail" in data
        assert data["detail"]["error_code"] == "PARSE_ERROR"

    def test_parse_xlsx_with_dates(self, client):
        """Test that dates are converted to ISO format."""
        from datetime import date

        wb = Workbook()
        ws = wb.active
        ws.append(["date", "value"])
        ws.append([date(2025, 1, 15), "Test"])

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        xlsx_data = buffer.read()

        response = client.post(
            "/api/v1/imports/parse-xlsx",
            files={
                "file": (
                    "test.xlsx",
                    xlsx_data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        assert response.status_code == 200
        data = response.json()
        assert data["rows"][0]["date"] == "2025-01-15"

    def test_parse_xlsx_duplicate_headers(self, client):
        """Test that duplicate headers are renamed and warned."""
        xlsx_data = create_test_xlsx(
            [
                ["name", "name", "value"],  # Duplicate header
                ["A", "B", "C"],
            ]
        )

        response = client.post(
            "/api/v1/imports/parse-xlsx",
            files={
                "file": (
                    "test.xlsx",
                    xlsx_data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        assert response.status_code == 200
        data = response.json()
        # Second "name" should be renamed
        assert "name" in data["columns"]
        assert any("name_" in col for col in data["columns"])
        assert len(data["warnings"]) > 0
        assert any("Duplicate" in w for w in data["warnings"])


class TestListXlsxSheetsEndpoint:
    """Tests for POST /imports/parse-xlsx/sheets endpoint."""

    def test_list_sheets(self, client):
        """Test listing sheets in a workbook."""
        wb = Workbook()
        wb.active.title = "Schedule"
        wb.create_sheet("Leave")
        wb.create_sheet("Personnel")

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        xlsx_data = buffer.read()

        response = client.post(
            "/api/v1/imports/parse-xlsx/sheets",
            files={
                "file": (
                    "test.xlsx",
                    xlsx_data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        assert response.status_code == 200
        data = response.json()
        assert data["success"] is True
        assert data["count"] == 3
        assert "Schedule" in data["sheets"]
        assert "Leave" in data["sheets"]
        assert "Personnel" in data["sheets"]
        assert data["default"] == "Schedule"
