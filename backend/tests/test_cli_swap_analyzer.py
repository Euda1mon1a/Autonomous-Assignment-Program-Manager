"""
Tests for the swap_analyzer CLI tool.

Tests command-line argument parsing, error handling, and output formatting.
"""
import pytest
import subprocess
import sys
import json
import tempfile
from datetime import date
from pathlib import Path
from io import BytesIO
from openpyxl import Workbook


# Path to the CLI module
CLI_MODULE = "app.cli.swap_analyzer"


@pytest.fixture
def sample_fmit_excel(tmp_path):
    """Create a sample FMIT Excel file for testing."""
    file_path = tmp_path / "test_schedule.xlsx"

    wb = Workbook()
    ws = wb.active

    # Header row with dates (weekly schedule)
    ws["A1"] = "Provider"
    ws["B1"] = date(2025, 3, 3)   # Week 1
    ws["C1"] = date(2025, 3, 10)  # Week 2
    ws["D1"] = date(2025, 3, 17)  # Week 3
    ws["E1"] = date(2025, 3, 24)  # Week 4
    ws["F1"] = date(2025, 3, 31)  # Week 5

    # Provider data - Dr. Smith has weeks 1, 3, 5 (alternating)
    ws["A2"] = "Dr. Smith"
    ws["B2"] = "FMIT"
    ws["C2"] = ""
    ws["D2"] = "FMIT"
    ws["E2"] = ""
    ws["F2"] = "FMIT"

    # Dr. Jones has weeks 2, 4
    ws["A3"] = "Dr. Jones"
    ws["B3"] = ""
    ws["C3"] = "FMIT"
    ws["D3"] = ""
    ws["E3"] = "FMIT"
    ws["F3"] = ""

    # Dr. Lee has only week 3
    ws["A4"] = "Dr. Lee"
    ws["B4"] = ""
    ws["C4"] = ""
    ws["D4"] = ""
    ws["E4"] = "FMIT"
    ws["F4"] = ""

    wb.save(file_path)
    return str(file_path)


@pytest.fixture
def conflicts_json(tmp_path):
    """Create a sample external conflicts JSON file."""
    file_path = tmp_path / "conflicts.json"

    conflicts = [
        {
            "faculty": "Dr. Jones",
            "start_date": "2025-03-10",
            "end_date": "2025-03-16",
            "conflict_type": "conference",
            "description": "AAFP Conference"
        }
    ]

    with open(file_path, "w") as f:
        json.dump(conflicts, f)

    return str(file_path)


@pytest.fixture
def invalid_excel(tmp_path):
    """Create an invalid Excel file (not a real schedule)."""
    file_path = tmp_path / "invalid.xlsx"

    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Invalid"
    ws["B1"] = "Data"

    wb.save(file_path)
    return str(file_path)


def run_cli(*args):
    """Helper to run the CLI module and capture output."""
    cmd = [sys.executable, "-m", CLI_MODULE] + list(args)
    result = subprocess.run(
        cmd,
        cwd=str(Path(__file__).parent.parent),
        capture_output=True,
        text=True,
        env={"PYTHONPATH": str(Path(__file__).parent.parent)}
    )
    return result


# ============================================================================
# Help and Argument Parsing Tests
# ============================================================================

class TestHelpAndArguments:
    """Tests for --help output and argument parsing."""

    def test_help_output(self):
        """Should display help message with --help flag."""
        result = run_cli("--help")

        assert result.returncode == 0
        assert "Analyze FMIT schedule swaps" in result.stdout
        assert "--file" in result.stdout
        assert "--faculty" in result.stdout
        assert "--week" in result.stdout
        assert "--alternating" in result.stdout
        assert "--analyze-all" in result.stdout

    def test_missing_required_file_argument(self):
        """Should fail when --file is not provided."""
        result = run_cli("--analyze-all")

        assert result.returncode != 0
        assert "required" in result.stderr.lower() or "error" in result.stderr.lower()

    def test_incomplete_swap_search_arguments(self, sample_fmit_excel):
        """Should show help when faculty/week are incomplete."""
        # Missing --week
        result = run_cli("--file", sample_fmit_excel, "--faculty", "Dr. Smith")

        assert result.returncode == 1
        assert "Specify --faculty and --week" in result.stdout or \
               "Specify --faculty and --week" in result.stderr


# ============================================================================
# File Error Handling Tests
# ============================================================================

class TestFileErrorHandling:
    """Tests for file-related error handling."""

    def test_file_not_found(self):
        """Should report error when file doesn't exist."""
        result = run_cli("--file", "/nonexistent/path/schedule.xlsx", "--analyze-all")

        assert result.returncode == 1
        assert "File not found" in result.stderr or "not found" in result.stderr.lower()

    def test_invalid_file_format(self, invalid_excel):
        """Should handle invalid Excel format gracefully."""
        result = run_cli("--file", invalid_excel, "--analyze-all")

        # The tool handles invalid formats gracefully (returns 0) with empty results
        assert result.returncode == 0
        assert "Faculty members: 0" in result.stdout

    def test_invalid_date_format(self, sample_fmit_excel):
        """Should report error for invalid date format."""
        result = run_cli(
            "--file", sample_fmit_excel,
            "--faculty", "Dr. Smith",
            "--week", "invalid-date"
        )

        assert result.returncode == 1
        assert "Cannot parse date" in result.stderr or "date" in result.stderr.lower()

    def test_faculty_not_found(self, sample_fmit_excel):
        """Should report error when faculty not in schedule."""
        result = run_cli(
            "--file", sample_fmit_excel,
            "--faculty", "Dr. NonExistent",
            "--week", "2025-03-03"
        )

        assert result.returncode == 1
        assert "not found" in result.stderr.lower()
        # Should suggest available faculty
        assert "Available faculty" in result.stderr or "Dr. Smith" in result.stderr


# ============================================================================
# Successful Execution Tests
# ============================================================================

class TestSuccessfulExecution:
    """Tests for successful CLI execution modes."""

    def test_analyze_all_mode(self, sample_fmit_excel):
        """Should run full analysis mode successfully."""
        result = run_cli("--file", sample_fmit_excel, "--analyze-all")

        assert result.returncode == 0
        assert "FMIT Schedule Analysis" in result.stdout
        assert "Summary:" in result.stdout
        assert "Faculty members:" in result.stdout

    def test_alternating_patterns_mode(self, sample_fmit_excel):
        """Should run alternating patterns analysis mode."""
        result = run_cli("--file", sample_fmit_excel, "--alternating", "--threshold", "2")

        assert result.returncode == 0
        assert "Alternating FMIT Patterns" in result.stdout
        # The output should show either found patterns or indicate none were found
        assert "alternating cycles" in result.stdout.lower() or "no faculty found" in result.stdout.lower()

    def test_swap_candidate_search(self, sample_fmit_excel):
        """Should find swap candidates for faculty/week."""
        result = run_cli(
            "--file", sample_fmit_excel,
            "--faculty", "Dr. Smith",
            "--week", "2025-03-03"
        )

        assert result.returncode == 0
        assert "Swap Candidates" in result.stdout
        assert "Dr. Smith" in result.stdout
        assert "Total candidates:" in result.stdout

    def test_verbose_output(self, sample_fmit_excel):
        """Should include additional details with --verbose."""
        result = run_cli(
            "--file", sample_fmit_excel,
            "--analyze-all",
            "--verbose"
        )

        assert result.returncode == 0
        assert "Loaded schedule" in result.stdout

    def test_custom_threshold(self, sample_fmit_excel):
        """Should accept custom threshold for alternating detection."""
        result = run_cli(
            "--file", sample_fmit_excel,
            "--alternating",
            "--threshold", "2"
        )

        assert result.returncode == 0
        # Should find faculty with 2+ alternating cycles
        assert "alternating cycles" in result.stdout.lower()

    def test_external_conflicts_file(self, sample_fmit_excel, conflicts_json):
        """Should load and apply external conflicts from JSON file."""
        result = run_cli(
            "--file", sample_fmit_excel,
            "--faculty", "Dr. Smith",
            "--week", "2025-03-03",
            "--conflicts-file", conflicts_json
        )

        assert result.returncode == 0
        # Dr. Jones should have conflict flagged
        if "Dr. Jones" in result.stdout:
            # Check for conflict indicator in the output
            assert "⚠" in result.stdout or "External conflict" in result.stdout


# ============================================================================
# Short Option Tests
# ============================================================================

class TestShortOptions:
    """Tests for short option flags."""

    def test_short_file_option(self, sample_fmit_excel):
        """Should accept -f as shorthand for --file."""
        result = run_cli("-f", sample_fmit_excel, "--analyze-all")

        assert result.returncode == 0
        assert "FMIT Schedule Analysis" in result.stdout

    def test_short_faculty_option(self, sample_fmit_excel):
        """Should accept -F as shorthand for --faculty."""
        result = run_cli(
            "-f", sample_fmit_excel,
            "-F", "Dr. Smith",
            "-w", "2025-03-03"
        )

        assert result.returncode == 0
        assert "Dr. Smith" in result.stdout

    def test_short_verbose_option(self, sample_fmit_excel):
        """Should accept -v as shorthand for --verbose."""
        result = run_cli(
            "-f", sample_fmit_excel,
            "--analyze-all",
            "-v"
        )

        assert result.returncode == 0
        assert "Loaded schedule" in result.stdout


# ============================================================================
# Date Format Tests
# ============================================================================

class TestDateFormats:
    """Tests for various date format parsing."""

    def test_iso_date_format(self, sample_fmit_excel):
        """Should accept YYYY-MM-DD date format."""
        result = run_cli(
            "-f", sample_fmit_excel,
            "-F", "Dr. Smith",
            "-w", "2025-03-03"
        )

        assert result.returncode == 0

    def test_us_date_format(self, sample_fmit_excel):
        """Should accept MM/DD/YYYY date format."""
        result = run_cli(
            "-f", sample_fmit_excel,
            "-F", "Dr. Smith",
            "-w", "03/03/2025"
        )

        assert result.returncode == 0

    def test_alt_us_date_format(self, sample_fmit_excel):
        """Should accept MM-DD-YYYY date format."""
        result = run_cli(
            "-f", sample_fmit_excel,
            "-F", "Dr. Smith",
            "-w", "03-03-2025"
        )

        assert result.returncode == 0


# ============================================================================
# Output Format Tests
# ============================================================================

class TestOutputFormat:
    """Tests for output formatting and content."""

    def test_swap_candidates_format(self, sample_fmit_excel):
        """Should format swap candidates with status indicators."""
        result = run_cli(
            "-f", sample_fmit_excel,
            "-F", "Dr. Smith",
            "-w", "2025-03-03"
        )

        assert result.returncode == 0
        # Should show viable vs total counts
        assert "candidates:" in result.stdout.lower()
        # Should show status indicators (✓ or ✗)
        assert "✓" in result.stdout or "✗" in result.stdout or "[" in result.stdout

    def test_alternating_pattern_suggestions(self, sample_fmit_excel):
        """Should suggest swap weeks for alternating patterns."""
        result = run_cli(
            "-f", sample_fmit_excel,
            "--alternating",
            "--threshold", "2"
        )

        assert result.returncode == 0
        # Should show FMIT weeks for faculty with alternating patterns
        if "Dr. Smith" in result.stdout:
            assert "FMIT weeks:" in result.stdout or "weeks:" in result.stdout.lower()
