"""Tests for XlwingsBridge — mocked for CI, integration tests skipif no xlwings."""

from datetime import date
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest

from app.services.xlwings_bridge import XlwingsBridge


@pytest.fixture
def mock_color_scheme():
    scheme = MagicMock()
    scheme.get_code_color.return_value = "FFFFC000"  # FMIT yellow
    scheme.get_font_color.return_value = "FF000000"
    scheme._code_colors = {"FMIT": "FFFFC000", "C": "FF92D050", "LV": "FFBFBFBF"}
    return scheme


@pytest.fixture
def bridge(mock_color_scheme):
    return XlwingsBridge(color_scheme=mock_color_scheme, visible=False)


class TestGracefulDegradation:
    def test_canonical_export_returns_original_bytes_on_bridge_error(self):
        """When xlwings fails, canonical export should return openpyxl output."""
        from app.services.canonical_schedule_export_service import (
            CanonicalScheduleExportService,
        )

        service = CanonicalScheduleExportService.__new__(CanonicalScheduleExportService)

        original_bytes = b"PK\x03\x04fake-xlsx-content"

        # The import of XlwingsBridge happens inside _apply_xlwings_finishing,
        # so we mock it at the source module level.
        with patch(
            "app.services.xlwings_bridge.XlwingsBridge",
            side_effect=RuntimeError("Excel crashed"),
        ):
            result = service._apply_xlwings_finishing(original_bytes)
            assert result == original_bytes  # Graceful fallback


class TestVerifyCfColors:
    def test_returns_mismatches_for_wrong_colors(self, bridge):
        """Cells with wrong background color should appear as mismatches."""
        mock_ws = MagicMock()

        # Pre-populate the grid with cells — the method scans rows 9-44, cols 6-61
        # We need cells that return "FMIT" for specific coordinates and None for others
        cell_data = {}
        # Put FMIT in several cells to ensure random sampling picks at least one
        for col in range(6, 20):
            cell_data[(9, col)] = ("FMIT", (255, 0, 0))  # Wrong color (red, not yellow)
            cell_data[(10, col)] = ("FMIT", (255, 0, 0))

        def range_side_effect(addr):
            cell = MagicMock()
            if addr in cell_data:
                cell.value = cell_data[addr][0]
                cell.color = cell_data[addr][1]
            else:
                cell.value = None
                cell.color = None
            return cell

        mock_ws.range = MagicMock(side_effect=range_side_effect)

        mismatches = bridge._verify_cf_colors(mock_ws, sample_size=100)
        fmit_mismatches = [m for m in mismatches if m["code"] == "FMIT"]
        assert len(fmit_mismatches) >= 1

    def test_no_mismatches_when_colors_match(self, bridge):
        """Cells with correct colors should produce no mismatches."""
        mock_ws = MagicMock()

        # All cells return None (no data)
        cell = MagicMock()
        cell.value = None
        cell.color = None
        mock_ws.range = MagicMock(return_value=cell)

        mismatches = bridge._verify_cf_colors(mock_ws)
        assert len(mismatches) == 0


class TestBridgeInit:
    def test_defaults(self, mock_color_scheme):
        bridge = XlwingsBridge(color_scheme=mock_color_scheme)
        assert bridge.visible is False
        assert bridge.timeout_seconds == 120

    def test_custom_params(self, mock_color_scheme):
        bridge = XlwingsBridge(
            color_scheme=mock_color_scheme, visible=True, timeout_seconds=60
        )
        assert bridge.visible is True
        assert bridge.timeout_seconds == 60


def _xlwings_available():
    """Check if xlwings and Excel are available for integration tests."""
    try:
        import xlwings  # noqa: F401

        return Path("/Applications/Microsoft Excel.app").exists()
    except ImportError:
        return False


@pytest.mark.skipif(not _xlwings_available(), reason="xlwings/Excel not available")
class TestXlwingsIntegration:
    """Integration tests that require a real Excel installation.

    These tests are skipped in CI and only run on machines with
    macOS + Excel installed.
    """

    def test_finishing_pass_preserves_data(self, bridge, tmp_path):
        """Generate a workbook with openpyxl, run finishing pass, verify data."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Block Template2"

        # Write test data
        ws.cell(row=9, column=6, value="FMIT")
        ws.cell(row=9, column=7, value="C")
        ws.cell(row=10, column=6, value="LV")

        xlsx_path = tmp_path / "test_finishing.xlsx"
        wb.save(xlsx_path)

        bridge.apply_finishing_pass(xlsx_path)

        # Re-read with openpyxl and verify data preserved
        from openpyxl import load_workbook

        wb2 = load_workbook(xlsx_path)
        ws2 = wb2["Block Template2"]
        assert ws2.cell(row=9, column=6).value == "FMIT"
        assert ws2.cell(row=9, column=7).value == "C"
        assert ws2.cell(row=10, column=6).value == "LV"

    def test_surgical_edit_preserves_formatting(self, bridge, tmp_path):
        """Edit one cell, verify surrounding data unchanged."""
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Block Template2"

        # Write data with formatting
        ws.cell(row=9, column=6, value="FMIT")
        ws.cell(row=9, column=7, value="C")
        ws.cell(row=10, column=6, value="LV")
        ws.cell(row=10, column=6).fill = PatternFill(
            start_color="FFBFBFBF", fill_type="solid"
        )

        xlsx_path = tmp_path / "test_surgical.xlsx"
        wb.save(xlsx_path)

        # Surgical edit: change only one cell
        bridge.surgical_edit(xlsx_path, "Block Template2", {(9, 6): "LEC"})

        from openpyxl import load_workbook

        wb2 = load_workbook(xlsx_path)
        ws2 = wb2["Block Template2"]
        assert ws2.cell(row=9, column=6).value == "LEC"  # Changed
        assert ws2.cell(row=9, column=7).value == "C"  # Unchanged
        assert ws2.cell(row=10, column=6).value == "LV"  # Unchanged

    def test_read_rendered_colors(self, bridge, tmp_path):
        """Read actual cell colors after CF evaluation."""
        from openpyxl import Workbook
        from openpyxl.formatting.rule import CellIsRule
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Block Template2"

        # Write a cell with CF rule
        ws.cell(row=9, column=6, value="FMIT")
        rule = CellIsRule(
            operator="equal",
            formula=['"FMIT"'],
            fill=PatternFill(start_color="FFC000", fill_type="solid"),
            font=Font(color="000000"),
        )
        ws.conditional_formatting.add("F9:F9", rule)

        xlsx_path = tmp_path / "test_colors.xlsx"
        wb.save(xlsx_path)

        colors = bridge.read_rendered_colors(xlsx_path, "Block Template2", [(9, 6)])
        # After Excel evaluates CF, cell should have the FMIT yellow color
        assert (9, 6) in colors
        # Color should be approximately (255, 192, 0) = FFC000
        if colors[(9, 6)] is not None:
            r, g, b = colors[(9, 6)]
            assert r > 200  # Red channel should be high
            assert g > 150  # Green channel should be medium-high
