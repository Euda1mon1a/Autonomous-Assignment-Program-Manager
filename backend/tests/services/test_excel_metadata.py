"""Tests for Excel metadata utilities (Phase 1: Phantom Database)."""

import json

import pytest
from openpyxl import Workbook

from app.services.excel_metadata import (
    EXPORT_VERSION,
    METADATA_SHEET_NAME,
    ExportMetadata,
    compute_row_hash,
    compute_schedule_checksum,
    normalize_for_hash,
    read_ref_codes,
    read_sys_meta,
    write_ref_sheet,
    write_sys_meta_sheet,
)


# ──────────────────────────────────────────────────────────────────────
# __SYS_META__ roundtrip
# ──────────────────────────────────────────────────────────────────────


def test_sys_meta_roundtrip():
    """Write and read back all metadata fields."""
    wb = Workbook()
    meta = ExportMetadata(
        academic_year=2026,
        block_number=10,
        export_timestamp="2026-02-24T10:30:00Z",
        exported_by="test_user",
        block_start_date="2026-03-04",
        block_end_date="2026-03-31",
        row_count=42,
        checksum="abc123",
    )

    write_sys_meta_sheet(wb, meta)

    assert METADATA_SHEET_NAME in wb.sheetnames
    assert wb[METADATA_SHEET_NAME].sheet_state == "veryHidden"

    read_meta = read_sys_meta(wb)

    assert read_meta is not None
    assert read_meta.academic_year == 2026
    assert read_meta.block_number == 10
    assert read_meta.export_timestamp == "2026-02-24T10:30:00Z"
    assert read_meta.export_version == EXPORT_VERSION
    assert read_meta.exported_by == "test_user"
    assert read_meta.block_start_date == "2026-03-04"
    assert read_meta.block_end_date == "2026-03-31"
    assert read_meta.row_count == 42
    assert read_meta.checksum == "abc123"


def test_sys_meta_legacy():
    """Workbooks without __SYS_META__ return None (backward compatible)."""
    wb = Workbook()
    read_meta = read_sys_meta(wb)
    assert read_meta is None


def test_sys_meta_empty_cell_returns_none():
    """An existing but empty __SYS_META__ sheet returns None."""
    wb = Workbook()
    ws = wb.create_sheet(METADATA_SHEET_NAME)
    ws.sheet_state = "veryHidden"
    # cell A1 is empty
    assert read_sys_meta(wb) is None


def test_sys_meta_malformed_json_returns_none():
    """Corrupted JSON in __SYS_META__ returns None instead of raising."""
    wb = Workbook()
    ws = wb.create_sheet(METADATA_SHEET_NAME)
    ws.sheet_state = "veryHidden"
    ws.cell(row=1, column=1, value="not valid json {{{")
    assert read_sys_meta(wb) is None


def test_sys_meta_overwrites_existing():
    """Writing metadata twice replaces the previous sheet."""
    wb = Workbook()
    meta1 = ExportMetadata(
        academic_year=2025,
        block_number=5,
        export_timestamp="2025-01-01T00:00:00Z",
    )
    meta2 = ExportMetadata(
        academic_year=2026,
        block_number=10,
        export_timestamp="2026-06-01T00:00:00Z",
    )
    write_sys_meta_sheet(wb, meta1)
    write_sys_meta_sheet(wb, meta2)

    # Only one __SYS_META__ sheet
    assert wb.sheetnames.count(METADATA_SHEET_NAME) == 1
    read_meta = read_sys_meta(wb)
    assert read_meta is not None
    assert read_meta.academic_year == 2026
    assert read_meta.block_number == 10


def test_sys_meta_forward_compatible():
    """from_json ignores unknown keys from a newer exporter."""
    future_json = json.dumps(
        {
            "academic_year": 2027,
            "export_timestamp": "2027-01-01T00:00:00Z",
            "block_number": 3,
            "export_version": "2.0",
            "some_future_field": True,
            "another_unknown": [1, 2, 3],
        }
    )
    meta = ExportMetadata.from_json(future_json)
    assert meta.academic_year == 2027
    assert meta.export_version == "2.0"
    # Unknown fields silently ignored
    assert not hasattr(meta, "some_future_field")


def test_sys_meta_legacy_int_version():
    """Legacy files with integer export_version=1 are converted to str."""
    legacy_json = json.dumps(
        {
            "academic_year": 2025,
            "export_timestamp": "2025-06-01T00:00:00Z",
            "export_version": 1,
        }
    )
    meta = ExportMetadata.from_json(legacy_json)
    assert meta.export_version == "1"


def test_sys_meta_defaults():
    """Fields with defaults are populated when absent from JSON."""
    minimal_json = json.dumps(
        {
            "academic_year": 2026,
            "export_timestamp": "2026-01-01T00:00:00Z",
        }
    )
    meta = ExportMetadata.from_json(minimal_json)
    assert meta.block_number is None
    assert meta.export_version == EXPORT_VERSION
    assert meta.exported_by is None
    assert meta.block_start_date is None
    assert meta.block_end_date is None
    assert meta.row_count == 0
    assert meta.checksum == ""


def test_sys_meta_year_level_no_block():
    """Year-level exports have block_number=None."""
    wb = Workbook()
    meta = ExportMetadata(
        academic_year=2026,
        export_timestamp="2026-01-01T00:00:00Z",
        block_map={"Block 0": "uuid-0", "Block 1": "uuid-1"},
    )
    write_sys_meta_sheet(wb, meta)
    read_meta = read_sys_meta(wb)
    assert read_meta is not None
    assert read_meta.block_number is None
    assert read_meta.block_map == {"Block 0": "uuid-0", "Block 1": "uuid-1"}


# ──────────────────────────────────────────────────────────────────────
# __REF__ sheet
# ──────────────────────────────────────────────────────────────────────


def test_ref_sheet_roundtrip():
    wb = Workbook()
    rotations = ["SURG", "MED", "PEDS"]
    activities = ["CLI", "OR", "LV"]

    write_ref_sheet(wb, rotations, activities)

    assert "__REF__" in wb.sheetnames
    assert wb["__REF__"].sheet_state == "veryHidden"

    # Check named ranges
    assert "ValidRotations" in wb.defined_names
    assert "ValidActivities" in wb.defined_names

    ref = read_ref_codes(wb)
    assert ref["rotations"] == sorted(rotations)
    assert ref["activities"] == sorted(activities)


def test_ref_sheet_legacy():
    """Workbooks without __REF__ return empty lists."""
    wb = Workbook()
    ref = read_ref_codes(wb)
    assert ref == {"rotations": [], "activities": []}


def test_ref_sheet_deduplicates():
    """Duplicate codes are collapsed."""
    wb = Workbook()
    write_ref_sheet(wb, ["A", "B", "A"], ["X", "X", "Y"])
    ref = read_ref_codes(wb)
    assert ref["rotations"] == ["A", "B"]
    assert ref["activities"] == ["X", "Y"]


def test_ref_sheet_filters_empty_codes():
    """Empty strings and None-ish codes are excluded."""
    wb = Workbook()
    write_ref_sheet(wb, ["", "A", ""], ["", "X"])
    ref = read_ref_codes(wb)
    assert ref["rotations"] == ["A"]
    assert ref["activities"] == ["X"]


# ──────────────────────────────────────────────────────────────────────
# compute_schedule_checksum (SHA-256)
# ──────────────────────────────────────────────────────────────────────


def test_compute_schedule_checksum_deterministic():
    """Same input always produces the same checksum."""
    rows = [["Alice", "CLI", "OR"], ["Bob", "LV", None]]
    c1 = compute_schedule_checksum(rows)
    c2 = compute_schedule_checksum(rows)
    assert c1 == c2
    assert len(c1) == 64  # SHA-256 hex digest


def test_compute_schedule_checksum_differs_on_change():
    """Changing any cell value changes the checksum."""
    rows_a = [["Alice", "CLI", "OR"]]
    rows_b = [["Alice", "CLI", "LV"]]  # OR -> LV
    assert compute_schedule_checksum(rows_a) != compute_schedule_checksum(rows_b)


def test_compute_schedule_checksum_differs_on_row_add():
    """Adding a row changes the checksum."""
    rows_a = [["Alice", "CLI"]]
    rows_b = [["Alice", "CLI"], ["Bob", "OR"]]
    assert compute_schedule_checksum(rows_a) != compute_schedule_checksum(rows_b)


def test_compute_schedule_checksum_empty():
    """Empty schedule produces a valid checksum."""
    c = compute_schedule_checksum([])
    assert isinstance(c, str)
    assert len(c) == 64


def test_compute_schedule_checksum_handles_non_string_types():
    """Checksum handles date objects, ints, Nones, etc."""
    from datetime import date

    rows = [["Alice", date(2026, 1, 1), 42, None, True]]
    c = compute_schedule_checksum(rows)
    assert isinstance(c, str)
    assert len(c) == 64


# ──────────────────────────────────────────────────────────────────────
# Block mismatch detection
# ──────────────────────────────────────────────────────────────────────


def test_block_mismatch_detected():
    """Import-side code should detect when file block != target block."""
    wb = Workbook()
    meta = ExportMetadata(
        academic_year=2026,
        block_number=8,
        export_timestamp="2026-01-01T00:00:00Z",
    )
    write_sys_meta_sheet(wb, meta)

    read_meta = read_sys_meta(wb)
    assert read_meta is not None

    target_block = 9
    assert read_meta.block_number != target_block


def test_block_match_accepted():
    """Matching blocks pass validation."""
    wb = Workbook()
    meta = ExportMetadata(
        academic_year=2026,
        block_number=9,
        export_timestamp="2026-01-01T00:00:00Z",
    )
    write_sys_meta_sheet(wb, meta)

    read_meta = read_sys_meta(wb)
    assert read_meta is not None

    target_block = 9
    assert read_meta.block_number == target_block


def test_academic_year_mismatch_detected():
    """Import-side code should detect when file AY != target AY."""
    wb = Workbook()
    meta = ExportMetadata(
        academic_year=2025,
        block_number=5,
        export_timestamp="2026-01-01T00:00:00Z",
    )
    write_sys_meta_sheet(wb, meta)

    read_meta = read_sys_meta(wb)
    assert read_meta is not None
    assert read_meta.academic_year != 2026


# ──────────────────────────────────────────────────────────────────────
# compute_row_hash (MD5 for row-level change detection)
# ──────────────────────────────────────────────────────────────────────


def test_compute_row_hash():
    hash1 = compute_row_hash("uuid-123", "SURG", None, ["CLI", "OR", None])
    hash2 = compute_row_hash("uuid-123", "SURG", None, ["CLI", "OR", None])
    hash3 = compute_row_hash("uuid-123", "MED", None, ["CLI", "OR", None])

    assert hash1 == hash2
    assert hash1 != hash3


def test_compute_row_hash_symmetric():
    """Export JSON values and Excel cell values should produce the same hash.

    Simulates: export writes 'CLI' as string, Excel reads it back as 'CLI'
    or ' CLI ' with whitespace.
    """
    export_hash = compute_row_hash("uuid-1", "SURG", None, ["CLI", "OR", None, "NF"])
    import_hash = compute_row_hash(
        "uuid-1", " surg ", None, [" cli ", "or", None, "nf"]
    )
    assert export_hash == import_hash


def test_compute_row_hash_float_coercion():
    """Excel may coerce integer activity codes to floats."""
    export_hash = compute_row_hash("uuid-1", "SURG", None, ["1", "2", None])
    import_hash = compute_row_hash("uuid-1", "SURG", None, ["1.0", "2.0", None])
    assert export_hash == import_hash


# ──────────────────────────────────────────────────────────────────────
# normalize_for_hash
# ──────────────────────────────────────────────────────────────────────


def test_normalize_for_hash_none_and_empty():
    assert normalize_for_hash(None) == ""
    assert normalize_for_hash("") == ""
    assert normalize_for_hash("  ") == ""


def test_normalize_for_hash_strips_and_uppercases():
    assert normalize_for_hash("  cli  ") == "CLI"
    assert normalize_for_hash("surg") == "SURG"
    assert normalize_for_hash("NF") == "NF"


def test_normalize_for_hash_removes_trailing_dot_zero():
    """Excel renders integer-like floats as '1.0' -- should normalize to '1'."""
    assert normalize_for_hash("1.0") == "1"
    assert normalize_for_hash(1.0) == "1"
    assert normalize_for_hash("10.0") == "10"


def test_normalize_for_hash_preserves_real_decimals():
    assert normalize_for_hash("1.5") == "1.5"
    assert normalize_for_hash("0.75") == "0.75"


def test_normalize_for_hash_idempotent():
    """Double-normalization should produce the same result."""
    values = [None, "", "  cli  ", "SURG", "1.0", 1.0, "NF"]
    for v in values:
        once = normalize_for_hash(v)
        twice = normalize_for_hash(once)
        assert once == twice, f"Not idempotent for {v!r}: {once!r} != {twice!r}"


# ──────────────────────────────────────────────────────────────────────
# to_json / from_json serialization
# ──────────────────────────────────────────────────────────────────────


def test_to_json_omits_none_values():
    """to_json strips None fields for cleaner JSON."""
    meta = ExportMetadata(
        academic_year=2026,
        export_timestamp="2026-01-01T00:00:00Z",
    )
    d = json.loads(meta.to_json())
    assert "block_number" not in d
    assert "exported_by" not in d
    assert "block_map" not in d
    assert "academic_year" in d


def test_to_json_includes_zero_and_empty_string():
    """Zero and empty-string are valid values, not stripped."""
    meta = ExportMetadata(
        academic_year=2026,
        export_timestamp="2026-01-01T00:00:00Z",
        row_count=0,
        checksum="",
    )
    d = json.loads(meta.to_json())
    # row_count=0 is falsy but not None — should be included
    assert "row_count" in d
    # checksum="" is falsy but not None — should be included
    assert "checksum" in d
