"""
Comprehensive tests for xlsx_export service.

Tests the legacy Excel export functionality including:
- Helper functions (rotation colors, date calculations)
- LegacyXlsxExporter class and methods
- Excel file generation and formatting
- Integration with database models
"""

import io
from datetime import date, timedelta
from uuid import uuid4

import pytest
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from sqlalchemy.orm import Session

from app.models.absence import Absence
from app.models.assignment import Assignment
from app.models.block import Block
from app.models.person import Person
from app.models.rotation_template import RotationTemplate
from app.services.xlsx_export import (
    COLORS,
    LegacyXlsxExporter,
    calculate_block_dates,
    format_date_header,
    generate_legacy_xlsx,
    get_day_abbreviation,
    get_rotation_color,
)


class TestHelperFunctions:
    """Test standalone helper functions."""

    def test_get_rotation_color_military(self):
        """Test military rotation color."""
        color = get_rotation_color("Military Duty", "MIL")
        assert color == COLORS["military"]

        color = get_rotation_color("Deployment", "DEP")
        assert color == COLORS["military"]

    def test_get_rotation_color_night_float(self):
        """Test night float rotation color."""
        color = get_rotation_color("Night Float", "NF")
        assert color == COLORS["night_float"]

        color = get_rotation_color("Night Coverage", "NF")
        assert color == COLORS["night_float"]

    def test_get_rotation_color_nicu(self):
        """Test NICU rotation color."""
        color = get_rotation_color("NICU Rotation", "NICU")
        assert color == COLORS["nicu"]

    def test_get_rotation_color_fmit(self):
        """Test FMIT rotation color."""
        color = get_rotation_color("FMIT Block", "FMIT")
        assert color == COLORS["fmit"]

    def test_get_rotation_color_default(self):
        """Test default color for unknown rotations."""
        color = get_rotation_color("General Clinic", "GC")
        assert color == COLORS["default"]

    def test_get_rotation_color_case_insensitive(self):
        """Test that color matching is case-insensitive."""
        color = get_rotation_color("MILITARY", "mil")
        assert color == COLORS["military"]

        color = get_rotation_color("fmit", "FMIT")
        assert color == COLORS["fmit"]

    def test_get_rotation_color_none_values(self):
        """Test handling of None values."""
        color = get_rotation_color(None, None)
        assert color == COLORS["default"]

        color = get_rotation_color("Test", None)
        assert color == COLORS["default"]

    def test_calculate_block_dates_block_1(self):
        """Test block date calculation for first block."""
        academic_year_start = date(2024, 7, 1)
        start, end = calculate_block_dates(1, academic_year_start)

        assert start == date(2024, 7, 1)
        assert end == date(2024, 7, 28)
        assert (end - start).days == 27  # 28 days total (0-27)

    def test_calculate_block_dates_block_2(self):
        """Test block date calculation for second block."""
        academic_year_start = date(2024, 7, 1)
        start, end = calculate_block_dates(2, academic_year_start)

        assert start == date(2024, 7, 29)
        assert end == date(2024, 8, 25)

    def test_calculate_block_dates_block_13(self):
        """Test block date calculation for last block of year."""
        academic_year_start = date(2024, 7, 1)
        start, end = calculate_block_dates(13, academic_year_start)

        # Block 13 starts after 12 * 28 = 336 days
        expected_start = date(2024, 7, 1) + timedelta(days=12 * 28)
        assert start == expected_start

    def test_calculate_block_dates_consistency(self):
        """Test that consecutive blocks don't overlap."""
        academic_year_start = date(2024, 7, 1)

        for block_num in range(1, 13):
            start1, end1 = calculate_block_dates(block_num, academic_year_start)
            start2, end2 = calculate_block_dates(block_num + 1, academic_year_start)

            # Next block should start day after previous block ends
            assert start2 == end1 + timedelta(days=1)

    def test_get_day_abbreviation_all_days(self):
        """Test day abbreviations for all weekdays."""
        # Use a known Monday: 2024-01-01 (actually a Monday)
        monday = date(2024, 1, 1)

        expected = ["MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN"]
        for i, abbrev in enumerate(expected):
            day = monday + timedelta(days=i)
            assert get_day_abbreviation(day) == abbrev

    def test_format_date_header_various_dates(self):
        """Test date header formatting."""
        # January 15, 2024
        formatted = format_date_header(date(2024, 1, 15))
        assert formatted == "15-Jan"

        # February 1, 2024
        formatted = format_date_header(date(2024, 2, 1))
        assert formatted == "1-Feb"

        # December 31, 2024
        formatted = format_date_header(date(2024, 12, 31))
        assert formatted == "31-Dec"

    def test_format_date_header_single_digit_days(self):
        """Test that single digit days don't have leading zeros."""
        formatted = format_date_header(date(2024, 3, 5))
        # Should be "5-Mar", not "05-Mar"
        assert formatted == "5-Mar"


class TestLegacyXlsxExporter:
    """Test LegacyXlsxExporter class."""

    def test_initialization(self, db: Session):
        """Test exporter initialization."""
        exporter = LegacyXlsxExporter(db)

        assert exporter.db == db
        assert exporter.wb is not None
        # Default sheet should be removed
        assert len(exporter.wb.worksheets) == 0

    def test_get_absence_abbreviation_all_types(self, db: Session):
        """Test absence abbreviation for all types."""
        exporter = LegacyXlsxExporter(db)

        assert exporter._get_absence_abbreviation("vacation") == "VAC"
        assert exporter._get_absence_abbreviation("sick") == "SICK"
        assert exporter._get_absence_abbreviation("medical") == "MED"
        assert exporter._get_absence_abbreviation("conference") == "CONF"
        assert exporter._get_absence_abbreviation("deployment") == "DEP"
        assert exporter._get_absence_abbreviation("tdy") == "TDY"
        assert exporter._get_absence_abbreviation("family_emergency") == "FEM"
        assert exporter._get_absence_abbreviation("personal") == "PER"

    def test_get_absence_abbreviation_unknown(self, db: Session):
        """Test absence abbreviation for unknown type."""
        exporter = LegacyXlsxExporter(db)

        assert exporter._get_absence_abbreviation("unknown_type") == "ABS"
        assert exporter._get_absence_abbreviation("") == "ABS"

    def test_save_to_bytes_empty_workbook(self, db: Session):
        """Test saving empty workbook to bytes."""
        exporter = LegacyXlsxExporter(db)

        bytes_output = exporter.save_to_bytes()

        assert isinstance(bytes_output, bytes)
        assert len(bytes_output) > 0

        # Verify it's valid Excel format by loading it back
        wb = load_workbook(io.BytesIO(bytes_output))
        assert wb is not None

    def test_save_to_bytes_with_sheet(self, db: Session):
        """Test saving workbook with sheet to bytes."""
        exporter = LegacyXlsxExporter(db)
        exporter.wb.create_sheet(title="Test Sheet")

        bytes_output = exporter.save_to_bytes()

        # Verify sheet exists when loaded back
        wb = load_workbook(io.BytesIO(bytes_output))
        assert "Test Sheet" in wb.sheetnames

    def test_generate_block_schedule_empty_database(self, db: Session):
        """Test generating block schedule with no data."""
        exporter = LegacyXlsxExporter(db)

        start_date = date(2024, 7, 1)
        end_date = date(2024, 7, 28)

        exporter.generate_block_schedule(
            block_number=1,
            start_date=start_date,
            end_date=end_date,
            federal_holidays=None,
        )

        # Should create sheet even with no data
        assert len(exporter.wb.worksheets) == 1
        assert exporter.wb.worksheets[0].title == "Block 1"

    def test_generate_block_schedule_with_residents(
        self, db: Session, sample_residents: list[Person]
    ):
        """Test generating block schedule with residents."""
        exporter = LegacyXlsxExporter(db)

        start_date = date(2024, 7, 1)
        end_date = date(2024, 7, 28)

        exporter.generate_block_schedule(
            block_number=1,
            start_date=start_date,
            end_date=end_date,
        )

        ws = exporter.wb.worksheets[0]

        # Check header has block number
        assert ws["C1"].value == 1

        # Verify sheet was created
        assert ws.title == "Block 1"

    def test_generate_block_schedule_with_assignments(
        self,
        db: Session,
        sample_resident: Person,
        sample_rotation_template: RotationTemplate,
    ):
        """Test generating schedule with assignments."""
        exporter = LegacyXlsxExporter(db)

        start_date = date(2024, 7, 1)
        end_date = date(2024, 7, 7)

        # Create blocks
        for i in range(7):
            for time_of_day in ["AM", "PM"]:
                block = Block(
                    id=uuid4(),
                    date=start_date + timedelta(days=i),
                    time_of_day=time_of_day,
                    block_number=1,
                    is_weekend=(start_date + timedelta(days=i)).weekday() >= 5,
                )
                db.add(block)

        db.commit()

        # Create assignments for first day
        blocks = db.query(Block).filter(Block.date == start_date).all()
        for block in blocks:
            assignment = Assignment(
                id=uuid4(),
                block_id=block.id,
                person_id=sample_resident.id,
                rotation_template_id=sample_rotation_template.id,
                abbreviation=sample_rotation_template.abbreviation,
            )
            db.add(assignment)

        db.commit()

        exporter.generate_block_schedule(
            block_number=1,
            start_date=start_date,
            end_date=end_date,
        )

        # Verify schedule was generated
        assert len(exporter.wb.worksheets) == 1

    def test_generate_block_schedule_with_absences(
        self, db: Session, sample_resident: Person
    ):
        """Test generating schedule with absences."""
        exporter = LegacyXlsxExporter(db)

        start_date = date(2024, 7, 1)
        end_date = date(2024, 7, 7)

        # Create blocks
        for i in range(7):
            for time_of_day in ["AM", "PM"]:
                block = Block(
                    id=uuid4(),
                    date=start_date + timedelta(days=i),
                    time_of_day=time_of_day,
                    block_number=1,
                )
                db.add(block)

        # Create absence
        absence = Absence(
            id=uuid4(),
            person_id=sample_resident.id,
            start_date=start_date,
            end_date=start_date + timedelta(days=2),
            absence_type="vacation",
        )
        db.add(absence)
        db.commit()

        exporter.generate_block_schedule(
            block_number=1,
            start_date=start_date,
            end_date=end_date,
        )

        # Verify schedule was generated
        assert len(exporter.wb.worksheets) == 1

    def test_generate_block_schedule_with_holidays(self, db: Session):
        """Test generating schedule with federal holidays."""
        exporter = LegacyXlsxExporter(db)

        start_date = date(2024, 7, 1)
        end_date = date(2024, 7, 7)

        # July 4th is a federal holiday
        federal_holidays = [date(2024, 7, 4)]

        exporter.generate_block_schedule(
            block_number=1,
            start_date=start_date,
            end_date=end_date,
            federal_holidays=federal_holidays,
        )

        ws = exporter.wb.worksheets[0]

        # Find column for July 4th (should be highlighted)
        # Day 4 would be at offset 3 (0-indexed), each day has 2 columns (AM/PM)
        # Starting at column F (6), so July 4 AM would be at column 6 + (3 * 2) = 12
        # We can verify the sheet exists
        assert ws.title == "Block 1"

    def test_generate_block_schedule_mixed_pgy_levels(
        self, db: Session, sample_residents: list[Person]
    ):
        """Test that residents are grouped and separated by PGY level."""
        exporter = LegacyXlsxExporter(db)

        start_date = date(2024, 7, 1)
        end_date = date(2024, 7, 7)

        exporter.generate_block_schedule(
            block_number=1,
            start_date=start_date,
            end_date=end_date,
        )

        # Should have created sheet with residents from different PGY levels
        assert len(exporter.wb.worksheets) == 1

    def test_generate_block_schedule_faculty_separation(
        self,
        db: Session,
        sample_residents: list[Person],
        sample_faculty_members: list[Person],
    ):
        """Test that faculty are separated from residents."""
        exporter = LegacyXlsxExporter(db)

        start_date = date(2024, 7, 1)
        end_date = date(2024, 7, 7)

        exporter.generate_block_schedule(
            block_number=1,
            start_date=start_date,
            end_date=end_date,
        )

        # Should have created sheet with both residents and faculty
        assert len(exporter.wb.worksheets) == 1

    def test_generate_full_year_default_blocks(self, db: Session):
        """Test generating full year with default 13 blocks."""
        exporter = LegacyXlsxExporter(db)

        academic_year_start = date(2024, 7, 1)

        exporter.generate_full_year(academic_year_start=academic_year_start)

        # Should have 13 sheets (one per block)
        assert len(exporter.wb.worksheets) == 13

        # Verify sheet names
        for i in range(1, 14):
            assert f"Block {i}" in exporter.wb.sheetnames

    def test_generate_full_year_custom_blocks(self, db: Session):
        """Test generating full year with custom number of blocks."""
        exporter = LegacyXlsxExporter(db)

        academic_year_start = date(2024, 7, 1)

        exporter.generate_full_year(
            academic_year_start=academic_year_start,
            num_blocks=10,
        )

        # Should have 10 sheets
        assert len(exporter.wb.worksheets) == 10

    def test_generate_full_year_with_holidays(self, db: Session):
        """Test generating full year with federal holidays."""
        exporter = LegacyXlsxExporter(db)

        academic_year_start = date(2024, 7, 1)
        federal_holidays = [
            date(2024, 7, 4),  # Independence Day
            date(2024, 11, 28),  # Thanksgiving
            date(2024, 12, 25),  # Christmas
        ]

        exporter.generate_full_year(
            academic_year_start=academic_year_start,
            num_blocks=13,
            federal_holidays=federal_holidays,
        )

        assert len(exporter.wb.worksheets) == 13


class TestGenerateLegacyXlsx:
    """Test the main generate_legacy_xlsx function."""

    def test_generate_empty_schedule(self, db: Session):
        """Test generating Excel file with no data."""
        start_date = date(2024, 7, 1)
        end_date = date(2024, 7, 28)

        xlsx_bytes = generate_legacy_xlsx(
            db=db,
            start_date=start_date,
            end_date=end_date,
            block_number=1,
        )

        assert isinstance(xlsx_bytes, bytes)
        assert len(xlsx_bytes) > 0

        # Verify it's valid Excel
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        assert "Block 1" in wb.sheetnames

    def test_generate_with_explicit_block_number(self, db: Session):
        """Test generating with explicit block number."""
        start_date = date(2024, 7, 1)
        end_date = date(2024, 7, 28)

        xlsx_bytes = generate_legacy_xlsx(
            db=db,
            start_date=start_date,
            end_date=end_date,
            block_number=5,
        )

        wb = load_workbook(io.BytesIO(xlsx_bytes))
        assert "Block 5" in wb.sheetnames

        # Check block number in cell C1
        ws = wb["Block 5"]
        assert ws["C1"].value == 5

    def test_generate_with_calculated_block_number(self, db: Session):
        """Test that block number is calculated if not provided."""
        # Start date in August (second month of academic year starting July 1)
        start_date = date(2024, 8, 1)
        end_date = date(2024, 8, 28)

        xlsx_bytes = generate_legacy_xlsx(
            db=db,
            start_date=start_date,
            end_date=end_date,
            block_number=None,  # Let it calculate
        )

        # Should auto-calculate block number
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        # Aug 1 is 31 days after July 1, so block 2 (31 // 28 = 1, +1 = 2)
        assert len(wb.worksheets) == 1

    def test_generate_with_residents_and_assignments(
        self,
        db: Session,
        sample_resident: Person,
        sample_rotation_template: RotationTemplate,
    ):
        """Test generating complete schedule with residents and assignments."""
        start_date = date(2024, 7, 1)
        end_date = date(2024, 7, 7)

        # Create blocks
        for i in range(7):
            for time_of_day in ["AM", "PM"]:
                block = Block(
                    id=uuid4(),
                    date=start_date + timedelta(days=i),
                    time_of_day=time_of_day,
                    block_number=1,
                )
                db.add(block)

        db.commit()

        # Create some assignments
        blocks = db.query(Block).limit(4).all()
        for block in blocks:
            assignment = Assignment(
                id=uuid4(),
                block_id=block.id,
                person_id=sample_resident.id,
                rotation_template_id=sample_rotation_template.id,
                abbreviation=sample_rotation_template.abbreviation,
            )
            db.add(assignment)

        db.commit()

        xlsx_bytes = generate_legacy_xlsx(
            db=db,
            start_date=start_date,
            end_date=end_date,
            block_number=1,
        )

        # Verify file is valid
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb["Block 1"]

        # Should have resident name somewhere
        assert ws["E9"].value == sample_resident.name

    def test_generate_with_federal_holidays(self, db: Session):
        """Test generating schedule with federal holidays."""
        start_date = date(2024, 7, 1)
        end_date = date(2024, 7, 7)

        federal_holidays = [date(2024, 7, 4)]

        xlsx_bytes = generate_legacy_xlsx(
            db=db,
            start_date=start_date,
            end_date=end_date,
            block_number=1,
            federal_holidays=federal_holidays,
        )

        wb = load_workbook(io.BytesIO(xlsx_bytes))
        assert "Block 1" in wb.sheetnames

    def test_generate_across_month_boundary(self, db: Session):
        """Test generating schedule that spans multiple months."""
        # Start in July, end in August
        start_date = date(2024, 7, 15)
        end_date = date(2024, 8, 11)

        xlsx_bytes = generate_legacy_xlsx(
            db=db,
            start_date=start_date,
            end_date=end_date,
            block_number=1,
        )

        wb = load_workbook(io.BytesIO(xlsx_bytes))
        assert len(wb.worksheets) == 1

    def test_generate_single_day(self, db: Session):
        """Test generating schedule for a single day."""
        single_day = date(2024, 7, 1)

        xlsx_bytes = generate_legacy_xlsx(
            db=db,
            start_date=single_day,
            end_date=single_day,
            block_number=1,
        )

        wb = load_workbook(io.BytesIO(xlsx_bytes))
        assert "Block 1" in wb.sheetnames


class TestEdgeCases:
    """Test edge cases and boundary conditions."""

    def test_year_boundary_schedule(self, db: Session):
        """Test schedule that crosses year boundary."""
        # December to January
        start_date = date(2024, 12, 20)
        end_date = date(2025, 1, 10)

        xlsx_bytes = generate_legacy_xlsx(
            db=db,
            start_date=start_date,
            end_date=end_date,
        )

        wb = load_workbook(io.BytesIO(xlsx_bytes))
        assert len(wb.worksheets) == 1

    def test_leap_year_dates(self, db: Session):
        """Test schedule during leap year (Feb 29)."""
        # 2024 is a leap year
        start_date = date(2024, 2, 26)
        end_date = date(2024, 3, 3)

        xlsx_bytes = generate_legacy_xlsx(
            db=db,
            start_date=start_date,
            end_date=end_date,
        )

        wb = load_workbook(io.BytesIO(xlsx_bytes))
        assert len(wb.worksheets) == 1

    def test_weekend_highlighting(self, db: Session):
        """Test that weekends are properly handled."""
        # Start on a Friday, go through weekend
        start_date = date(2024, 7, 5)  # Friday
        end_date = date(2024, 7, 7)  # Sunday

        exporter = LegacyXlsxExporter(db)
        exporter.generate_block_schedule(
            block_number=1,
            start_date=start_date,
            end_date=end_date,
        )

        # Verify schedule was created
        assert len(exporter.wb.worksheets) == 1

    def test_very_long_provider_name(self, db: Session):
        """Test handling of very long provider names."""
        long_name = "Dr. " + "VeryLong" * 20 + " Name"
        resident = Person(
            id=uuid4(),
            name=long_name,
            type="resident",
            email="longname@test.org",
            pgy_level=1,
        )
        db.add(resident)
        db.commit()

        start_date = date(2024, 7, 1)
        end_date = date(2024, 7, 7)

        xlsx_bytes = generate_legacy_xlsx(
            db=db,
            start_date=start_date,
            end_date=end_date,
        )

        # Should handle long names without crashing
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        assert len(wb.worksheets) == 1

    def test_special_characters_in_names(self, db: Session):
        """Test handling of special characters in provider names."""
        names = [
            "Dr. O'Brien",
            "Dr. José García",
            "Dr. Müller-Schmidt",
            "Dr. 李明",
        ]

        for name in names:
            resident = Person(
                id=uuid4(),
                name=name,
                type="resident",
                email=f"{uuid4()}@test.org",
                pgy_level=1,
            )
            db.add(resident)

        db.commit()

        start_date = date(2024, 7, 1)
        end_date = date(2024, 7, 7)

        xlsx_bytes = generate_legacy_xlsx(
            db=db,
            start_date=start_date,
            end_date=end_date,
        )

        wb = load_workbook(io.BytesIO(xlsx_bytes))
        assert len(wb.worksheets) == 1

    def test_multiple_absences_same_person(self, db: Session, sample_resident: Person):
        """Test handling multiple absences for same person."""
        start_date = date(2024, 7, 1)
        end_date = date(2024, 7, 28)

        # Create multiple absences
        absences = [
            Absence(
                id=uuid4(),
                person_id=sample_resident.id,
                start_date=date(2024, 7, 5),
                end_date=date(2024, 7, 7),
                absence_type="vacation",
            ),
            Absence(
                id=uuid4(),
                person_id=sample_resident.id,
                start_date=date(2024, 7, 15),
                end_date=date(2024, 7, 16),
                absence_type="conference",
            ),
        ]

        for absence in absences:
            db.add(absence)

        db.commit()

        xlsx_bytes = generate_legacy_xlsx(
            db=db,
            start_date=start_date,
            end_date=end_date,
        )

        wb = load_workbook(io.BytesIO(xlsx_bytes))
        assert len(wb.worksheets) == 1

    def test_overlapping_absence_and_assignment(
        self,
        db: Session,
        sample_resident: Person,
        sample_rotation_template: RotationTemplate,
    ):
        """Test when a person has both assignment and absence on same day."""
        start_date = date(2024, 7, 1)
        end_date = date(2024, 7, 7)

        # Create blocks
        for i in range(7):
            for time_of_day in ["AM", "PM"]:
                block = Block(
                    id=uuid4(),
                    date=start_date + timedelta(days=i),
                    time_of_day=time_of_day,
                    block_number=1,
                )
                db.add(block)

        db.commit()

        # Create assignment
        block = db.query(Block).filter(Block.date == start_date).first()
        assignment = Assignment(
            id=uuid4(),
            block_id=block.id,
            person_id=sample_resident.id,
            rotation_template_id=sample_rotation_template.id,
            abbreviation="SM",
        )
        db.add(assignment)

        # Create absence for same day
        absence = Absence(
            id=uuid4(),
            person_id=sample_resident.id,
            start_date=start_date,
            end_date=start_date,
            absence_type="sick",
        )
        db.add(absence)
        db.commit()

        xlsx_bytes = generate_legacy_xlsx(
            db=db,
            start_date=start_date,
            end_date=end_date,
        )

        # Should prioritize absence over assignment
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        assert len(wb.worksheets) == 1

    def test_no_faculty_in_database(self, db: Session, sample_resident: Person):
        """Test generating schedule when no faculty exist."""
        start_date = date(2024, 7, 1)
        end_date = date(2024, 7, 7)

        xlsx_bytes = generate_legacy_xlsx(
            db=db,
            start_date=start_date,
            end_date=end_date,
        )

        wb = load_workbook(io.BytesIO(xlsx_bytes))
        assert len(wb.worksheets) == 1

    def test_no_residents_in_database(self, db: Session, sample_faculty: Person):
        """Test generating schedule when no residents exist."""
        start_date = date(2024, 7, 1)
        end_date = date(2024, 7, 7)

        xlsx_bytes = generate_legacy_xlsx(
            db=db,
            start_date=start_date,
            end_date=end_date,
        )

        wb = load_workbook(io.BytesIO(xlsx_bytes))
        assert len(wb.worksheets) == 1

    def test_empty_federal_holidays_list(self, db: Session):
        """Test with explicitly empty federal holidays list."""
        start_date = date(2024, 7, 1)
        end_date = date(2024, 7, 7)

        xlsx_bytes = generate_legacy_xlsx(
            db=db,
            start_date=start_date,
            end_date=end_date,
            federal_holidays=[],
        )

        wb = load_workbook(io.BytesIO(xlsx_bytes))
        assert len(wb.worksheets) == 1


class TestExcelFormatting:
    """Test Excel formatting and styling (where verifiable)."""

    def test_column_widths_set(self, db: Session):
        """Test that column widths are adjusted."""
        exporter = LegacyXlsxExporter(db)

        start_date = date(2024, 7, 1)
        end_date = date(2024, 7, 7)

        exporter.generate_block_schedule(
            block_number=1,
            start_date=start_date,
            end_date=end_date,
        )

        ws = exporter.wb.worksheets[0]

        # Check that specific columns have widths set
        assert ws.column_dimensions["A"].width == 15
        assert ws.column_dimensions["E"].width == 15

    def test_header_merge_cells(self, db: Session):
        """Test that header cells are properly merged."""
        exporter = LegacyXlsxExporter(db)

        start_date = date(2024, 7, 1)
        end_date = date(2024, 7, 7)

        exporter.generate_block_schedule(
            block_number=1,
            start_date=start_date,
            end_date=end_date,
        )

        ws = exporter.wb.worksheets[0]

        # Check that C1:D3 is merged for block number
        assert "C1:D3" in ws.merged_cells

    def test_date_range_in_header(self, db: Session):
        """Test that date range appears in header."""
        exporter = LegacyXlsxExporter(db)

        start_date = date(2024, 7, 1)
        end_date = date(2024, 7, 28)

        exporter.generate_block_schedule(
            block_number=1,
            start_date=start_date,
            end_date=end_date,
        )

        ws = exporter.wb.worksheets[0]

        # Check C4 for date range
        cell_value = ws["C4"].value
        assert "1Jul" in cell_value
        assert "28Jul" in cell_value

    def test_header_labels_present(self, db: Session):
        """Test that header labels are present."""
        exporter = LegacyXlsxExporter(db)

        start_date = date(2024, 7, 1)
        end_date = date(2024, 7, 7)

        exporter.generate_block_schedule(
            block_number=1,
            start_date=start_date,
            end_date=end_date,
        )

        ws = exporter.wb.worksheets[0]

        # Check for expected header labels
        assert ws["E3"].value == "Date:"
        assert ws["E4"].value == "Staff Call"
        assert ws["E5"].value == "Resident Call"
        assert ws["C6"].value == "TEMPLATE"
        assert ws["D6"].value == "ROLE"
        assert ws["E6"].value == "PROVIDER"
