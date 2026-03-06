"""Tests for canonical_schedule_export_service.py - Export orchestration.

These tests use mocking to avoid JSONB/SQLite compatibility issues.
"""

import io
from datetime import date
from pathlib import Path
from unittest.mock import MagicMock, patch
from uuid import uuid4

import pytest
from openpyxl import Workbook

from app.services.canonical_schedule_export_service import (
    CanonicalScheduleExportService,
)
from app.services.excel_metadata import read_ref_codes, read_sys_meta


def _mock_db_with_codes():
    """Create a mock DB session that returns rotation/activity codes."""
    mock_db = MagicMock()
    rot_query = MagicMock()
    rot_query.distinct.return_value.all.return_value = [("SURG",), ("MED",)]
    act_query = MagicMock()
    act_query.distinct.return_value.all.return_value = [("CLI",), ("LV",)]
    person_query = MagicMock()
    person_query.filter.return_value.all.return_value = []

    # Needs to handle export_block_xlsx: Rotation, Activity, Person
    mock_db.query.side_effect = [
        rot_query,
        act_query,
        person_query,
        rot_query,
        act_query,
        person_query,
    ]
    return mock_db


class TestCanonicalScheduleExportService:
    """Tests for CanonicalScheduleExportService."""

    def _make_template_bytes(self) -> bytes:
        """Create a minimal Block Template2 workbook for export_year_xlsx."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Block Template2"
        ws.cell(row=1, column=1, value="ok")
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    @patch("app.services.canonical_schedule_export_service.get_block_dates")
    @patch("app.services.canonical_schedule_export_service.HalfDayJSONExporter")
    @patch("app.services.canonical_schedule_export_service.TAMCBlockExporter")
    def test_export_block_xlsx_returns_bytes(
        self, mock_exporter_class, mock_json_exporter_class, mock_get_block_dates
    ):
        """export_block_xlsx() should return bytes."""
        mock_db = _mock_db_with_codes()

        # Mock block dates
        mock_block_dates = MagicMock()
        mock_block_dates.start_date = date(2026, 3, 12)
        mock_block_dates.end_date = date(2026, 4, 8)
        mock_get_block_dates.return_value = mock_block_dates

        # Mock JSON exporter
        mock_json_exporter = MagicMock()
        mock_json_exporter.export.return_value = {"residents": [], "faculty": []}
        mock_json_exporter_class.return_value = mock_json_exporter

        # Mock TAMCBlockExporter
        mock_block_exporter = MagicMock()
        mock_block_exporter.export.return_value = b"xlsx_content"
        mock_exporter_class.return_value = mock_block_exporter

        service = CanonicalScheduleExportService(mock_db)
        result = service.export_block_xlsx(block_number=10, academic_year=2025)

        assert isinstance(result, bytes)
        mock_exporter_class.assert_called_once()

    @patch("app.services.canonical_schedule_export_service.get_block_dates")
    @patch("app.services.canonical_schedule_export_service.TAMCBlockExporter")
    def test_export_uses_block_dates(self, mock_exporter_class, mock_get_block_dates):
        """export_block_xlsx() should use correct block dates."""
        mock_db = _mock_db_with_codes()

        mock_block_dates = MagicMock()
        mock_block_dates.start_date = date(2026, 3, 12)
        mock_block_dates.end_date = date(2026, 4, 8)
        mock_get_block_dates.return_value = mock_block_dates

        mock_block_exporter = MagicMock()
        mock_block_exporter.export.return_value = b"xlsx"
        mock_exporter_class.return_value = mock_block_exporter

        service = CanonicalScheduleExportService(mock_db)

        with patch.object(service, "_export_json_data", return_value={}):
            service.export_block_xlsx(block_number=10, academic_year=2025)

        mock_get_block_dates.assert_called_once_with(10, 2025)

    @patch("app.services.canonical_schedule_export_service.get_block_dates")
    @patch("app.services.canonical_schedule_export_service.HalfDayJSONExporter")
    @patch("app.services.canonical_schedule_export_service.TAMCBlockExporter")
    def test_export_can_disable_qa_sheet(
        self, mock_exporter_class, mock_json_exporter_class, mock_get_block_dates
    ):
        """export_block_xlsx() should call TAMCBlockExporter.export()."""
        mock_db = _mock_db_with_codes()

        mock_block_dates = MagicMock()
        mock_block_dates.start_date = date(2026, 3, 12)
        mock_block_dates.end_date = date(2026, 4, 8)
        mock_get_block_dates.return_value = mock_block_dates

        mock_json_exporter = MagicMock()
        mock_json_exporter.export.return_value = {"residents": [], "faculty": []}
        mock_json_exporter_class.return_value = mock_json_exporter

        mock_block_exporter = MagicMock()
        mock_block_exporter.export.return_value = b"xlsx_content"
        mock_exporter_class.return_value = mock_block_exporter

        service = CanonicalScheduleExportService(mock_db)
        service.export_block_xlsx(
            block_number=10,
            academic_year=2025,
            include_qa_sheet=False,
        )

        mock_block_exporter.export.assert_called_once()

    @patch("app.services.canonical_schedule_export_service.get_block_dates")
    @patch("app.services.canonical_schedule_export_service.HalfDayJSONExporter")
    @patch("app.services.canonical_schedule_export_service.TAMCBlockExporter")
    def test_export_can_override_identity_and_profile(
        self, mock_exporter_class, mock_json_exporter_class, mock_get_block_dates
    ):
        """export_block_xlsx() should call TAMCBlockExporter with block_config."""
        mock_db = _mock_db_with_codes()

        mock_block_dates = MagicMock()
        mock_block_dates.start_date = date(2026, 3, 12)
        mock_block_dates.end_date = date(2026, 4, 8)
        mock_get_block_dates.return_value = mock_block_dates

        mock_json_exporter = MagicMock()
        mock_json_exporter.export.return_value = {"residents": [], "faculty": []}
        mock_json_exporter_class.return_value = mock_json_exporter

        mock_block_exporter = MagicMock()
        mock_block_exporter.export.return_value = b"xlsx_content"
        mock_exporter_class.return_value = mock_block_exporter

        service = CanonicalScheduleExportService(mock_db)
        service.export_block_xlsx(
            block_number=10,
            academic_year=2025,
            preserve_template_identity_fields=False,
            presentation_profile="none",
        )

        # TAMCBlockExporter was created with block_config kwarg
        mock_exporter_class.assert_called_once()
        assert "block_config" in mock_exporter_class.call_args.kwargs

    @patch("app.services.canonical_schedule_export_service.get_block_dates")
    @patch("app.services.canonical_schedule_export_service.HalfDayJSONExporter")
    @patch("app.services.canonical_schedule_export_service.TAMCBlockExporter")
    def test_export_passes_metadata_to_exporter(
        self, mock_exporter_class, mock_json_exporter_class, mock_get_block_dates
    ):
        """export_block_xlsx() should pass ExportMetadata to TAMCBlockExporter."""
        mock_db = _mock_db_with_codes()

        mock_block_dates = MagicMock()
        mock_block_dates.start_date = date(2026, 3, 12)
        mock_block_dates.end_date = date(2026, 4, 8)
        mock_get_block_dates.return_value = mock_block_dates

        mock_json_exporter = MagicMock()
        mock_json_exporter.export.return_value = {"residents": []}
        mock_json_exporter_class.return_value = mock_json_exporter

        mock_block_exporter = MagicMock()
        mock_block_exporter.export.return_value = b"raw_xlsx"
        mock_exporter_class.return_value = mock_block_exporter

        service = CanonicalScheduleExportService(mock_db)
        result = service.export_block_xlsx(block_number=10, academic_year=2025)

        # TAMCBlockExporter.export() should receive metadata kwargs
        call_kwargs = mock_block_exporter.export.call_args.kwargs
        assert call_kwargs["export_metadata"] is not None
        assert call_kwargs["export_metadata"].academic_year == 2025
        assert call_kwargs["export_metadata"].block_number == 10
        assert call_kwargs["rotation_codes"] is not None
        assert call_kwargs["activity_codes"] is not None
        assert result == b"raw_xlsx"

    def test_export_year_xlsx_raises_when_no_blocks(self):
        """export_year_xlsx() should raise if no academic blocks found."""
        mock_db = MagicMock()
        mock_db.execute.return_value.scalars.return_value.all.return_value = []

        service = CanonicalScheduleExportService(mock_db)

        with pytest.raises(ValueError, match="No academic blocks found"):
            service.export_year_xlsx(academic_year=2026)

    @patch("app.services.canonical_schedule_export_service.TAMCBlockExporter")
    def test_export_year_xlsx_builds_block_map_and_metadata(self, mock_exporter_class):
        """export_year_xlsx() should build block_map and embed inline metadata."""
        mock_db = MagicMock()
        block_one = MagicMock()
        block_one.block_number = 1
        block_one.start_date = date(2026, 7, 1)
        block_one.end_date = date(2026, 7, 28)
        block_one.id = uuid4()

        block_two = MagicMock()
        block_two.block_number = 13
        block_two.start_date = date(2027, 6, 3)
        block_two.end_date = date(2027, 6, 30)
        block_two.id = uuid4()

        mock_db.execute.return_value.scalars.return_value.all.return_value = [
            block_one,
            block_two,
        ]
        # Mock the .query().distinct().all() calls for rotation/activity codes
        mock_db.query.return_value.distinct.return_value.all.return_value = []

        mock_exporter = MagicMock()
        mock_exporter.export.return_value = self._make_template_bytes()
        mock_exporter.summary_col_start = 62
        mock_exporter_class.return_value = mock_exporter

        service = CanonicalScheduleExportService(mock_db)
        with (
            patch.object(
                service,
                "_export_json_data",
                return_value={
                    "faculty": [],
                    "block_start": "2026-07-01",
                    "block_end": "2026-07-28",
                },
            ),
            patch.object(service, "_copy_worksheet"),
            patch.object(service, "_apply_phantom_columns"),
            patch.object(service, "_build_ytd_summary_sheet"),
        ):
            result = service.export_year_xlsx(academic_year=2026)

        # Result should be valid xlsx bytes (starts with PK zip header)
        assert result[:2] == b"PK"

        # Verify metadata was written inline by loading the workbook
        from openpyxl import load_workbook
        from app.services.excel_metadata import read_sys_meta

        wb = load_workbook(io.BytesIO(result))
        meta = read_sys_meta(wb)
        assert meta is not None
        assert meta.academic_year == 2026
        expected_block_map = {
            "Block 1": str(block_one.id),
            "Block 13": str(block_two.id),
        }
        assert meta.block_map == expected_block_map
        wb.close()


class TestTemplatePaths:
    """Tests for template path resolution."""

    def test_data_dir_returns_path(self):
        """_data_dir() should return a Path."""
        mock_db = MagicMock()
        service = CanonicalScheduleExportService(mock_db)

        result = service._data_dir()

        assert isinstance(result, Path)

    def test_template_path_returns_none_if_missing(self):
        """_template_path() should return None if template missing (graceful)."""
        mock_db = MagicMock()
        service = CanonicalScheduleExportService(mock_db)

        with patch.object(service, "_data_dir", return_value=Path("/nonexistent")):
            result = service._template_path()
            assert result is None

    def test_structure_path_returns_none_if_missing(self):
        """_structure_path() should return None if structure missing (graceful)."""
        mock_db = MagicMock()
        service = CanonicalScheduleExportService(mock_db)

        with patch.object(service, "_data_dir", return_value=Path("/nonexistent")):
            result = service._structure_path()
            assert result is None


class TestPhase3DataValidation:
    """Tests that Phase 3 DataValidation survives the single-save pipeline.

    Note: DataValidation requires row_mappings (from the real template) to
    bind DV objects to cell ranges. Without a template, DV objects are created
    but have no cell ranges and are not serialized. This test verifies the
    metadata sheets survive alongside the DV setup code path.
    """

    def test_metadata_survives_single_save_with_dv_code_path(self):
        """__SYS_META__ and __REF__ survive when DV code path is exercised."""
        from openpyxl import load_workbook
        from app.services.excel_metadata import ExportMetadata
        from app.services.xml_to_xlsx_converter import XMLToXlsxConverter

        data = {
            "block_start": "2026-03-12",
            "block_end": "2026-04-08",
            "residents": [
                {
                    "name": "Test Resident",
                    "id": "00000000-0000-0000-0000-000000000001",
                    "pgy": "2",
                    "rotation1": "SURG",
                    "rotation2": "",
                    "block_assignment_id": "00000000-0000-0000-0000-000000000002",
                    "days": [{"am": "C", "pm": "C"}] * 28,
                }
            ],
            "faculty": [],
        }

        meta = ExportMetadata(
            academic_year=2026,
            block_number=10,
            export_timestamp="2026-02-25T10:00:00Z",
        )

        converter = XMLToXlsxConverter(
            use_block_template2=True,
            apply_colors=False,
            strict_row_mapping=False,
        )
        xlsx_bytes = converter.convert_from_data(
            data,
            export_metadata=meta,
            rotation_codes=["SURG", "MED"],
            activity_codes=["C", "NF", "LV"],
        )

        wb = load_workbook(io.BytesIO(xlsx_bytes))

        # Phase 1: metadata sheets survive single-save
        assert "__SYS_META__" in wb.sheetnames
        assert "__REF__" in wb.sheetnames
        assert wb["__SYS_META__"].sheet_state == "veryHidden"
        assert wb["__REF__"].sheet_state == "veryHidden"

        # Verify metadata content
        meta_read = read_sys_meta(wb)
        assert meta_read is not None
        assert meta_read.academic_year == 2026
        assert meta_read.block_number == 10

        # Verify ref codes
        ref = read_ref_codes(wb)
        assert sorted(ref["rotations"]) == ["MED", "SURG"]
        assert sorted(ref["activities"]) == ["C", "LV", "NF"]

        wb.close()


class TestDynamicMappings:
    """Tests for UUID-based dynamic row allocation."""

    def test_dynamic_mappings_assigns_rows_by_pgy(self):
        """Residents get rows by PGY desc; faculty start at row 31."""
        from openpyxl import load_workbook
        from app.services.xml_to_xlsx_converter import XMLToXlsxConverter

        data = {
            "block_start": "2026-05-07",
            "block_end": "2026-06-03",
            "residents": [
                {
                    "name": "Intern Alpha",
                    "id": "aaa-1",
                    "pgy": 1,
                    "rotation1": "FMC",
                    "rotation2": "",
                    "days": [],
                },
                {
                    "name": "Senior Beta",
                    "id": "aaa-3",
                    "pgy": 3,
                    "rotation1": "Hilo",
                    "rotation2": "",
                    "days": [],
                },
                {
                    "name": "Junior Gamma",
                    "id": "aaa-2",
                    "pgy": 2,
                    "rotation1": "SM",
                    "rotation2": "",
                    "days": [],
                },
            ],
            "faculty": [
                {
                    "name": "Dr. Zulu",
                    "id": "bbb-1",
                    "rotation1": "",
                    "rotation2": "",
                    "days": [],
                },
                {
                    "name": "Dr. Alpha",
                    "id": "bbb-2",
                    "rotation1": "",
                    "rotation2": "",
                    "days": [],
                },
            ],
        }

        converter = XMLToXlsxConverter(
            use_block_template2=True,
            apply_colors=False,
            strict_row_mapping=False,
            preserve_template_identity_fields=False,
        )
        xlsx_bytes = converter.convert_from_data(data)
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        sheet = wb["Block Template2"]

        # PGY 3 first (row 9), then PGY 2 (row 10), then PGY 1 (row 11)
        assert sheet.cell(row=9, column=5).value == "Beta, Senior"  # PGY 3
        assert sheet.cell(row=10, column=5).value == "Gamma, Junior"  # PGY 2
        assert sheet.cell(row=11, column=5).value == "Alpha, Intern"  # PGY 1

        # Faculty alphabetical: Dr. Alpha (row 31), Dr. Zulu (row 32)
        assert sheet.cell(row=31, column=5).value == "Alpha, Dr."
        assert sheet.cell(row=32, column=5).value == "Zulu, Dr."

        # Unused rows hidden
        assert sheet.row_dimensions[12].hidden is True  # no 4th resident
        assert sheet.row_dimensions[33].hidden is True  # no 3rd faculty

        wb.close()

    def test_dynamic_mappings_with_fat_template(self):
        """Dynamic mappings work when loading the Fat Template XLSX."""
        from openpyxl import load_workbook
        from app.services.xml_to_xlsx_converter import XMLToXlsxConverter

        template_path = (
            Path(__file__).resolve().parents[2]
            / "data"
            / "BlockTemplate2_Official.xlsx"
        )
        if not template_path.exists():
            pytest.skip("Fat Template not generated yet")

        data = {
            "block_start": "2026-05-07",
            "block_end": "2026-06-03",
            "residents": [
                {
                    "name": "Test Resident",
                    "id": "uuid-r1",
                    "pgy": 2,
                    "rotation1": "SURG",
                    "rotation2": "",
                    "days": [{"date": "2026-05-07", "am": "C", "pm": "NF"}],
                },
            ],
            "faculty": [
                {
                    "name": "Test Faculty",
                    "id": "uuid-f1",
                    "rotation1": "",
                    "rotation2": "",
                    "days": [{"date": "2026-05-07", "am": "C", "pm": "AT"}],
                },
            ],
        }

        converter = XMLToXlsxConverter(
            template_path=template_path,
            use_block_template2=True,
            apply_colors=True,
            strict_row_mapping=False,
            preserve_template_identity_fields=False,
        )
        xlsx_bytes = converter.convert_from_data(data)
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        sheet = wb["Block Template2"]

        # Resident at row 9 (first slot), faculty at row 31
        assert sheet.cell(row=9, column=5).value == "Resident, Test"
        assert sheet.cell(row=31, column=5).value == "Faculty, Test"

        # Schedule data written (col 6=AM day 1, col 7=PM day 1)
        assert sheet.cell(row=9, column=6).value == "C"
        assert sheet.cell(row=9, column=7).value == "NF"
        assert sheet.cell(row=31, column=6).value == "C"
        assert sheet.cell(row=31, column=7).value == "AT"

        wb.close()


class TestPhase1Metadata:
    """Tests for Phase 1 metadata stamping (_inject_metadata).

    Note: _inject_metadata is retained as a utility for year-level export
    but the primary block export now uses single-save via converter kwargs.
    """

    def _make_blank_xlsx(self) -> bytes:
        """Create a minimal valid XLSX for _inject_metadata to load."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Schedule"
        ws.cell(row=1, column=1, value="test")
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def test_inject_metadata_adds_sys_meta(self):
        """_inject_metadata() should add __SYS_META__ sheet."""
        mock_db = MagicMock()
        mock_db.query.return_value.distinct.return_value.all.return_value = []

        service = CanonicalScheduleExportService(mock_db)
        xlsx_bytes = self._make_blank_xlsx()

        result = service._inject_metadata(
            xlsx_bytes, academic_year=2026, block_number=10
        )

        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(result))
        assert "__SYS_META__" in wb.sheetnames

        meta = read_sys_meta(wb)
        assert meta is not None
        assert meta.academic_year == 2026
        assert meta.block_number == 10
        assert meta.export_version == 1
        wb.close()

    def test_inject_metadata_adds_ref_sheet(self):
        """_inject_metadata() should add __REF__ sheet with rotation/activity codes."""
        mock_db = MagicMock()

        # First call returns rotation codes, second returns activity codes
        rot_query = MagicMock()
        rot_query.distinct.return_value.all.return_value = [
            ("SURG",),
            ("MED",),
            ("PEDS",),
        ]
        act_query = MagicMock()
        act_query.distinct.return_value.all.return_value = [("CLI",), ("OR",), ("LV",)]
        mock_db.query.side_effect = [rot_query, act_query]

        service = CanonicalScheduleExportService(mock_db)
        xlsx_bytes = self._make_blank_xlsx()

        result = service._inject_metadata(
            xlsx_bytes, academic_year=2026, block_number=10
        )

        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(result))
        assert "__REF__" in wb.sheetnames

        ref = read_ref_codes(wb)
        assert sorted(ref["rotations"]) == ["MED", "PEDS", "SURG"]
        assert sorted(ref["activities"]) == ["CLI", "LV", "OR"]
        wb.close()

    def test_inject_metadata_sys_meta_is_veryhidden(self):
        """Metadata sheets should be veryHidden."""
        mock_db = MagicMock()
        mock_db.query.return_value.distinct.return_value.all.return_value = []

        service = CanonicalScheduleExportService(mock_db)
        xlsx_bytes = self._make_blank_xlsx()

        result = service._inject_metadata(
            xlsx_bytes, academic_year=2026, block_number=5
        )

        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(result))
        assert wb["__SYS_META__"].sheet_state == "veryHidden"
        assert wb["__REF__"].sheet_state == "veryHidden"
        wb.close()


# ==================== YTD_SUMMARY Tests ====================


class TestBuildYtdSummarySheet:
    """Tests for _build_ytd_summary_sheet() and faculty union logic."""

    def test_ytd_summary_headers(self):
        """YTD_SUMMARY should have 10 column headers."""
        mock_db = MagicMock()
        service = CanonicalScheduleExportService(mock_db)

        wb = Workbook()
        ws = wb.active
        ws.title = "YTD_SUMMARY"

        block = MagicMock()
        block.block_number = 1
        faculty = [{"name": "Dr. Test"}]

        service._build_ytd_summary_sheet(ws, [block], faculty)

        expected_headers = [
            "Faculty Name",
            "YTD Clinic (C+SM)",
            "YTD CC",
            "YTD CV",
            "YTD Total Clinic",
            "YTD AT (AT+PCAT+DO)",
            "YTD Admin",
            "YTD Leave",
            "YTD FMIT Weeks",
            "YTD Call Nights",
        ]
        for col_idx, header in enumerate(expected_headers, start=1):
            assert ws.cell(row=1, column=col_idx).value == header

        wb.close()

    def test_ytd_summary_faculty_name_in_column_a(self):
        """Faculty names should be in column A starting at row 2."""
        mock_db = MagicMock()
        service = CanonicalScheduleExportService(mock_db)

        wb = Workbook()
        ws = wb.active

        block = MagicMock()
        block.block_number = 1
        faculty = [
            {"name": "Alpha, Dr."},
            {"name": "Beta, Dr."},
        ]

        service._build_ytd_summary_sheet(ws, [block], faculty)

        assert ws.cell(row=2, column=1).value == "Alpha, Dr."
        assert ws.cell(row=3, column=1).value == "Beta, Dr."
        wb.close()

    def test_ytd_summary_sumif_formulas_reference_block_sheets(self):
        """Cross-sheet SUMIF formulas should reference all block sheets."""
        mock_db = MagicMock()
        service = CanonicalScheduleExportService(mock_db)

        wb = Workbook()
        ws = wb.active

        block1 = MagicMock()
        block1.block_number = 0
        block2 = MagicMock()
        block2.block_number = 1
        faculty = [{"name": "Dr. Test"}]

        service._build_ytd_summary_sheet(ws, [block1, block2], faculty)

        # Column B (col 2) = YTD Clinic → SUMIF on BJ across both blocks
        formula = ws.cell(row=2, column=2).value
        assert formula is not None
        assert "'Block 0'" in formula
        assert "'Block 1'" in formula
        assert "SUMIF" in formula
        assert "$E$31:$E$80" in formula  # name lookup range
        assert "BJ$31:BJ$80" in formula  # data range

        wb.close()

    def test_ytd_summary_total_clinic_formula(self):
        """Column E (Total Clinic) should be =B+C+D."""
        mock_db = MagicMock()
        service = CanonicalScheduleExportService(mock_db)

        wb = Workbook()
        ws = wb.active

        block = MagicMock()
        block.block_number = 1
        faculty = [{"name": "Dr. Test"}]

        service._build_ytd_summary_sheet(ws, [block], faculty)

        # Column E = Total Clinic = B + C + D
        formula = ws.cell(row=2, column=5).value
        assert formula == "=B2+C2+D2"
        wb.close()

    def test_ytd_summary_fmit_divides_by_14(self):
        """FMIT weeks formula should divide SUMIF total by 14."""
        mock_db = MagicMock()
        service = CanonicalScheduleExportService(mock_db)

        wb = Workbook()
        ws = wb.active

        block = MagicMock()
        block.block_number = 5
        faculty = [{"name": "Dr. Test"}]

        service._build_ytd_summary_sheet(ws, [block], faculty)

        # Column I (col 9) = FMIT weeks = SUMIF(BQ) / 14
        formula = ws.cell(row=2, column=9).value
        assert formula is not None
        assert "/14" in formula
        assert "BQ" in formula
        wb.close()

    def test_ytd_summary_empty_name_skipped(self):
        """Faculty with empty name should not get formulas."""
        mock_db = MagicMock()
        service = CanonicalScheduleExportService(mock_db)

        wb = Workbook()
        ws = wb.active

        block = MagicMock()
        block.block_number = 1
        faculty = [{"name": ""}, {"name": "Dr. Real"}]

        service._build_ytd_summary_sheet(ws, [block], faculty)

        # Row 2 has empty name — no formula in col B
        assert ws.cell(row=2, column=2).value is None
        # Row 3 has real name — formula in col B
        assert ws.cell(row=3, column=2).value is not None
        assert "SUMIF" in str(ws.cell(row=3, column=2).value)
        wb.close()

    @patch("app.services.canonical_schedule_export_service.TAMCBlockExporter")
    def test_export_year_xlsx_uses_faculty_union(self, mock_exporter_class):
        """export_year_xlsx() should pass union of all blocks' faculty to YTD_SUMMARY."""
        mock_db = MagicMock()

        block_one = MagicMock()
        block_one.block_number = 1
        block_one.start_date = date(2026, 7, 1)
        block_one.end_date = date(2026, 7, 28)
        block_one.id = uuid4()

        block_two = MagicMock()
        block_two.block_number = 2
        block_two.start_date = date(2026, 7, 29)
        block_two.end_date = date(2026, 8, 25)
        block_two.id = uuid4()

        mock_db.execute.return_value.scalars.return_value.all.return_value = [
            block_one,
            block_two,
        ]
        mock_db.query.return_value.distinct.return_value.all.return_value = []

        mock_exporter = MagicMock()

        def _make_template():
            wb = Workbook()
            ws = wb.active
            ws.title = "Block Template2"
            buf = io.BytesIO()
            wb.save(buf)
            return buf.getvalue()

        mock_exporter.export.return_value = _make_template()
        mock_exporter.summary_col_start = 62
        mock_exporter_class.return_value = mock_exporter

        # Block 1 has Faculty A and B; Block 2 has Faculty B and C
        export_calls = [0]
        faculty_block_1 = [{"name": "Alpha, Dr."}, {"name": "Beta, Dr."}]
        faculty_block_2 = [{"name": "Beta, Dr."}, {"name": "Charlie, Dr."}]

        def mock_export_json(*args, **kwargs):
            export_calls[0] += 1
            if export_calls[0] == 1:
                return {"faculty": faculty_block_1}
            return {"faculty": faculty_block_2}

        service = CanonicalScheduleExportService(mock_db)
        captured_faculty = []

        original_build = service._build_ytd_summary_sheet

        def capture_build(ws, blocks, faculty, block_summary_cols=None):
            captured_faculty.extend(faculty)
            return original_build(ws, blocks, faculty, block_summary_cols)

        with (
            patch.object(service, "_export_json_data", side_effect=mock_export_json),
            patch.object(service, "_copy_worksheet"),
            patch.object(service, "_apply_phantom_columns"),
            patch.object(
                service, "_build_ytd_summary_sheet", side_effect=capture_build
            ),
        ):
            service.export_year_xlsx(academic_year=2026)

        # Should have 3 unique faculty (union), sorted alphabetically
        names = [f["name"] for f in captured_faculty]
        assert len(names) == 3
        assert names == ["Alpha, Dr.", "Beta, Dr.", "Charlie, Dr."]


class TestPhantomColumnBoundary:
    """Verify phantom columns don't erase summary formulas in short blocks."""

    def test_phantom_stops_before_summary_columns(self):
        """For a 3-day stub block, phantom fill must not touch summary cols."""
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill

        wb = Workbook()
        ws = wb.active

        # Simulate summary formula in column 12 (where a 3-day block's
        # summary starts: col 6 + 3*2 = 12)
        summary_col = 12
        ws.cell(row=9, column=summary_col, value="=SUMPRODUCT(...)")
        ws.cell(row=30, column=summary_col, value="C")

        mock_block = MagicMock()
        mock_block.start_date = date(2026, 7, 1)
        mock_block.end_date = date(2026, 7, 3)  # 3 days

        service = CanonicalScheduleExportService(MagicMock())
        service._apply_phantom_columns(ws, mock_block, summary_col_start=summary_col)

        # Summary cells should NOT be erased
        assert ws.cell(row=9, column=summary_col).value == "=SUMPRODUCT(...)"
        assert ws.cell(row=30, column=summary_col).value == "C"

        # Phantom cells before summary should be grayed out
        phantom_cell = ws.cell(row=9, column=summary_col - 1)
        assert phantom_cell.value is None
        assert phantom_cell.fill.fgColor.rgb is not None
