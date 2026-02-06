"""Diagnostics-focused tests for XMLToXlsxConverter."""

from __future__ import annotations

import pytest
from openpyxl import Workbook, load_workbook

from app.services.xml_to_xlsx_converter import XMLToXlsxConverter


def _write_template(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Block Template2"
    wb.save(path)


def _write_structure_xml(path):
    path.write_text(
        """<?xml version="1.0" encoding="UTF-8"?>
<block_template2>
  <layout><call_row row="4"/></layout>
  <residents>
    <resident name="Mapped Resident" row="9" pgy="2"/>
  </residents>
  <faculty>
    <person name="Mapped Faculty" row="31" template="C19"/>
  </faculty>
</block_template2>
""",
        encoding="utf-8",
    )


def _one_day_payload():
    return {
        "block_start": "2026-03-12",
        "block_end": "2026-03-12",
        "residents": [
            {
                "name": "Mapped Resident",
                "pgy": 2,
                "rotation1": "FMC",
                "rotation2": "",
                "days": [{"date": "2026-03-12", "am": "C", "pm": "AT"}],
            },
            {
                "name": "Unmapped Resident",
                "pgy": 2,
                "rotation1": "FMC",
                "rotation2": "",
                "days": [{"date": "2026-03-12", "am": "C", "pm": ""}],
            },
        ],
        "faculty": [
            {
                "name": "Mapped Faculty",
                "pgy": None,
                "rotation1": "",
                "rotation2": "",
                "days": [{"date": "2026-03-12", "am": "GME", "pm": "C"}],
            }
        ],
        "call": {"nights": []},
    }


def _one_day_payload_only_mapped():
    payload = _one_day_payload()
    payload["residents"] = [payload["residents"][0]]
    return payload


def test_conversion_stats_tracks_mapped_and_unmapped_people(tmp_path):
    template_path = tmp_path / "template.xlsx"
    structure_path = tmp_path / "structure.xml"
    output_path = tmp_path / "out.xlsx"
    _write_template(template_path)
    _write_structure_xml(structure_path)

    converter = XMLToXlsxConverter(
        template_path=template_path,
        structure_xml_path=structure_path,
        use_block_template2=True,
        apply_colors=False,
        strict_row_mapping=False,
    )

    converter.convert_from_data(_one_day_payload(), output_path=output_path)

    stats = converter.last_conversion_stats
    assert stats["residents"]["people_input"] == 2
    assert stats["residents"]["people_written"] == 1
    assert stats["residents"]["unmapped_names"] == ["Unmapped Resident"]
    assert stats["residents"]["codes_written"]["C"] == 1
    assert stats["residents"]["codes_written"]["AT"] == 1
    assert stats["faculty"]["people_input"] == 1
    assert stats["faculty"]["people_written"] == 1
    assert stats["faculty"]["unmapped_names"] == []
    assert stats["faculty"]["codes_written"]["GME"] == 1
    assert stats["faculty"]["codes_written"]["C"] == 1

    wb = load_workbook(output_path, data_only=False)
    ws = wb["Block Template2"]
    assert ws.cell(row=9, column=6).value == "C"
    assert ws.cell(row=9, column=7).value == "AT"
    assert ws.cell(row=31, column=6).value == "GME"
    assert ws.cell(row=31, column=7).value == "C"

    assert "Export_QA" in wb.sheetnames
    qa = wb["Export_QA"]
    assert qa["A1"].value == "Export QA Summary"
    assert qa["B3"].value == "2026-03-12"
    assert qa["B4"].value == "2026-03-12"
    # Code totals from payload: residents C=1, faculty C=1 => combined C=2.
    c_row = None
    for row in range(14, 40):
        if qa.cell(row=row, column=1).value == "C":
            c_row = row
            break
    assert c_row is not None
    assert qa.cell(row=c_row, column=2).value == 1
    assert qa.cell(row=c_row, column=3).value == 1
    assert qa.cell(row=c_row, column=4).value == 2


def test_strict_row_mapping_raises_for_unmapped_people(tmp_path):
    template_path = tmp_path / "template.xlsx"
    structure_path = tmp_path / "structure.xml"
    _write_template(template_path)
    _write_structure_xml(structure_path)

    converter = XMLToXlsxConverter(
        template_path=template_path,
        structure_xml_path=structure_path,
        use_block_template2=True,
        apply_colors=False,
        strict_row_mapping=True,
    )

    with pytest.raises(ValueError, match="No row mapping for: Unmapped Resident"):
        converter.convert_from_data(_one_day_payload())


def test_converter_unmerges_schedule_cells_before_writing(tmp_path):
    template_path = tmp_path / "template.xlsx"
    structure_path = tmp_path / "structure.xml"
    output_path = tmp_path / "out.xlsx"

    _write_template(template_path)
    _write_structure_xml(structure_path)

    # Simulate Block Template2 merged AM/PM cells.
    wb = load_workbook(template_path)
    ws = wb["Block Template2"]
    ws.merge_cells(start_row=9, start_column=6, end_row=9, end_column=7)
    wb.save(template_path)

    converter = XMLToXlsxConverter(
        template_path=template_path,
        structure_xml_path=structure_path,
        use_block_template2=True,
        apply_colors=False,
        strict_row_mapping=True,
    )

    converter.convert_from_data(_one_day_payload_only_mapped(), output_path=output_path)

    stats = converter.last_conversion_stats
    assert stats["schedule_cells_unmerged"] >= 1

    out_wb = load_workbook(output_path, data_only=False)
    out_ws = out_wb["Block Template2"]
    assert out_ws.cell(row=9, column=6).value == "C"
    assert out_ws.cell(row=9, column=7).value == "AT"


def test_can_disable_export_qa_sheet(tmp_path):
    template_path = tmp_path / "template.xlsx"
    structure_path = tmp_path / "structure.xml"
    output_path = tmp_path / "out.xlsx"
    _write_template(template_path)
    _write_structure_xml(structure_path)

    converter = XMLToXlsxConverter(
        template_path=template_path,
        structure_xml_path=structure_path,
        use_block_template2=True,
        apply_colors=False,
        strict_row_mapping=True,
        include_qa_sheet=False,
    )
    converter.convert_from_data(_one_day_payload_only_mapped(), output_path=output_path)

    wb = load_workbook(output_path, data_only=False)
    assert "Export_QA" not in wb.sheetnames


def test_prunes_empty_placeholder_sheets(tmp_path):
    template_path = tmp_path / "template.xlsx"
    structure_path = tmp_path / "structure.xml"
    output_path = tmp_path / "out.xlsx"
    _write_template(template_path)
    _write_structure_xml(structure_path)

    # Add an empty placeholder sheet similar to template artifacts.
    wb = load_workbook(template_path)
    wb.create_sheet("Sheet1")
    wb.save(template_path)

    converter = XMLToXlsxConverter(
        template_path=template_path,
        structure_xml_path=structure_path,
        use_block_template2=True,
        apply_colors=False,
        strict_row_mapping=True,
    )
    converter.convert_from_data(_one_day_payload_only_mapped(), output_path=output_path)

    out_wb = load_workbook(output_path, data_only=False)
    assert "Block Template2" in out_wb.sheetnames
    assert "Export_QA" in out_wb.sheetnames
    assert "Sheet1" not in out_wb.sheetnames
