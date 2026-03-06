"""Tests for TAMCBlockExporter — pixel-perfect block schedule export."""

from __future__ import annotations

from datetime import date
from io import BytesIO

import pytest
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from app.services.export.tamc_block_exporter import (
    NBSP,
    TAMCBlockExporter,
    RowAllocation,
)


def _make_person(
    pid: str = "aaa-111",
    name: str = "Smith, John",
    pgy: int | None = 2,
    faculty_role: str | None = None,
    rotation1: str = "FMC",
    rotation2: str = "",
    days: list | None = None,
):
    """Create a minimal person dict for testing."""
    if days is None:
        days = [
            {"date": "2026-05-07", "weekday": "Thu", "am": "C", "pm": "CC"},
            {"date": "2026-05-08", "weekday": "Fri", "am": "W", "pm": "W"},
        ]
    return {
        "id": pid,
        "name": name,
        "pgy": pgy,
        "faculty_role": faculty_role,
        "rotation1": rotation1,
        "rotation2": rotation2,
        "days": days,
    }


def _make_data(
    residents: list | None = None,
    faculty: list | None = None,
    call: dict | None = None,
    block_start: str = "2026-05-07",
    block_end: str = "2026-06-03",
):
    """Create minimal schedule data dict."""
    return {
        "block_start": block_start,
        "block_end": block_end,
        "source": "half_day_assignments",
        "residents": residents or [],
        "faculty": faculty or [],
        "call": call or {"nights": []},
    }


def _export_wb(data=None, block_config=None):
    """Helper: export and return openpyxl Workbook."""
    if data is None:
        data = _make_data(
            residents=[
                _make_person("r1", "Alpha, Zack", pgy=3, rotation1="HILO"),
                _make_person("r2", "Bravo, Charlie", pgy=2, rotation1="FMC"),
                _make_person("r3", "Charlie, Adam", pgy=1, rotation1="NF"),
            ],
            faculty=[
                _make_person(
                    "f1",
                    "Faculty, Core",
                    pgy=None,
                    faculty_role="core",
                    rotation1="",
                ),
            ],
        )
    exporter = TAMCBlockExporter(block_config=block_config or {})
    xlsx_bytes = exporter.export(data)
    return load_workbook(BytesIO(xlsx_bytes))


class TestExportBasics:
    def test_export_returns_bytes(self):
        data = _make_data()
        exporter = TAMCBlockExporter()
        result = exporter.export(data)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_sheet_name(self):
        wb = _export_wb()
        assert "Block Template2" in wb.sheetnames

    def test_column_widths(self):
        wb = _export_wb()
        ws = wb["Block Template2"]
        assert abs(ws.column_dimensions["A"].width - 10.16) < 0.1
        assert abs(ws.column_dimensions["E"].width - 40.5) < 0.1
        assert abs(ws.column_dimensions["F"].width - 10.16) < 0.1

    def test_row_heights(self):
        wb = _export_wb()
        ws = wb["Block Template2"]
        assert ws.row_dimensions[1].height == 15
        assert ws.row_dimensions[4].height == 25
        assert ws.row_dimensions[9].height == 21


class TestBlackSeparators:
    def test_black_separator_rows(self):
        wb = _export_wb()
        ws = wb["Block Template2"]
        separator_rows = {7, 8, 48, 55, 66, 69}
        for row in separator_rows:
            cell = ws.cell(row=row, column=1)
            assert cell.value == NBSP, f"Row {row} col A should have NBSP"
            fill_color = cell.fill.fgColor.rgb if cell.fill.fgColor else None
            assert fill_color == "FF000000", (
                f"Row {row} col A should be black, got {fill_color}"
            )


class TestHeaderColors:
    def test_header_day_colors(self):
        """Thursday should be light blue, Sat/Sun should be weekend blue."""
        wb = _export_wb()
        ws = wb["Block Template2"]
        # Block 12 starts 2026-05-07 (Thursday)
        # Col F (6) = day 0 = Thursday
        thu_cell = ws.cell(row=1, column=6)
        assert thu_cell.value == "THURSDAY"
        thu_color = thu_cell.fill.fgColor.rgb if thu_cell.fill.fgColor else None
        assert thu_color == "FF99CCFF", f"Thursday should be FF99CCFF, got {thu_color}"

        # Day 2 = Saturday (col 10)
        sat_cell = ws.cell(row=1, column=10)
        assert sat_cell.value == "SATURDAY"
        sat_color = sat_cell.fill.fgColor.rgb if sat_cell.fill.fgColor else None
        assert sat_color == "FF9BC2E6", f"Saturday should be FF9BC2E6, got {sat_color}"


class TestResidentSortOrder:
    def test_pgy3_first(self):
        wb = _export_wb()
        ws = wb["Block Template2"]
        # Row 9 should be PGY 3, row 10 PGY 2, row 11 PGY 1
        assert ws.cell(row=9, column=4).value == "PGY 3"
        assert ws.cell(row=10, column=4).value == "PGY 2"
        assert ws.cell(row=11, column=4).value == "PGY 1"


class TestFacultyFormatting:
    def test_faculty_col_ab_black(self):
        wb = _export_wb()
        ws = wb["Block Template2"]
        for row in range(31, 48):
            cell_a = ws.cell(row=row, column=1)
            assert cell_a.value == NBSP, f"Faculty row {row} col A should have NBSP"
            fill = cell_a.fill.fgColor.rgb if cell_a.fill.fgColor else None
            assert fill == "FF000000", (
                f"Faculty row {row} col A should be black, got {fill}"
            )


class TestRotationColors:
    def test_rotation_colors(self):
        wb = _export_wb()
        ws = wb["Block Template2"]
        # Row 9 = PGY 3 resident with rotation1="HILO" -> red
        hilo_cell = ws.cell(row=9, column=1)
        assert hilo_cell.value == "HILO"
        fill = hilo_cell.fill.fgColor.rgb if hilo_cell.fill.fgColor else None
        assert fill == "FFFF0000", f"HILO should be red, got {fill}"

        # Row 11 = PGY 1 with rotation1="NF" -> black
        nf_cell = ws.cell(row=11, column=1)
        assert nf_cell.value == "NF"
        nf_fill = nf_cell.fill.fgColor.rgb if nf_cell.fill.fgColor else None
        assert nf_fill == "FF000000", f"NF should be black, got {nf_fill}"


class TestSummaryFormulas:
    def test_summary_formulas_present(self):
        wb = _export_wb()
        ws = wb["Block Template2"]
        # Row 8 headers
        assert ws.cell(row=8, column=62).value == "C"
        assert ws.cell(row=8, column=63).value == "CC"
        assert ws.cell(row=8, column=64).value == "CV"

        # Row 9 formula check
        val = ws.cell(row=9, column=62).value
        assert val is not None
        assert "SUMPRODUCT" in str(val)
        assert "COUNTIF" in str(val)

    def test_leave_column_formula(self):
        wb = _export_wb()
        ws = wb["Block Template2"]
        assert ws.cell(row=8, column=71).value == "LV Days"
        val = ws.cell(row=9, column=71).value
        assert val is not None
        assert "LV" in str(val)


class TestMetadataSheets:
    def test_metadata_sheets_present(self):
        from app.services.excel_metadata import ExportMetadata

        data = _make_data()
        meta = ExportMetadata(
            academic_year=2025,
            block_number=12,
            export_timestamp="2026-03-06T00:00:00Z",
        )
        exporter = TAMCBlockExporter()
        xlsx_bytes = exporter.export(
            data,
            export_metadata=meta,
            rotation_codes=["FMC", "NF"],
            activity_codes=["C", "CC", "CV"],
        )
        wb = load_workbook(BytesIO(xlsx_bytes))
        assert "__SYS_META__" in wb.sheetnames
        assert "__REF__" in wb.sheetnames
        assert "__ANCHORS__" in wb.sheetnames


class TestNoBorders:
    def test_no_borders_on_body(self):
        wb = _export_wb()
        ws = wb["Block Template2"]
        # Check a body cell (row 9, col 6) has no borders
        cell = ws.cell(row=9, column=6)
        border = cell.border
        # openpyxl may return None for border sides, which means no border
        assert border.left is None or border.left.style is None
        assert border.right is None or border.right.style is None
        assert border.top is None or border.top.style is None
        assert border.bottom is None or border.bottom.style is None


class TestStaffCallYellow:
    def test_staff_call_yellow(self):
        wb = _export_wb()
        ws = wb["Block Template2"]
        # Row 4 schedule cells should have FFFFF2CC fill
        cell = ws.cell(row=4, column=6)
        fill = cell.fill.fgColor.rgb if cell.fill.fgColor else None
        assert fill == "FFFFF2CC", f"Staff call should be yellow, got {fill}"


class TestAppointmentsSection:
    def test_appointments_labels(self):
        wb = _export_wb()
        ws = wb["Block Template2"]
        assert ws.cell(row=72, column=5).value == "Screeners Needed"
        assert ws.cell(row=73, column=5).value == "Providers Virtual (CV)"

    def test_appointments_formulas(self):
        wb = _export_wb()
        ws = wb["Block Template2"]
        val = ws.cell(row=72, column=6).value
        assert val is not None
        assert "COUNTIF" in str(val)


class TestHiddenRows:
    def test_hidden_unused_rows(self):
        wb = _export_wb()
        ws = wb["Block Template2"]
        # With 3 residents (rows 9-11), rows 12-28 should be hidden
        for row in range(12, 29):
            dim = ws.row_dimensions.get(row)
            assert dim is not None and dim.hidden, f"Row {row} should be hidden"
