"""
Comprehensive coverage tests for xlsx_import.py module.

Tests SlotType classification, ProviderSchedule, date parsing,
and ClinicScheduleImporter functionality.
"""
from datetime import date, datetime
from io import BytesIO

import pytest
from openpyxl import Workbook

from app.services.xlsx_import import (
    ClinicScheduleImporter,
    ImportResult,
    ProviderSchedule,
    ScheduleSlot,
    SlotType,
)

# ============================================================================
# SlotType Classification Tests
# ============================================================================

class TestSlotTypeClassification:
    """Tests for slot type classification logic."""

    @pytest.fixture
    def importer(self):
        """Create a ClinicScheduleImporter instance without database."""
        return ClinicScheduleImporter(db=None)

    def test_classify_fmit_variants(self, importer):
        """Should correctly classify FMIT slot type variants."""
        fmit_values = ["fmit", "FMIT", "inpt", "inpatient", "ward", "wards", "Inpatient"]
        for value in fmit_values:
            assert importer.classify_slot(value) == SlotType.FMIT, f"Failed for: {value}"

    def test_classify_clinic_variants(self, importer):
        """Should correctly classify clinic slot type variants."""
        clinic_values = ["c", "C", "clinic", "Clinic", "pts", "patient", "appt", "sm", "sports"]
        for value in clinic_values:
            assert importer.classify_slot(value) == SlotType.CLINIC, f"Failed for: {value}"

    def test_classify_off_variants(self, importer):
        """Should correctly classify off/unavailable slot variants."""
        off_values = ["off", "OFF", "x", "X", "-", ""]
        for value in off_values:
            assert importer.classify_slot(value) == SlotType.OFF, f"Failed for: {value}"

    def test_classify_none_value(self, importer):
        """Should treat None as OFF."""
        assert importer.classify_slot(None) == SlotType.OFF

    def test_classify_vacation_variants(self, importer):
        """Should correctly classify vacation/leave variants."""
        vacation_values = ["vac", "vacation", "lv", "leave", "al"]
        for value in vacation_values:
            assert importer.classify_slot(value) == SlotType.VACATION, f"Failed for: {value}"

    def test_classify_conference_variants(self, importer):
        """Should correctly classify conference variants."""
        conference_values = ["conf", "conference", "cme", "mtg"]
        for value in conference_values:
            assert importer.classify_slot(value) == SlotType.CONFERENCE, f"Failed for: {value}"

    def test_classify_admin_variants(self, importer):
        """Should correctly classify admin variants."""
        admin_values = ["admin", "adm", "office"]
        for value in admin_values:
            assert importer.classify_slot(value) == SlotType.ADMIN, f"Failed for: {value}"

    def test_classify_unknown_value(self, importer):
        """Should return UNKNOWN for truly unrecognized values."""
        assert importer.classify_slot("random") == SlotType.UNKNOWN
        assert importer.classify_slot("123") == SlotType.UNKNOWN
        assert importer.classify_slot("???") == SlotType.UNKNOWN
        # "xyz" is now UNKNOWN - prefix matching requires 2+ char keys
        # and "x" is only 1 char, so no false positives
        assert importer.classify_slot("xyz") == SlotType.UNKNOWN

    def test_classify_prefix_match(self, importer):
        """Should match values starting with known prefixes (2+ chars)."""
        # Prefix matching for codes starting with known keys
        # This avoids false positives from substring matching

        # Starts with "fmit" -> FMIT
        assert importer.classify_slot("fmit-rotation") == SlotType.FMIT
        assert importer.classify_slot("fmit2") == SlotType.FMIT

        # Starts with "ward"/"wards" -> FMIT
        assert importer.classify_slot("wards-only") == SlotType.FMIT

        # Starts with "appt" -> CLINIC
        assert importer.classify_slot("appt-only") == SlotType.CLINIC

        # Starts with "conf" -> CONFERENCE
        assert importer.classify_slot("conference-room") == SlotType.CONFERENCE

    def test_classify_number_stripping(self, importer):
        """Should strip trailing numbers and match base code."""
        # C30 -> C -> CLINIC
        assert importer.classify_slot("C30") == SlotType.CLINIC
        assert importer.classify_slot("FMIT2") == SlotType.FMIT
        assert importer.classify_slot("ward1") == SlotType.FMIT

    def test_classify_compound_codes(self, importer):
        """Should handle slash-separated compound codes."""
        # PC/OFF -> should prioritize OFF (restrictive)
        assert importer.classify_slot("PC/OFF") == SlotType.OFF
        # C/CV -> both are CLINIC, first wins
        assert importer.classify_slot("C/CV") == SlotType.CLINIC
        # FMIT/C -> FMIT is restrictive, should win
        assert importer.classify_slot("FMIT/C") == SlotType.FMIT

    def test_classify_real_schedule_codes(self, importer):
        """Should correctly classify actual codes from real schedule files."""
        # These codes appeared in actual faculty schedule Excel files
        real_codes = {
            "AT": SlotType.ADMIN,      # Admin Time
            "CV": SlotType.CLINIC,     # Virtual Clinic
            "DO": SlotType.OFF,        # Day Off
            "GME": SlotType.ADMIN,     # GME Time
            "HOL": SlotType.OFF,       # Holiday
            "LEC": SlotType.CONFERENCE,  # Lecture
            "NF": SlotType.FMIT,       # Night Float
            "OIC": SlotType.FMIT,      # Officer In Charge
            "PC": SlotType.OFF,        # Post Call (NOT clinic!)
            "PCAT": SlotType.ADMIN,    # Patient Care Admin Team
            "W": SlotType.OFF,         # Weekend
        }
        for code, expected in real_codes.items():
            result = importer.classify_slot(code)
            assert result == expected, f"Code '{code}' classified as {result}, expected {expected}"

    def test_classify_case_insensitive(self, importer):
        """Should be case-insensitive in classification."""
        assert importer.classify_slot("FmIt") == SlotType.FMIT
        assert importer.classify_slot("ClInIc") == SlotType.CLINIC
        assert importer.classify_slot("VaCaTiOn") == SlotType.VACATION


# ============================================================================
# ProviderSchedule Tests
# ============================================================================

class TestProviderSchedule:
    """Tests for ProviderSchedule class."""

    def test_add_slot(self):
        """Should add slots to provider schedule."""
        schedule = ProviderSchedule(name="Dr. Smith")
        slot = ScheduleSlot(
            provider_name="Dr. Smith",
            date=date(2025, 3, 10),
            time_of_day="AM",
            slot_type=SlotType.CLINIC,
            raw_value="clinic"
        )

        schedule.add_slot(slot)
        assert len(schedule.slots) == 1
        assert schedule.get_slot(date(2025, 3, 10), "AM") == slot

    def test_add_multiple_slots(self):
        """Should handle multiple slots for different dates/times."""
        schedule = ProviderSchedule(name="Dr. Jones")

        slots = [
            ScheduleSlot("Dr. Jones", date(2025, 3, 10), "AM", SlotType.CLINIC, "clinic"),
            ScheduleSlot("Dr. Jones", date(2025, 3, 10), "PM", SlotType.FMIT, "fmit"),
            ScheduleSlot("Dr. Jones", date(2025, 3, 11), "AM", SlotType.OFF, "off"),
        ]

        for slot in slots:
            schedule.add_slot(slot)

        assert len(schedule.slots) == 3
        assert schedule.get_slot(date(2025, 3, 10), "AM").slot_type == SlotType.CLINIC
        assert schedule.get_slot(date(2025, 3, 10), "PM").slot_type == SlotType.FMIT
        assert schedule.get_slot(date(2025, 3, 11), "AM").slot_type == SlotType.OFF

    def test_get_slot_nonexistent(self):
        """Should return None for non-existent slot."""
        schedule = ProviderSchedule(name="Dr. Smith")
        assert schedule.get_slot(date(2025, 3, 10), "AM") is None

    def test_get_fmit_weeks_empty(self):
        """Should return empty list when no FMIT slots."""
        schedule = ProviderSchedule(name="Dr. Smith")
        schedule.add_slot(ScheduleSlot("Dr. Smith", date(2025, 3, 10), "AM", SlotType.CLINIC, "clinic"))

        assert schedule.get_fmit_weeks() == []

    def test_get_fmit_weeks_single_week(self):
        """Should identify a single FMIT week."""
        schedule = ProviderSchedule(name="Dr. Smith")

        # Add FMIT slots for Mon-Fri of a single week (March 3-7, 2025)
        for day in range(3, 8):  # Mon-Fri
            schedule.add_slot(ScheduleSlot(
                "Dr. Smith",
                date(2025, 3, day),
                "AM",
                SlotType.FMIT,
                "fmit"
            ))

        weeks = schedule.get_fmit_weeks()
        assert len(weeks) == 1
        # Week should be Monday (Mar 3) to Sunday (Mar 9)
        assert weeks[0] == (date(2025, 3, 3), date(2025, 3, 9))

    def test_get_fmit_weeks_multiple_weeks(self):
        """Should identify multiple distinct FMIT weeks."""
        schedule = ProviderSchedule(name="Dr. Smith")

        # Week 1: March 3-5, 2025 (Mon-Wed)
        for day in [3, 4, 5]:
            schedule.add_slot(ScheduleSlot(
                "Dr. Smith",
                date(2025, 3, day),
                "AM",
                SlotType.FMIT,
                "fmit"
            ))

        # Week 2: March 17-19, 2025 (Mon-Wed)
        for day in [17, 18, 19]:
            schedule.add_slot(ScheduleSlot(
                "Dr. Smith",
                date(2025, 3, day),
                "AM",
                SlotType.FMIT,
                "fmit"
            ))

        weeks = schedule.get_fmit_weeks()
        assert len(weeks) == 2
        assert weeks[0] == (date(2025, 3, 3), date(2025, 3, 9))
        assert weeks[1] == (date(2025, 3, 17), date(2025, 3, 23))

    def test_has_alternating_pattern_insufficient_weeks(self):
        """Should not detect alternating pattern with less than 3 weeks."""
        schedule = ProviderSchedule(name="Dr. Smith")

        # Only 2 weeks
        schedule.add_slot(ScheduleSlot("Dr. Smith", date(2025, 3, 3), "AM", SlotType.FMIT, "fmit"))
        schedule.add_slot(ScheduleSlot("Dr. Smith", date(2025, 3, 17), "AM", SlotType.FMIT, "fmit"))

        assert not schedule.has_alternating_pattern()

    def test_has_alternating_pattern_detected(self):
        """Should detect week-on/week-off alternating pattern."""
        schedule = ProviderSchedule(name="Dr. Smith")

        # Week on: Mar 3, Week off, Week on: Mar 17, Week off, Week on: Mar 31
        # This creates an alternating pattern with 1-week gaps
        schedule.add_slot(ScheduleSlot("Dr. Smith", date(2025, 3, 3), "AM", SlotType.FMIT, "fmit"))
        schedule.add_slot(ScheduleSlot("Dr. Smith", date(2025, 3, 17), "AM", SlotType.FMIT, "fmit"))
        schedule.add_slot(ScheduleSlot("Dr. Smith", date(2025, 3, 31), "AM", SlotType.FMIT, "fmit"))

        assert schedule.has_alternating_pattern()

    def test_has_alternating_pattern_not_detected(self):
        """Should not detect pattern when weeks are consolidated."""
        schedule = ProviderSchedule(name="Dr. Smith")

        # 3 consecutive weeks (no gaps)
        schedule.add_slot(ScheduleSlot("Dr. Smith", date(2025, 3, 3), "AM", SlotType.FMIT, "fmit"))
        schedule.add_slot(ScheduleSlot("Dr. Smith", date(2025, 3, 10), "AM", SlotType.FMIT, "fmit"))
        schedule.add_slot(ScheduleSlot("Dr. Smith", date(2025, 3, 17), "AM", SlotType.FMIT, "fmit"))

        # Gap calculation: 7 days between each week (consecutive)
        # has_alternating_pattern looks for gaps of 6-8 days but needs at least 2 such gaps
        # Here we have consecutive weeks, not alternating
        assert not schedule.has_alternating_pattern()


# ============================================================================
# Date Parsing Tests
# ============================================================================

class TestDateParsing:
    """Tests for date parsing from various formats."""

    @pytest.fixture
    def importer(self):
        """Create a ClinicScheduleImporter instance."""
        return ClinicScheduleImporter(db=None)

    def test_parse_date_from_datetime(self, importer):
        """Should convert datetime to date."""
        dt = datetime(2025, 3, 15, 10, 30)
        result = importer.parse_date_from_header(dt)
        assert result == date(2025, 3, 15)

    def test_parse_date_from_date(self, importer):
        """Should pass through date objects."""
        d = date(2025, 3, 15)
        result = importer.parse_date_from_header(d)
        assert result == date(2025, 3, 15)

    def test_parse_date_from_none(self, importer):
        """Should return None for None input."""
        assert importer.parse_date_from_header(None) is None

    def test_parse_date_iso_format(self, importer):
        """Should parse ISO format dates (YYYY-MM-DD)."""
        result = importer.parse_date_from_header("2025-03-15")
        assert result == date(2025, 3, 15)

    def test_parse_date_us_format(self, importer):
        """Should parse US format dates (MM/DD/YYYY)."""
        result = importer.parse_date_from_header("03/15/2025")
        assert result == date(2025, 3, 15)

    def test_parse_date_us_short_year(self, importer):
        """Should parse US format with 2-digit year (MM/DD/YY)."""
        result = importer.parse_date_from_header("03/15/25")
        assert result == date(2025, 3, 15)

    def test_parse_date_day_month_abbrev(self, importer):
        """Should parse day-month abbreviation format (DD-Mon)."""
        # Note: This format assumes current academic year
        result = importer.parse_date_from_header("15-Mar")
        assert result is not None
        assert result.month == 3
        assert result.day == 15

    def test_parse_date_invalid_format(self, importer):
        """Should return None for unparseable strings."""
        assert importer.parse_date_from_header("not a date") is None
        assert importer.parse_date_from_header("xyz") is None
        assert importer.parse_date_from_header("13/32/2025") is None


# ============================================================================
# ClinicScheduleImporter Excel Parsing Tests
# ============================================================================

class TestClinicScheduleImporter:
    """Tests for Excel file import functionality."""

    def create_simple_excel(self):
        """Create a simple test Excel workbook in memory."""
        wb = Workbook()
        ws = wb.active

        # Header row with dates
        ws['A1'] = "Provider"
        ws['B1'] = date(2025, 3, 10)
        ws['C1'] = date(2025, 3, 11)
        ws['D1'] = date(2025, 3, 12)

        # Data rows
        ws['A2'] = "Dr. Smith"
        ws['B2'] = "clinic"
        ws['C2'] = "fmit"
        ws['D2'] = "off"

        ws['A3'] = "Dr. Jones"
        ws['B3'] = "fmit"
        ws['C3'] = "clinic"
        ws['D3'] = "vacation"

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def create_invalid_excel(self):
        """Create an Excel file with no recognizable date headers."""
        wb = Workbook()
        ws = wb.active

        # No dates, just random text
        ws['A1'] = "Random"
        ws['B1'] = "Data"
        ws['C1'] = "Here"
        ws['A2'] = "No"
        ws['B2'] = "Dates"
        ws['C2'] = "Found"

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def test_import_valid_excel_file(self):
        """Should successfully import valid Excel file."""
        importer = ClinicScheduleImporter(db=None)
        excel_bytes = self.create_simple_excel()

        result = importer.import_file(file_bytes=excel_bytes)

        assert result.success
        assert len(result.providers) == 2
        assert "Dr. Smith" in result.providers
        assert "Dr. Jones" in result.providers
        assert result.total_slots > 0
        assert result.clinic_slots > 0
        assert result.fmit_slots > 0

    def test_import_invalid_excel_structure(self):
        """Should handle Excel files with invalid structure gracefully."""
        importer = ClinicScheduleImporter(db=None)
        excel_bytes = self.create_invalid_excel()

        result = importer.import_file(file_bytes=excel_bytes)

        # Should still succeed but may have limited data
        # The importer tries to detect format and may find nothing
        assert result is not None

    def test_import_no_file_provided(self):
        """Should return error when no file is provided."""
        importer = ClinicScheduleImporter(db=None)

        result = importer.import_file()

        assert not result.success
        assert len(result.errors) > 0
        assert "No file path or bytes provided" in result.errors[0]

    def test_import_specific_sheet_not_found(self):
        """Should return error when requested sheet doesn't exist."""
        importer = ClinicScheduleImporter(db=None)
        excel_bytes = self.create_simple_excel()

        result = importer.import_file(file_bytes=excel_bytes, sheet_name="NonExistentSheet")

        assert not result.success
        assert len(result.errors) > 0
        assert "not found" in result.errors[0]

    def test_import_tracks_date_range(self):
        """Should track min and max dates in imported schedule."""
        importer = ClinicScheduleImporter(db=None)
        excel_bytes = self.create_simple_excel()

        result = importer.import_file(file_bytes=excel_bytes)

        assert result.date_range[0] is not None
        assert result.date_range[1] is not None
        assert result.date_range[0] <= result.date_range[1]
        assert result.date_range[0] == date(2025, 3, 10)
        assert result.date_range[1] == date(2025, 3, 12)

    def test_import_counts_slot_types(self):
        """Should correctly count different slot types."""
        importer = ClinicScheduleImporter(db=None)
        excel_bytes = self.create_simple_excel()

        result = importer.import_file(file_bytes=excel_bytes)

        # Dr. Smith: clinic, fmit, off
        # Dr. Jones: fmit, clinic, vacation
        assert result.total_slots == 6
        assert result.clinic_slots == 2  # 2 clinic slots
        assert result.fmit_slots == 2    # 2 fmit slots

    def test_provider_schedule_slot_retrieval(self):
        """Should allow retrieving specific slots from provider schedules."""
        importer = ClinicScheduleImporter(db=None)
        excel_bytes = self.create_simple_excel()

        result = importer.import_file(file_bytes=excel_bytes)

        smith_schedule = result.providers["Dr. Smith"]

        # Dr. Smith has clinic on 3/10, fmit on 3/11, off on 3/12
        slot_310 = smith_schedule.get_slot(date(2025, 3, 10), "AM")
        assert slot_310 is not None
        assert slot_310.slot_type == SlotType.CLINIC

        slot_311 = smith_schedule.get_slot(date(2025, 3, 11), "AM")
        assert slot_311 is not None
        assert slot_311.slot_type == SlotType.FMIT


# ============================================================================
# Edge Cases and Integration Tests
# ============================================================================

class TestEdgeCases:
    """Tests for edge cases and boundary conditions."""

    def test_schedule_slot_key_property(self):
        """Should generate unique keys for slots."""
        slot1 = ScheduleSlot("Dr. Smith", date(2025, 3, 10), "AM", SlotType.CLINIC, "clinic")
        slot2 = ScheduleSlot("Dr. Smith", date(2025, 3, 10), "PM", SlotType.CLINIC, "clinic")
        slot3 = ScheduleSlot("Dr. Jones", date(2025, 3, 10), "AM", SlotType.CLINIC, "clinic")

        # Same provider, date, time should have same key
        slot1_dup = ScheduleSlot("Dr. Smith", date(2025, 3, 10), "AM", SlotType.FMIT, "fmit")
        assert slot1.key == slot1_dup.key

        # Different times should have different keys
        assert slot1.key != slot2.key

        # Different providers should have different keys
        assert slot1.key != slot3.key

    def test_import_result_add_conflict(self):
        """Should be able to add conflicts to import results."""
        from app.services.xlsx_import import ScheduleConflict

        result = ImportResult(success=True)
        conflict = ScheduleConflict(
            provider_name="Dr. Smith",
            date=date(2025, 3, 10),
            time_of_day="AM",
            conflict_type="double_book"
        )

        result.add_conflict(conflict)
        assert len(result.conflicts) == 1
        assert result.get_conflicts_by_provider("Dr. Smith") == [conflict]

    def test_import_result_filter_conflicts_by_type(self):
        """Should filter conflicts by type."""
        from app.services.xlsx_import import ScheduleConflict

        result = ImportResult(success=True)
        conflict1 = ScheduleConflict(
            provider_name="Dr. Smith",
            date=date(2025, 3, 10),
            time_of_day="AM",
            conflict_type="double_book"
        )
        conflict2 = ScheduleConflict(
            provider_name="Dr. Jones",
            date=date(2025, 3, 11),
            time_of_day="PM",
            conflict_type="fmit_clinic_overlap"
        )

        result.add_conflict(conflict1)
        result.add_conflict(conflict2)

        double_books = result.get_conflicts_by_type("double_book")
        assert len(double_books) == 1
        assert double_books[0].provider_name == "Dr. Smith"

    def test_empty_provider_schedule(self):
        """Should handle empty provider schedules gracefully."""
        schedule = ProviderSchedule(name="Dr. Empty")

        assert len(schedule.slots) == 0
        assert schedule.get_fmit_weeks() == []
        assert not schedule.has_alternating_pattern()
        assert schedule.get_slot(date(2025, 3, 10), "AM") is None
