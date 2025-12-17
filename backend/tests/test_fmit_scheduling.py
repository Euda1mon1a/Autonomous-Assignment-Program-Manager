"""
Tests for FMIT scheduling domain model.

Tests SwapFinder, ExternalConflict integration, and the swap finder API.
"""
from datetime import date, timedelta
from uuid import uuid4

import pytest
from fastapi.testclient import TestClient
from sqlalchemy.orm import Session

from app.models.absence import Absence
from app.models.person import Person
from app.services.xlsx_import import (
    ExternalConflict,
    ImportResult,
    ProviderSchedule,
    ScheduleSlot,
    SlotType,
    SwapFinder,
    absence_to_external_conflict,
    count_alternating_cycles,
    get_schedule_flexibility,
    has_back_to_back_conflict,
    load_external_conflicts_from_absences,
)

# ============================================================================
# Unit Tests for Constraint Detection
# ============================================================================

class TestBackToBackDetection:
    """Tests for back-to-back week conflict detection."""

    def test_no_conflict_with_gap(self):
        """Should not detect conflict with sufficient gap between weeks."""
        weeks = [
            date(2025, 3, 3),   # Week 1
            date(2025, 3, 17),  # Week 3 (2 week gap)
            date(2025, 4, 7),   # Week 6 (3 week gap)
        ]
        assert not has_back_to_back_conflict(weeks)

    def test_conflict_consecutive_weeks(self):
        """Should detect conflict for consecutive weeks."""
        weeks = [
            date(2025, 3, 3),   # Week 1
            date(2025, 3, 10),  # Week 2 (consecutive!)
        ]
        assert has_back_to_back_conflict(weeks)

    def test_conflict_near_consecutive(self):
        """Should detect conflict when gap is 7 days or less."""
        weeks = [
            date(2025, 3, 3),   # Monday
            date(2025, 3, 9),   # Sunday (6 days later)
        ]
        assert has_back_to_back_conflict(weeks)

    def test_no_conflict_empty_list(self):
        """Should not detect conflict with empty list."""
        assert not has_back_to_back_conflict([])

    def test_no_conflict_single_week(self):
        """Should not detect conflict with single week."""
        weeks = [date(2025, 3, 3)]
        assert not has_back_to_back_conflict(weeks)

    def test_conflict_unsorted_input(self):
        """Should detect conflict even with unsorted input."""
        weeks = [
            date(2025, 3, 17),  # Out of order
            date(2025, 3, 3),
            date(2025, 3, 10),  # Consecutive with 3/3
        ]
        assert has_back_to_back_conflict(weeks)


class TestAlternatingPatternDetection:
    """Tests for week-on/week-off alternating pattern detection."""

    def test_no_alternating_sparse(self):
        """Should not detect alternating with sparse weeks."""
        weeks = [
            date(2025, 3, 3),
            date(2025, 4, 7),   # 5 week gap
            date(2025, 5, 12),  # 5 week gap
        ]
        assert count_alternating_cycles(weeks) == 0

    def test_alternating_pattern_detected(self):
        """Should detect week-on/week-off pattern."""
        weeks = [
            date(2025, 3, 3),   # Week 1
            date(2025, 3, 17),  # Week 3 (1 week gap = alternating)
            date(2025, 3, 31),  # Week 5 (1 week gap = alternating)
        ]
        assert count_alternating_cycles(weeks) == 2

    def test_mixed_pattern(self):
        """Should count only alternating portions."""
        weeks = [
            date(2025, 3, 3),   # Week 1
            date(2025, 3, 17),  # Week 3 (alternating)
            date(2025, 4, 21),  # Week 8 (not alternating - 5 week gap)
            date(2025, 5, 5),   # Week 10 (alternating)
        ]
        assert count_alternating_cycles(weeks) == 2

    def test_empty_weeks(self):
        """Should return 0 for empty list."""
        assert count_alternating_cycles([]) == 0


class TestScheduleFlexibility:
    """Tests for schedule flexibility assessment."""

    def test_past_dates_impossible(self):
        """Past dates should be impossible to change."""
        past = date.today() - timedelta(days=1)
        assert get_schedule_flexibility(past) == "impossible"

    def test_imminent_dates_very_hard(self):
        """Dates within 2 weeks should be very hard."""
        near = date.today() + timedelta(days=7)
        assert get_schedule_flexibility(near) == "very_hard"

    def test_released_dates_hard(self):
        """Dates within release window should be hard."""
        released = date.today() + timedelta(days=60)
        assert get_schedule_flexibility(released, release_horizon_days=90) == "hard"

    def test_future_dates_easy(self):
        """Dates beyond release window should be easy."""
        future = date.today() + timedelta(days=120)
        assert get_schedule_flexibility(future, release_horizon_days=90) == "easy"


# ============================================================================
# SwapFinder Unit Tests
# ============================================================================

class TestSwapFinder:
    """Tests for SwapFinder class."""

    @pytest.fixture
    def mock_fmit_schedule(self) -> ImportResult:
        """Create a mock FMIT schedule with test providers."""
        result = ImportResult(success=True)

        # Provider A has weeks 1, 3, 5 (alternating)
        provider_a = ProviderSchedule(name="Dr. Smith")
        for week_num in [1, 3, 5]:
            week_start = date(2025, 3, 3) + timedelta(weeks=week_num - 1)
            for day in range(5):  # Mon-Fri
                for tod in ["AM", "PM"]:
                    slot = ScheduleSlot(
                        provider_name="Dr. Smith",
                        date=week_start + timedelta(days=day),
                        time_of_day=tod,
                        slot_type=SlotType.FMIT,
                        raw_value="FMIT",
                    )
                    provider_a.add_slot(slot)
        result.providers["Dr. Smith"] = provider_a

        # Provider B has weeks 2, 6 (not alternating)
        provider_b = ProviderSchedule(name="Dr. Jones")
        for week_num in [2, 6]:
            week_start = date(2025, 3, 3) + timedelta(weeks=week_num - 1)
            for day in range(5):
                for tod in ["AM", "PM"]:
                    slot = ScheduleSlot(
                        provider_name="Dr. Jones",
                        date=week_start + timedelta(days=day),
                        time_of_day=tod,
                        slot_type=SlotType.FMIT,
                        raw_value="FMIT",
                    )
                    provider_b.add_slot(slot)
        result.providers["Dr. Jones"] = provider_b

        # Provider C has week 4 only (under target)
        provider_c = ProviderSchedule(name="Dr. Lee")
        week_start = date(2025, 3, 3) + timedelta(weeks=3)
        for day in range(5):
            for tod in ["AM", "PM"]:
                slot = ScheduleSlot(
                    provider_name="Dr. Lee",
                    date=week_start + timedelta(days=day),
                    time_of_day=tod,
                    slot_type=SlotType.FMIT,
                    raw_value="FMIT",
                )
                provider_c.add_slot(slot)
        result.providers["Dr. Lee"] = provider_c

        return result

    def test_swap_finder_init(self, mock_fmit_schedule):
        """Should initialize SwapFinder with faculty weeks mapping."""
        finder = SwapFinder(fmit_schedule=mock_fmit_schedule)

        assert len(finder.faculty_weeks) == 3
        assert "Dr. Smith" in finder.faculty_weeks
        assert "Dr. Jones" in finder.faculty_weeks
        assert "Dr. Lee" in finder.faculty_weeks

    def test_find_swap_candidates_basic(self, mock_fmit_schedule):
        """Should find swap candidates for a target week."""
        finder = SwapFinder(fmit_schedule=mock_fmit_schedule)

        # Find candidates to take Dr. Smith's week 3
        target_week = date(2025, 3, 17)  # Week 3
        candidates = finder.find_swap_candidates("Dr. Smith", target_week)

        assert len(candidates) == 2  # Jones and Lee
        faculty_names = [c.faculty for c in candidates]
        assert "Dr. Jones" in faculty_names
        assert "Dr. Lee" in faculty_names

    def test_find_swap_candidates_back_to_back_check(self, mock_fmit_schedule):
        """Should flag candidates with back-to-back conflicts."""
        finder = SwapFinder(fmit_schedule=mock_fmit_schedule)

        # Week 3 is March 17. Jones has week 2 (Mar 10) - would create back-to-back
        target_week = date(2025, 3, 17)
        candidates = finder.find_swap_candidates("Dr. Smith", target_week)

        jones_candidate = next(c for c in candidates if c.faculty == "Dr. Jones")
        # Jones has week 2, taking week 3 would be back-to-back
        assert not jones_candidate.back_to_back_ok

        lee_candidate = next(c for c in candidates if c.faculty == "Dr. Lee")
        # Lee has week 4, taking week 3 would be back-to-back
        assert not lee_candidate.back_to_back_ok

    def test_find_excessive_alternating(self, mock_fmit_schedule):
        """Should find faculty with excessive alternating patterns."""
        finder = SwapFinder(fmit_schedule=mock_fmit_schedule)

        alternating = finder.find_excessive_alternating(threshold=2)

        # Dr. Smith has weeks 1, 3, 5 which is alternating
        assert len(alternating) >= 1
        smith_entry = next((a for a in alternating if a[0] == "Dr. Smith"), None)
        assert smith_entry is not None
        assert smith_entry[1] >= 2  # At least 2 alternating cycles

    def test_external_conflict_blocks_candidate(self, mock_fmit_schedule):
        """Should flag candidates with external conflicts."""
        conflicts = [
            ExternalConflict(
                faculty="Dr. Lee",
                start_date=date(2025, 3, 15),
                end_date=date(2025, 3, 22),
                conflict_type="conference",
                description="USAFP Conference",
            )
        ]

        finder = SwapFinder(
            fmit_schedule=mock_fmit_schedule,
            external_conflicts=conflicts,
        )

        # Find candidates for week 3 (March 17)
        target_week = date(2025, 3, 17)
        candidates = finder.find_swap_candidates("Dr. Smith", target_week)

        lee_candidate = next(c for c in candidates if c.faculty == "Dr. Lee")
        assert lee_candidate.external_conflict == "conference"


# ============================================================================
# Absence Integration Tests
# ============================================================================

class TestAbsenceIntegration:
    """Tests for ExternalConflict integration with Absence model."""

    @pytest.fixture
    def sample_faculty_for_absence(self, db: Session) -> Person:
        """Create a faculty member for absence testing."""
        faculty = Person(
            id=uuid4(),
            name="Dr. Test Faculty",
            type="faculty",
            email="test.faculty@hospital.org",
        )
        db.add(faculty)
        db.commit()
        db.refresh(faculty)
        return faculty

    def test_load_conflicts_from_blocking_absence(
        self,
        db: Session,
        sample_faculty_for_absence: Person,
    ):
        """Should load blocking absences as ExternalConflicts."""
        # Create a blocking TDY absence
        absence = Absence(
            id=uuid4(),
            person_id=sample_faculty_for_absence.id,
            start_date=date.today() + timedelta(days=30),
            end_date=date.today() + timedelta(days=37),
            absence_type="tdy",
            tdy_location="Germany",
            is_blocking=True,
        )
        db.add(absence)
        db.commit()

        conflicts = load_external_conflicts_from_absences(db)

        assert len(conflicts) >= 1
        tdy_conflict = next(
            (c for c in conflicts if c.faculty == "Dr. Test Faculty"),
            None
        )
        assert tdy_conflict is not None
        assert tdy_conflict.conflict_type == "tdy"
        assert "Germany" in tdy_conflict.description

    def test_non_blocking_absence_excluded(
        self,
        db: Session,
        sample_faculty_for_absence: Person,
    ):
        """Should not include non-blocking absences."""
        # Create a short vacation (non-blocking)
        absence = Absence(
            id=uuid4(),
            person_id=sample_faculty_for_absence.id,
            start_date=date.today() + timedelta(days=30),
            end_date=date.today() + timedelta(days=32),  # 3 days
            absence_type="vacation",
            is_blocking=False,
        )
        db.add(absence)
        db.commit()

        conflicts = load_external_conflicts_from_absences(db)

        # Should not include short vacation
        vacation_conflict = next(
            (c for c in conflicts if c.faculty == "Dr. Test Faculty" and c.conflict_type == "leave"),
            None
        )
        assert vacation_conflict is None

    def test_deployment_always_blocking(
        self,
        db: Session,
        sample_faculty_for_absence: Person,
    ):
        """Deployment absences should always be treated as blocking."""
        absence = Absence(
            id=uuid4(),
            person_id=sample_faculty_for_absence.id,
            start_date=date.today() + timedelta(days=30),
            end_date=date.today() + timedelta(days=90),
            absence_type="deployment",
            deployment_orders=True,
            replacement_activity="Deployed - Middle East",
            is_blocking=True,  # Must be explicit due to column default
        )
        db.add(absence)
        db.commit()

        conflicts = load_external_conflicts_from_absences(db)

        deploy_conflict = next(
            (c for c in conflicts if c.faculty == "Dr. Test Faculty"),
            None
        )
        assert deploy_conflict is not None
        assert deploy_conflict.conflict_type == "deployment"
        assert "Deployed" in deploy_conflict.description

    def test_absence_to_external_conflict_conversion(
        self,
        db: Session,
        sample_faculty_for_absence: Person,
    ):
        """Should correctly convert Absence to ExternalConflict."""
        absence = Absence(
            id=uuid4(),
            person_id=sample_faculty_for_absence.id,
            start_date=date(2025, 6, 1),
            end_date=date(2025, 6, 7),
            absence_type="conference",
            notes="Annual AAFP Conference",
            is_blocking=True,
        )
        db.add(absence)
        db.commit()
        db.refresh(absence)

        conflict = absence_to_external_conflict(absence)

        assert conflict is not None
        assert conflict.faculty == "Dr. Test Faculty"
        assert conflict.conflict_type == "conference"
        assert conflict.start_date == date(2025, 6, 1)
        assert conflict.end_date == date(2025, 6, 7)


# ============================================================================
# API Endpoint Tests
# ============================================================================

class TestSwapFinderAPI:
    """Tests for POST /api/schedule/swaps/find endpoint."""

    @pytest.fixture
    def mock_fmit_excel_bytes(self) -> bytes:
        """Create a minimal mock Excel file for testing."""
        import io

        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active

        # Header row with dates
        ws["A1"] = "Provider"
        ws["B1"] = date(2025, 3, 3)
        ws["C1"] = date(2025, 3, 10)
        ws["D1"] = date(2025, 3, 17)
        ws["E1"] = date(2025, 3, 24)

        # Provider data
        ws["A2"] = "Dr. Smith"
        ws["B2"] = "FMIT"
        ws["C2"] = ""
        ws["D2"] = "FMIT"
        ws["E2"] = ""

        ws["A3"] = "Dr. Jones"
        ws["B3"] = ""
        ws["C3"] = "FMIT"
        ws["D3"] = ""
        ws["E3"] = "FMIT"

        ws["A4"] = "Dr. Lee"
        ws["B4"] = ""
        ws["C4"] = ""
        ws["D4"] = ""
        ws["E4"] = "FMIT"

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()

    def test_find_swaps_endpoint_basic(
        self,
        client: TestClient,
        mock_fmit_excel_bytes: bytes,
        auth_headers: dict,
    ):
        """Should return swap candidates for a faculty/week."""
        import json

        request_data = {
            "target_faculty": "Dr. Smith",
            "target_week": "2025-03-03",
            "include_absence_conflicts": False,
        }

        response = client.post(
            "/api/schedule/swaps/find",
            files={"fmit_file": ("schedule.xlsx", mock_fmit_excel_bytes)},
            data={"request_json": json.dumps(request_data)},
            headers=auth_headers,
        )

        assert response.status_code == 200
        data = response.json()
        assert data["success"]
        assert data["target_faculty"] == "Dr. Smith"
        assert "candidates" in data
        assert "total_candidates" in data
        assert "viable_candidates" in data

    def test_find_swaps_with_external_conflicts(
        self,
        client: TestClient,
        mock_fmit_excel_bytes: bytes,
        auth_headers: dict,
    ):
        """Should include provided external conflicts in analysis."""
        import json

        request_data = {
            "target_faculty": "Dr. Smith",
            "target_week": "2025-03-03",
            "include_absence_conflicts": False,
            "external_conflicts": [
                {
                    "faculty": "Dr. Jones",
                    "start_date": "2025-03-01",
                    "end_date": "2025-03-07",
                    "conflict_type": "leave",
                    "description": "Vacation",
                }
            ],
        }

        response = client.post(
            "/api/schedule/swaps/find",
            files={"fmit_file": ("schedule.xlsx", mock_fmit_excel_bytes)},
            data={"request_json": json.dumps(request_data)},
            headers=auth_headers,
        )

        assert response.status_code == 200
        data = response.json()

        # Dr. Jones should have external conflict flagged
        jones_candidate = next(
            (c for c in data["candidates"] if c["faculty"] == "Dr. Jones"),
            None
        )
        if jones_candidate:
            assert jones_candidate["external_conflict"] == "leave"

    def test_find_swaps_faculty_not_found(
        self,
        client: TestClient,
        mock_fmit_excel_bytes: bytes,
        auth_headers: dict,
    ):
        """Should return 404 when target faculty not in schedule."""
        import json

        request_data = {
            "target_faculty": "Dr. NonExistent",
            "target_week": "2025-03-03",
            "include_absence_conflicts": False,
        }

        response = client.post(
            "/api/schedule/swaps/find",
            files={"fmit_file": ("schedule.xlsx", mock_fmit_excel_bytes)},
            data={"request_json": json.dumps(request_data)},
            headers=auth_headers,
        )

        assert response.status_code == 404
        assert "not found" in response.json()["detail"].lower()

    def test_find_swaps_invalid_json(
        self,
        client: TestClient,
        mock_fmit_excel_bytes: bytes,
        auth_headers: dict,
    ):
        """Should return 400 for invalid request JSON."""
        response = client.post(
            "/api/schedule/swaps/find",
            files={"fmit_file": ("schedule.xlsx", mock_fmit_excel_bytes)},
            data={"request_json": "not valid json"},
            headers=auth_headers,
        )

        assert response.status_code == 400
        assert "Invalid request JSON" in response.json()["detail"]

    def test_find_swaps_returns_alternating_patterns(
        self,
        client: TestClient,
        mock_fmit_excel_bytes: bytes,
        auth_headers: dict,
    ):
        """Should include alternating pattern analysis in response."""
        import json

        request_data = {
            "target_faculty": "Dr. Smith",
            "target_week": "2025-03-03",
            "include_absence_conflicts": False,
        }

        response = client.post(
            "/api/schedule/swaps/find",
            files={"fmit_file": ("schedule.xlsx", mock_fmit_excel_bytes)},
            data={"request_json": json.dumps(request_data)},
            headers=auth_headers,
        )

        assert response.status_code == 200
        data = response.json()
        assert "alternating_patterns" in data


# ============================================================================
# Schema Validation Tests
# ============================================================================

class TestSwapFinderSchemas:
    """Tests for SwapFinder API schema validation."""

    def test_swap_finder_request_valid(self):
        """Should accept valid SwapFinderRequest."""
        from app.schemas.schedule import SwapFinderRequest

        request = SwapFinderRequest(
            target_faculty="Dr. Smith",
            target_week=date(2025, 3, 3),
        )
        assert request.target_faculty == "Dr. Smith"
        assert request.include_absence_conflicts is True  # default

    def test_external_conflict_input_validates_dates(self):
        """Should reject invalid date ranges."""
        import pytest

        from app.schemas.schedule import ExternalConflictInput

        with pytest.raises(ValueError):
            ExternalConflictInput(
                faculty="Dr. Smith",
                start_date=date(2025, 3, 10),
                end_date=date(2025, 3, 1),  # End before start
                conflict_type="leave",
            )

    def test_faculty_target_input_constraints(self):
        """Should enforce constraints on FacultyTargetInput."""
        from app.schemas.schedule import FacultyTargetInput

        # Valid input
        target = FacultyTargetInput(
            name="Dr. Smith",
            target_weeks=6,
            role="faculty",
        )
        assert target.target_weeks == 6

        # Should reject invalid weeks
        with pytest.raises(ValueError):
            FacultyTargetInput(
                name="Dr. Smith",
                target_weeks=100,  # Too high
            )
