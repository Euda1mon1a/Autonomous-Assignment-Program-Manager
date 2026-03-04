"""Parse leave requests from coordinator Excel files.

Reads the AY leave request spreadsheet format:
  Column A = resident last name
  Columns B-G = free-text leave requests with dates and descriptions

Uses strict name matching (Unmatched Queue) instead of fuzzy matching
to prevent silent misassignment of annual leave.
"""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from uuid import UUID

import openpyxl

from app.scheduling.annual.context import LeaveRequest, map_leave_to_blocks
from app.utils.academic_blocks import BlockDates

logger = logging.getLogger(__name__)

MONTHS: dict[str, int] = {
    "jan": 1,
    "january": 1,
    "feb": 2,
    "february": 2,
    "mar": 3,
    "march": 3,
    "apr": 4,
    "april": 4,
    "may": 5,
    "jun": 6,
    "june": 6,
    "jul": 7,
    "july": 7,
    "aug": 8,
    "august": 8,
    "sep": 9,
    "sept": 9,
    "september": 9,
    "oct": 10,
    "october": 10,
    "nov": 11,
    "november": 11,
    "dec": 12,
    "december": 12,
}


@dataclass
class UnmatchedEntry:
    """A leave request row that couldn't be matched to a resident."""

    row_number: int
    name_text: str
    leave_texts: list[str]
    reason: str  # "no_match" or "ambiguous"


@dataclass
class ParseResult:
    """Result of parsing a leave request Excel file."""

    leave_requests: list[LeaveRequest]
    unmatched: list[UnmatchedEntry]
    parse_errors: list[str]  # Free-text dates that couldn't be parsed


def parse_dates_from_text(text: str) -> tuple[date | None, date | None]:
    """Extract start and end dates from free-text leave request.

    Handles formats like:
    - "July 14-20 PNW, Flights booked"
    - "Dec 28-Jan3"
    - "August 29-Sep 4, 2026 (Friend's wedding)"
    - "Aug 12th 1st day skool"
    """
    text_lower = text.lower().strip()

    # Try cross-month: "Month DD - Month DD[, YYYY]"
    m = re.search(
        r"(\w+)\s+(\d{1,2})(?:st|nd|rd|th)?\s*[-–]\s*(\w+)\s*(\d{1,2})(?:st|nd|rd|th)?(?:\s*,?\s*(\d{4}))?",
        text_lower,
    )
    if m:
        m1_str, d1, m2_str, d2 = (
            m.group(1),
            int(m.group(2)),
            m.group(3),
            int(m.group(4)),
        )
        year = int(m.group(5)) if m.group(5) else None
        month1, month2 = MONTHS.get(m1_str), MONTHS.get(m2_str)
        if month1 and month2:
            y1 = year or (2026 if month1 >= 7 else 2027)
            y2 = year or (2026 if month2 >= 7 else 2027)
            if month2 < month1 and year is None:
                y2 = y1 + 1
            try:
                return date(y1, month1, d1), date(y2, month2, d2)
            except ValueError:
                pass

    # Try same-month range: "Month DD-DD[, YYYY]"
    m = re.search(
        r"(\w+)\s+(\d{1,2})(?:st|nd|rd|th)?\s*[-–]\s*(\d{1,2})(?:st|nd|rd|th)?(?:\s*,?\s*(\d{4}))?",
        text_lower,
    )
    if m:
        month_str, start_day, end_day = m.group(1), int(m.group(2)), int(m.group(3))
        year = int(m.group(4)) if m.group(4) else None
        month = MONTHS.get(month_str)
        if month:
            if year is None:
                year = 2026 if month >= 7 else 2027
            try:
                return date(year, month, start_day), date(year, month, end_day)
            except ValueError:
                pass

    # Try single date: "Month DD[th][, YYYY]"
    m = re.search(
        r"(\w+)\s+(\d{1,2})(?:st|nd|rd|th)?(?:\s*,?\s*(\d{4}))?",
        text_lower,
    )
    if m:
        month_str, day = m.group(1), int(m.group(2))
        year = int(m.group(3)) if m.group(3) else None
        month = MONTHS.get(month_str)
        if month:
            if year is None:
                year = 2026 if month >= 7 else 2027
            try:
                d = date(year, month, day)
                return d, d
            except ValueError:
                pass

    return None, None


def parse_leave_excel(
    file_path: str | Path,
    resident_lookup: dict[str, tuple[UUID, int]],
    blocks: list[BlockDates],
) -> ParseResult:
    """Parse a leave request Excel file.

    Args:
        file_path: Path to .xlsx file
        resident_lookup: Dict mapping lowercase last name -> (person_id, pgy_level).
            Built from the Person table by the caller.
        blocks: List of BlockDates for the target AY (Block 0 excluded).

    Returns:
        ParseResult with leave_requests, unmatched entries, and parse errors.
    """
    wb = openpyxl.load_workbook(str(file_path), data_only=True)
    ws = wb[wb.sheetnames[0]]

    leave_requests: list[LeaveRequest] = []
    unmatched: list[UnmatchedEntry] = []
    parse_errors: list[str] = []
    current_pgy: int | None = None

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        a_val = row[0].value
        if not a_val:
            continue

        a_str = str(a_val).strip()

        # Detect PGY header rows
        if "PGY" in a_str.upper():
            m = re.search(r"PGY[- ]?(\d)", a_str, re.IGNORECASE)
            if m:
                current_pgy = int(m.group(1))
            continue

        # Skip notes/headers
        if a_str.lower().startswith("notes") or a_str.lower().startswith("leave"):
            continue

        if current_pgy is None:
            continue

        # This is a resident row
        name_text = a_str.rstrip()
        name_lower = name_text.lower().rstrip()

        # Strict name matching — exact last name, case-insensitive
        match = resident_lookup.get(name_lower)
        if match is None:
            # Try without trailing spaces
            match = resident_lookup.get(name_lower.strip())

        leave_texts = [str(c.value).strip() for c in row[1:] if c.value]

        if match is None:
            unmatched.append(
                UnmatchedEntry(
                    row_number=row[0].row,
                    name_text=name_text,
                    leave_texts=leave_texts,
                    reason="no_match",
                )
            )
            continue

        person_id, pgy = match

        # Parse each leave request cell
        for leave_text in leave_texts:
            start, end = parse_dates_from_text(leave_text)
            if start is None or end is None:
                parse_errors.append(
                    f"{name_text}: could not parse dates from '{leave_text}'"
                )
                continue

            touched_blocks = map_leave_to_blocks(start, end, blocks)
            if not touched_blocks:
                parse_errors.append(
                    f"{name_text}: dates {start}-{end} don't overlap any blocks"
                )
                continue

            leave_requests.append(
                LeaveRequest(
                    person_id=person_id,
                    resident_name=name_text,
                    pgy=pgy,
                    start_date=start,
                    end_date=end,
                    blocks=touched_blocks,
                    raw_text=leave_text,
                )
            )

    wb.close()

    logger.info(
        "Parsed %d leave requests from %s (%d unmatched, %d parse errors)",
        len(leave_requests),
        file_path,
        len(unmatched),
        len(parse_errors),
    )
    return ParseResult(
        leave_requests=leave_requests,
        unmatched=unmatched,
        parse_errors=parse_errors,
    )
