"""
Export scheduler service.

Manages scheduled export jobs including:
- Creating and updating export jobs
- Calculating next run times
- Executing export jobs
- Managing export templates
"""

import csv
import io
import json
import logging
from datetime import date, datetime, timedelta
from typing import Any

from croniter import croniter
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import joinedload

from app.exports.delivery import ExportDeliveryService
from app.models.absence import Absence
from app.models.assignment import Assignment
from app.models.block import Block
from app.models.certification import PersonCertification
from app.models.export_job import (
    ExportDeliveryMethod,
    ExportFormat,
    ExportJob,
    ExportJobExecution,
    ExportJobStatus,
    ExportTemplate,
)
from app.models.person import Person
from app.models.swap import SwapRecord
from app.services.xlsx_export import generate_legacy_xlsx

logger = logging.getLogger(__name__)


class ExportSchedulerService:
    """Service for managing export job scheduling and execution."""

    def __init__(self, db: AsyncSession):
        """
        Initialize export scheduler service.

        Args:
            db: Database session
        """
        self.db = db
        self.delivery_service = ExportDeliveryService()

    async def create_job(
        self,
        job_data: dict[str, Any],
        created_by: str
    ) -> ExportJob:
        """
        Create a new export job.

        Args:
            job_data: Job configuration data
            created_by: Username of creator

        Returns:
            ExportJob: Created job instance
        """
        job = ExportJob(**job_data, created_by=created_by)

        # Calculate next run time if scheduled
        if job.schedule_enabled and job.schedule_cron:
            job.next_run_at = self._calculate_next_run_time(job.schedule_cron)

        self.db.add(job)
        await self.db.commit()
        await self.db.refresh(job)

        logger.info(f"Created export job: {job.name} (ID: {job.id})")
        return job

    async def update_job(
        self,
        job_id: str,
        update_data: dict[str, Any]
    ) -> ExportJob | None:
        """
        Update an existing export job.

        Args:
            job_id: Job ID
            update_data: Fields to update

        Returns:
            ExportJob | None: Updated job or None if not found
        """
        result = await self.db.execute(
            select(ExportJob).where(ExportJob.id == job_id)
        )
        job = result.scalar_one_or_none()

        if not job:
            return None

        # Update fields
        for field, value in update_data.items():
            if hasattr(job, field) and value is not None:
                setattr(job, field, value)

        # Recalculate next run time if schedule changed
        if "schedule_cron" in update_data or "schedule_enabled" in update_data:
            if job.schedule_enabled and job.schedule_cron:
                job.next_run_at = self._calculate_next_run_time(job.schedule_cron)
            else:
                job.next_run_at = None

        job.updated_at = datetime.utcnow()
        await self.db.commit()
        await self.db.refresh(job)

        logger.info(f"Updated export job: {job.name} (ID: {job.id})")
        return job

    async def delete_job(self, job_id: str) -> bool:
        """
        Delete an export job.

        Args:
            job_id: Job ID

        Returns:
            bool: True if deleted, False if not found
        """
        result = await self.db.execute(
            select(ExportJob).where(ExportJob.id == job_id)
        )
        job = result.scalar_one_or_none()

        if not job:
            return False

        await self.db.delete(job)
        await self.db.commit()

        logger.info(f"Deleted export job: {job.name} (ID: {job_id})")
        return True

    async def get_job(self, job_id: str) -> ExportJob | None:
        """
        Get an export job by ID.

        Args:
            job_id: Job ID

        Returns:
            ExportJob | None: Job or None if not found
        """
        result = await self.db.execute(
            select(ExportJob).where(ExportJob.id == job_id)
        )
        return result.scalar_one_or_none()

    async def list_jobs(
        self,
        page: int = 1,
        page_size: int = 20,
        enabled_only: bool = False
    ) -> tuple[list[ExportJob], int]:
        """
        List export jobs with pagination.

        Args:
            page: Page number (1-indexed)
            page_size: Items per page
            enabled_only: Only return enabled jobs

        Returns:
            tuple[list[ExportJob], int]: (jobs, total_count)
        """
        # Build query
        query = select(ExportJob)
        if enabled_only:
            query = query.where(ExportJob.enabled == True)

        # Get total count
        count_query = select(func.count()).select_from(ExportJob)
        if enabled_only:
            count_query = count_query.where(ExportJob.enabled == True)
        total_result = await self.db.execute(count_query)
        total = total_result.scalar()

        # Get paginated results
        query = query.order_by(ExportJob.created_at.desc())
        query = query.offset((page - 1) * page_size).limit(page_size)
        result = await self.db.execute(query)
        jobs = result.scalars().all()

        return list(jobs), total

    async def get_due_jobs(self) -> list[ExportJob]:
        """
        Get jobs that are due to run.

        Returns:
            list[ExportJob]: Jobs due for execution
        """
        now = datetime.utcnow()
        result = await self.db.execute(
            select(ExportJob)
            .where(
                ExportJob.enabled == True,
                ExportJob.schedule_enabled == True,
                ExportJob.next_run_at <= now
            )
            .order_by(ExportJob.next_run_at)
        )
        return list(result.scalars().all())

    async def execute_job(
        self,
        job_id: str,
        triggered_by: str = "scheduled",
        override_filters: dict[str, Any] | None = None,
        override_recipients: list[str] | None = None
    ) -> ExportJobExecution:
        """
        Execute an export job.

        Args:
            job_id: Job ID to execute
            triggered_by: Who/what triggered this execution
            override_filters: Temporary filter overrides
            override_recipients: Temporary recipient overrides

        Returns:
            ExportJobExecution: Execution record
        """
        # Get job
        job = await self.get_job(job_id)
        if not job:
            raise ValueError(f"Export job not found: {job_id}")

        # Create execution record
        execution = ExportJobExecution(
            job_id=job.id,
            job_name=job.name,
            started_at=datetime.utcnow(),
            scheduled_run_time=job.next_run_at,
            status=ExportJobStatus.RUNNING,
            triggered_by=triggered_by,
            email_recipients=override_recipients or job.email_recipients,
        )
        self.db.add(execution)
        await self.db.commit()
        await self.db.refresh(execution)

        try:
            # Generate export data
            filters = override_filters or job.filters or {}
            file_data, row_count = await self._generate_export_data(
                job.template,
                job.format,
                filters,
                job.columns,
                job.include_headers
            )

            # Generate filename
            timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            extension = job.format.value
            filename = f"{job.name.replace(' ', '_')}_{timestamp}.{extension}"

            # Determine content type
            content_types = {
                ExportFormat.CSV: "text/csv",
                ExportFormat.JSON: "application/json",
                ExportFormat.XLSX: "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ExportFormat.XML: "application/xml",
            }
            content_type = content_types.get(job.format, "application/octet-stream")

            # Deliver export
            delivery_result = self.delivery_service.deliver(
                file_data=file_data,
                filename=filename,
                delivery_method=job.delivery_method.value,
                email_recipients=execution.email_recipients,
                email_subject=job.email_subject_template or f"Export: {job.name}",
                email_body=job.email_body_template,
                s3_bucket=job.s3_bucket,
                s3_key_prefix=job.s3_key_prefix,
                content_type=content_type,
                metadata={
                    "job_id": str(job.id),
                    "job_name": job.name,
                    "execution_id": str(execution.id),
                    "generated_at": datetime.utcnow().isoformat(),
                }
            )

            # Update execution record
            execution.finished_at = datetime.utcnow()
            execution.runtime_seconds = int(
                (execution.finished_at - execution.started_at).total_seconds()
            )
            execution.status = ExportJobStatus.COMPLETED if delivery_result.success else ExportJobStatus.FAILED
            execution.row_count = row_count
            execution.file_size_bytes = len(file_data) if isinstance(file_data, bytes) else None
            execution.email_sent = delivery_result.email_sent
            execution.s3_url = delivery_result.s3_url

            if not delivery_result.success:
                execution.error_message = delivery_result.message

            # Update job
            job.last_run_at = execution.started_at
            job.run_count += 1
            if job.schedule_enabled and job.schedule_cron:
                job.next_run_at = self._calculate_next_run_time(job.schedule_cron)

            await self.db.commit()
            await self.db.refresh(execution)

            logger.info(
                f"Export job executed: {job.name} (ID: {job.id}), "
                f"Status: {execution.status}, Rows: {row_count}"
            )
            return execution

        except Exception as e:
            # Update execution with error
            execution.finished_at = datetime.utcnow()
            execution.runtime_seconds = int(
                (execution.finished_at - execution.started_at).total_seconds()
            )
            execution.status = ExportJobStatus.FAILED
            execution.error_message = str(e)
            execution.error_traceback = logging.traceback.format_exc() if hasattr(logging, "traceback") else None

            # Update job
            job.last_run_at = execution.started_at
            job.run_count += 1

            await self.db.commit()
            await self.db.refresh(execution)

            logger.error(f"Export job failed: {job.name} (ID: {job.id}): {e}", exc_info=True)
            return execution

    async def _generate_export_data(
        self,
        template: str,
        format: ExportFormat,
        filters: dict[str, Any],
        columns: list[str] | None,
        include_headers: bool
    ) -> tuple[bytes, int]:
        """
        Generate export data based on template and format.

        Args:
            template: Export template
            format: Export format
            filters: Data filters
            columns: Columns to include
            include_headers: Whether to include headers

        Returns:
            tuple[bytes, int]: (file_data, row_count)
        """
        # Get data based on template
        data, headers = await self._get_template_data(template, filters, columns)

        # Format data based on format
        if format == ExportFormat.CSV:
            return self._format_as_csv(data, headers, include_headers), len(data)
        elif format == ExportFormat.JSON:
            return self._format_as_json(data, headers), len(data)
        elif format == ExportFormat.XLSX:
            return self._format_as_xlsx(data, headers, include_headers), len(data)
        elif format == ExportFormat.XML:
            return self._format_as_xml(data, headers), len(data)
        else:
            raise ValueError(f"Unsupported export format: {format}")

    async def _get_template_data(
        self,
        template: str,
        filters: dict[str, Any],
        columns: list[str] | None
    ) -> tuple[list[dict], list[str]]:
        """
        Get data for a specific template.

        Args:
            template: Template name
            filters: Data filters
            columns: Columns to include

        Returns:
            tuple[list[dict], list[str]]: (data_rows, headers)
        """
        if template == ExportTemplate.PERSONNEL.value:
            return await self._get_personnel_data(filters, columns)
        elif template == ExportTemplate.ABSENCES.value:
            return await self._get_absences_data(filters, columns)
        elif template == ExportTemplate.FULL_SCHEDULE.value:
            return await self._get_schedule_data(filters, columns)
        elif template == ExportTemplate.CERTIFICATIONS.value:
            return await self._get_certifications_data(filters, columns)
        elif template == ExportTemplate.SWAP_HISTORY.value:
            return await self._get_swap_history_data(filters, columns)
        else:
            raise ValueError(f"Unsupported template: {template}")

    async def _get_personnel_data(
        self,
        filters: dict[str, Any],
        columns: list[str] | None
    ) -> tuple[list[dict], list[str]]:
        """Get personnel data."""
        query = select(Person).order_by(Person.name)

        # Apply filters
        if filters.get("type"):
            query = query.where(Person.type == filters["type"])
        if filters.get("pgy_level"):
            query = query.where(Person.pgy_level == filters["pgy_level"])

        result = await self.db.execute(query)
        people = result.scalars().all()

        headers = ["id", "name", "type", "pgy_level", "email", "specialties", "performs_procedures"]
        if columns:
            headers = [h for h in headers if h in columns]

        data = []
        for person in people:
            row = {
                "id": str(person.id),
                "name": person.name,
                "type": person.type,
                "pgy_level": person.pgy_level,
                "email": person.email,
                "specialties": person.specialties,
                "performs_procedures": person.performs_procedures,
            }
            if columns:
                row = {k: v for k, v in row.items() if k in columns}
            data.append(row)

        return data, headers

    async def _get_absences_data(
        self,
        filters: dict[str, Any],
        columns: list[str] | None
    ) -> tuple[list[dict], list[str]]:
        """Get absences data."""
        query = select(Absence).options(joinedload(Absence.person)).order_by(Absence.start_date)

        # Apply filters
        if filters.get("start_date"):
            start = datetime.fromisoformat(filters["start_date"]).date()
            query = query.where(Absence.end_date >= start)
        if filters.get("end_date"):
            end = datetime.fromisoformat(filters["end_date"]).date()
            query = query.where(Absence.start_date <= end)
        if filters.get("absence_type"):
            query = query.where(Absence.absence_type == filters["absence_type"])

        result = await self.db.execute(query)
        absences = result.scalars().all()

        headers = ["person_name", "absence_type", "start_date", "end_date", "notes"]
        if columns:
            headers = [h for h in headers if h in columns]

        data = []
        for absence in absences:
            row = {
                "person_name": absence.person.name if absence.person else "Unknown",
                "absence_type": absence.absence_type,
                "start_date": absence.start_date.isoformat() if absence.start_date else None,
                "end_date": absence.end_date.isoformat() if absence.end_date else None,
                "notes": absence.notes,
            }
            if columns:
                row = {k: v for k, v in row.items() if k in columns}
            data.append(row)

        return data, headers

    async def _get_schedule_data(
        self,
        filters: dict[str, Any],
        columns: list[str] | None
    ) -> tuple[list[dict], list[str]]:
        """Get schedule data."""
        query = (
            select(Assignment)
            .options(
                joinedload(Assignment.block),
                joinedload(Assignment.person),
                joinedload(Assignment.rotation_template)
            )
            .join(Block)
            .order_by(Block.date, Block.time_of_day)
        )

        # Apply filters
        if filters.get("start_date"):
            start = datetime.fromisoformat(filters["start_date"]).date()
            query = query.where(Block.date >= start)
        if filters.get("end_date"):
            end = datetime.fromisoformat(filters["end_date"]).date()
            query = query.where(Block.date <= end)

        result = await self.db.execute(query)
        assignments = result.scalars().all()

        headers = ["date", "time_of_day", "person_name", "person_type", "role", "activity"]
        if columns:
            headers = [h for h in headers if h in columns]

        data = []
        for assignment in assignments:
            row = {
                "date": assignment.block.date.isoformat() if assignment.block else None,
                "time_of_day": assignment.block.time_of_day if assignment.block else None,
                "person_name": assignment.person.name if assignment.person else "Unknown",
                "person_type": assignment.person.type if assignment.person else None,
                "role": assignment.role,
                "activity": assignment.activity_name,
            }
            if columns:
                row = {k: v for k, v in row.items() if k in columns}
            data.append(row)

        return data, headers

    async def _get_certifications_data(
        self,
        filters: dict[str, Any],
        columns: list[str] | None
    ) -> tuple[list[dict], list[str]]:
        """Get certifications data."""
        query = select(PersonCertification).options(
            joinedload(PersonCertification.person),
            joinedload(PersonCertification.certification_type)
        )

        result = await self.db.execute(query)
        certs = result.scalars().all()

        headers = ["person_name", "certification_name", "expiration_date", "status"]
        if columns:
            headers = [h for h in headers if h in columns]

        data = []
        for cert in certs:
            row = {
                "person_name": cert.person.name if cert.person else "Unknown",
                "certification_name": cert.certification_type.name if cert.certification_type else "Unknown",
                "expiration_date": cert.expiration_date.isoformat() if cert.expiration_date else None,
                "status": "Active" if cert.expiration_date and cert.expiration_date > date.today() else "Expired",
            }
            if columns:
                row = {k: v for k, v in row.items() if k in columns}
            data.append(row)

        return data, headers

    async def _get_swap_history_data(
        self,
        filters: dict[str, Any],
        columns: list[str] | None
    ) -> tuple[list[dict], list[str]]:
        """Get swap history data."""
        query = select(SwapRecord).order_by(SwapRecord.created_at.desc())

        result = await self.db.execute(query)
        swaps = result.scalars().all()

        headers = ["swap_type", "status", "requested_date", "created_at", "executed_at"]
        if columns:
            headers = [h for h in headers if h in columns]

        data = []
        for swap in swaps:
            row = {
                "swap_type": swap.swap_type,
                "status": swap.status,
                "requested_date": swap.requested_date.isoformat() if swap.requested_date else None,
                "created_at": swap.created_at.isoformat() if swap.created_at else None,
                "executed_at": swap.executed_at.isoformat() if swap.executed_at else None,
            }
            if columns:
                row = {k: v for k, v in row.items() if k in columns}
            data.append(row)

        return data, headers

    def _format_as_csv(
        self,
        data: list[dict],
        headers: list[str],
        include_headers: bool
    ) -> bytes:
        """Format data as CSV."""
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=headers)

        if include_headers:
            writer.writeheader()

        for row in data:
            writer.writerow(row)

        return output.getvalue().encode("utf-8")

    def _format_as_json(self, data: list[dict], headers: list[str]) -> bytes:
        """Format data as JSON."""
        return json.dumps(data, indent=2, default=str).encode("utf-8")

    def _format_as_xlsx(
        self,
        data: list[dict],
        headers: list[str],
        include_headers: bool
    ) -> bytes:
        """Format data as XLSX."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Export"

        # Write headers
        if include_headers:
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_idx, value=header)
            start_row = 2
        else:
            start_row = 1

        # Write data
        for row_idx, row_data in enumerate(data, start_row):
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=row_data.get(header))

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def _format_as_xml(self, data: list[dict], headers: list[str]) -> bytes:
        """Format data as XML."""
        from xml.etree.ElementTree import Element, SubElement, tostring

        root = Element("export")
        for row in data:
            item = SubElement(root, "item")
            for header in headers:
                field = SubElement(item, header)
                field.text = str(row.get(header, ""))

        return tostring(root, encoding="utf-8")

    def _calculate_next_run_time(self, cron_expression: str) -> datetime:
        """
        Calculate next run time from cron expression.

        Args:
            cron_expression: Cron expression

        Returns:
            datetime: Next run time
        """
        try:
            cron = croniter(cron_expression, datetime.utcnow())
            return cron.get_next(datetime)
        except Exception as e:
            logger.error(f"Invalid cron expression: {cron_expression}: {e}")
            # Default to 1 day from now
            return datetime.utcnow() + timedelta(days=1)
