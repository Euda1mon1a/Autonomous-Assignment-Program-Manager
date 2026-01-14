"""
ROSETTA Parser - Extract ground truth from Block10_ROSETTA_CORRECT.xlsx.

Parses the ROSETTA xlsx file (the CORRECT Block 10 schedule) and returns
structured assignment data that can be used as pytest fixtures for TDD.

The ROSETTA file is the "answer key" - expansion service output must match it.
"""

from datetime import date, timedelta
from pathlib import Path
from typing import NamedTuple

from openpyxl import load_workbook


class RosettaAssignment(NamedTuple):
    """A single ground truth assignment from ROSETTA."""

    resident: str
    pgy: int
    rotation1: str
    rotation2: str | None
    date: date
    time_of_day: str  # "AM" or "PM"
    expected_code: str  # "C", "KAP", "LEC", etc.


def parse_rosetta(xlsx_path: Path | str) -> list[RosettaAssignment]:
    """
    Extract ground truth assignments from ROSETTA xlsx.

    Args:
        xlsx_path: Path to Block10_ROSETTA_CORRECT.xlsx

    Returns:
        List of RosettaAssignment namedtuples, one per cell in the schedule

    Structure of ROSETTA xlsx:
        Row 1: Headers
        Rows 2-10: Resident data
        Column A: Resident name
        Column B: PGY level
        Column C: Primary rotation
        Column D: Secondary rotation (mid-block transition)
        Column E: Notes
        Columns F-BI: Schedule slots (56 half-days: 28 days * 2)
    """
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        raise FileNotFoundError(f"ROSETTA file not found: {xlsx_path}")

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    assignments: list[RosettaAssignment] = []
    start_date = date(2026, 3, 12)  # Block 10 start (Thu Mar 12, 2026)

    # Parse rows 2-10 (data rows, row 1 is header)
    for row in range(2, 11):
        resident = ws.cell(row=row, column=1).value  # Column A
        if not resident:
            continue

        pgy_value = ws.cell(row=row, column=2).value  # Column B
        pgy = int(pgy_value) if pgy_value else 0

        rot1 = ws.cell(row=row, column=3).value or ""  # Column C
        rot2_value = ws.cell(row=row, column=4).value  # Column D
        rot2 = str(rot2_value).strip() if rot2_value else None

        # Parse columns F-BI (columns 6-61, 56 half-day slots)
        for col in range(6, 62):
            code = ws.cell(row=row, column=col).value
            if not code:
                continue

            # Column offset from F (col 6)
            col_offset = col - 6
            day_offset = col_offset // 2
            is_am = col_offset % 2 == 0
            slot_date = start_date + timedelta(days=day_offset)

            assignments.append(
                RosettaAssignment(
                    resident=str(resident).strip(),
                    pgy=pgy,
                    rotation1=str(rot1).strip(),
                    rotation2=rot2,
                    date=slot_date,
                    time_of_day="AM" if is_am else "PM",
                    expected_code=str(code).strip(),
                )
            )

    wb.close()
    return assignments


def get_rosetta_by_resident(
    assignments: list[RosettaAssignment],
) -> dict[str, list[RosettaAssignment]]:
    """Group ROSETTA assignments by resident name."""
    result: dict[str, list[RosettaAssignment]] = {}
    for a in assignments:
        result.setdefault(a.resident, []).append(a)
    return result


def get_rosetta_by_date(
    assignments: list[RosettaAssignment],
) -> dict[date, list[RosettaAssignment]]:
    """Group ROSETTA assignments by date."""
    result: dict[date, list[RosettaAssignment]] = {}
    for a in assignments:
        result.setdefault(a.date, []).append(a)
    return result


def get_wednesday_am_assignments(
    assignments: list[RosettaAssignment],
    pgy_filter: int | None = None,
) -> list[RosettaAssignment]:
    """Get all Wednesday AM assignments (for testing intern continuity rule)."""
    result = [
        a
        for a in assignments
        if a.date.weekday() == 2  # Wednesday
        and a.time_of_day == "AM"
    ]
    if pgy_filter is not None:
        result = [a for a in result if a.pgy == pgy_filter]
    return result


def get_last_wednesday_assignments(
    assignments: list[RosettaAssignment],
) -> list[RosettaAssignment]:
    """Get assignments for last Wednesday (Apr 8) - should be LEC/ADV."""
    last_wed = date(2026, 4, 8)
    return [a for a in assignments if a.date == last_wed]


# Default path to ROSETTA file
DEFAULT_ROSETTA_PATH = (
    Path(__file__).parent.parent.parent.parent.parent
    / "docs"
    / "scheduling"
    / "Block10_ROSETTA_CORRECT.xlsx"
)


def load_rosetta() -> list[RosettaAssignment]:
    """Load ROSETTA from default path."""
    return parse_rosetta(DEFAULT_ROSETTA_PATH)
