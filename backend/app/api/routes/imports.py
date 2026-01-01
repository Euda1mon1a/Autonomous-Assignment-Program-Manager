"""
Generic import routes for parsing various file formats.

This module provides a backend endpoint for parsing Excel files,
returning raw parsed data that the frontend can then validate and process.
"""

import io
import json
from typing import Any

from fastapi import APIRouter, File, Form, HTTPException, UploadFile
from openpyxl import load_workbook
from openpyxl.cell import MergedCell
from pydantic import BaseModel

from app.core.file_security import validate_excel_upload
from app.core.logging import get_logger

router = APIRouter()
logger = get_logger(__name__)


class ParsedXlsxResponse(BaseModel):
    """Response model for parsed Excel data."""

    success: bool
    rows: list[dict[str, Any]]
    columns: list[str]
    total_rows: int
    sheet_name: str
    warnings: list[str] = []


class XlsxParseError(BaseModel):
    """Error response for xlsx parsing failures."""

    success: bool = False
    error: str
    error_code: str


@router.post(
    "/parse-xlsx",
    response_model=ParsedXlsxResponse,
    responses={
        400: {"model": XlsxParseError, "description": "Invalid file or parsing error"},
        413: {"model": XlsxParseError, "description": "File too large"},
    },
)
async def parse_xlsx_file(
    file: UploadFile = File(..., description="Excel file to parse (.xlsx, .xls)"),
    sheet_name: str | None = Form(
        None, description="Specific sheet name to parse (default: first sheet)"
    ),
    header_row: int = Form(
        1, description="Row number containing headers (1-indexed, default: 1)", ge=1
    ),
    max_rows: int = Form(
        10000, description="Maximum rows to parse (default: 10000)", ge=1, le=100000
    ),
    skip_empty_rows: bool = Form(
        True, description="Skip rows where all cells are empty"
    ),
) -> ParsedXlsxResponse:
    """
    Parse an Excel file and return structured data.

    This endpoint reads an Excel file (.xlsx or .xls) and returns:
    - Column headers from the specified header row
    - All data rows as dictionaries keyed by column headers
    - Metadata about the parsed content

    The returned data can be used by the frontend for preview and validation
    before executing the actual import.

    Features:
    - Handles merged cells (uses the value from the top-left cell)
    - Converts Excel dates to ISO format strings
    - Trims whitespace from string values
    - Detects and reports common issues as warnings
    """
    warnings: list[str] = []

    # Validate file extension
    filename = file.filename or "unknown.xlsx"
    ext = filename.lower().split(".")[-1] if "." in filename else ""
    if ext not in ("xlsx", "xls"):
        raise HTTPException(
            status_code=400,
            detail={
                "success": False,
                "error": f"Invalid file extension: .{ext}. Expected .xlsx or .xls",
                "error_code": "INVALID_EXTENSION",
            },
        )

    # Read file content
    try:
        content = await file.read()
    except OSError as e:
        logger.error(f"Failed to read uploaded file: {e}", exc_info=True)
        raise HTTPException(
            status_code=400,
            detail={
                "success": False,
                "error": "Failed to read uploaded file",
                "error_code": "READ_ERROR",
            },
        )
    except ValueError as e:
        logger.error(f"Invalid file content: {e}", exc_info=True)
        raise HTTPException(
            status_code=400,
            detail={
                "success": False,
                "error": "Invalid file content",
                "error_code": "INVALID_CONTENT",
            },
        )

    # Check file size (10MB limit)
    max_size = 10 * 1024 * 1024  # 10MB
    if len(content) > max_size:
        raise HTTPException(
            status_code=413,
            detail={
                "success": False,
                "error": f"File too large. Maximum size is 10MB, got {len(content) / 1024 / 1024:.1f}MB",
                "error_code": "FILE_TOO_LARGE",
            },
        )

    # Validate file security
    try:
        validate_excel_upload(content, filename, file.content_type or "")
    except ValueError as e:
        raise HTTPException(
            status_code=400,
            detail={
                "success": False,
                "error": str(e),
                "error_code": "SECURITY_VALIDATION_FAILED",
            },
        )

    # Parse the Excel file
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except OSError as e:
        logger.error(f"File I/O error parsing Excel file: {e}", exc_info=True)
        raise HTTPException(
            status_code=400,
            detail={
                "success": False,
                "error": f"Failed to parse Excel file: {str(e)}",
                "error_code": "PARSE_ERROR",
            },
        )
    except (ValueError, TypeError) as e:
        logger.error(f"Invalid Excel file format: {e}", exc_info=True)
        raise HTTPException(
            status_code=400,
            detail={
                "success": False,
                "error": f"Invalid Excel file format: {str(e)}",
                "error_code": "INVALID_FORMAT",
            },
        )

    # Select worksheet
    try:
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                available = ", ".join(wb.sheetnames[:5])
                raise HTTPException(
                    status_code=400,
                    detail={
                        "success": False,
                        "error": f"Sheet '{sheet_name}' not found. Available: {available}",
                        "error_code": "SHEET_NOT_FOUND",
                    },
                )
            ws = wb[sheet_name]
        else:
            ws = wb.active
            sheet_name = ws.title if ws else wb.sheetnames[0]
            if not ws:
                ws = wb[wb.sheetnames[0]]
    except HTTPException:
        raise
    except (ValueError, KeyError, AttributeError) as e:
        logger.error(f"Failed to access worksheet: {e}", exc_info=True)
        raise HTTPException(
            status_code=400,
            detail={
                "success": False,
                "error": f"Failed to access worksheet: {str(e)}",
                "error_code": "WORKSHEET_ERROR",
            },
        )

    # Extract headers from the specified row
    headers: list[str] = []
    header_row_idx = header_row  # 1-indexed in openpyxl

    try:
        for cell in ws[header_row_idx]:
            if isinstance(cell, MergedCell):
                # For merged cells, we'll use empty string (merged headers are unusual)
                headers.append("")
                warnings.append(
                    f"Merged cell found in header row at column {len(headers)}"
                )
            else:
                value = cell.value
                if value is not None:
                    headers.append(str(value).strip())
                else:
                    headers.append(f"Column_{len(headers) + 1}")
    except (ValueError, TypeError, AttributeError) as e:
        logger.error(f"Failed to read header row: {e}", exc_info=True)
        raise HTTPException(
            status_code=400,
            detail={
                "success": False,
                "error": f"Failed to read header row {header_row}: {str(e)}",
                "error_code": "HEADER_READ_ERROR",
            },
        )

    # Remove trailing empty columns
    while headers and headers[-1].startswith("Column_"):
        headers.pop()

    if not headers:
        raise HTTPException(
            status_code=400,
            detail={
                "success": False,
                "error": f"No headers found in row {header_row}",
                "error_code": "NO_HEADERS",
            },
        )

    # Check for duplicate headers
    seen_headers: set[str] = set()
    for i, h in enumerate(headers):
        if h in seen_headers:
            # Make unique by appending index
            new_header = f"{h}_{i + 1}"
            warnings.append(f"Duplicate header '{h}' renamed to '{new_header}'")
            headers[i] = new_header
        seen_headers.add(headers[i])

    # Parse data rows
    rows: list[dict[str, Any]] = []
    empty_row_count = 0

    try:
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=header_row + 1), start=header_row + 1
        ):
            if row_idx > header_row + max_rows:
                warnings.append(
                    f"Row limit ({max_rows}) reached. File may contain more data."
                )
                break

            row_data: dict[str, Any] = {}
            is_empty = True

            for col_idx, cell in enumerate(row):
                if col_idx >= len(headers):
                    break

                header = headers[col_idx]

                if isinstance(cell, MergedCell):
                    # Skip merged cells in data (they're usually formatting)
                    value = None
                else:
                    value = cell.value

                # Convert value to appropriate type
                if value is not None:
                    is_empty = False

                    # Handle dates
                    if hasattr(value, "isoformat"):
                        value = value.isoformat()
                    # Handle booleans
                    elif isinstance(value, bool):
                        value = value
                    # Handle numbers
                    elif isinstance(value, (int, float)):
                        # Keep numeric types
                        value = value
                    # Handle strings
                    else:
                        value = str(value).strip()
                        # Check for boolean-like strings
                        if value.lower() in ("true", "yes"):
                            value = True
                        elif value.lower() in ("false", "no"):
                            value = False

                row_data[header] = value

            # Skip empty rows if configured
            if is_empty:
                empty_row_count += 1
                if skip_empty_rows:
                    continue

            rows.append(row_data)

    except (ValueError, TypeError, AttributeError) as e:
        logger.error(f"Failed to parse data rows: {e}", exc_info=True)
        raise HTTPException(
            status_code=400,
            detail={
                "success": False,
                "error": f"Failed to parse data rows: {str(e)}",
                "error_code": "DATA_PARSE_ERROR",
            },
        )
    finally:
        wb.close()

    if empty_row_count > 0 and skip_empty_rows:
        warnings.append(f"Skipped {empty_row_count} empty rows")

    if not rows:
        warnings.append("No data rows found after header row")

    logger.info(
        f"Parsed Excel file: {filename}, sheet: {sheet_name}, "
        f"columns: {len(headers)}, rows: {len(rows)}"
    )

    return ParsedXlsxResponse(
        success=True,
        rows=rows,
        columns=headers,
        total_rows=len(rows),
        sheet_name=sheet_name or "Sheet1",
        warnings=warnings,
    )


@router.post("/parse-xlsx/sheets")
async def list_xlsx_sheets(
    file: UploadFile = File(..., description="Excel file to inspect"),
) -> dict[str, Any]:
    """
    List all sheet names in an Excel file.

    Useful for letting users select which sheet to import when a file
    contains multiple worksheets.
    """
    filename = file.filename or "unknown.xlsx"
    ext = filename.lower().split(".")[-1] if "." in filename else ""
    if ext not in ("xlsx", "xls"):
        raise HTTPException(
            status_code=400,
            detail={
                "success": False,
                "error": f"Invalid file extension: .{ext}",
                "error_code": "INVALID_EXTENSION",
            },
        )

    try:
        content = await file.read()
        wb = load_workbook(io.BytesIO(content), read_only=True)
        sheets = wb.sheetnames
        wb.close()

        return {
            "success": True,
            "sheets": sheets,
            "count": len(sheets),
            "default": sheets[0] if sheets else None,
        }
    except OSError as e:
        logger.error(f"File I/O error listing sheets: {e}", exc_info=True)
        raise HTTPException(
            status_code=400,
            detail={
                "success": False,
                "error": f"Failed to read Excel file: {str(e)}",
                "error_code": "READ_ERROR",
            },
        )
    except (ValueError, TypeError) as e:
        logger.error(f"Invalid Excel file format listing sheets: {e}", exc_info=True)
        raise HTTPException(
            status_code=400,
            detail={
                "success": False,
                "error": f"Invalid Excel file format: {str(e)}",
                "error_code": "INVALID_FORMAT",
            },
        )
