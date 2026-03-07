"""
XML to XLSX Converter (legacy/validation).

Converts schedule XML to Excel format.
Canonical export now uses JSON → JSONToXlsxConverter, but XML remains
for validation and legacy tooling.

Supports two formats:
1. ROSETTA format (validation): Simple layout for testing
2. Block Template2 format (production): Full TAMC layout with all rows/columns

Legacy pipeline:
    DB → HalfDayXMLExporter → XML → XMLToXlsxConverter → xlsx
"""

from __future__ import annotations

from collections import Counter
from copy import copy
from datetime import date, datetime, timedelta, UTC
from io import BytesIO
from pathlib import Path
from typing import Any
from uuid import UUID
from defusedxml import ElementTree

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.core.logging import get_logger
from app.services.tamc_color_scheme import get_color_scheme
from app.utils.academic_blocks import get_block_number_for_date

logger = get_logger(__name__)


# ═══════════════════════════════════════════════════════════════════════════════
# ROSETTA FORMAT (validation-focused, simple layout)
# ═══════════════════════════════════════════════════════════════════════════════
# Column Layout: Name, PGY, Rotation1, Rotation2, Notes, Schedule...
ROSETTA_COL_NAME = 1
ROSETTA_COL_PGY = 2
ROSETTA_COL_ROTATION1 = 3
ROSETTA_COL_ROTATION2 = 4
ROSETTA_COL_NOTES = 5
ROSETTA_COL_SCHEDULE_START = 6

# ═══════════════════════════════════════════════════════════════════════════════
# BLOCK TEMPLATE2 FORMAT (production, full TAMC layout)
# ═══════════════════════════════════════════════════════════════════════════════
# Column Layout: Rotation1, Rotation2, Template, Role, Name, Schedule...
BT2_COL_ROTATION1 = 1  # First-half rotation (e.g., "Hilo", "FMC")
BT2_COL_ROTATION2 = 2  # Second-half rotation (if mid-block switch)
BT2_COL_TEMPLATE = 3  # Template code (R1, R2, R3, C19, ADJ)
BT2_COL_ROLE = 4  # Role (PGY 1, PGY 2, PGY 3, FAC)
BT2_COL_NAME = 5  # Name ("Last, First")
BT2_COL_SCHEDULE_START = 6  # First schedule column

# Special rows in Block Template2
BT2_ROW_STAFF_CALL = 4
BT2_ROW_RESIDENT_CALL = 5

# Date-to-column calculation (same for both formats)
# Each date uses 2 columns: AM (even col), PM (odd col+1)
# Block 10: Mar 12-Apr 8 = 28 days = 56 columns (6-61)
COLS_PER_DAY = 2
TOTAL_DAYS = 28

# Legacy aliases for backward compatibility
ROW_HEADERS = 1
COL_RESIDENT_NAME = ROSETTA_COL_NAME
COL_PGY = ROSETTA_COL_PGY
COL_ROTATION1 = ROSETTA_COL_ROTATION1
COL_ROTATION2 = ROSETTA_COL_ROTATION2
COL_NOTES = ROSETTA_COL_NOTES
COL_SCHEDULE_START = ROSETTA_COL_SCHEDULE_START
COL_SCHEDULE_END = COL_SCHEDULE_START + (TOTAL_DAYS * COLS_PER_DAY) - 1  # 61


class XMLToXlsxConverter:
    """
    Convert HalfDayXMLExporter XML to Excel xlsx format.

    Supports two formats:
    1. ROSETTA format (validation): Simple layout - Name, PGY, Rotation, Schedule
    2. Block Template2 format (production): Full TAMC layout with correct columns

    Block Template2 format uses:
    - Column layout: Rotation1, Rotation2, Template, Role, Name, Schedule...
    - Row mappings from structure XML for exact positioning
    - Name format: "Last, First"
    """

    def __init__(
        self,
        template_path: Path | str | None = None,
        apply_colors: bool = True,
        structure_xml_path: Path | str | None = None,
        use_block_template2: bool = True,
        strict_row_mapping: bool = False,
        include_qa_sheet: bool = True,
        preserve_template_identity_fields: bool = True,
        presentation_profile: str = "tamc_handjam_v2",
    ) -> None:
        """
        Initialize converter with optional template.

        Args:
            template_path: Path to Excel template (e.g., Block_Template2.xlsx).
                          If not provided, creates new workbook.
            apply_colors: Whether to apply TAMC color scheme to cells.
            structure_xml_path: Path to BlockTemplate2_Structure.xml for row mappings.
                               If provided, uses name → row lookup instead of sequential.
            use_block_template2: If True, use Block Template2 column layout (production).
                                If False, use ROSETTA layout (validation).
            strict_row_mapping: If True, fail export when a person name has no row mapping.
            include_qa_sheet: If True, add an Export_QA worksheet with explicit
                code totals and per-faculty clinic counts.
            preserve_template_identity_fields: If True, keeps template values in
                columns A-E (rotations/template/role/name) and only writes schedule
                cells. This preserves full handjam labels/names.
            presentation_profile: Post-processing profile for summary formulas and
                freeze panes. "tamc_handjam_v2" aligns output with manual Block 10.
        """
        self.template_path = Path(template_path) if template_path else None
        self.structure_xml_path = (
            Path(structure_xml_path) if structure_xml_path else None
        )
        self.apply_colors = apply_colors
        self.color_scheme = get_color_scheme() if apply_colors else None
        self.use_block_template2 = use_block_template2
        self.strict_row_mapping = strict_row_mapping
        self.include_qa_sheet = include_qa_sheet
        self.preserve_template_identity_fields = preserve_template_identity_fields
        self.presentation_profile = presentation_profile

        # Row mappings: key is person name (legacy XML) or person UUID (dynamic).
        # Value is always the Excel row number.
        self.row_mappings: dict[str, int] = {}
        self.pgy_mappings: dict[str, int] = {}  # key → pgy level
        self.template_mappings: dict[str, str] = {}  # key → template code
        self.call_row: int = BT2_ROW_STAFF_CALL
        self.resident_call_row: int = BT2_ROW_RESIDENT_CALL
        self.last_conversion_stats: dict[str, Any] = {}
        if structure_xml_path:
            self._load_structure_xml(Path(structure_xml_path))

    def _load_structure_xml(self, xml_path: Path) -> None:
        """Load row mappings from BlockTemplate2_Structure.xml."""
        if not xml_path.exists():
            logger.warning(f"Structure XML not found: {xml_path}")
            return

        tree = ElementTree.parse(xml_path)
        root = tree.getroot()

        # Load layout settings
        layout = root.find("layout")
        if layout is not None:
            call_row_elem = layout.find("call_row")
            if call_row_elem is not None:
                self.call_row = int(call_row_elem.get("row", "4"))

                # Load resident row mappings (name → row, pgy, template)
        for resident in root.findall(".//resident"):
            name = resident.get("name", "")
            row = resident.get("row", "")
            pgy = resident.get("pgy", "")
            if name and row:
                # Normalize name (remove asterisks, extra spaces)
                normalized = name.replace("*", "").strip()
                self.row_mappings[normalized] = int(row)
                if pgy:
                    self.pgy_mappings[normalized] = int(pgy)
                    # Template code from PGY: R1, R2, R3
                    self.template_mappings[normalized] = f"R{pgy}"

                    # Load faculty row mappings
        for person in root.findall(".//faculty/person"):
            name = person.get("name", "")
            row = person.get("row", "")
            template = person.get("template", "C19")
            if name and row:
                normalized = name.replace("*", "").strip()
                self.row_mappings[normalized] = int(row)
                self.template_mappings[normalized] = template

        logger.info(f"Loaded {len(self.row_mappings)} row mappings from {xml_path}")

    # ─── Fat-Band Dynamic Mapping (UUID-keyed) ─────────────────────────
    # Resident band: rows 9-30 (22 slots)
    _FAT_BAND_RESIDENT_START = 9
    _FAT_BAND_RESIDENT_END = 30
    # Faculty band: rows 31-80 (50 slots)
    _FAT_BAND_FACULTY_START = 31
    _FAT_BAND_FACULTY_END = 80

    def _build_dynamic_mappings(self, data: dict[str, Any], sheet: Any) -> None:
        """Dynamically assign Excel rows from JSON data, keyed by person UUID.

        Replaces static Structure XML with in-memory row allocation:
        - Residents sorted by PGY desc (3→2→1), then alphabetically
        - Faculty sorted alphabetically
        - Unused rows in the Fat Band are hidden for clean presentation
        """
        self.row_mappings = {}
        self.template_mappings = {}
        self.pgy_mappings = {}

        # --- Residents (rows 9-30) ---
        residents = data.get("residents", []) or []

        def _pgy_sort_key(r: dict[str, Any]) -> tuple[int, str]:
            pgy_val = r.get("pgy")
            pgy_int = int(pgy_val) if pgy_val else 0
            return (-pgy_int, r.get("name", ""))

        residents_sorted = sorted(residents, key=_pgy_sort_key)

        used_res_rows: set[int] = set()
        current_row = self._FAT_BAND_RESIDENT_START
        for res in residents_sorted:
            pid = str(res.get("id", ""))
            if not pid:
                continue
            if current_row > self._FAT_BAND_RESIDENT_END:
                raise ValueError(
                    f"Resident band overflow: {len(residents_sorted)} residents "
                    f"exceed {self._FAT_BAND_RESIDENT_END - self._FAT_BAND_RESIDENT_START + 1} "
                    f"available slots (rows {self._FAT_BAND_RESIDENT_START}-"
                    f"{self._FAT_BAND_RESIDENT_END}). "
                    f"First skipped: {res.get('name', '?')}"
                )
            self.row_mappings[pid] = current_row
            pgy = res.get("pgy") or 1
            self.template_mappings[pid] = f"R{pgy}"
            self.pgy_mappings[pid] = int(pgy)
            used_res_rows.add(current_row)
            current_row += 1

        # --- Faculty (rows 31-80) --- core first, adjunct below, alphabetical within
        faculty = data.get("faculty", []) or []
        faculty_sorted = sorted(
            faculty,
            key=lambda f: (
                1 if f.get("faculty_role") == "adjunct" else 0,
                f.get("name", ""),
            ),
        )

        used_fac_rows: set[int] = set()
        current_row = self._FAT_BAND_FACULTY_START
        for fac in faculty_sorted:
            pid = str(fac.get("id", ""))
            if not pid:
                continue
            if current_row > self._FAT_BAND_FACULTY_END:
                raise ValueError(
                    f"Faculty band overflow: {len(faculty_sorted)} faculty "
                    f"exceed {self._FAT_BAND_FACULTY_END - self._FAT_BAND_FACULTY_START + 1} "
                    f"available slots (rows {self._FAT_BAND_FACULTY_START}-"
                    f"{self._FAT_BAND_FACULTY_END}). "
                    f"First skipped: {fac.get('name', '?')}"
                )
            self.row_mappings[pid] = current_row
            self.template_mappings[pid] = fac.get("template", "C19")
            used_fac_rows.add(current_row)
            current_row += 1

        # --- Hide unused rows for clean presentation ---
        for r in range(self._FAT_BAND_RESIDENT_START, self._FAT_BAND_RESIDENT_END + 1):
            if r not in used_res_rows:
                sheet.row_dimensions[r].hidden = True
        for r in range(self._FAT_BAND_FACULTY_START, self._FAT_BAND_FACULTY_END + 1):
            if r not in used_fac_rows:
                sheet.row_dimensions[r].hidden = True

        logger.info(
            "Dynamic mappings: %d residents (rows %d-%d), %d faculty (rows %d-%d)",
            len(used_res_rows),
            self._FAT_BAND_RESIDENT_START,
            max(used_res_rows) if used_res_rows else 0,
            len(used_fac_rows),
            self._FAT_BAND_FACULTY_START,
            max(used_fac_rows) if used_fac_rows else 0,
        )

    def convert_from_string(
        self,
        xml_string: str,
        output_path: Path | str | None = None,
    ) -> bytes:
        """
        Convert XML schedule string to xlsx.

        Args:
            xml_string: XML from ScheduleXMLExporter
            output_path: Optional path to save xlsx

        Returns:
            xlsx bytes (openpyxl workbook)
        """
        root = ElementTree.fromstring(xml_string)
        data = self._parse_xml_to_data(root)
        return self.convert_from_data(data, output_path=output_path)

    def convert_from_data(
        self,
        data: dict[str, Any],
        output_path: Path | str | None = None,
        export_metadata: Any | None = None,
        rotation_codes: list[str] | None = None,
        activity_codes: list[str] | None = None,
    ) -> bytes:
        """Convert schedule data dict to xlsx.

        Args:
            data: Schedule data dictionary
            output_path: Optional path to save xlsx
            export_metadata: Optional ExportMetadata for __SYS_META__ (Phase 1)
            rotation_codes: Optional rotation codes for __REF__ (Phase 1)
            activity_codes: Optional activity codes for __REF__ (Phase 1)
        """
        block_start = self._coerce_date(data.get("block_start"))
        block_end = self._coerce_date(data.get("block_end"))
        if not block_start or not block_end:
            raise ValueError("Missing block_start/block_end in schedule data")

        logger.info(f"Converting schedule data to xlsx: {block_start} to {block_end}")

        # Load template or create new workbook
        if self.template_path and self.template_path.exists():
            wb = load_workbook(self.template_path)
            # Use "Block Template2" sheet if available, else active
            if "Block Template2" in wb.sheetnames:
                sheet = wb["Block Template2"]
            else:
                sheet = wb.active
            logger.info(f"Loaded template from {self.template_path}")
        else:
            # Create new workbook — BT2 layout if enabled, else ROSETTA
            wb = self._create_new_workbook(block_start, block_end)
            sheet = wb.active
            if self.use_block_template2:
                sheet.title = "Block Template2"

        # Dynamic UUID-based row allocation when no structure XML was loaded.
        # Always rebuild per call so year-export gets fresh mappings per block.
        if self.use_block_template2 and not self.structure_xml_path:
            self._build_dynamic_mappings(data, sheet)

        # Fill header row (always - helps with readability)
        self._fill_header_row(sheet, block_start, block_end)
        unmerged_cells = self._prepare_schedule_grid(sheet, block_start, block_end)

        # Fill residents and faculty
        residents = data.get("residents", [])
        faculty = data.get("faculty", [])
        stats: dict[str, Any] = {
            "block_start": block_start.isoformat(),
            "block_end": block_end.isoformat(),
            "residents": {
                "people_input": len(residents),
                "people_written": 0,
                "unmapped_names": [],
                "codes_written": {},
            },
            "faculty": {
                "people_input": len(faculty),
                "people_written": 0,
                "unmapped_names": [],
                "codes_written": {},
            },
        }
        stats["schedule_cells_unmerged"] = unmerged_cells
        self._fill_residents(
            sheet,
            residents,
            block_start,
            is_faculty=False,
            stats=stats["residents"],
        )
        self._fill_residents(
            sheet,
            faculty,
            block_start,
            is_faculty=True,
            stats=stats["faculty"],
        )

        # Fill call row (Row 4) - single cells for user to merge
        call_section = data.get("call", {})
        call_rows: list[dict[str, Any]] = []
        if isinstance(call_section, dict):
            call_rows = call_section.get("nights", []) or []
        elif isinstance(call_section, list):
            call_rows = call_section
        if call_rows:
            self._fill_call_row(sheet, call_rows, block_start, block_end)
        stats["call_nights_input"] = len(call_rows)
        self._apply_presentation_profile(sheet, block_start, block_end)
        self.last_conversion_stats = stats

        # Spatial UUID Anchoring (Phase 2)
        if self.use_block_template2:
            self._write_anchor_sheet(wb, data)

        # Strict UI Contracts (Phase 3)
        if self.use_block_template2:
            self._add_data_validation(sheet, block_start, block_end)

        # Stateful Overlays (Phase 4)
        if self.use_block_template2:
            self._add_dynamic_cf(sheet, block_start, block_end)
            self._add_leave_formula_column(sheet, block_start, block_end)

        if self.include_qa_sheet and self.use_block_template2:
            self._write_export_qa_sheet(wb, data, stats)

        # Phase 1: Phantom Database (__SYS_META__ + __REF__) — single-save
        # Written here (not in a separate load/save cycle) to avoid stripping
        # DataValidation, conditional formatting, and veryHidden sheets.
        if export_metadata is not None:
            from app.services.excel_metadata import (
                write_sys_meta_sheet,
                write_ref_sheet,
            )

            write_sys_meta_sheet(wb, export_metadata)
            if rotation_codes is not None and activity_codes is not None:
                write_ref_sheet(wb, rotation_codes, activity_codes)

        if self.use_block_template2:
            self._prune_empty_sheets(wb, keep={"Block Template2", "Export_QA"})

        # ── Final presentation pass: match TAMC hand-jammed format ──
        if self.use_block_template2:
            main_sheet = (
                wb["Block Template2"]
                if "Block Template2" in wb.sheetnames
                else wb.active
            )
        else:
            main_sheet = wb.active
        actual_days = (block_end - block_start).days + 1
        schedule_end_col = BT2_COL_SCHEDULE_START + (actual_days * COLS_PER_DAY) - 1
        if self.use_block_template2:
            self._apply_tamc_presentation(main_sheet, schedule_end_col)

        # Apply Arial 16 base font to non-schedule sheets (QA, etc.)
        base_font = Font(name="Arial", size=16)
        for ws in wb.worksheets:
            if ws is main_sheet:
                continue
            for row in ws.iter_rows():
                for cell in row:
                    cell.font = base_font

        # Save to bytes
        buffer = BytesIO()
        wb.save(buffer)
        xlsx_bytes = buffer.getvalue()

        # Save to file if path provided
        if output_path:
            Path(output_path).write_bytes(xlsx_bytes)
            logger.info(f"Saved xlsx to {output_path}")

        return xlsx_bytes

    def _prepare_schedule_grid(self, sheet, block_start: date, block_end: date) -> int:
        """Unmerge schedule-grid cells in mapped rows so AM/PM can be written independently."""
        if not self.use_block_template2 or not self.row_mappings:
            return 0

        schedule_start_col = COL_SCHEDULE_START
        total_days = (block_end - block_start).days + 1
        schedule_end_col = schedule_start_col + (total_days * COLS_PER_DAY) - 1

        target_rows = set(self.row_mappings.values())
        if not target_rows:
            return 0

        merged_ranges = list(sheet.merged_cells.ranges)
        unmerged = 0

        for merged_range in merged_ranges:
            if merged_range.max_col < schedule_start_col:
                continue
            if merged_range.min_col > schedule_end_col:
                continue
            if merged_range.max_row < min(target_rows):
                continue
            if merged_range.min_row > max(target_rows):
                continue

            overlaps_target_row = any(
                merged_range.min_row <= row <= merged_range.max_row
                for row in target_rows
            )
            if not overlaps_target_row:
                continue

            top_left = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)

            # Preserve style/format from top-left before unmerge.
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    if row == merged_range.min_row and col == merged_range.min_col:
                        continue
                    cell = sheet.cell(row=row, column=col)
                    cell._style = copy(top_left._style)
                    cell.number_format = top_left.number_format
                    cell.protection = copy(top_left.protection)
                    cell.alignment = copy(top_left.alignment)

            sheet.unmerge_cells(str(merged_range))
            unmerged += 1

        if unmerged:
            logger.info(f"Unmerged {unmerged} schedule cell ranges for AM/PM fidelity")
        return unmerged

    def _normalize_code(self, value: Any) -> str:
        if value is None:
            return ""
        code = str(value).strip().upper()
        return code

    def _collect_people_code_counts(
        self, people: list[dict[str, Any]]
    ) -> list[tuple[str, Counter[str]]]:
        rows: list[tuple[str, Counter[str]]] = []
        for person in sorted(people, key=lambda p: str(p.get("name", ""))):
            name = str(person.get("name", "") or "").strip()
            if not name:
                continue
            counter: Counter[str] = Counter()
            for day in person.get("days", []) or []:
                am = self._normalize_code(day.get("am", ""))
                pm = self._normalize_code(day.get("pm", ""))
                if am:
                    counter[am] += 1
                if pm:
                    counter[pm] += 1
            rows.append((name, counter))
        return rows

    def _apply_presentation_profile(
        self, sheet, block_start: date, block_end: date
    ) -> None:
        """Apply output formatting profile overlays for Block Template2."""
        if not self.use_block_template2:
            return
        if self.presentation_profile != "tamc_handjam_v2":
            return
        self._apply_tamc_handjam_summary_layout(sheet, block_start, block_end)

    def _apply_tamc_handjam_summary_layout(
        self, sheet, block_start: date, block_end: date
    ) -> None:
        """Align summary headers/formulas/freeze panes with handjam Block 10."""
        # Keep navigation locked to schedule start while exposing summary rows.
        sheet.freeze_panes = "E8"

        summary_headers = [
            "C",
            "CC",
            "CV",
            "(C+CC+CV)",
            "NF",
            "CC",
            "PC/OFF",
            "LV",
            "FMIT",
        ]
        for idx, header in enumerate(summary_headers):
            sheet.cell(row=8, column=62 + idx).value = header

        for row in range(9, 29):
            c_terms = '{"C","C-I","SM"}' if row == 9 else '{"C","CV","C-I","SM"}'
            sheet.cell(
                row=row, column=62
            ).value = f"=SUMPRODUCT(COUNTIF(F{row}:BI{row}, {c_terms}))"
            sheet.cell(row=row, column=63).value = f'=COUNTIF(F{row}:BI{row}, "CC")'
            sheet.cell(row=row, column=64).value = f'=COUNTIF(F{row}:BI{row}, "CV")'
            sheet.cell(row=row, column=65).value = f"=BJ{row}+BK{row}+BL{row}"
            sheet.cell(row=row, column=66).value = f'=COUNTIF(F{row}:BI{row}, "NF")'
            sheet.cell(row=row, column=67).value = f'=COUNTIF(F{row}:BI{row}, "CC")'
            sheet.cell(
                row=row, column=68
            ).value = f'=SUMPRODUCT(COUNTIF(F{row}:BI{row}, {{"PC","OFF","W"}}))'
            sheet.cell(row=row, column=69).value = f'=COUNTIF(F{row}:BI{row}, "LV")'
            sheet.cell(row=row, column=70).value = f'=COUNTIF(F{row}:BI{row}, "FMIT")'

        sheet.cell(row=29, column=62).value = "=SUM(BJ9:BJ26)"
        sheet.cell(row=29, column=63).value = "=SUM(BK9:BK26)"
        sheet.cell(row=29, column=64).value = "=SUM(BL9:BL26)"
        sheet.cell(row=29, column=65).value = "=SUM(BM9:BM26)"
        sheet.cell(row=29, column=66).value = "=SUM(BN9:BN28)"
        sheet.cell(row=29, column=67).value = "=SUM(BO9:BO28)"
        sheet.cell(row=29, column=68).value = "=SUM(BP9:BP28)"
        sheet.cell(row=29, column=69).value = "=SUM(BQ9:BQ28)"
        sheet.cell(row=29, column=70).value = None

        row30_headers = [
            "C",
            "CC",
            "CV",
            "(C+CC+CV)",
            "AT",
            "ADM",
            "LV",
            "FMIT",
            "CALL",
        ]
        for idx, header in enumerate(row30_headers):
            sheet.cell(row=30, column=62 + idx).value = header
        # SUN header at col 72/BT (col 71/BS is LV Days)
        sheet.cell(row=30, column=72).value = "SUN"

        # Compute Sunday column letters for SUN call count
        from datetime import timedelta

        from openpyxl.utils import get_column_letter

        sunday_col_letters: list[str] = []
        num_days = (block_end - block_start).days + 1
        for day_offset in range(num_days):
            d = block_start + timedelta(days=day_offset)
            if d.weekday() == 6:  # Sunday
                am_col = COL_SCHEDULE_START + (day_offset * 2)
                sunday_col_letters.append(get_column_letter(am_col))
                sunday_col_letters.append(get_column_letter(am_col + 1))

        for row in range(31, 81):
            sheet.cell(
                row=row, column=62
            ).value = f'=SUMPRODUCT(COUNTIF(F{row}:BI{row}, {{"C","SM"}}))'
            sheet.cell(row=row, column=63).value = f'=COUNTIF(F{row}:BI{row}, "CC")'
            sheet.cell(row=row, column=64).value = f'=COUNTIF(F{row}:BI{row}, "CV")'
            sheet.cell(row=row, column=65).value = f"=BJ{row}+BK{row}+BL{row}"
            sheet.cell(
                row=row, column=66
            ).value = f'=SUMPRODUCT(COUNTIF(F{row}:BI{row}, {{"AT","PCAT","DO"}}))'
            sheet.cell(
                row=row, column=67
            ).value = f'=SUMPRODUCT(COUNTIF(F{row}:BI{row}, {{"GME","DFM","DOFM"}}))'
            sheet.cell(row=row, column=68).value = f'=COUNTIF(F{row}:BI{row}, "LV")'
            sheet.cell(row=row, column=69).value = f'=COUNTIF(F{row}:BI{row}, "FMIT")'
            call_name = self._call_last_name_token(sheet.cell(row=row, column=5).value)
            sheet.cell(row=row, column=70).value = (
                f'=COUNTIF(F4:BI4, "{call_name}")' if call_name else None
            )
            # SUN column (col 72/BT): count calls on Sunday columns only
            # Col 71/BS is reserved for LV Days (Phase 4 leave formula)
            if call_name and sunday_col_letters:
                sun_terms = "+".join(
                    f'COUNTIF({c}4,"{call_name}")' for c in sunday_col_letters
                )
                sheet.cell(row=row, column=72).value = f"={sun_terms}"
            else:
                sheet.cell(row=row, column=72).value = 0

        sheet.cell(row=81, column=62).value = "=SUM(BJ31:BJ80)"
        sheet.cell(row=81, column=63).value = "=SUM(BK31:BK80)"
        sheet.cell(row=81, column=64).value = "=SUM(BL31:BL80)"
        sheet.cell(row=81, column=65).value = "=SUM(BM31:BM80)"
        sheet.cell(row=81, column=66).value = "=SUM(BN31:BN80)"
        sheet.cell(row=81, column=67).value = "=SUM(BO31:BO80)"
        sheet.cell(row=81, column=68).value = "=SUM(BP31:BP80)"
        sheet.cell(row=81, column=69).value = "=SUM(BQ31:BQ80)"
        sheet.cell(row=81, column=70).value = "=SUM(BR31:BR80)"
        sheet.cell(row=81, column=72).value = "=SUM(BT31:BT80)"
        sheet.cell(row=82, column=62).value = "%CVf"
        sheet.cell(row=82, column=64).value = "=BL81/(BJ81+BK81+BL81)*100"
        # %CVc: combined (resident + faculty) CV percentage
        sheet.cell(row=82, column=66).value = "%CVc"
        sheet.cell(
            row=82, column=68
        ).value = "=(BL29+BL81)/((BJ29+BJ81)+(BK29+BK81)+(BL29+BL81))*100"
        # PGY1 CV Total
        sheet.cell(row=82, column=71).value = "PGY1 CV"
        sheet.cell(row=82, column=72).value = '=SUMIF($D$9:$D$30,"PGY 1",BL$9:BL$30)'

        self._apply_bottom_calc_rows(sheet, block_start, block_end)

    def _apply_bottom_calc_rows(
        self, sheet, block_start: date, block_end: date
    ) -> None:
        """Add 13 per-column calculation rows below the faculty summary.

        Rows 84-96 provide daily staffing tallies per AM/PM slot:
        screeners needed, clinic counts by PGY, supervision ratios,
        procedure counts, and appointment estimates.
        """
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter

        LABEL_COL = 5  # Column E
        CALC_START_ROW = 84
        SCHEDULE_START = 6  # Column F
        num_days = (block_end - block_start).days + 1
        schedule_end = SCHEDULE_START + (num_days * 2) - 1

        labels = [
            "Screeners Needed",
            "Providers Virtual (CV)",
            "Interns in Clinic",
            "Residents in Clinic",
            "Attendings Needed Clinic",
            "Residents in PROC",
            "Residents in V Clinic",
            "Residents on HV",
            "Total Attendings Needed",
            "# Attendings Assigned",
            "# Primary Care Appts",
            "PR Count",
            "VAS Count",
            "PR/VAS Conflict",
        ]

        bold = Font(name="Arial", bold=True, size=10)
        sheet.cell(row=83, column=LABEL_COL, value="DAILY STAFFING").font = bold

        for i, label in enumerate(labels):
            sheet.cell(
                row=CALC_START_ROW + i, column=LABEL_COL, value=label
            ).font = bold

        # Range covers all people rows (residents 9-30 + faculty 31-80)
        R_ALL = "$9:${c}$80"  # noqa: N806 — placeholder, used via f-string

        for col_idx in range(SCHEDULE_START, schedule_end + 1):
            c = get_column_letter(col_idx)
            r_all = f"{c}$9:{c}$80"  # all people rows
            r_res = f"{c}$9:{c}$30"  # resident rows
            r_fac = f"{c}$31:{c}$80"  # faculty rows

            # Row 84: Screeners Needed — total providers in clinic
            # Matches reference: counts C, C30, C40, C60, CC, C-I, C-N,
            # VAS, VASc, PR, OPR, SM, GAC, PE, C10, V1-V3, PAP(×2),
            # CF2V/CV2F (×0.5)
            sheet.cell(row=84, column=col_idx).value = (
                f'=COUNTIF({r_all},"C")'
                f'+COUNTIF({r_all},"C30")'
                f'+COUNTIF({r_all},"C40")'
                f'+COUNTIF({r_all},"C60")'
                f'+COUNTIF({r_all},"CC")'
                f'+COUNTIF({r_all},"C-I")'
                f'+COUNTIF({r_all},"C-N")'
                f'+COUNTIF({r_all},"VAS")'
                f'+COUNTIF({r_all},"VASc")'
                f'+COUNTIF({r_all},"PR")'
                f'+COUNTIF({r_all},"OPR")'
                f'+COUNTIF({r_all},"SM")'
                f'+COUNTIF({r_all},"GAC")'
                f'+COUNTIF({r_all},"PE")'
                f'+COUNTIF({r_all},"C10")'
                f'+COUNTIF({r_all},"V1")'
                f'+COUNTIF({r_all},"V2")'
                f'+COUNTIF({r_all},"V3")'
                f'+COUNTIF({r_all},"PAP")*2'
                f'+COUNTIF({r_all},"CF2V")*0.5'
                f'+COUNTIF({r_all},"CV2F")*0.5'
            )

            # Row 85: Providers Virtual (CV, CV10, CF2V/CV2F half-weight)
            sheet.cell(row=85, column=col_idx).value = (
                f'=COUNTIF({r_all},"CV")'
                f'+COUNTIF({r_all},"CV10")'
                f'+COUNTIF({r_all},"CF2V")*0.5'
                f'+COUNTIF({r_all},"CV2F")*0.5'
            )

            # Row 86: Interns in Clinic (PGY 1 rows with clinic codes)
            sheet.cell(row=86, column=col_idx).value = (
                f'=COUNTIFS($D$9:$D$30,"PGY 1",{r_res},"C")'
                f'+COUNTIFS($D$9:$D$30,"PGY 1",{r_res},"C30")'
                f'+COUNTIFS($D$9:$D$30,"PGY 1",{r_res},"C40")'
                f'+COUNTIFS($D$9:$D$30,"PGY 1",{r_res},"C60")'
                f'+COUNTIFS($D$9:$D$30,"PGY 1",{r_res},"CC")'
                f'+COUNTIFS($D$9:$D$30,"PGY 1",{r_res},"C-I")'
            )

            # Row 87: Residents in Clinic (PGY 2/3 with clinic codes + CV)
            sheet.cell(row=87, column=col_idx).value = (
                f'=COUNTIFS($D$9:$D$30,"PGY 2",{r_res},"C")'
                f'+COUNTIFS($D$9:$D$30,"PGY 3",{r_res},"C")'
                f'+COUNTIFS($D$9:$D$30,"PGY 2",{r_res},"C30")'
                f'+COUNTIFS($D$9:$D$30,"PGY 3",{r_res},"C30")'
                f'+COUNTIFS($D$9:$D$30,"PGY 2",{r_res},"C40")'
                f'+COUNTIFS($D$9:$D$30,"PGY 3",{r_res},"C40")'
                f'+COUNTIFS($D$9:$D$30,"PGY 2",{r_res},"C60")'
                f'+COUNTIFS($D$9:$D$30,"PGY 3",{r_res},"C60")'
                f'+COUNTIFS($D$9:$D$30,"PGY 2",{r_res},"CC")'
                f'+COUNTIFS($D$9:$D$30,"PGY 3",{r_res},"CC")'
                f'+COUNTIFS($D$9:$D$30,"PGY 2",{r_res},"C-I")'
                f'+COUNTIFS($D$9:$D$30,"PGY 3",{r_res},"C-I")'
                f'+COUNTIFS($D$9:$D$30,"PGY 2",{r_res},"SM")'
                f'+COUNTIFS($D$9:$D$30,"PGY 3",{r_res},"SM")'
                f'+COUNTIFS($D$9:$D$30,"PGY 2",{r_res},"CV")'
                f'+COUNTIFS($D$9:$D$30,"PGY 3",{r_res},"CV")'
            )

            # Row 88: Attendings Needed = interns×0.5 + residents×0.25
            sheet.cell(row=88, column=col_idx).value = f"={c}86*0.5+{c}87*0.25"

            # Row 89: Residents in PROC (VAS + PR + OPR + PROC in resident rows)
            sheet.cell(row=89, column=col_idx).value = (
                f'=COUNTIF({r_res},"VAS")+COUNTIF({r_res},"PR")'
                f'+COUNTIF({r_res},"OPR")+COUNTIF({r_res},"PROC")'
            )

            # Row 90: Residents in V Clinic
            sheet.cell(row=90, column=col_idx).value = (
                f'=COUNTIF({r_res},"V1")'
                f'+COUNTIF({r_res},"V2")'
                f'+COUNTIF({r_res},"V3")'
                f'+COUNTIF({r_res},"CV")'
            )

            # Row 91: Residents on HV
            sheet.cell(row=91, column=col_idx).value = f'=COUNTIF({r_res},"HV")'

            # Row 92: Total Attendings Needed = ROUNDUP(clinic + proc)
            sheet.cell(row=92, column=col_idx).value = f"=ROUNDUP(SUM({c}88:{c}89),0)"

            # Row 93: # Attendings Assigned (AT + PCAT + DO in faculty rows)
            sheet.cell(
                row=93, column=col_idx
            ).value = (
                f'=COUNTIF({r_fac},"AT")+COUNTIF({r_fac},"PCAT")+COUNTIF({r_fac},"DO")'
            )

            # Row 94: # Primary Care Appts (placeholder — manual input)
            sheet.cell(row=94, column=col_idx).value = ""

            # Row 95: PR Count (resident rows)
            sheet.cell(row=95, column=col_idx).value = f'=COUNTIF({r_res},"PR")'

            # Row 96: VAS Count (resident rows)
            sheet.cell(row=96, column=col_idx).value = f'=COUNTIF({r_res},"VAS")'

            # Row 97: PR/VAS Conflict detector
            sheet.cell(
                row=97, column=col_idx
            ).value = f'=IF(OR({c}95>1,{c}96>1,AND({c}95>=1,{c}96>=1)),"\u26a0","")'

    def _call_last_name_token(self, raw_name: Any) -> str:
        if not raw_name:
            return ""
        text = str(raw_name).replace("*", "").strip()
        if not text:
            return ""
        if "," in text:
            last = text.split(",", 1)[0]
        else:
            last = text.split()[-1]
        return " ".join(last.split()).upper()

    def _write_export_qa_sheet(
        self, wb: Workbook, data: dict[str, Any], stats: dict[str, Any]
    ) -> None:
        """Write a human-readable QA sheet with explicit code breakdowns."""
        if "Export_QA" in wb.sheetnames:
            del wb["Export_QA"]
        qa = wb.create_sheet("Export_QA")

        bold = Font(bold=True)
        qa["A1"] = "Export QA Summary"
        qa["A1"].font = bold
        qa["A2"] = "Generated UTC"
        qa["B2"] = f"{datetime.now(UTC).isoformat(timespec='seconds')}Z"
        qa["A3"] = "Block Start"
        qa["B3"] = str(stats.get("block_start", ""))
        qa["A4"] = "Block End"
        qa["B4"] = str(stats.get("block_end", ""))
        qa["A5"] = "Source"
        qa["B5"] = str(data.get("source", "unknown"))
        qa["A6"] = "Unmerged Schedule Ranges"
        qa["B6"] = int(stats.get("schedule_cells_unmerged", 0))
        qa["A7"] = "Presentation Profile"
        qa["B7"] = self.presentation_profile
        qa["C7"] = "Identity Fields"
        qa["D7"] = (
            "template-preserved"
            if self.preserve_template_identity_fields
            else "db-written"
        )

        qa["A9"] = "Section"
        qa["B9"] = "People Input"
        qa["C9"] = "People Written"
        qa["D9"] = "Unmapped Names"
        qa["E9"] = "Non-empty Cells"
        for cell in ("A9", "B9", "C9", "D9", "E9"):
            qa[cell].font = bold

        for idx, section in enumerate(("residents", "faculty"), start=10):
            section_stats = stats.get(section, {})
            qa.cell(row=idx, column=1, value=section.title())
            qa.cell(row=idx, column=2, value=int(section_stats.get("people_input", 0)))
            qa.cell(
                row=idx, column=3, value=int(section_stats.get("people_written", 0))
            )
            unmapped = section_stats.get("unmapped_names", []) or []
            qa.cell(row=idx, column=4, value=", ".join(str(n) for n in unmapped))
            codes_written = section_stats.get("codes_written", {}) or {}
            qa.cell(
                row=idx, column=5, value=sum(int(v) for v in codes_written.values())
            )

        resident_codes = Counter(
            {
                k: int(v)
                for k, v in (
                    stats.get("residents", {}).get("codes_written", {}) or {}
                ).items()
            }
        )
        faculty_codes = Counter(
            {
                k: int(v)
                for k, v in (
                    stats.get("faculty", {}).get("codes_written", {}) or {}
                ).items()
            }
        )
        combined_codes = resident_codes + faculty_codes

        qa["A14"] = "Code"
        qa["B14"] = "Residents"
        qa["C14"] = "Faculty"
        qa["D14"] = "Combined"
        for cell in ("A14", "B14", "C14", "D14"):
            qa[cell].font = bold

        code_row = 15
        for code in sorted(combined_codes):
            qa.cell(row=code_row, column=1, value=code)
            qa.cell(row=code_row, column=2, value=int(resident_codes.get(code, 0)))
            qa.cell(row=code_row, column=3, value=int(faculty_codes.get(code, 0)))
            qa.cell(row=code_row, column=4, value=int(combined_codes.get(code, 0)))
            code_row += 1

        qa["F9"] = "Note"
        qa["F9"].font = bold
        qa["F10"] = (
            "Template summary columns on Block Template2 are composite in places "
            "(e.g. BJ may include C with SM/C-I). Use this sheet for explicit counts."
        )

        qa["F14"] = "Faculty Clinic Detail"
        qa["F14"].font = bold
        qa["F15"] = "Name"
        qa["G15"] = "C"
        qa["H15"] = "SM"
        qa["I15"] = "C+SM"
        for cell in ("F15", "G15", "H15", "I15"):
            qa[cell].font = bold

        faculty_people = self._collect_people_code_counts(data.get("faculty", []) or [])
        faculty_row = 16
        for name, counter in faculty_people:
            c_count = int(counter.get("C", 0))
            sm_count = int(counter.get("SM", 0))
            qa.cell(row=faculty_row, column=6, value=name)
            qa.cell(row=faculty_row, column=7, value=c_count)
            qa.cell(row=faculty_row, column=8, value=sm_count)
            qa.cell(row=faculty_row, column=9, value=c_count + sm_count)
            faculty_row += 1

        for col, width in (
            ("A", 28),
            ("B", 16),
            ("C", 16),
            ("D", 40),
            ("E", 16),
            ("F", 42),
            ("G", 10),
            ("H", 10),
            ("I", 10),
        ):
            qa.column_dimensions[col].width = width

    def _write_anchor_sheet(self, wb: Workbook, data: dict[str, Any]) -> None:
        """Create a veryHidden sheet __ANCHORS__ with UUIDs and row hashes."""
        if "__ANCHORS__" in wb.sheetnames:
            del wb["__ANCHORS__"]
        ws = wb.create_sheet("__ANCHORS__")
        ws.sheet_state = "veryHidden"

        # Headers
        ws.cell(row=1, column=1, value="person_id")
        ws.cell(row=1, column=2, value="block_assignment_id")
        ws.cell(row=1, column=3, value="row_hash")

        from app.services.excel_metadata import compute_row_hash

        all_people = (data.get("residents", []) or []) + (data.get("faculty", []) or [])
        for person in all_people:
            person_id = person.get("id")
            if not person_id:
                continue

            # UUID-based lookup first (dynamic), then name-based (legacy XML)
            pid_str = str(person_id)
            row = self.row_mappings.get(pid_str)
            if not row:
                name = str(person.get("name", "") or "").replace("*", "").strip()
                row = self.row_mappings.get(name)
            if not row:
                continue

            # Compute hash for Phase 2 O(1) change detection
            # compute_row_hash internally normalizes via normalize_for_hash()
            rotation1 = person.get("rotation1")
            rotation2 = person.get("rotation2")
            days_codes = []
            for day in person.get("days", []):
                days_codes.append(day.get("am"))
                days_codes.append(day.get("pm"))

            try:
                pid_uuid = UUID(person_id)
            except (ValueError, AttributeError):
                pid_uuid = UUID(int=0)  # fallback for non-UUID test IDs
            row_hash = compute_row_hash(pid_uuid, rotation1, rotation2, days_codes)

            # Write to anchor sheet at matching row (Spatial Anchoring)
            ws.cell(row=row, column=1, value=str(person_id))
            ws.cell(row=row, column=2, value=str(person.get("block_assignment_id", "")))
            ws.cell(row=row, column=3, value=row_hash)

    def _add_data_validation(self, sheet, block_start: date, block_end: date) -> None:
        """Add Excel DataValidation dropdowns referencing __REF__ Named Ranges.

        Note: The Named Ranges (ValidRotations, ValidActivities) are created globally
        in canonical_schedule_export_service.py. The dropdowns are applied here.
        """
        from openpyxl.worksheet.datavalidation import DataValidation

        # Rotation columns: dropdown from ValidRotations named range
        rot_dv = DataValidation(
            type="list",
            formula1="ValidRotations",
            allow_blank=True,
            showErrorMessage=False,  # Allow free-text overrides
        )
        sheet.add_data_validation(rot_dv)

        # Activity columns: dropdown from ValidActivities named range
        act_dv = DataValidation(
            type="list",
            formula1="ValidActivities",
            allow_blank=True,
            showErrorMessage=False,  # Allow free-text overrides
        )
        sheet.add_data_validation(act_dv)

        schedule_start_col = COL_SCHEDULE_START
        total_days = (block_end - block_start).days + 1
        schedule_end_col = schedule_start_col + (total_days * COLS_PER_DAY) - 1

        # Apply to all mapped resident/faculty rows
        target_rows = sorted(set(self.row_mappings.values()))
        if not target_rows:
            return

        for row in target_rows:
            # Rotation columns (A and B)
            rot_dv.add(sheet.cell(row=row, column=BT2_COL_ROTATION1))
            rot_dv.add(sheet.cell(row=row, column=BT2_COL_ROTATION2))

            # Schedule columns (F through schedule end)
            for col in range(schedule_start_col, schedule_end_col + 1):
                act_dv.add(sheet.cell(row=row, column=col))

    def _add_dynamic_cf(self, sheet, block_start: date, block_end: date) -> None:
        """Add dynamic conditional formatting rules based on tamc_color_scheme."""
        from openpyxl.formatting.rule import CellIsRule
        from openpyxl.styles import Font, PatternFill

        if not self.color_scheme:
            return

        schedule_start_col = COL_SCHEDULE_START
        total_days = (block_end - block_start).days + 1
        schedule_end_col = schedule_start_col + (total_days * COLS_PER_DAY) - 1

        # Standard grid range for Block Template 2
        min_row = 9
        max_row = 69
        grid_range = f"{get_column_letter(schedule_start_col)}{min_row}:{get_column_letter(schedule_end_col)}{max_row}"

        # Add rules for each code in scheme
        for code, bg_color in self.color_scheme._code_colors.items():
            fg_color = self.color_scheme._font_colors.get(code, "000000")

            rule = CellIsRule(
                operator="equal",
                formula=[f'"{code}"'],
                fill=PatternFill(start_color=bg_color, fill_type="solid"),
                font=Font(color=fg_color),
            )
            sheet.conditional_formatting.add(grid_range, rule)

    def _add_leave_formula_column(
        self, sheet, block_start: date, block_end: date
    ) -> None:
        """Add a column that auto-calculates leave days from the grid."""
        # Standard summary position: Column BS (71)
        leave_col = 71
        sheet.cell(row=8, column=leave_col, value="LV Days")

        schedule_start_col_letter = get_column_letter(COL_SCHEDULE_START)

        # Calculate end column dynamically based on block length
        actual_days = (block_end - block_start).days + 1
        schedule_end_col_letter = get_column_letter(
            COL_SCHEDULE_START + (actual_days * COLS_PER_DAY) - 1
        )

        for row in range(9, 70):
            # Formula: count LV codes and divide by 2 (AM/PM slots)
            sheet.cell(
                row=row,
                column=leave_col,
                value=f'=COUNTIF({schedule_start_col_letter}{row}:{schedule_end_col_letter}{row}, "LV")/2',
            )

    def _prune_empty_sheets(self, wb: Workbook, keep: set[str]) -> None:
        """Remove placeholder sheets that contain no data."""
        for name in list(wb.sheetnames):
            if name in keep:
                continue
            ws = wb[name]
            if ws.max_row == 1 and ws.max_column == 1 and ws["A1"].value in (None, ""):
                del wb[name]

    def _parse_xml_to_data(self, root: ElementTree.Element) -> dict[str, Any]:
        """Parse XML schedule into a JSON-like dict."""
        block_start = root.get("block_start", "")
        block_end = root.get("block_end", "")

        def _parse_people(elements: list[ElementTree.Element]) -> list[dict[str, Any]]:
            people: list[dict[str, Any]] = []
            for elem in elements:
                person = {
                    "name": elem.get("name", ""),
                    "pgy": elem.get("pgy", ""),
                    "rotation1": elem.get("rotation1", ""),
                    "rotation2": elem.get("rotation2", ""),
                    "days": [],
                }
                for day in elem.findall("day"):
                    person["days"].append(
                        {
                            "date": day.get("date", ""),
                            "am": day.get("am", ""),
                            "pm": day.get("pm", ""),
                        }
                    )
                people.append(person)
            return people

        residents = _parse_people(root.findall(".//resident"))
        faculty = _parse_people(root.findall(".//faculty"))

        call_rows: list[dict[str, Any]] = []
        call_section = root.find(".//call")
        if call_section is not None:
            for night in call_section.findall("night"):
                call_rows.append(
                    {
                        "date": night.get("date", ""),
                        "staff": night.get("staff", ""),
                    }
                )

        return {
            "block_start": block_start,
            "block_end": block_end,
            "source": root.get("source", "xml"),
            "residents": residents,
            "faculty": faculty,
            "call": {"nights": call_rows},
        }

    def _coerce_date(self, value: Any) -> date | None:
        """Coerce ISO date string or datetime/date into date."""
        if isinstance(value, date) and not isinstance(value, datetime):
            return value
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, str) and value:
            return date.fromisoformat(value)
        return None

    def _apply_tamc_presentation(self, sheet, schedule_end_col: int) -> None:
        """Apply TAMC hand-jammed formatting to match the reference workbook.

        Matches the formatting from 'Current AY 25-26' Block 11:
        - Arial 16 body, Arial 18 day names, Arial 20 call names
        - Column widths matched to reference
        - Row heights: 15-25 headers, 21 residents, 20 faculty
        - No borders on body cells (reference has none)
        - AM/PM column merges on header rows 1-5
        """
        no_border = Border()
        base_font = Font(name="Arial", size=16)
        bold_font = Font(name="Arial", size=16, bold=True)

        # --- Column widths (match reference Block 11) ---
        sheet.column_dimensions["A"].width = 10.2  # Rotation1
        sheet.column_dimensions["B"].width = 12.2  # Rotation2
        sheet.column_dimensions["C"].width = 9.2  # Template
        sheet.column_dimensions["D"].width = 11.8  # Role/PGY
        sheet.column_dimensions["E"].width = 40.5  # Name
        # Schedule columns: reference defaultColWidth = 10.2
        for col in range(BT2_COL_SCHEDULE_START, schedule_end_col + 1):
            sheet.column_dimensions[get_column_letter(col)].width = 10.2

        # --- Row heights ---
        sheet.row_dimensions[1].height = 15  # Day full names
        sheet.row_dimensions[2].height = 15  # Day abbreviations
        sheet.row_dimensions[3].height = 20  # Dates
        sheet.row_dimensions[4].height = 25  # Staff call
        sheet.row_dimensions[5].height = 15  # Resident call
        for row in range(6, 8):
            sheet.row_dimensions[row].height = 20  # Label/spacer rows
        sheet.row_dimensions[8].height = 21  # Row 8 matches resident height
        for row in range(9, 31):
            sheet.row_dimensions[row].height = 21  # Resident rows
        for row in range(31, 81):
            sheet.row_dimensions[row].height = 20  # Faculty rows

        # --- Fonts and borders for all cells ---
        for row in sheet.iter_rows(
            min_row=1,
            max_row=sheet.max_row,
            min_col=1,
            max_col=max(schedule_end_col, sheet.max_column),
        ):
            for cell in row:
                r = cell.row
                # Preserve existing font color from color scheme
                existing_color = (
                    cell.font.color if cell.font and cell.font.color else None
                )
                # Row-specific font sizes (match reference Block 11)
                if r in (1, 2) and cell.column >= BT2_COL_SCHEDULE_START:
                    # Day names + abbreviations: Arial 18, not bold
                    cell.font = Font(name="Arial", size=18, color=existing_color)
                elif r == 4 and cell.column >= BT2_COL_SCHEDULE_START:
                    # Staff call names: Arial 20
                    cell.font = Font(name="Arial", size=20, color=existing_color)
                elif r == 5 and cell.column >= BT2_COL_SCHEDULE_START:
                    # Resident call: Arial 14
                    cell.font = Font(name="Arial", size=14, color=existing_color)
                elif r == 6 and cell.column <= 5:
                    # Row 6 label columns (TEMPLATE, ROLE, PROVIDER): Arial 11 bold
                    cell.font = Font(
                        name="Arial",
                        size=11,
                        bold=True,
                        color=existing_color,
                    )
                elif r <= 5:
                    # Header rows 1-5 label columns: Arial 16 bold
                    cell.font = Font(
                        name="Arial",
                        size=16,
                        bold=True,
                        color=existing_color,
                    )
                elif r <= 8:
                    # Rows 6-8 schedule area: Arial 16, not bold
                    cell.font = Font(name="Arial", size=16, color=existing_color)
                else:
                    # Body cells: Arial 16 with existing color
                    cell.font = Font(name="Arial", size=16, color=existing_color)

                # Remove borders from body cells (reference has none)
                if r >= 6:
                    cell.border = no_border

        # --- Merge AM/PM columns in header rows 1-5 ---
        for col in range(BT2_COL_SCHEDULE_START, schedule_end_col + 1, COLS_PER_DAY):
            am_letter = get_column_letter(col)
            pm_letter = get_column_letter(col + 1)
            for row in range(1, 6):
                merge_range = f"{am_letter}{row}:{pm_letter}{row}"
                try:
                    sheet.merge_cells(merge_range)
                except ValueError:
                    pass  # Already merged

    def _apply_cell_color(self, cell, code: str) -> None:
        """Apply background and font color to cell based on schedule code.

        Font colors have semantic meaning:
        - Red: procedures needing supervision (PR adds +1 AT), or high-visibility roles
        - Light gray: Night Float (NF, Peds NF)
        - White: Contrast on dark backgrounds
        """
        if not self.apply_colors or not self.color_scheme:
            return

            # Apply fill color
        hex_color = self.color_scheme.get_code_color(code)
        if hex_color:
            cell.fill = PatternFill(start_color=hex_color, fill_type="solid")

            # Apply font color (priority: explicit mapping > contrast fallback)
        font_color = self.color_scheme.get_font_color(code)
        if font_color:
            cell.font = Font(color=font_color)
        elif hex_color:
            # Fallback: white text on dark fills (e.g., red).
            # Colors are 8-char ARGB (FF prefix = opaque).
            raw = hex_color[-6:]
            if raw == "FF0000":
                cell.font = Font(color="FFFFFFFF")

    def _apply_header_color(self, cell, day_of_week: int) -> None:
        """Apply header background color based on day of week."""
        if not self.apply_colors or not self.color_scheme:
            return

        hex_color = self.color_scheme.get_header_color(day_of_week)
        if hex_color and hex_color != "FFFFFF":  # Skip white (default)
            cell.fill = PatternFill(start_color=hex_color, fill_type="solid")

    def _apply_rotation_color(self, cell, rotation: str) -> None:
        """Apply color to rotation column cell."""
        if not self.apply_colors or not self.color_scheme:
            return

        hex_color = self.color_scheme.get_rotation_color(rotation)
        if hex_color:
            cell.fill = PatternFill(start_color=hex_color, fill_type="solid")
            # Use white text for dark backgrounds
            if hex_color in ("000000", "FF0000"):
                cell.font = Font(color="FFFFFF")

    def _create_new_workbook(
        self,
        block_start: date,
        block_end: date,
    ) -> Workbook:
        """Create new workbook matching ROSETTA validation format."""
        wb = Workbook()
        sheet = wb.active

        # Calculate block number
        block_number = self._calculate_block_number(block_start)
        sheet.title = f"Block{block_number}_GENERATED"

        # Set column widths
        sheet.column_dimensions["A"].width = 20  # Resident name
        sheet.column_dimensions["B"].width = 5  # PGY
        sheet.column_dimensions["C"].width = 12  # Rotation1
        sheet.column_dimensions["D"].width = 12  # Rotation2
        sheet.column_dimensions["E"].width = 30  # Notes
        for col in range(COL_SCHEDULE_START, COL_SCHEDULE_END + 1):
            sheet.column_dimensions[get_column_letter(col)].width = 6

        return wb

    def _get_writable_cell(self, sheet, row: int, column: int) -> Cell:
        """Return a writable cell, resolving merged ranges to their top-left cell."""
        cell = sheet.cell(row=row, column=column)
        if isinstance(cell, MergedCell):
            coord = f"{get_column_letter(column)}{row}"
            for merged_range in sheet.merged_cells.ranges:
                if coord in merged_range:
                    return sheet.cell(
                        row=merged_range.min_row, column=merged_range.min_col
                    )
        return cell

    def _fill_block_template2_headers(
        self,
        sheet,
        block_start: date,
        block_end: date,
    ) -> None:
        """Update day-of-week and date headers for Block Template2."""
        day_full = {
            0: "MON",
            1: "TUE",
            2: "WED",
            3: "THURS",
            4: "FRI",
            5: "SAT",
            6: "SUN",
        }
        day_abbrev = {
            0: "MON",
            1: "TUE",
            2: "WED",
            3: "THU",
            4: "FRI",
            5: "SAT",
            6: "SUN",
        }

        block_number = self._calculate_block_number(block_start)
        if block_number:
            block_cell = self._get_writable_cell(sheet, row=1, column=BT2_COL_TEMPLATE)
            block_cell.value = block_number

        current = block_start
        col = BT2_COL_SCHEDULE_START
        while current <= block_end:
            weekday = current.weekday()
            day_cell = self._get_writable_cell(sheet, row=1, column=col)
            abbrev_cell = self._get_writable_cell(sheet, row=2, column=col)
            date_cell = self._get_writable_cell(sheet, row=3, column=col)

            day_cell.value = day_full.get(weekday, "")
            abbrev_cell.value = day_abbrev.get(weekday, "")
            date_cell.value = datetime(current.year, current.month, current.day)
            date_cell.number_format = "d-mmm"

            current += timedelta(days=1)
            col += 2

    def _fill_header_row(
        self,
        sheet,
        block_start: date,
        block_end: date,
    ) -> None:
        """Fill header row based on format (Block Template2 or ROSETTA)."""
        if self.use_block_template2:
            # Use the template's own layout; only update day/date headers.
            self._fill_block_template2_headers(sheet, block_start, block_end)
            return
        else:
            # ROSETTA format headers
            sheet.cell(row=ROW_HEADERS, column=ROSETTA_COL_NAME).value = "Resident"
            sheet.cell(row=ROW_HEADERS, column=ROSETTA_COL_PGY).value = "PGY"
            sheet.cell(
                row=ROW_HEADERS, column=ROSETTA_COL_ROTATION1
            ).value = "Rotation 1"
            sheet.cell(
                row=ROW_HEADERS, column=ROSETTA_COL_ROTATION2
            ).value = "Rotation 2"
            sheet.cell(row=ROW_HEADERS, column=ROSETTA_COL_NOTES).value = "Notes"
            col_schedule_start = ROSETTA_COL_SCHEDULE_START

            # Schedule column headers (e.g., "Thu Mar 12 AM", "Thu Mar 12 PM")
        current = block_start
        col = col_schedule_start
        while current <= block_end:
            day_str = current.strftime("%a %b %d").replace(" 0", " ")  # "Thu Mar 12"

            am_cell = self._get_writable_cell(sheet, row=ROW_HEADERS, column=col)
            pm_cell = self._get_writable_cell(sheet, row=ROW_HEADERS, column=col + 1)

            am_cell.value = f"{day_str} AM"
            pm_cell.value = f"{day_str} PM"

            # Apply header colors based on day of week
            self._apply_header_color(am_cell, current.weekday())
            self._apply_header_color(pm_cell, current.weekday())

            current += timedelta(days=1)
            col += 2

    def _to_last_first(self, name: str) -> str:
        """Convert 'First Last' to 'Last, First' format."""
        if not name or "," in name:
            return name  # Already in Last, First format or empty
        parts = name.strip().split()
        if len(parts) >= 2:
            return f"{parts[-1]}, {' '.join(parts[:-1])}"
        return name

    def _get_role(self, pgy: str, is_faculty: bool) -> str:
        """Get role string for Block Template2 format."""
        if is_faculty:
            return "FAC"
        if pgy:
            return f"PGY {pgy}"
        return ""

    def _resolve_row(
        self, resident: dict[str, Any], index: int, stats: dict[str, Any] | None
    ) -> int | None:
        """Resolve the Excel row for a person, trying UUID first then name."""
        person_id = str(resident.get("id", ""))
        name = str(resident.get("name", "") or "")

        # 1. UUID-based lookup (dynamic mappings)
        if person_id and person_id in self.row_mappings:
            return self.row_mappings[person_id]

        # 2. Name-based lookup (legacy structure XML mappings)
        if self.row_mappings:
            normalized = name.replace("*", "").strip()
            lookup_name = normalized

            if lookup_name not in self.row_mappings and "," in lookup_name:
                last, first = [part.strip() for part in lookup_name.split(",", 1)]
                swapped = f"{first} {last}".strip()
                if swapped in self.row_mappings:
                    lookup_name = swapped

            row = self.row_mappings.get(lookup_name)

            # Fuzzy match by first name
            if not row:
                first_name = lookup_name.split()[0] if lookup_name else ""
                for mapping_name, mapping_row in self.row_mappings.items():
                    if mapping_name.startswith(first_name):
                        row = mapping_row
                        break

            if row:
                return row

            if stats is not None:
                stats["unmapped_names"].append(name)
            if self.strict_row_mapping:
                raise ValueError(
                    f"No row mapping for: {name}. "
                    "Update BlockTemplate2_Structure.xml or use dynamic mappings."
                )
            logger.warning(f"No row mapping for: {name}")
            return None

        # 3. Sequential fallback (ROSETTA)
        return index + 2  # Start at row 2 (row 1 is headers)

    def _resolve_template_code(
        self, resident: dict[str, Any], pgy: Any, is_faculty: bool
    ) -> str:
        """Resolve the template code for a person, trying UUID first then name."""
        person_id = str(resident.get("id", ""))

        # UUID-based lookup
        if person_id and person_id in self.template_mappings:
            return self.template_mappings[person_id]

        # Name-based fallback (legacy XML)
        name = str(resident.get("name", "") or "")
        normalized = name.replace("*", "").strip()
        lookup_name = normalized
        if lookup_name not in self.template_mappings and "," in lookup_name:
            last, first = [part.strip() for part in lookup_name.split(",", 1)]
            swapped = f"{first} {last}".strip()
            if swapped in self.template_mappings:
                lookup_name = swapped

        template_code = self.template_mappings.get(lookup_name, "")
        if not template_code and pgy:
            template_code = f"R{pgy}"
        elif not template_code and is_faculty:
            template_code = "C19"
        return template_code

    def _fill_residents(
        self,
        sheet,
        residents: list[dict[str, Any]],
        block_start: date,
        is_faculty: bool,
        stats: dict[str, Any] | None = None,
    ) -> None:
        """Fill resident/faculty rows from schedule dicts.

        Row resolution order: UUID (dynamic) → name (legacy XML) → sequential (ROSETTA).

        Column layout depends on use_block_template2 flag:
        - Block Template2: Rotation1, Rotation2, Template, Role, Name, Schedule...
        - ROSETTA: Name, PGY, Rotation1, Rotation2, Notes, Schedule...
        """
        # Sort by name when no mappings (ROSETTA order)
        sorted_residents = sorted(residents, key=lambda r: r.get("name", ""))

        for i, resident in enumerate(sorted_residents):
            name = resident.get("name", "")
            name = str(name) if name is not None else ""

            row = self._resolve_row(resident, i, stats)
            if row is None:
                continue

            if stats is not None:
                stats["people_written"] += 1

            pgy = resident.get("pgy", "")
            rotation1 = resident.get("rotation1", "") or ""
            rotation2 = resident.get("rotation2", "") or ""

            if self.use_block_template2:
                template_code = self._resolve_template_code(resident, pgy, is_faculty)
                role = self._get_role(pgy, is_faculty)
                display_name = self._to_last_first(name)

                if not self.preserve_template_identity_fields:
                    rot1_cell = self._get_writable_cell(
                        sheet, row=row, column=BT2_COL_ROTATION1
                    )
                    rot1_cell.value = rotation1
                    self._apply_rotation_color(rot1_cell, rotation1)

                    rot2_cell = self._get_writable_cell(
                        sheet, row=row, column=BT2_COL_ROTATION2
                    )
                    rot2_cell.value = rotation2 or ""
                    if rotation2:
                        self._apply_rotation_color(rot2_cell, rotation2)

                    self._get_writable_cell(
                        sheet, row=row, column=BT2_COL_TEMPLATE
                    ).value = template_code
                    self._get_writable_cell(
                        sheet, row=row, column=BT2_COL_ROLE
                    ).value = role
                    self._get_writable_cell(
                        sheet, row=row, column=BT2_COL_NAME
                    ).value = display_name
            else:
                # ROSETTA format: Name, PGY, Rotation1, Rotation2, Notes
                self._get_writable_cell(
                    sheet, row=row, column=ROSETTA_COL_NAME
                ).value = name
                self._get_writable_cell(
                    sheet, row=row, column=ROSETTA_COL_PGY
                ).value = pgy

                rot1_cell = self._get_writable_cell(
                    sheet, row=row, column=ROSETTA_COL_ROTATION1
                )
                rot1_cell.value = rotation1
                self._apply_rotation_color(rot1_cell, rotation1)

                rot2_cell = self._get_writable_cell(
                    sheet, row=row, column=ROSETTA_COL_ROTATION2
                )
                rot2_cell.value = rotation2 or ""
                if rotation2:
                    self._apply_rotation_color(rot2_cell, rotation2)

                    # Fill schedule columns from day elements
            days = resident.get("days", []) or []
            sorted_days = sorted(days, key=lambda d: str(d.get("date", "")))
            for day in sorted_days:
                day_date = self._coerce_date(day.get("date"))
                if not day_date:
                    continue
                am_code = day.get("am", "") or ""
                pm_code = day.get("pm", "") or ""

                # Calculate column from date
                day_offset = (day_date - block_start).days
                am_col = COL_SCHEDULE_START + (day_offset * 2)
                pm_col = am_col + 1

                am_cell = self._get_writable_cell(sheet, row=row, column=am_col)
                pm_cell = self._get_writable_cell(sheet, row=row, column=pm_col)

                am_cell.value = am_code
                pm_cell.value = pm_code

                if stats is not None and am_code:
                    codes_written = stats["codes_written"]
                    codes_written[am_code] = codes_written.get(am_code, 0) + 1
                if stats is not None and pm_code:
                    codes_written = stats["codes_written"]
                    codes_written[pm_code] = codes_written.get(pm_code, 0) + 1

                # Apply colors based on schedule code
                if am_code:
                    self._apply_cell_color(am_cell, am_code)
                if pm_code:
                    self._apply_cell_color(pm_cell, pm_code)

    def _fill_call_row(
        self,
        sheet,
        call_rows: list[dict[str, Any]],
        block_start: date,
        block_end: date,
    ) -> None:
        """Fill call rows with staff names, routing faculty→row 4, residents→row 5.

        Writes staff name to AM column only (even columns 6, 8, 10, ...).
        User can manually merge AM/PM cells in Excel if desired.

        Row positions: self.call_row (faculty, default 4),
                       self.resident_call_row (residents, default 5).
        """

        # Build date -> (staff, target_row) lookup
        # Faculty go to staff call row (4), residents to resident call row (5)
        call_lookup: dict[tuple[date, int], str] = {}
        for night in call_rows:
            night_date_val = night.get("date")
            night_date = self._coerce_date(night_date_val)
            if not night_date:
                continue
            staff_name = night.get("staff", "")
            if not staff_name:
                continue
            person_type = night.get("person_type", "faculty")
            target_row = (
                self.resident_call_row if person_type == "resident" else self.call_row
            )
            call_lookup[(night_date, target_row)] = staff_name

        # Write to call rows, AM column only (col 6, 8, 10, ...)
        current = block_start
        col = COL_SCHEDULE_START  # Column 6
        while current <= block_end:
            for target_row in (self.call_row, self.resident_call_row):
                staff = call_lookup.get((current, target_row), "")
                if staff:
                    cell = self._get_writable_cell(sheet, row=target_row, column=col)
                    cell.value = staff
            current += timedelta(days=1)
            col += 2  # Skip PM column (write to AM only)

    def _calculate_block_number(self, block_start: date) -> int:
        """Calculate block number from start date."""
        block_number, _ = get_block_number_for_date(block_start)
        return block_number


def convert_xml_to_xlsx(
    xml_string: str,
    output_path: Path | str | None = None,
    template_path: Path | str | None = None,
    apply_colors: bool = True,
) -> bytes:
    """
    Convenience function to convert XML to xlsx.

    Args:
        xml_string: XML from ScheduleXMLExporter
        output_path: Optional path to save xlsx
        template_path: Optional template xlsx
        apply_colors: Whether to apply TAMC color scheme (default: True)

    Returns:
        xlsx bytes
    """
    converter = XMLToXlsxConverter(template_path, apply_colors=apply_colors)
    return converter.convert_from_string(xml_string, output_path)
