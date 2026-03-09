"""Excel metadata utility for stateful round-trips.

Phase 1 of the Stateful Round-Trip roadmap ("Phantom Database").
Hidden metadata sheets enable import-side validation (block/year mismatch
rejection, stale-file detection) and schedule-level checksumming for
smart diff.
"""

from __future__ import annotations

import json
import hashlib
from dataclasses import dataclass, asdict, fields
from datetime import datetime, UTC
from typing import Any
from uuid import UUID

from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName

METADATA_SHEET_NAME = "__SYS_META__"
EXPORT_VERSION = "1.0"


@dataclass
class ExportMetadata:
    """Metadata blob for exported Excel files.

    Stored as a JSON string in cell A1 of the veryHidden __SYS_META__ sheet.
    Fields are additive — new fields MUST have defaults so that older
    readers (``from_json``) can ignore them gracefully.
    """

    academic_year: int
    export_timestamp: str
    block_number: int | None = None  # Optional for year-level exports
    export_version: str = EXPORT_VERSION
    exported_by: str | None = None  # Username or system identifier
    block_start_date: str | None = None  # ISO 8601 date
    block_end_date: str | None = None  # ISO 8601 date
    row_count: int = 0  # Number of data rows exported
    checksum: str = ""  # SHA-256 of schedule data (compute_schedule_checksum)
    block_map: dict[str, str] | None = None  # sheet_name -> block_uuid
    llm_rules_of_engagement: str | None = (
        None  # AI agent instructions embedded in the workbook
    )

    def to_json(self) -> str:
        """Convert metadata to JSON string."""
        d = asdict(self)
        # Clean up nulls for cleaner JSON
        return json.dumps({k: v for k, v in d.items() if v is not None})

    @classmethod
    def from_json(cls, data: str) -> ExportMetadata:
        """Load metadata from JSON string.

        Tolerates unknown keys so that workbooks written by a newer
        version of the exporter can still be read by older code.
        """
        raw = json.loads(data)
        # Filter to only known fields for forward compatibility
        known = {f.name for f in fields(cls)}
        filtered = {k: v for k, v in raw.items() if k in known}
        # Handle legacy int export_version -> str conversion
        if "export_version" in filtered and isinstance(filtered["export_version"], int):
            filtered["export_version"] = str(filtered["export_version"])
        return cls(**filtered)


def write_sys_meta_sheet(wb: Workbook, meta: ExportMetadata) -> None:
    """Create a veryHidden sheet __SYS_META__ with metadata JSON.

    ``veryHidden`` sheets cannot be unhidden via the Excel UI — only VBA
    or openpyxl can access them.  This prevents coordinators from
    accidentally deleting the metadata.
    """
    if METADATA_SHEET_NAME in wb.sheetnames:
        del wb[METADATA_SHEET_NAME]
    ws = wb.create_sheet(METADATA_SHEET_NAME)
    ws.sheet_state = "veryHidden"
    ws.cell(row=1, column=1, value=meta.to_json())


def read_sys_meta(wb: Workbook) -> ExportMetadata | None:
    """Read metadata from __SYS_META__ sheet if present.

    Returns ``None`` for legacy workbooks without metadata (backward
    compatible — callers gate new behaviour on ``meta is not None``).
    """
    if METADATA_SHEET_NAME not in wb.sheetnames:
        return None
    ws = wb[METADATA_SHEET_NAME]
    val = ws.cell(row=1, column=1).value
    if not val:
        return None
    try:
        return ExportMetadata.from_json(str(val))
    except Exception:
        return None


def write_ref_sheet(
    wb: Workbook, rotation_codes: list[str], activity_codes: list[str]
) -> None:
    """
    Create a veryHidden sheet __REF__ with Named Ranges for validation.

    Named Ranges:
    - ValidRotations: range of valid rotation abbreviations
    - ValidActivities: range of valid activity codes
    """
    if "__REF__" in wb.sheetnames:
        del wb["__REF__"]
    ws = wb.create_sheet("__REF__")
    ws.sheet_state = "veryHidden"

    # Header
    ws.cell(row=1, column=1, value="ValidRotations")
    ws.cell(row=1, column=2, value="ValidActivities")

    # Data
    sorted_rotations = sorted(set(code for code in rotation_codes if code))
    sorted_activities = sorted(set(code for code in activity_codes if code))

    for i, code in enumerate(sorted_rotations, start=2):
        ws.cell(row=i, column=1, value=code)

    for i, code in enumerate(sorted_activities, start=2):
        ws.cell(row=i, column=2, value=code)

    # Named Ranges using DefinedName
    # ValidRotations: Col A
    rot_range = f"'{ws.title}'!$A$2:$A${max(2, len(sorted_rotations) + 1)}"
    wb.defined_names.add(DefinedName("ValidRotations", attr_text=rot_range))

    # ValidActivities: Col B
    act_range = f"'{ws.title}'!$B$2:$B${max(2, len(sorted_activities) + 1)}"
    wb.defined_names.add(DefinedName("ValidActivities", attr_text=act_range))


def read_ref_codes(wb: Workbook) -> dict[str, list[str]]:
    """Read reference codes from __REF__ sheet."""
    res = {"rotations": [], "activities": []}
    if "__REF__" not in wb.sheetnames:
        return res

    ws = wb["__REF__"]

    # Read rotations (Col A)
    i = 2
    while True:
        val = ws.cell(row=i, column=1).value
        if not val:
            break
        res["rotations"].append(str(val))
        i += 1

    # Read activities (Col B)
    i = 2
    while True:
        val = ws.cell(row=i, column=2).value
        if not val:
            break
        res["activities"].append(str(val))
        i += 1

    return res


def write_baseline_sheet(
    wb: Workbook,
    sheet_name: str,
    cell_data: list[dict[str, Any]],
) -> None:
    """Write a veryHidden __BASELINE__ sheet with cell fingerprints.

    Each entry records the system-generated value for a data cell so that
    hand-jammed edits can be detected on reimport.

    Args:
        wb: Target workbook.
        sheet_name: Block sheet name (e.g. "Block 12").
        cell_data: List of dicts with keys: cell_ref, value, row_hash, source.
    """
    baseline_name = f"__BASELINE_{sheet_name}__"
    if baseline_name in wb.sheetnames:
        del wb[baseline_name]
    ws = wb.create_sheet(baseline_name)
    ws.sheet_state = "veryHidden"

    headers = ["cell_ref", "value", "row_hash", "source"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)

    for i, entry in enumerate(cell_data, start=2):
        ws.cell(row=i, column=1, value=entry.get("cell_ref", ""))
        ws.cell(row=i, column=2, value=entry.get("value"))
        ws.cell(row=i, column=3, value=entry.get("row_hash", ""))
        ws.cell(row=i, column=4, value=entry.get("source", ""))


def read_baseline(wb: Workbook, sheet_name: str) -> dict[str, dict[str, Any]]:
    """Read baseline data for a block sheet.

    Returns:
        Dict mapping cell_ref -> {value, row_hash, source}.
    """
    baseline_name = f"__BASELINE_{sheet_name}__"
    if baseline_name not in wb.sheetnames:
        return {}
    ws = wb[baseline_name]
    result: dict[str, dict[str, Any]] = {}
    row = 2
    while True:
        cell_ref = ws.cell(row=row, column=1).value
        if not cell_ref:
            break
        result[str(cell_ref)] = {
            "value": ws.cell(row=row, column=2).value,
            "row_hash": ws.cell(row=row, column=3).value,
            "source": ws.cell(row=row, column=4).value,
        }
        row += 1
    return result


def write_overrides_sheet(
    wb: Workbook,
    sheet_name: str,
    overrides: list[dict[str, Any]],
) -> None:
    """Write a veryHidden __OVERRIDES__ sheet tracking hand-jammed cells.

    Args:
        wb: Target workbook.
        sheet_name: Block sheet name (e.g. "Block 12").
        overrides: List of dicts with keys: cell_ref, original_value,
            new_value, person_name, date, time_of_day.
    """
    override_name = f"__OVERRIDES_{sheet_name}__"
    if override_name in wb.sheetnames:
        del wb[override_name]
    if not overrides:
        return  # Don't create empty sheet
    ws = wb.create_sheet(override_name)
    ws.sheet_state = "veryHidden"

    headers = [
        "cell_ref",
        "original_value",
        "new_value",
        "person_name",
        "date",
        "time_of_day",
    ]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)

    for i, entry in enumerate(overrides, start=2):
        for col, key in enumerate(headers, start=1):
            ws.cell(row=i, column=col, value=entry.get(key))


def normalize_for_hash(val: Any) -> str:
    """Aggressively normalize cell values for symmetric export/import hashes.

    Excel may coerce types (e.g. "1.0" for integer 1), add trailing spaces,
    or change case. This ensures the same logical value hashes identically
    regardless of whether it comes from JSON export or Excel cell read.
    """
    if val is None or str(val).strip() == "":
        return ""
    s = str(val).strip().upper()
    # Excel renders integer-like floats as "1.0" — normalize to "1"
    if s.endswith(".0"):
        s = s[:-2]
    return s


def compute_row_hash(
    person_id: UUID,
    rotation1: str | None,
    rotation2: str | None,
    days: list[str | None],
) -> str:
    """
    Compute MD5 hash of row content for change detection.

    Used to skip unchanged rows during import. Uses normalize_for_hash()
    to ensure symmetric hashing between export (JSON strings) and import
    (Excel cell values subject to type coercion).
    """
    nr1 = normalize_for_hash(rotation1)
    nr2 = normalize_for_hash(rotation2)
    nd = "|".join(normalize_for_hash(d) for d in days)
    data = f"{person_id}|{nr1}|{nr2}|{nd}"
    return hashlib.md5(data.encode()).hexdigest()  # nosec B324 — not for security


def compute_schedule_checksum(data_rows: list[list[Any]]) -> str:
    """SHA-256 checksum of schedule data for change detection.

    Produces a deterministic fingerprint of the entire schedule grid so
    that the import side can detect whether the file's content has been
    modified since export (beyond what row-level hashes catch, e.g.
    added/removed rows).

    Args:
        data_rows: List of rows, where each row is a list of cell values.
            Typically the schedule grid (names + activity codes).

    Returns:
        Hex-encoded SHA-256 digest string.
    """
    content = json.dumps(data_rows, sort_keys=True, default=str)
    return hashlib.sha256(content.encode()).hexdigest()
