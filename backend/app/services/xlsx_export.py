"""
Legacy Excel export service.

Generates Excel files in the historical format used for schedule distribution.
Format: 4-week block view with AM/PM columns per day, color-coded rotations.
"""

import io
from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session, joinedload

from app.models.absence import Absence
from app.models.assignment import Assignment
from app.models.block import Block
from app.models.person import Person

# Color definitions (ARGB format for openpyxl)
COLORS = {
    # Rotation type colors (for column A labels)
    "military": PatternFill(
        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
    ),  # Yellow
    "night_float": PatternFill(
        start_color="FF0000", end_color="FF0000", fill_type="solid"
    ),  # Red
    "nicu": PatternFill(
        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
    ),  # Yellow
    "fmit": PatternFill(
        start_color="00B0F0", end_color="00B0F0", fill_type="solid"
    ),  # Blue
    "default": PatternFill(
        start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
    ),  # White
    # Special cell colors
    "header_gray": PatternFill(
        start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"
    ),
    "weekend": PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid"),
    "holiday": PatternFill(
        start_color="92D050", end_color="92D050", fill_type="solid"
    ),  # Green
    "highlight_yellow": PatternFill(
        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
    ),
    "highlight_red": PatternFill(
        start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"
    ),
}

# Font styles
FONTS = {
    "header": Font(bold=True, size=11),
    "block_number": Font(bold=True, size=36),
    "date": Font(size=10),
    "data": Font(size=9),
    "red_text": Font(size=9, color="FF0000"),
}

# Border styles
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def get_rotation_color(rotation_name: str, abbreviation: str) -> PatternFill:
    """Get the background color for a rotation based on name/abbreviation."""
    name_lower = rotation_name.lower() if rotation_name else ""
    abbrev_upper = abbreviation.upper() if abbreviation else ""

    if "military" in name_lower or abbrev_upper in ["MIL", "DEP"]:
        return COLORS["military"]
    elif "night" in name_lower or abbrev_upper == "NF":
        return COLORS["night_float"]
    elif "nicu" in name_lower or abbrev_upper == "NICU":
        return COLORS["nicu"]
    elif "fmit" in name_lower or abbrev_upper == "FMIT":
        return COLORS["fmit"]
    return COLORS["default"]


def calculate_block_dates(
    block_number: int, academic_year_start: date
) -> tuple[date, date]:
    """
    Calculate start and end dates for a given block number.
    Each block is 4 weeks (28 days).
    Academic year typically starts July 1.
    """
    block_start = academic_year_start + timedelta(days=(block_number - 1) * 28)
    block_end = block_start + timedelta(days=27)
    return block_start, block_end


def get_day_abbreviation(d: date) -> str:
    """Get short day name."""
    days = ["MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN"]
    return days[d.weekday()]


def format_date_header(d: date) -> str:
    """Format date for header row (e.g., '12-Feb')."""
    return d.strftime("%-d-%b")


class LegacyXlsxExporter:
    """Generates Excel files in the legacy schedule format."""

    def __init__(self, db: Session):
        self.db = db
        self.wb = Workbook()
        # Remove default sheet
        self.wb.remove(self.wb.active)

    def generate_block_schedule(
        self,
        block_number: int,
        start_date: date,
        end_date: date,
        federal_holidays: list[date] | None = None,
    ) -> None:
        """Generate a single block sheet in legacy format."""

        if federal_holidays is None:
            federal_holidays = []

        # Create sheet
        sheet_name = f"Block {block_number}"
        ws = self.wb.create_sheet(title=sheet_name)

        # Get all people grouped by type and PGY level
        residents = (
            self.db.query(Person)
            .filter(Person.type == "resident")
            .order_by(Person.pgy_level.desc(), Person.name)
            .all()
        )

        faculty = (
            self.db.query(Person)
            .filter(Person.type == "faculty")
            .order_by(Person.name)
            .all()
        )

        # Get assignments for this date range
        assignments = (
            self.db.query(Assignment)
            .options(
                joinedload(Assignment.block),
                joinedload(Assignment.person),
                joinedload(Assignment.rotation_template),
            )
            .join(Block)
            .filter(Block.date >= start_date, Block.date <= end_date)
            .all()
        )

        # Build assignment lookup: {(person_id, date, time_of_day): assignment}
        assignment_lookup = {}
        for a in assignments:
            if a.block:
                key = (str(a.person_id), a.block.date, a.block.time_of_day)
                assignment_lookup[key] = a

        # Get absences for this date range
        absences = (
            self.db.query(Absence)
            .filter(Absence.start_date <= end_date, Absence.end_date >= start_date)
            .all()
        )

        # Build absence lookup: {(person_id, date): absence_type}
        absence_lookup = {}
        for absence in absences:
            current = absence.start_date
            while current <= absence.end_date:
                if start_date <= current <= end_date:
                    absence_lookup[(str(absence.person_id), current)] = (
                        absence.absence_type
                    )
                current += timedelta(days=1)

        # Calculate number of days
        num_days = (end_date - start_date).days + 1

        # Write header section
        self._write_header(ws, block_number, start_date, num_days, federal_holidays)

        # Write resident data (starting at row 9)
        current_row = 9
        current_pgy = None

        for resident in residents:
            # Add gap row between PGY levels
            if current_pgy is not None and resident.pgy_level != current_pgy:
                current_row += 1
            current_pgy = resident.pgy_level

            self._write_person_row(
                ws,
                current_row,
                resident,
                start_date,
                num_days,
                assignment_lookup,
                absence_lookup,
                federal_holidays,
            )
            current_row += 1

        # Add gap before faculty section
        current_row += 2

        # Write faculty data
        for fac in faculty:
            self._write_person_row(
                ws,
                current_row,
                fac,
                start_date,
                num_days,
                assignment_lookup,
                absence_lookup,
                federal_holidays,
                is_faculty=True,
            )
            current_row += 1

        # Adjust column widths
        self._adjust_column_widths(ws, num_days)

    def _write_header(
        self,
        ws,
        block_number: int,
        start_date: date,
        num_days: int,
        federal_holidays: list[date],
    ) -> None:
        """Write the header section (rows 1-8)."""

        # Block number in merged cell C1:D3
        ws.merge_cells("C1:D3")
        ws["C1"] = block_number
        ws["C1"].font = FONTS["block_number"]
        ws["C1"].alignment = Alignment(horizontal="center", vertical="center")

        # Date range label
        end_date = start_date + timedelta(days=num_days - 1)
        ws["C4"] = f"{start_date.strftime('%-d%b')}-{end_date.strftime('%-d%b')}"
        ws["C4"].font = FONTS["header"]

        # Row labels
        ws["E3"] = "Date:"
        ws["E4"] = "Staff Call"
        ws["E5"] = "Resident Call"
        ws["C6"] = "TEMPLATE"
        ws["D6"] = "ROLE"
        ws["E6"] = "PROVIDER"

        for cell in ["E3", "E4", "E5", "C6", "D6", "E6"]:
            ws[cell].font = FONTS["header"]

        # Day headers and dates
        col = 6  # Start at column F
        for day_offset in range(num_days):
            current_date = start_date + timedelta(days=day_offset)

            # Each day has 2 columns (AM/PM)
            am_col = col
            pm_col = col + 1
            am_letter = get_column_letter(am_col)
            pm_letter = get_column_letter(pm_col)

            # Row 1-2: Day of week (merged)
            ws.merge_cells(f"{am_letter}1:{pm_letter}1")
            ws[f"{am_letter}1"] = get_day_abbreviation(current_date)
            ws[f"{am_letter}1"].font = FONTS["header"]
            ws[f"{am_letter}1"].alignment = Alignment(horizontal="center")

            # Row 2: AM/PM labels
            ws[f"{am_letter}2"] = "am"
            ws[f"{pm_letter}2"] = "pm"

            # Row 3: Date (merged)
            ws.merge_cells(f"{am_letter}3:{pm_letter}3")
            ws[f"{am_letter}3"] = current_date
            ws[f"{am_letter}3"].number_format = "D-MMM"
            ws[f"{am_letter}3"].alignment = Alignment(horizontal="center")

            # Highlight weekends and holidays
            if current_date.weekday() >= 5:  # Weekend
                ws[f"{am_letter}1"].fill = COLORS["weekend"]
                ws[f"{pm_letter}1"].fill = COLORS["weekend"]
            if current_date in federal_holidays:
                ws[f"{am_letter}3"].fill = COLORS["holiday"]
                ws[f"{pm_letter}3"].fill = COLORS["holiday"]

            col += 2

    def _write_person_row(
        self,
        ws,
        row: int,
        person: Person,
        start_date: date,
        num_days: int,
        assignment_lookup: dict,
        absence_lookup: dict,
        federal_holidays: list[date],
        is_faculty: bool = False,
    ) -> None:
        """Write a single person's row with their schedule."""

        # Column A: Rotation label (with color) - use their current rotation if any
        # For simplicity, we'll determine this from their first assignment
        first_assignment = None
        for day_offset in range(num_days):
            d = start_date + timedelta(days=day_offset)
            key = (str(person.id), d, "AM")
            if key in assignment_lookup:
                first_assignment = assignment_lookup[key]
                break

        rotation_name = ""
        if first_assignment and first_assignment.rotation_template:
            rotation_name = first_assignment.rotation_template.name
            ws["A" + str(row)] = rotation_name
            ws["A" + str(row)].fill = get_rotation_color(
                rotation_name, first_assignment.rotation_template.abbreviation
            )

        # Column C: PGY level code (R1, R2, R3) or C19 for faculty
        if is_faculty:
            ws["C" + str(row)] = "C19"
        else:
            ws["C" + str(row)] = f"R{person.pgy_level}" if person.pgy_level else ""

        # Column D: PGY level text or FAC
        if is_faculty:
            ws["D" + str(row)] = "FAC"
        else:
            ws["D" + str(row)] = f"PGY {person.pgy_level}" if person.pgy_level else ""

        # Column E: Provider name
        ws["E" + str(row)] = person.name
        ws["E" + str(row)].font = FONTS["data"]

        # Data columns (F onwards)
        col = 6
        for day_offset in range(num_days):
            current_date = start_date + timedelta(days=day_offset)

            for time_of_day in ["AM", "PM"]:
                cell = ws.cell(row=row, column=col)

                # Set default font first (avoids StyleProxy hashing issue)
                cell.font = FONTS["data"]

                # Check for absence first
                absence_key = (str(person.id), current_date)
                if absence_key in absence_lookup:
                    absence_type = absence_lookup[absence_key]
                    cell.value = self._get_absence_abbreviation(absence_type)
                    cell.fill = COLORS["highlight_yellow"]
                else:
                    # Check for assignment
                    assignment_key = (str(person.id), current_date, time_of_day)
                    if assignment_key in assignment_lookup:
                        assignment = assignment_lookup[assignment_key]
                        cell.value = assignment.abbreviation

                        # Color code certain assignments
                        if assignment.rotation_template:
                            abbrev = assignment.rotation_template.abbreviation or ""
                            if abbrev.upper() in ["PR", "VAS", "EPIC"]:
                                cell.font = FONTS["red_text"]
                            elif abbrev.upper() in ["OFF", "HOL", "FED"]:
                                cell.fill = COLORS["holiday"]
                    else:
                        # Default to W (working) or empty
                        cell.value = "W"

                cell.alignment = Alignment(horizontal="center")
                col += 1

    def _get_absence_abbreviation(self, absence_type: str) -> str:
        """Get abbreviation for absence type."""
        mapping = {
            "vacation": "VAC",
            "sick": "SICK",
            "medical": "MED",
            "conference": "CONF",
            "deployment": "DEP",
            "tdy": "TDY",
            "family_emergency": "FEM",
            "personal": "PER",
        }
        return mapping.get(absence_type, "ABS")

    def _adjust_column_widths(self, ws, num_days: int) -> None:
        """Adjust column widths for readability."""
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 8
        ws.column_dimensions["C"].width = 6
        ws.column_dimensions["D"].width = 8
        ws.column_dimensions["E"].width = 15

        # Data columns (narrower for AM/PM)
        col = 6
        for _ in range(num_days * 2):
            ws.column_dimensions[get_column_letter(col)].width = 5
            col += 1

    def generate_full_year(
        self,
        academic_year_start: date,
        num_blocks: int = 13,
        federal_holidays: list[date] | None = None,
    ) -> None:
        """Generate all block sheets for a full academic year."""

        for block_num in range(1, num_blocks + 1):
            block_start, block_end = calculate_block_dates(
                block_num, academic_year_start
            )
            self.generate_block_schedule(
                block_number=block_num,
                start_date=block_start,
                end_date=block_end,
                federal_holidays=federal_holidays or [],
            )

    def save_to_bytes(self) -> bytes:
        """Save workbook to bytes for HTTP response."""
        output = io.BytesIO()
        self.wb.save(output)
        output.seek(0)
        return output.getvalue()


def generate_legacy_xlsx(
    db: Session,
    start_date: date,
    end_date: date,
    block_number: int | None = None,
    federal_holidays: list[date] | None = None,
) -> bytes:
    """
    Generate legacy format Excel export.

    Args:
        db: Database session
        start_date: Start date of the export
        end_date: End date of the export
        block_number: Block number to display in header (calculated if not provided)
        federal_holidays: List of federal holiday dates for highlighting

    Returns:
        Excel file as bytes
    """
    exporter = LegacyXlsxExporter(db)

    # Calculate block number if not provided (assume 28-day blocks starting July 1)
    if block_number is None:
        # Rough calculation - days since July 1 divided by 28
        july_1 = date(
            start_date.year if start_date.month >= 7 else start_date.year - 1, 7, 1
        )
        days_since_start = (start_date - july_1).days
        block_number = (days_since_start // 28) + 1

    exporter.generate_block_schedule(
        block_number=block_number,
        start_date=start_date,
        end_date=end_date,
        federal_holidays=federal_holidays or [],
    )

    return exporter.save_to_bytes()
