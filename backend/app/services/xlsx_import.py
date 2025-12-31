"""
Excel import service for clinic schedules and block rotation schedules.

Two main parsers:
1. ClinicScheduleImporter - Parses clinic day schedules, detects conflicts
2. BlockScheduleParser - Fuzzy-tolerant parser for block rotation sheets

Block schedule parsing handles human-induced drift:
- Column shifts from copy/paste
- Row insertions/deletions
- Name typos and variations
- Merged cells
- Missing/malformed data

Uses semantic anchors rather than hard-coded positions.
"""

import io
import re
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from difflib import SequenceMatcher
from enum import Enum
from pathlib import Path
from typing import Optional
import logging

from openpyxl import load_workbook
from openpyxl.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session

from app.models.absence import Absence as AbsenceModel
from app.models.person import Person

logger = logging.getLogger(__name__)


class SlotType(Enum):
    """Types of schedule slots."""

    CLINIC = "clinic"
    FMIT = "fmit"
    OFF = "off"
    VACATION = "vacation"
    CONFERENCE = "conference"
    ADMIN = "admin"
    UNKNOWN = "unknown"


@dataclass
class ScheduleSlot:
    """Represents a single half-day slot in the schedule."""

    provider_name: str
    date: date
    time_of_day: str  # "AM" or "PM"
    slot_type: SlotType
    raw_value: str
    location: str | None = None
    notes: str | None = None

    @property
    def key(self) -> tuple:
        """Unique key for this slot."""
        return (self.provider_name, self.date, self.time_of_day)


@dataclass
class ScheduleConflict:
    """Represents a scheduling conflict."""

    provider_name: str
    date: date
    time_of_day: str
    conflict_type: str  # "double_book", "fmit_clinic_overlap", "specialty_unavailable"
    fmit_assignment: str | None = None
    clinic_assignment: str | None = None
    severity: str = "warning"  # "error", "warning", "info"
    message: str = ""

    def __post_init__(self) -> None:
        if not self.message:
            self.message = self._generate_message()

    def _generate_message(self) -> str:
        """Generate a human-readable conflict message."""
        date_str = self.date.strftime("%a %b %d")
        if self.conflict_type == "double_book":
            return f"{self.provider_name} double-booked on {date_str} {self.time_of_day}: FMIT={self.fmit_assignment}, Clinic={self.clinic_assignment}"
        elif self.conflict_type == "fmit_clinic_overlap":
            return f"{self.provider_name} has FMIT during clinic on {date_str} {self.time_of_day}"
        elif self.conflict_type == "specialty_unavailable":
            return f"{self.provider_name} (specialty provider) unavailable for clinic on {date_str} {self.time_of_day}"
        elif self.conflict_type == "consecutive_weeks":
            return f"{self.provider_name} has alternating week pattern causing family hardship"
        return f"Conflict for {self.provider_name} on {date_str} {self.time_of_day}"


@dataclass
class ProviderSchedule:
    """Complete schedule for a single provider."""

    name: str
    specialties: list[str] = field(default_factory=list)
    slots: dict[tuple, ScheduleSlot] = field(
        default_factory=dict
    )  # key: (date, time_of_day)

    def add_slot(self, slot: ScheduleSlot) -> None:
        """Add a slot to this provider's schedule."""
        key = (slot.date, slot.time_of_day)
        self.slots[key] = slot

    def get_slot(self, d: date, time_of_day: str) -> ScheduleSlot | None:
        """Get slot for a specific date/time."""
        return self.slots.get((d, time_of_day))

    def get_fmit_weeks(self) -> list[tuple[date, date]]:
        """Get list of FMIT week ranges (Mon-Sun)."""
        fmit_dates = sorted(
            [
                slot.date
                for slot in self.slots.values()
                if slot.slot_type == SlotType.FMIT
            ]
        )

        if not fmit_dates:
            return []

        weeks = []
        current_week_start = None
        current_week_end = None

        for d in fmit_dates:
            # Get Monday of this week
            week_start = d - timedelta(days=d.weekday())
            week_end = week_start + timedelta(days=6)

            if current_week_start is None:
                current_week_start = week_start
                current_week_end = week_end
            elif week_start != current_week_start:
                weeks.append((current_week_start, current_week_end))
                current_week_start = week_start
                current_week_end = week_end

        if current_week_start:
            weeks.append((current_week_start, current_week_end))

        return weeks

    def has_alternating_pattern(self) -> bool:
        """Check if FMIT schedule has week-on/week-off pattern (hard on families)."""
        weeks = self.get_fmit_weeks()
        if len(weeks) < 3:
            return False

        # Check for alternating pattern (gap of exactly 1 week between FMIT weeks)
        gaps = []
        for i in range(1, len(weeks)):
            prev_end = weeks[i - 1][1]
            curr_start = weeks[i][0]
            gap_days = (curr_start - prev_end).days
            gaps.append(gap_days)

        # Alternating = gaps of ~7 days (1 week off between FMIT weeks)
        alternating_count = sum(1 for g in gaps if 6 <= g <= 8)
        return alternating_count >= 2  # At least 2 alternating gaps


@dataclass
class ImportResult:
    """Result of importing a schedule file."""

    success: bool
    providers: dict[str, ProviderSchedule] = field(default_factory=dict)
    conflicts: list[ScheduleConflict] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)

    # Statistics
    total_slots: int = 0
    clinic_slots: int = 0
    fmit_slots: int = 0
    date_range: tuple[date | None, date | None] = (None, None)

    def add_conflict(self, conflict: ScheduleConflict) -> None:
        """Add a conflict to the result."""
        self.conflicts.append(conflict)

    def get_conflicts_by_provider(self, provider_name: str) -> list[ScheduleConflict]:
        """Get all conflicts for a specific provider."""
        return [c for c in self.conflicts if c.provider_name == provider_name]

    def get_conflicts_by_type(self, conflict_type: str) -> list[ScheduleConflict]:
        """Get all conflicts of a specific type."""
        return [c for c in self.conflicts if c.conflict_type == conflict_type]


class ClinicScheduleImporter:
    """
    Imports clinic schedules from Excel files.

    Supports multiple formats:
    - Standard format: providers in rows, dates in columns
    - Transposed format: dates in rows, providers in columns

    Auto-detects format based on header structure.
    """

    # Common abbreviations for slot types
    # -------------------------------------------------------------------------
    # SLOT TYPE MAPPING
    # Maps cell values to schedule slot types. Direct lookups are checked
    # first, then fuzzy matching (number stripping, prefix matching).
    # -------------------------------------------------------------------------
    SLOT_TYPE_MAPPING = {
        # --- CLINIC (Schedulable patient care) ---
        "c": SlotType.CLINIC,
        "cc": SlotType.CLINIC,  # Continuity Clinic
        "clc": SlotType.CLINIC,
        "clinic": SlotType.CLINIC,
        "cv": SlotType.CLINIC,  # Virtual Clinic
        "pts": SlotType.CLINIC,
        "patient": SlotType.CLINIC,
        "appt": SlotType.CLINIC,
        "sm": SlotType.CLINIC,  # Sports Medicine
        "asm": SlotType.CLINIC,  # Advanced Sports Medicine
        "sports": SlotType.CLINIC,
        "pr": SlotType.CLINIC,  # Procedures
        "vas": SlotType.CLINIC,  # Vasectomy Clinic
        "pedc": SlotType.CLINIC,  # Peds Clinic
        "pedsp": SlotType.CLINIC,  # Peds Specialty
        "hv": SlotType.CLINIC,  # Home Visit
        "hc": SlotType.CLINIC,  # Home Care
        "c-i": SlotType.CLINIC,  # Clinic Internship
        "rcc": SlotType.CLINIC,  # Resident Continuity Clinic
        # --- FMIT / INPATIENT (Blocks clinic availability) ---
        "fmit": SlotType.FMIT,
        "nf": SlotType.FMIT,  # Night Float
        "nicu": SlotType.FMIT,  # NICU rotation
        "pedw": SlotType.FMIT,  # Peds Ward
        "imw": SlotType.FMIT,  # IM Ward
        "inpt": SlotType.FMIT,
        "inpatient": SlotType.FMIT,
        "ward": SlotType.FMIT,
        "wards": SlotType.FMIT,
        "er": SlotType.FMIT,  # ER shifts block clinic
        "kap": SlotType.FMIT,  # Kapiolani (offsite hospital)
        "straub": SlotType.FMIT,  # Straub (offsite hospital)
        "oic": SlotType.FMIT,  # Officer In Charge (on-call)
        # --- OFF / UNAVAILABLE (Do not schedule) ---
        "off": SlotType.OFF,
        "pc": SlotType.OFF,  # Post Call
        "do": SlotType.OFF,  # Day Off
        "w": SlotType.OFF,  # Weekend (context: Sat/Sun cells)
        "x": SlotType.OFF,
        "-": SlotType.OFF,
        "": SlotType.OFF,
        "fed": SlotType.OFF,  # Federal Holiday
        "hol": SlotType.OFF,  # Holiday
        # --- VACATION / LEAVE ---
        "vac": SlotType.VACATION,
        "vacation": SlotType.VACATION,
        "lv": SlotType.VACATION,
        "leave": SlotType.VACATION,
        "al": SlotType.VACATION,  # Annual Leave
        "dep": SlotType.VACATION,  # Deployment
        "tdy": SlotType.VACATION,  # TDY (temporary duty)
        # --- CONFERENCE / ACADEMIC (Protected time) ---
        "conf": SlotType.CONFERENCE,
        "conference": SlotType.CONFERENCE,
        "cme": SlotType.CONFERENCE,
        "mtg": SlotType.CONFERENCE,
        "lec": SlotType.CONFERENCE,  # Lecture
        "sim": SlotType.CONFERENCE,  # Simulation
        "usafp": SlotType.CONFERENCE,
        "hafp": SlotType.CONFERENCE,
        "facdev": SlotType.CONFERENCE,  # Faculty Development
        # --- ADMIN (Administrative time) ---
        "admin": SlotType.ADMIN,
        "adm": SlotType.ADMIN,
        "office": SlotType.ADMIN,
        "gme": SlotType.ADMIN,  # GME Time
        "rsh": SlotType.ADMIN,  # Research
        "pi": SlotType.ADMIN,  # Process Improvement
        "at": SlotType.ADMIN,  # Admin Time
        "pcat": SlotType.ADMIN,  # Patient Care Admin Team
        "fac": SlotType.ADMIN,  # Faculty (administrative)
        "dm": SlotType.ADMIN,  # Department Meeting
        "dfm": SlotType.ADMIN,  # DFM administrative
    }

    def __init__(self, db: Session | None = None) -> None:
        self.db = db
        self.known_providers: dict[str, Person] = {}

        if db:
            self._load_known_providers()

    def _load_known_providers(self) -> None:
        """Load known providers from database for matching."""
        if not self.db:
            return

        people = self.db.query(Person).filter(Person.type == "faculty").all()
        for person in people:
            # Store by lowercase name for fuzzy matching
            self.known_providers[person.name.lower()] = person

    def classify_slot(self, value: str) -> SlotType:
        """
        Classify a cell value into a slot type with fuzzy matching.

        Matching order:
        1. Handle compound codes (e.g., "PC/OFF") - prioritize restrictive types
        2. Direct lookup in mapping
        3. Strip trailing numbers (e.g., "C30" -> "C")
        4. Prefix matching for longer keys (2+ chars)
        5. Fallback heuristics for common terms
        """
        if value is None:
            return SlotType.OFF

        # Normalize: lowercase, strip whitespace
        clean = str(value).strip().lower()

        if not clean:
            return SlotType.OFF

        # 1. Handle compound codes like "PC/OFF" or "C/CV"
        if "/" in clean:
            parts = clean.split("/")
            # Prioritize restrictive types (safety first)
            for part in parts:
                part = part.strip()
                if part in self.SLOT_TYPE_MAPPING:
                    slot_type = self.SLOT_TYPE_MAPPING[part]
                    if slot_type in (SlotType.FMIT, SlotType.OFF, SlotType.VACATION):
                        return slot_type
            # Otherwise classify first part
            first_part = parts[0].strip()
            if first_part:
                return self.classify_slot(first_part)

        # 2. Direct lookup
        if clean in self.SLOT_TYPE_MAPPING:
            return self.SLOT_TYPE_MAPPING[clean]

        # 3. Strip trailing numbers (e.g., "C30" -> "C", "FMIT2" -> "FMIT")
        clean_no_nums = "".join(c for c in clean if not c.isdigit())
        if clean_no_nums and clean_no_nums in self.SLOT_TYPE_MAPPING:
            return self.SLOT_TYPE_MAPPING[clean_no_nums]

        # 4. Prefix matching - only for keys with 2+ chars to avoid false positives
        for key, slot_type in self.SLOT_TYPE_MAPPING.items():
            if len(key) >= 2 and clean.startswith(key):
                return slot_type

        # 5. Fallback heuristics for common terms
        if "call" in clean:
            return SlotType.FMIT
        if "clinic" in clean:
            return SlotType.CLINIC
        if "leave" in clean or "vacation" in clean:
            return SlotType.VACATION
        if "conf" in clean:
            return SlotType.CONFERENCE

        return SlotType.UNKNOWN

    def parse_date_from_header(self, value) -> date | None:
        """Parse a date from various header formats."""
        if value is None:
            return None

        # Already a date/datetime
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value

        # String formats
        text = str(value).strip()

        # Try common formats
        formats = [
            "%Y-%m-%d",
            "%m/%d/%Y",
            "%m/%d/%y",
            "%d-%b",
            "%d-%b-%Y",
            "%b %d",
            "%b %d, %Y",
        ]

        for fmt in formats:
            try:
                parsed = datetime.strptime(text, fmt)
                # If year missing, assume current academic year
                if parsed.year == 1900:
                    today = date.today()
                    if parsed.month >= 7:
                        parsed = parsed.replace(
                            year=today.year if today.month >= 7 else today.year - 1
                        )
                    else:
                        parsed = parsed.replace(
                            year=today.year if today.month < 7 else today.year + 1
                        )
                return parsed.date()
            except ValueError:
                continue

        return None

    def detect_format(self, ws: Worksheet) -> dict:
        """
        Detect the format of the schedule worksheet.

        Returns dict with:
        - orientation: "providers_in_rows" or "providers_in_columns"
        - header_row: row number containing date headers
        - data_start_row: first row of actual data
        - provider_col: column containing provider names
        - date_cols: list of (col_index, date, time_of_day) tuples
        """
        # Sample first 10 rows/cols to detect structure
        max_sample = 15

        # Look for dates in row 1-5
        date_row = None
        date_cols = []

        for row_idx in range(1, min(6, ws.max_row + 1)):
            found_dates = []
            for col_idx in range(1, min(max_sample, ws.max_column + 1)):
                cell = ws.cell(row=row_idx, column=col_idx)
                parsed_date = self.parse_date_from_header(cell.value)
                if parsed_date:
                    found_dates.append((col_idx, parsed_date))

            if len(found_dates) >= 3:  # At least 3 dates = likely header row
                date_row = row_idx
                date_cols = found_dates
                break

        if not date_row:
            # Try transposed format (dates in column A)
            found_dates = []
            for row_idx in range(1, min(max_sample, ws.max_row + 1)):
                cell = ws.cell(row=row_idx, column=1)
                parsed_date = self.parse_date_from_header(cell.value)
                if parsed_date:
                    found_dates.append((row_idx, parsed_date))

            if len(found_dates) >= 3:
                return {
                    "orientation": "providers_in_columns",
                    "date_col": 1,
                    "header_row": 1,
                    "date_rows": found_dates,
                }

        # Find provider column (usually leftmost non-date column)
        provider_col = 1
        for col_idx in range(1, min(6, ws.max_column + 1)):
            cell = ws.cell(row=date_row + 1 if date_row else 2, column=col_idx)
            if (
                cell.value
                and isinstance(cell.value, str)
                and not self.parse_date_from_header(cell.value)
            ):
                # Check if it looks like a name
                if len(str(cell.value)) > 2:
                    provider_col = col_idx
                    break

        # Detect AM/PM columns (dates might have 2 columns each)
        expanded_date_cols = []
        for i, (col_idx, d) in enumerate(date_cols):
            # Check if next column is also for same date (AM/PM split)
            if i + 1 < len(date_cols):
                next_col, next_date = date_cols[i + 1]
                if next_date == d:
                    expanded_date_cols.append((col_idx, d, "AM"))
                    expanded_date_cols.append((next_col, d, "PM"))
                    continue

            # Check row above or below for AM/PM labels
            am_pm_row = date_row - 1 if date_row > 1 else date_row + 1
            time_label = ws.cell(row=am_pm_row, column=col_idx).value
            if time_label and str(time_label).lower().strip() in ["am", "pm"]:
                expanded_date_cols.append((col_idx, d, str(time_label).upper().strip()))
            else:
                # Single column per date - assume full day
                expanded_date_cols.append((col_idx, d, "AM"))

        return {
            "orientation": "providers_in_rows",
            "header_row": date_row or 1,
            "data_start_row": (date_row or 1) + 1,
            "provider_col": provider_col,
            "date_cols": (
                expanded_date_cols
                if expanded_date_cols
                else [(c, d, "AM") for c, d in date_cols]
            ),
        }

    def import_worksheet(
        self, ws: Worksheet, format_hint: dict | None = None
    ) -> ImportResult:
        """Import a single worksheet."""
        result = ImportResult(success=True)

        # Detect or use provided format
        fmt = format_hint or self.detect_format(ws)

        if fmt["orientation"] == "providers_in_rows":
            return self._import_providers_in_rows(ws, fmt, result)
        else:
            return self._import_providers_in_columns(ws, fmt, result)

    def _import_providers_in_rows(
        self, ws: Worksheet, fmt: dict, result: ImportResult
    ) -> ImportResult:
        """Import format where providers are in rows, dates in columns."""

        provider_col = fmt["provider_col"]
        date_cols = fmt["date_cols"]
        data_start_row = fmt["data_start_row"]

        min_date = None
        max_date = None

        for row_idx in range(data_start_row, ws.max_row + 1):
            provider_cell = ws.cell(row=row_idx, column=provider_col)
            provider_name = (
                str(provider_cell.value).strip() if provider_cell.value else ""
            )

            if not provider_name or provider_name.lower() in ["", "none", "null"]:
                continue

            # Get or create provider schedule
            if provider_name not in result.providers:
                result.providers[provider_name] = ProviderSchedule(name=provider_name)

            provider_schedule = result.providers[provider_name]

            # Process each date column
            for col_idx, slot_date, time_of_day in date_cols:
                cell = ws.cell(row=row_idx, column=col_idx)
                raw_value = str(cell.value).strip() if cell.value else ""

                slot_type = self.classify_slot(raw_value)

                slot = ScheduleSlot(
                    provider_name=provider_name,
                    date=slot_date,
                    time_of_day=time_of_day,
                    slot_type=slot_type,
                    raw_value=raw_value,
                )

                provider_schedule.add_slot(slot)
                result.total_slots += 1

                if slot_type == SlotType.CLINIC:
                    result.clinic_slots += 1
                elif slot_type == SlotType.FMIT:
                    result.fmit_slots += 1

                # Track date range
                if min_date is None or slot_date < min_date:
                    min_date = slot_date
                if max_date is None or slot_date > max_date:
                    max_date = slot_date

        result.date_range = (min_date, max_date)
        return result

    def _import_providers_in_columns(
        self, ws: Worksheet, fmt: dict, result: ImportResult
    ) -> ImportResult:
        """Import format where dates are in rows, providers in columns."""

        date_rows = fmt.get("date_rows", [])
        header_row = fmt.get("header_row", 1)

        # Get provider names from header row
        providers = []
        for col_idx in range(2, ws.max_column + 1):
            cell = ws.cell(row=header_row, column=col_idx)
            if cell.value:
                provider_name = str(cell.value).strip()
                providers.append((col_idx, provider_name))
                result.providers[provider_name] = ProviderSchedule(name=provider_name)

        min_date = None
        max_date = None

        # Process each date row
        for row_idx, slot_date in date_rows:
            for col_idx, provider_name in providers:
                cell = ws.cell(row=row_idx, column=col_idx)
                raw_value = str(cell.value).strip() if cell.value else ""

                slot_type = self.classify_slot(raw_value)

                slot = ScheduleSlot(
                    provider_name=provider_name,
                    date=slot_date,
                    time_of_day="AM",  # Assume full day
                    slot_type=slot_type,
                    raw_value=raw_value,
                )

                result.providers[provider_name].add_slot(slot)
                result.total_slots += 1

                if slot_type == SlotType.CLINIC:
                    result.clinic_slots += 1
                elif slot_type == SlotType.FMIT:
                    result.fmit_slots += 1

                if min_date is None or slot_date < min_date:
                    min_date = slot_date
                if max_date is None or slot_date > max_date:
                    max_date = slot_date

        result.date_range = (min_date, max_date)
        return result

    def import_file(
        self,
        file_path: str | None = None,
        file_bytes: bytes | None = None,
        sheet_name: str | None = None,
    ) -> ImportResult:
        """
        Import a schedule from an Excel file.

        Args:
            file_path: Path to the Excel file
            file_bytes: Raw bytes of Excel file (for API uploads)
            sheet_name: Specific sheet to import (None = first sheet)

        Returns:
            ImportResult with parsed data and any conflicts
        """
        try:
            if file_bytes:
                wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
            elif file_path:
                wb = load_workbook(file_path, data_only=True)
            else:
                return ImportResult(
                    success=False, errors=["No file path or bytes provided"]
                )

            # Get worksheet
            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    return ImportResult(
                        success=False,
                        errors=[
                            f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
                        ],
                    )
                ws = wb[sheet_name]
            else:
                ws = wb.active

            return self.import_worksheet(ws)

        except Exception as e:
            return ImportResult(
                success=False, errors=[f"Failed to parse Excel file: {str(e)}"]
            )


@dataclass
class FacultyTarget:
    """Target FMIT week allocation for a faculty member."""

    name: str
    target_weeks: int = 6
    role: str = "faculty"  # "chief", "pd", "adjunct", "faculty"
    current_weeks: int = 0

    @property
    def flexibility(self) -> str:
        """How flexible this faculty is for taking additional weeks."""
        if self.role == "chief":
            return "low"  # Already at minimum
        elif self.role == "pd":
            return "emergency_only"  # Secret weapon
        elif self.role == "adjunct":
            return "unreliable"
        elif self.current_weeks < self.target_weeks:
            return "high"  # Under target, should take more
        else:
            return "medium"  # At target, can swap 1:1


@dataclass
class SwapCandidate:
    """A potential swap partner for a target week."""

    faculty: str
    can_take_week: date
    gives_week: date | None = None  # None if absorbing, date if 1:1 swap
    back_to_back_ok: bool = True
    external_conflict: str | None = None
    flexibility: str = "medium"
    reason: str = ""


@dataclass
class ExternalConflict:
    """External constraint preventing assignment."""

    faculty: str
    start_date: date
    end_date: date
    conflict_type: str  # "leave", "conference", "tdy", "training"
    description: str = ""


def has_back_to_back_conflict(faculty_weeks: list[date]) -> bool:
    """
    Check if any two FMIT weeks are consecutive.

    FMIT weeks include mandatory Fri/Sat call, creating continuous duty.
    Back-to-back weeks = 14+ consecutive days with embedded 24hr shifts.
    """
    if len(faculty_weeks) < 2:
        return False

    sorted_weeks = sorted(faculty_weeks)
    for i in range(len(sorted_weeks) - 1):
        gap_days = (sorted_weeks[i + 1] - sorted_weeks[i]).days
        if gap_days <= 7:
            return True
    return False


def count_alternating_cycles(weeks: list[date]) -> int:
    """
    Count consecutive week-on/week-off patterns.

    Returns the number of times FMIT weeks alternate with 1-2 week gaps.
    3+ alternating cycles in a row = family hardship pattern.
    """
    if len(weeks) < 2:
        return 0

    sorted_weeks = sorted(weeks)
    cycles = 0

    for i in range(len(sorted_weeks) - 1):
        gap_days = (sorted_weeks[i + 1] - sorted_weeks[i]).days
        # 7-21 days = 1-2 week gap = alternating pattern
        if 7 < gap_days <= 21:
            cycles += 1

    return cycles


def get_schedule_flexibility(target_date: date, release_horizon_days: int = 90) -> str:
    """
    Determine how easy it is to change a scheduled date.

    Args:
        target_date: The date being evaluated
        release_horizon_days: Days ahead clinic schedules are released

    Returns:
        "impossible", "very_hard", "hard", or "easy"
    """
    days_until = (target_date - date.today()).days

    if days_until < 0:
        return "impossible"  # Past
    elif days_until < 14:
        return "very_hard"  # Imminent
    elif days_until < release_horizon_days:
        return "hard"  # Already released to patients
    else:
        return "easy"  # Not yet released


class ConflictDetector:
    """
    Detects scheduling conflicts between FMIT and clinic schedules.
    """

    def __init__(
        self,
        fmit_schedule: ImportResult,
        clinic_schedule: ImportResult,
        specialty_providers: dict[str, list[str]] | None = None,
    ):
        """
        Initialize conflict detector.

        Args:
            fmit_schedule: Parsed FMIT rotation schedule
            clinic_schedule: Parsed clinic schedule
            specialty_providers: Dict of specialty -> provider names (e.g., {"Sports Medicine": ["FAC-SPORTS"]})
        """
        self.fmit = fmit_schedule
        self.clinic = clinic_schedule
        self.specialty_providers = specialty_providers or {}
        self.conflicts: list[ScheduleConflict] = []

    def detect_all_conflicts(self) -> list[ScheduleConflict]:
        """Run all conflict detection checks."""
        self.conflicts = []

        self._detect_double_bookings()
        self._detect_specialty_unavailability()
        self._detect_alternating_patterns()

        return self.conflicts

    def _detect_double_bookings(self) -> None:
        """Detect when provider is scheduled for both FMIT and clinic."""
        # Get all providers in both schedules
        common_providers = set(self.fmit.providers.keys()) & set(
            self.clinic.providers.keys()
        )

        for provider_name in common_providers:
            fmit_schedule = self.fmit.providers[provider_name]
            clinic_schedule = self.clinic.providers[provider_name]

            for _key, fmit_slot in fmit_schedule.slots.items():
                if fmit_slot.slot_type != SlotType.FMIT:
                    continue

                clinic_slot = clinic_schedule.get_slot(
                    fmit_slot.date, fmit_slot.time_of_day
                )

                if clinic_slot and clinic_slot.slot_type == SlotType.CLINIC:
                    self.conflicts.append(
                        ScheduleConflict(
                            provider_name=provider_name,
                            date=fmit_slot.date,
                            time_of_day=fmit_slot.time_of_day,
                            conflict_type="double_book",
                            fmit_assignment=fmit_slot.raw_value,
                            clinic_assignment=clinic_slot.raw_value,
                            severity="error",
                        )
                    )

    def _detect_specialty_unavailability(self) -> None:
        """Detect when specialty provider is unavailable for their specialty clinic."""
        for specialty, providers in self.specialty_providers.items():
            for provider_name in providers:
                if provider_name not in self.fmit.providers:
                    continue

                fmit_schedule = self.fmit.providers[provider_name]

                for _key, slot in fmit_schedule.slots.items():
                    if slot.slot_type == SlotType.FMIT:
                        self.conflicts.append(
                            ScheduleConflict(
                                provider_name=provider_name,
                                date=slot.date,
                                time_of_day=slot.time_of_day,
                                conflict_type="specialty_unavailable",
                                fmit_assignment=slot.raw_value,
                                severity="warning",
                                message=f"{provider_name} ({specialty} specialist) on FMIT on {slot.date.strftime('%b %d')} - no specialty coverage",
                            )
                        )

    def _detect_alternating_patterns(self) -> None:
        """Detect week-on/week-off patterns that are hard on families."""
        for provider_name, schedule in self.fmit.providers.items():
            if schedule.has_alternating_pattern():
                weeks = schedule.get_fmit_weeks()
                week_strs = [
                    f"{w[0].strftime('%b %d')}-{w[1].strftime('%b %d')}" for w in weeks
                ]

                self.conflicts.append(
                    ScheduleConflict(
                        provider_name=provider_name,
                        date=weeks[0][0] if weeks else date.today(),
                        time_of_day="AM",
                        conflict_type="consecutive_weeks",
                        severity="warning",
                        message=f"{provider_name} has alternating week FMIT pattern: {', '.join(week_strs)}. Consider consolidating.",
                    )
                )


class SwapFinder:
    """
    Finds valid swap candidates for FMIT weeks.

    Given a target faculty member and week to offload, identifies which
    other faculty could take that week, considering:
    - Back-to-back constraints
    - External conflicts (leave, conferences)
    - Target week distribution
    - Schedule flexibility (clinic release dates)
    """

    def __init__(
        self,
        fmit_schedule: ImportResult,
        faculty_targets: dict[str, FacultyTarget] | None = None,
        external_conflicts: list[ExternalConflict] | None = None,
        schedule_release_date: date | None = None,
    ):
        self.fmit = fmit_schedule
        self.faculty_targets = faculty_targets or {}
        self.external_conflicts = external_conflicts or []
        self.schedule_release_date = schedule_release_date or (
            date.today() + timedelta(days=90)
        )

        # Build faculty week mapping
        self.faculty_weeks: dict[str, list[date]] = {}
        for name, schedule in self.fmit.providers.items():
            weeks = schedule.get_fmit_weeks()
            # Use Monday of each week as the canonical date
            self.faculty_weeks[name] = [w[0] for w in weeks]

    def _check_external_conflict(
        self, faculty: str, week_start: date
    ) -> ExternalConflict | None:
        """Check if faculty has external conflict during week."""
        week_end = week_start + timedelta(days=6)

        for conflict in self.external_conflicts:
            if conflict.faculty == faculty:
                if conflict.start_date <= week_end and conflict.end_date >= week_start:
                    return conflict
        return None

    def _can_take_week(
        self, faculty: str, week_to_take: date, excluding_week: date | None = None
    ) -> tuple[bool, str]:
        """
        Check if faculty can take a specific week.

        Args:
            faculty: Faculty name
            week_to_take: Monday of the week to potentially take
            excluding_week: If doing 1:1 swap, the week being given up

        Returns:
            (can_take, reason) tuple
        """
        current_weeks = self.faculty_weeks.get(faculty, [])

        # Build test schedule
        test_weeks = [w for w in current_weeks if w != excluding_week]
        test_weeks.append(week_to_take)

        # Check back-to-back
        if has_back_to_back_conflict(test_weeks):
            return False, "back_to_back_conflict"

        # Check external conflicts
        if conflict := self._check_external_conflict(faculty, week_to_take):
            return False, f"{conflict.conflict_type}: {conflict.description}"

        return True, "ok"

    def find_swap_candidates(
        self,
        target_faculty: str,
        target_week: date,
    ) -> list[SwapCandidate]:
        """
        Find all faculty who could take a given week.

        Args:
            target_faculty: Faculty wanting to offload the week
            target_week: Monday of the week to offload

        Returns:
            List of SwapCandidate objects, sorted by viability
        """
        candidates = []

        for faculty, weeks in self.faculty_weeks.items():
            if faculty == target_faculty:
                continue

            # Get faculty metadata
            meta = self.faculty_targets.get(
                faculty, FacultyTarget(name=faculty, current_weeks=len(weeks))
            )

            # Skip unreliable faculty
            if meta.flexibility == "unreliable":
                continue

            # Check if can take the target week (absorb scenario)
            can_take, reason = self._can_take_week(faculty, target_week)

            # Find potential give-back weeks for 1:1 swap
            give_weeks = []
            if can_take:
                for w in weeks:
                    # Only consider weeks after release date (flexible)
                    if w <= self.schedule_release_date:
                        continue

                    # Check if target can take this week
                    target_current = self.faculty_weeks.get(target_faculty, [])
                    target_test = [tw for tw in target_current if tw != target_week]
                    target_test.append(w)

                    if not has_back_to_back_conflict(target_test):
                        # Check target doesn't have external conflict
                        if not self._check_external_conflict(target_faculty, w):
                            give_weeks.append(w)

            # Check external conflicts for target week
            ext_conflict = self._check_external_conflict(faculty, target_week)

            candidates.append(
                SwapCandidate(
                    faculty=faculty,
                    can_take_week=target_week,
                    gives_week=give_weeks[0] if give_weeks else None,
                    back_to_back_ok=can_take,
                    external_conflict=(
                        ext_conflict.conflict_type if ext_conflict else None
                    ),
                    flexibility=get_schedule_flexibility(target_week),
                    reason=reason if not can_take else "",
                )
            )

        # Sort by viability (best candidates first)
        return sorted(
            candidates,
            key=lambda c: (
                not c.back_to_back_ok,  # Back-to-back issues (worst)
                c.external_conflict is not None,  # External conflicts
                c.gives_week is None,  # Absorb-only (prefer 1:1)
                c.flexibility == "low",  # Low flexibility
            ),
        )

    def find_excessive_alternating(self, threshold: int = 3) -> list[tuple[str, int]]:
        """
        Find faculty with excessive alternating patterns.

        Args:
            threshold: Number of alternating cycles to flag

        Returns:
            List of (faculty_name, cycle_count) tuples
        """
        results = []

        for faculty, weeks in self.faculty_weeks.items():
            cycles = count_alternating_cycles(weeks)
            if cycles >= threshold:
                results.append((faculty, cycles))

        return sorted(results, key=lambda x: -x[1])  # Most cycles first

    @classmethod
    def from_fmit_file(
        cls,
        file_path: str | None = None,
        file_bytes: bytes | None = None,
        faculty_targets: dict[str, "FacultyTarget"] | None = None,
        external_conflicts: list["ExternalConflict"] | None = None,
        db: Session | None = None,
        include_absence_conflicts: bool = True,
        schedule_release_days: int = 90,
    ) -> "SwapFinder":
        """
        Factory method to create SwapFinder from an FMIT schedule file.

        Args:
            file_path: Path to Excel file
            file_bytes: Raw bytes of Excel file
            faculty_targets: Optional faculty target configurations
            external_conflicts: Optional external conflicts list
            db: Database session for loading absence conflicts
            include_absence_conflicts: Whether to include absence records
            schedule_release_days: Days ahead for schedule flexibility

        Returns:
            Configured SwapFinder instance
        """
        importer = ClinicScheduleImporter(db)
        result = importer.import_file(file_path=file_path, file_bytes=file_bytes)

        if not result.success:
            raise ValueError(f"Failed to import schedule: {result.errors}")

        # Build external conflicts list
        all_conflicts = list(external_conflicts or [])

        # Add absence-based conflicts if requested
        if include_absence_conflicts and db:
            absence_conflicts = load_external_conflicts_from_absences(db)
            all_conflicts.extend(absence_conflicts)

        return cls(
            fmit_schedule=result,
            faculty_targets=faculty_targets,
            external_conflicts=all_conflicts,
            schedule_release_date=date.today() + timedelta(days=schedule_release_days),
        )

    def suggest_swaps_for_alternating(
        self, faculty: str
    ) -> list[tuple[date, list[SwapCandidate]]]:
        """
        Suggest swaps to reduce alternating pattern for a faculty member.

        Returns list of (week_to_offload, candidates) tuples.
        """
        weeks = self.faculty_weeks.get(faculty, [])
        if not weeks:
            return []

        sorted_weeks = sorted(weeks)
        suggestions = []

        # Find weeks in the middle of alternating stretches
        for i in range(1, len(sorted_weeks) - 1):
            prev_gap = (sorted_weeks[i] - sorted_weeks[i - 1]).days
            next_gap = (sorted_weeks[i + 1] - sorted_weeks[i]).days

            # This week is in an alternating stretch
            if 7 < prev_gap <= 21 and 7 < next_gap <= 21:
                candidates = self.find_swap_candidates(faculty, sorted_weeks[i])
                valid_candidates = [c for c in candidates if c.back_to_back_ok]
                if valid_candidates:
                    suggestions.append((sorted_weeks[i], valid_candidates))

        return suggestions


def analyze_schedule_conflicts(
    fmit_file: str | None = None,
    fmit_bytes: bytes | None = None,
    clinic_file: str | None = None,
    clinic_bytes: bytes | None = None,
    specialty_providers: dict[str, list[str]] | None = None,
) -> dict:
    """
    Analyze conflicts between FMIT rotation and clinic schedules.

    Args:
        fmit_file: Path to FMIT rotation Excel file
        fmit_bytes: Bytes of FMIT rotation Excel file
        clinic_file: Path to clinic schedule Excel file
        clinic_bytes: Bytes of clinic schedule Excel file
        specialty_providers: Dict mapping specialty -> provider names

    Returns:
        Analysis results including conflicts, recommendations
    """
    importer = ClinicScheduleImporter()

    # Import FMIT schedule
    fmit_result = importer.import_file(file_path=fmit_file, file_bytes=fmit_bytes)

    if not fmit_result.success:
        return {
            "success": False,
            "error": f"Failed to import FMIT schedule: {fmit_result.errors}",
        }

    # Import clinic schedule if provided
    clinic_result = None
    if clinic_file or clinic_bytes:
        clinic_result = importer.import_file(
            file_path=clinic_file, file_bytes=clinic_bytes
        )

        if not clinic_result.success:
            return {
                "success": False,
                "error": f"Failed to import clinic schedule: {clinic_result.errors}",
            }

    # Detect conflicts
    conflicts = []
    if clinic_result:
        detector = ConflictDetector(
            fmit_schedule=fmit_result,
            clinic_schedule=clinic_result,
            specialty_providers=specialty_providers,
        )
        conflicts = detector.detect_all_conflicts()
    else:
        # Just check FMIT schedule for alternating patterns
        for provider_name, schedule in fmit_result.providers.items():
            if schedule.has_alternating_pattern():
                weeks = schedule.get_fmit_weeks()
                conflicts.append(
                    ScheduleConflict(
                        provider_name=provider_name,
                        date=weeks[0][0] if weeks else date.today(),
                        time_of_day="AM",
                        conflict_type="consecutive_weeks",
                        severity="warning",
                        message=f"{provider_name} has alternating FMIT pattern - hard on families",
                    )
                )

    # Build recommendations
    recommendations = []

    alternating = [c for c in conflicts if c.conflict_type == "consecutive_weeks"]
    if alternating:
        recommendations.append(
            {
                "type": "consolidate_fmit",
                "providers": [c.provider_name for c in alternating],
                "message": "Consider consolidating FMIT weeks to reduce alternating patterns",
            }
        )

    specialty_conflicts = [
        c for c in conflicts if c.conflict_type == "specialty_unavailable"
    ]
    if specialty_conflicts:
        recommendations.append(
            {
                "type": "specialty_coverage",
                "providers": list({c.provider_name for c in specialty_conflicts}),
                "message": "Specialty providers on FMIT create clinic coverage gaps",
            }
        )

    double_books = [c for c in conflicts if c.conflict_type == "double_book"]
    if double_books:
        recommendations.append(
            {
                "type": "resolve_double_booking",
                "count": len(double_books),
                "message": f"{len(double_books)} double-booking conflicts must be resolved",
            }
        )

    return {
        "success": True,
        "fmit_schedule": {
            "providers": list(fmit_result.providers.keys()),
            "date_range": [
                (
                    fmit_result.date_range[0].isoformat()
                    if fmit_result.date_range[0]
                    else None
                ),
                (
                    fmit_result.date_range[1].isoformat()
                    if fmit_result.date_range[1]
                    else None
                ),
            ],
            "total_slots": fmit_result.total_slots,
            "fmit_slots": fmit_result.fmit_slots,
        },
        "clinic_schedule": (
            {
                "providers": (
                    list(clinic_result.providers.keys()) if clinic_result else []
                ),
                "date_range": [
                    (
                        clinic_result.date_range[0].isoformat()
                        if clinic_result and clinic_result.date_range[0]
                        else None
                    ),
                    (
                        clinic_result.date_range[1].isoformat()
                        if clinic_result and clinic_result.date_range[1]
                        else None
                    ),
                ],
                "total_slots": clinic_result.total_slots if clinic_result else 0,
                "clinic_slots": clinic_result.clinic_slots if clinic_result else 0,
            }
            if clinic_result
            else None
        ),
        "conflicts": [
            {
                "provider": c.provider_name,
                "date": c.date.isoformat(),
                "time": c.time_of_day,
                "type": c.conflict_type,
                "severity": c.severity,
                "message": c.message,
            }
            for c in conflicts
        ],
        "recommendations": recommendations,
        "summary": {
            "total_conflicts": len(conflicts),
            "errors": len([c for c in conflicts if c.severity == "error"]),
            "warnings": len([c for c in conflicts if c.severity == "warning"]),
        },
    }


def load_external_conflicts_from_absences(
    db: Session,
    start_date: date | None = None,
    end_date: date | None = None,
) -> list[ExternalConflict]:
    """
    Load external conflicts from the absence model in the database.

    This integrates the Absence model with the ExternalConflict dataclass,
    allowing SwapFinder to automatically consider approved leave, TDY,
    deployments, and conferences when finding swap candidates.

    Args:
        db: Database session
        start_date: Optional start of date range to load (defaults to today)
        end_date: Optional end of date range (defaults to 1 year from now)

    Returns:
        List of ExternalConflict objects built from absence records
    """
    if start_date is None:
        start_date = date.today()
    if end_date is None:
        end_date = date.today() + timedelta(days=365)

    # Query blocking absences within the date range
    absences = (
        db.query(AbsenceModel)
        .join(Person)
        .filter(
            AbsenceModel.end_date >= start_date,
            AbsenceModel.start_date <= end_date,
        )
        .all()
    )

    conflicts = []
    for absence in absences:
        # Only include blocking absences
        if not absence.should_block_assignment:
            continue

        # Map absence_type to conflict_type
        conflict_type_map = {
            "vacation": "leave",
            "deployment": "deployment",
            "tdy": "tdy",
            "medical": "medical",
            "family_emergency": "leave",
            "conference": "conference",
        }
        conflict_type = conflict_type_map.get(absence.absence_type, "leave")

        # Build description
        description = absence.notes or ""
        if absence.tdy_location:
            description = f"TDY - {absence.tdy_location}"
        if absence.replacement_activity:
            description = absence.replacement_activity

        conflicts.append(
            ExternalConflict(
                faculty=absence.person.name,
                start_date=absence.start_date,
                end_date=absence.end_date,
                conflict_type=conflict_type,
                description=description,
            )
        )

    return conflicts


def absence_to_external_conflict(absence: AbsenceModel) -> ExternalConflict | None:
    """
    Convert a single Absence model to an ExternalConflict.

    Returns None if the absence is non-blocking.
    """
    if not absence.should_block_assignment:
        return None

    conflict_type_map = {
        "vacation": "leave",
        "deployment": "deployment",
        "tdy": "tdy",
        "medical": "medical",
        "family_emergency": "leave",
        "conference": "conference",
    }
    conflict_type = conflict_type_map.get(absence.absence_type, "leave")

    description = absence.notes or ""
    if absence.tdy_location:
        description = f"TDY - {absence.tdy_location}"
    if absence.replacement_activity:
        description = absence.replacement_activity

    return ExternalConflict(
        faculty=absence.person.name,
        start_date=absence.start_date,
        end_date=absence.end_date,
        conflict_type=conflict_type,
        description=description,
    )


# =============================================================================
# BLOCK SCHEDULE PARSER - Fuzzy-tolerant parsing for rotation spreadsheets
# =============================================================================


@dataclass
class ParsedBlockAssignment:
    """A single assignment extracted from block schedule spreadsheet."""

    person_name: str
    date: date
    template: str  # R1, R2, R3, C19, etc.
    role: str  # PGY 1, PGY 2, PGY 3, FAC
    slot_am: str | None = None  # Value in AM column
    slot_pm: str | None = None  # Value in PM column
    row_idx: int = 0
    confidence: float = 1.0  # Lower if fuzzy matched


@dataclass
class ParsedFMITWeek:
    """FMIT attending assignment for a week."""

    block_number: int
    week_number: int
    start_date: date | None = None
    end_date: date | None = None
    faculty_name: str = ""
    is_holiday_call: bool = False


@dataclass
class BlockParseResult:
    """Result of parsing a block schedule sheet."""

    block_number: int
    start_date: date | None = None
    end_date: date | None = None
    assignments: list[ParsedBlockAssignment] = field(default_factory=list)
    residents: list[dict] = field(default_factory=list)  # {name, template, role}
    fmit_schedule: list[ParsedFMITWeek] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)

    def get_residents_by_template(self) -> dict[str, list[dict]]:
        """Group residents by rotation template."""
        by_template: dict[str, list[dict]] = {}
        for r in self.residents:
            t = r.get("template", "")
            if t not in by_template:
                by_template[t] = []
            by_template[t].append(r)
        return by_template


class BlockScheduleParser:
    """
    Fuzzy-tolerant parser for block schedule Excel files.

    Handles the chaos of human-edited spreadsheets:
    - Finds anchors by content (TEMPLATE, PROVIDER), not position
    - Discovers date columns by finding datetime values
    - Extracts person rows by "Last, First" pattern
    - Fuzzy matches names against known people list
    - Handles merged cells gracefully

    Example usage:
        parser = BlockScheduleParser(known_people=["Doria, Russell", "Evans, Amber"])
        result = parser.parse_block_sheet("schedule.xlsx", "Block 10")
        for r in result.residents:
            print(f"{r['template']}: {r['name']} ({r['role']})")
    """

    # Patterns for finding column headers
    HEADER_PATTERNS = {
        "template": re.compile(r"^template$", re.I),
        "role": re.compile(r"^role$", re.I),
        "provider": re.compile(r"^provider$", re.I),
    }

    # Name pattern: "Last, First" or "Last, First Middle"
    NAME_PATTERN = re.compile(r"^([A-Z][a-z-]+),\s+([A-Z][a-z-]+)", re.I)

    # Known rotation templates
    VALID_TEMPLATES = {
        "R1",
        "R2",
        "R3",  # Resident rotations
        "C19",
        "Ca9",
        "CP",  # Clinic/faculty
        "TY",
        "PSYCH",
        "IPAP",  # Transitional/specialty
        "NF",  # Night float
    }

    # Known roles
    VALID_ROLES = {
        "PGY 1",
        "PGY 2",
        "PGY 3",
        "FAC",
        "FLT 1",
        "FLT 2",
        "FLT 3",
        "CP",
        "IPAP",
    }

    def __init__(self, known_people: list[str] | None = None) -> None:
        """
        Initialize parser.

        Args:
            known_people: List of known person names for fuzzy matching.
                          Format: "Last, First" or full name variants.
        """
        self.known_people = known_people or []
        self._name_cache: dict[
            str, tuple[str, float]
        ] = {}  # raw -> (matched, confidence)

    def parse_block_sheet(
        self,
        filepath: str | Path,
        sheet_name: str,
        expected_block: int | None = None,
    ) -> BlockParseResult:
        """
        Parse a single block schedule sheet.

        Args:
            filepath: Path to Excel file
            sheet_name: Name of sheet to parse (e.g., "Block 10")
            expected_block: If provided, validate block number matches

        Returns:
            BlockParseResult with assignments and any warnings/errors
        """
        wb = load_workbook(filepath, data_only=True)

        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
            )

        ws = wb[sheet_name]

        # Extract block number from sheet name or content
        block_num = self._extract_block_number(sheet_name, ws)
        if expected_block and block_num != expected_block:
            logger.warning(
                f"Block mismatch: expected {expected_block}, found {block_num}"
            )

        result = BlockParseResult(block_number=block_num)

        # Step 1: Find anchor positions (TEMPLATE, ROLE, PROVIDER columns)
        anchors = self._find_anchors(ws)
        if not anchors.get("provider_col"):
            result.errors.append("Could not find PROVIDER column header")
            return result

        # Step 2: Find date columns by scanning for datetime values
        date_cols = self._find_date_columns(ws)
        if not date_cols:
            result.warnings.append("Could not find date columns - roster-only mode")

        # Update date range
        if date_cols:
            all_dates = sorted(date_cols.values())
            result.start_date = all_dates[0]
            result.end_date = all_dates[-1]

        # Step 3: Find person rows and extract roster
        person_rows = self._find_person_rows(ws, anchors)

        for row_info in person_rows:
            raw_name = row_info["name"]

            # Fuzzy match name
            matched_name, confidence = self._match_name(raw_name)
            if confidence < 0.8:
                result.warnings.append(
                    f"Row {row_info['row']}: Low confidence name match "
                    f"'{raw_name}' -> '{matched_name}' ({confidence:.0%})"
                )

            # Add to roster
            result.residents.append(
                {
                    "name": matched_name or raw_name,
                    "template": row_info["template"],
                    "role": row_info["role"],
                    "row": row_info["row"],
                    "confidence": confidence,
                }
            )

            # Extract daily assignments if we have date columns
            if date_cols:
                for col_idx, assign_date in date_cols.items():
                    cell_val = self._get_cell_value(ws, row_info["row"], col_idx)
                    # Check adjacent column for PM value
                    pm_val = self._get_cell_value(ws, row_info["row"], col_idx + 1)

                    # Skip if both empty
                    if not cell_val and not pm_val:
                        continue

                    result.assignments.append(
                        ParsedBlockAssignment(
                            person_name=matched_name or raw_name,
                            date=assign_date,
                            template=row_info["template"],
                            role=row_info["role"],
                            slot_am=cell_val,
                            slot_pm=pm_val if pm_val != cell_val else None,
                            row_idx=row_info["row"],
                            confidence=confidence,
                        )
                    )

        return result

    def _find_anchors(self, ws: Worksheet) -> dict:
        """Find column positions of key headers by content matching."""
        anchors: dict = {}

        # Scan first 10 rows for headers
        for row_idx in range(1, 11):
            for col_idx in range(1, 20):
                cell_val = self._get_cell_value(ws, row_idx, col_idx)
                if not cell_val:
                    continue

                cell_str = str(cell_val).strip()

                for key, pattern in self.HEADER_PATTERNS.items():
                    if pattern.match(cell_str):
                        anchors[f"{key}_col"] = col_idx
                        anchors[f"{key}_row"] = row_idx
                        logger.debug(
                            f"Found {key} anchor at row={row_idx}, col={col_idx}"
                        )

        return anchors

    def _find_date_columns(self, ws: Worksheet) -> dict[int, date]:
        """
        Find columns containing dates.

        Scans first few rows for datetime values, returns {col_idx: date}.
        """
        date_cols: dict[int, date] = {}

        # Scan first 5 rows, columns 1-100
        for row_idx in range(1, 6):
            for col_idx in range(1, 100):
                try:
                    cell = ws.cell(row=row_idx, column=col_idx)

                    # Skip merged cells
                    if isinstance(cell, MergedCell):
                        continue

                    val = cell.value
                    if isinstance(val, datetime):
                        date_cols[col_idx] = val.date()
                    elif isinstance(val, date) and not isinstance(val, datetime):
                        date_cols[col_idx] = val
                except Exception:
                    continue

        return date_cols

    def _find_person_rows(self, ws: Worksheet, anchors: dict) -> list[dict]:
        """Find rows containing person data based on name pattern."""
        persons = []

        # Use discovered anchors or sensible defaults
        provider_col = anchors.get("provider_col", 4)
        template_col = anchors.get("template_col", provider_col - 2)
        role_col = anchors.get("role_col", provider_col - 1)
        header_row = anchors.get("provider_row", 6)

        # Scan rows after header
        for row_idx in range(header_row + 1, min(ws.max_row + 1, 200)):
            name_val = self._get_cell_value(ws, row_idx, provider_col)

            if not name_val:
                continue

            # Check if it looks like a name (Last, First pattern)
            name_str = str(name_val).strip().replace("*", "")
            if not self.NAME_PATTERN.match(name_str):
                continue

            template = str(
                self._get_cell_value(ws, row_idx, template_col) or ""
            ).strip()
            role = str(self._get_cell_value(ws, row_idx, role_col) or "").strip()

            # Log unknown templates for debugging
            if template and template not in self.VALID_TEMPLATES:
                logger.debug(f"Row {row_idx}: Unknown template '{template}'")

            persons.append(
                {
                    "row": row_idx,
                    "name": name_str,
                    "template": template,
                    "role": role,
                }
            )

        return persons

    def _match_name(self, raw_name: str) -> tuple[str, float]:
        """
        Fuzzy match a name against known people.

        Returns:
            (matched_name, confidence) where confidence is 0.0-1.0
        """
        if not self.known_people:
            return raw_name, 1.0

        # Check cache
        if raw_name in self._name_cache:
            return self._name_cache[raw_name]

        # Normalize for comparison
        normalized = raw_name.lower().replace(",", "").replace("*", "").strip()

        best_match = None
        best_score = 0.0

        for known in self.known_people:
            known_norm = known.lower().replace(",", "").strip()

            # Exact match
            if normalized == known_norm:
                self._name_cache[raw_name] = (known, 1.0)
                return known, 1.0

            # Fuzzy match using SequenceMatcher
            score = SequenceMatcher(None, normalized, known_norm).ratio()
            if score > best_score:
                best_score = score
                best_match = known

        if best_match and best_score >= 0.7:
            self._name_cache[raw_name] = (best_match, best_score)
            return best_match, best_score

        # No good match - return original
        self._name_cache[raw_name] = (raw_name, 1.0)
        return raw_name, 1.0

    def _get_cell_value(self, ws: Worksheet, row: int, col: int) -> str | None:
        """Get cell value, handling merged cells gracefully."""
        try:
            cell = ws.cell(row=row, column=col)

            if isinstance(cell, MergedCell):
                # Find the parent cell of the merge range
                for merge_range in ws.merged_cells.ranges:
                    if cell.coordinate in merge_range:
                        parent_cell = ws.cell(
                            row=merge_range.min_row,
                            column=merge_range.min_col,
                        )
                        val = parent_cell.value
                        if val is None:
                            return None
                        if isinstance(val, (datetime, date)):
                            return val.isoformat()
                        return str(val).strip() if str(val).strip() else None
                return None

            val = cell.value
            if val is None:
                return None

            if isinstance(val, (datetime, date)):
                return val.isoformat()

            return str(val).strip() if str(val).strip() else None

        except Exception as e:
            logger.warning(f"Error reading cell ({row}, {col}): {e}")
            return None

    def _extract_block_number(self, sheet_name: str, ws: Worksheet) -> int:
        """Extract block number from sheet name or cell content."""
        # Try sheet name first: "Block 10", "Block10", "10"
        match = re.search(r"block\s*(\d+)", sheet_name, re.I)
        if match:
            return int(match.group(1))

        # Just a number
        match = re.search(r"^(\d+)", sheet_name.strip())
        if match:
            return int(match.group(1))

        # Try cells A1, B1, B2 for block number
        for row, col in [(1, 1), (1, 2), (2, 2)]:
            val = self._get_cell_value(ws, row, col)
            if val:
                match = re.search(r"^(\d+)$", str(val).strip())
                if match:
                    return int(match.group(1))

        return 0  # Unknown


def parse_block_schedule(
    filepath: str | Path,
    block_number: int,
    known_people: list[str] | None = None,
) -> BlockParseResult:
    """
    Convenience function to parse a specific block.

    Args:
        filepath: Path to Excel file
        block_number: Block to parse (1-13)
        known_people: Optional list of known names for fuzzy matching

    Returns:
        BlockParseResult with residents and assignments
    """
    parser = BlockScheduleParser(known_people=known_people)

    # Try common sheet name patterns
    sheet_names = [
        f"Block {block_number}",
        f"Block{block_number}",
        f"{block_number}",
        f"{block_number} with formatting",
        f"{block_number} with formatting (2)",
    ]

    wb = load_workbook(filepath, data_only=True)

    for name in sheet_names:
        if name in wb.sheetnames:
            return parser.parse_block_sheet(filepath, name, expected_block=block_number)

    # Fuzzy match sheet name
    for sheet in wb.sheetnames:
        if str(block_number) in sheet:
            return parser.parse_block_sheet(
                filepath, sheet, expected_block=block_number
            )

    raise ValueError(
        f"Could not find sheet for Block {block_number}. Available: {wb.sheetnames}"
    )


def parse_fmit_attending(
    filepath: str | Path,
    sheet_name: str = "FMIT Attending (2025-2026)",
) -> list[ParsedFMITWeek]:
    """
    Parse FMIT attending schedule sheet.

    The FMIT sheet typically has structure:
    - Column 0: Block number
    - Column 1: Date range
    - Columns 2-5: Week 1-4 faculty names
    - Faculty names in row AFTER the dates row

    Returns:
        List of ParsedFMITWeek assignments
    """
    wb = load_workbook(filepath, data_only=True)

    if sheet_name not in wb.sheetnames:
        # Try to find it by partial match
        for name in wb.sheetnames:
            if "fmit" in name.lower():
                sheet_name = name
                break
        else:
            raise ValueError(f"FMIT sheet not found. Available: {wb.sheetnames}")

    ws = wb[sheet_name]
    results = []

    # Scan for block rows
    for row_idx in range(1, ws.max_row + 1):
        col0_val = ws.cell(row=row_idx, column=1).value

        # Check if this is a block header row (has block number)
        if col0_val and str(col0_val).strip().isdigit():
            block_num = int(str(col0_val).strip())

            # Faculty names are in the next row, columns 3-6
            faculty_row = row_idx + 1
            if faculty_row <= ws.max_row:
                for week_num, col in enumerate([3, 4, 5, 6], start=1):
                    faculty = ws.cell(row=faculty_row, column=col).value
                    if faculty and str(faculty).strip():
                        results.append(
                            ParsedFMITWeek(
                                block_number=block_num,
                                week_number=week_num,
                                faculty_name=str(faculty).strip(),
                            )
                        )

    return results
