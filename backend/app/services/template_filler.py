"""Template filler -- write schedule values into BlockTemplate2.

Values only. The template's conditional formatting rules handle all colors.
Display rules transform DB codes to match handjam conventions.

Pipeline:
    DB -> HalfDayJSONExporter -> JSON -> TemplateFiller (+ display rules) -> xlsx
"""

from __future__ import annotations

from datetime import date
from io import BytesIO
from pathlib import Path
from typing import Any

try:
    from defusedxml import ElementTree
except ImportError:
    from xml.etree import ElementTree  # type: ignore[no-redef]  # nosec B314

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

from app.core.logging import get_logger
from app.services.schedule_display_rules import transform_code

logger = get_logger(__name__)

COL_ROTATION1 = 1
COL_ROTATION2 = 2
COL_SCHEDULE_START = 6
COLS_PER_DAY = 2
MAX_DAYS = 28
ROW_STAFF_CALL = 4


class TemplateFiller:
    """Fill BlockTemplate2 with schedule data. Values only -- template owns colors."""

    def __init__(self, template_path: Path, structure_xml_path: Path) -> None:
        self.template_path = template_path
        self.structure_xml_path = structure_xml_path

    def fill(
        self,
        data: dict[str, Any],
        output_path: Path | str | None = None,
        absences: list[dict] | None = None,
    ) -> bytes:
        """Write schedule values into a copy of the template.

        Args:
            data: JSON dict from HalfDayJSONExporter.export()
            output_path: If provided, also writes to disk.
            absences: Optional list of absence dicts with keys:
                name, type, start (ISO), end (ISO).

        Returns:
            XLSX bytes.
        """
        wb = load_workbook(self.template_path)
        ws = wb.active

        name_map = self._load_name_map()
        block_start = date.fromisoformat(data["block_start"])
        absence_lookup = self._build_absence_lookup(absences or [])

        matched = 0
        skipped = 0
        transforms_applied = 0

        faculty_names = {p["name"] for p in data.get("faculty", [])}

        for person in data.get("residents", []) + data.get("faculty", []):
            row = self._resolve_row(person["name"], name_map)
            if row is None:
                logger.warning("No row for %s -- skipping", person["name"])
                skipped += 1
                continue

            matched += 1
            is_faculty = person["name"] in faculty_names
            rot1 = person.get("rotation1") or ""
            rot2 = person.get("rotation2") or ""

            # Rotation columns (A, B) -- some rows have A:B merged
            cell_a = ws.cell(row=row, column=COL_ROTATION1)
            if not isinstance(cell_a, MergedCell):
                cell_a.value = rot1 or None
            cell_b = ws.cell(row=row, column=COL_ROTATION2)
            if not isinstance(cell_b, MergedCell):
                cell_b.value = rot2 or None

            # Daily schedule codes with display rule transformation
            person_last = self._last_name(person["name"])
            for i, day in enumerate(person.get("days", [])):
                if i >= MAX_DAYS:
                    break

                day_date = date.fromisoformat(day["date"])
                is_weekend = day_date.weekday() in (5, 6)
                abs_type = absence_lookup.get((person_last, day_date))

                raw_am = day.get("am") or ("W" if is_weekend else "")
                raw_pm = day.get("pm") or ("W" if is_weekend else "")

                am = transform_code(
                    raw_am,
                    rotation=rot1,
                    rotation2=rot2,
                    is_weekend=is_weekend,
                    is_faculty=is_faculty,
                    absence_type=abs_type,
                    day_date=day_date,
                )
                pm = transform_code(
                    raw_pm,
                    rotation=rot1,
                    rotation2=rot2,
                    is_weekend=is_weekend,
                    is_faculty=is_faculty,
                    absence_type=abs_type,
                    day_date=day_date,
                )

                if am != raw_am or pm != raw_pm:
                    transforms_applied += 1

                am_col = COL_SCHEDULE_START + i * COLS_PER_DAY
                am_cell = ws.cell(row=row, column=am_col)
                if not isinstance(am_cell, MergedCell):
                    am_cell.value = am or None
                pm_cell = ws.cell(row=row, column=am_col + 1)
                if not isinstance(pm_cell, MergedCell):
                    pm_cell.value = pm or None

        # Call row (row 4): staff last names in AM columns
        call_written = 0
        for night in data.get("call", {}).get("nights", []):
            night_date = date.fromisoformat(night["date"])
            offset = (night_date - block_start).days
            if 0 <= offset < MAX_DAYS:
                staff = night.get("staff", "")
                if staff:
                    last = staff.split()[-1] if " " in staff else staff
                    col = COL_SCHEDULE_START + offset * COLS_PER_DAY
                    cell = ws.cell(row=ROW_STAFF_CALL, column=col)
                    if not isinstance(cell, MergedCell):
                        cell.value = last
                        call_written += 1

        logger.info(
            "Filled %d people (%d skipped), %d call nights, %d display transforms",
            matched,
            skipped,
            call_written,
            transforms_applied,
        )

        buf = BytesIO()
        wb.save(buf)
        xlsx_bytes = buf.getvalue()

        if output_path:
            out = Path(output_path)
            out.parent.mkdir(parents=True, exist_ok=True)
            out.write_bytes(xlsx_bytes)
            logger.info("Wrote %d bytes to %s", len(xlsx_bytes), out)

        wb.close()
        return xlsx_bytes

    @staticmethod
    def _build_absence_lookup(
        absences: list[dict],
    ) -> dict[tuple[str, date], str]:
        """Build (last_name, date) -> absence_type lookup."""
        from datetime import timedelta

        lookup: dict[tuple[str, date], str] = {}
        for a in absences:
            name = a.get("name", "")
            ln = name.split()[-1].lower() if " " in name else name.lower()
            try:
                start = date.fromisoformat(a["start"])
                end = date.fromisoformat(a["end"])
            except (KeyError, ValueError):
                continue
            d = start
            while d <= end:
                lookup[(ln, d)] = a.get("type", "")
                d += timedelta(days=1)
        return lookup

    @staticmethod
    def _last_name(name: str) -> str:
        """Extract lowercase last name from 'First Last' format."""
        return name.split()[-1].lower() if " " in name else name.lower()

    def _load_name_map(self) -> dict[str, int]:
        """Parse BlockTemplate2_Structure.xml -> {name: row}."""
        tree = ElementTree.parse(str(self.structure_xml_path))  # nosec B314 - local trusted XML
        root = tree.getroot()

        mapping: dict[str, int] = {}
        for tag in ("resident", "person"):
            for elem in root.iter(tag):
                name = elem.get("name", "")
                row = elem.get("row")
                if name and row and name not in ("Unassigned", "Unassigned-NF"):
                    mapping[name] = int(row)

        return mapping

    def _resolve_row(self, db_name: str, name_map: dict[str, int]) -> int | None:
        """Match DB name to row. Exact match first, then last-name fallback."""
        if db_name in name_map:
            return name_map[db_name]

        # Last-name fallback (handles nicknames: Cam/Cameron, Katie/Katherine)
        db_last = db_name.split()[-1].lower() if " " in db_name else db_name.lower()
        for xml_name, row in name_map.items():
            xml_last = (
                xml_name.split()[-1].lower() if " " in xml_name else xml_name.lower()
            )
            if db_last == xml_last:
                return row

        return None
