"""TAMCBlockExporter — Ground-up block schedule exporter.

Produces pixel-perfect Block Template2 workbooks from JSON schedule data,
matching the coordinator's hand-jammed reference format.

Pipeline:
    DB -> HalfDayJSONExporter -> JSON dict
                                   |
                         TAMCBlockExporter.export(data)
                           Phase 1: Structure (workbook, dims, freeze)
                           Phase 2: Formatting (headers, black bars, rotation colors)
                           Phase 3: Data injection (codes into cells, call rows)
                           Phase 4: Formulas + metadata (summary, CF, DV, anchors)
                                   |
                              openpyxl bytes
"""

from __future__ import annotations

import json
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path
from typing import Any
from uuid import UUID

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from app.core.logging import get_logger
from app.services.excel_metadata import (
    ExportMetadata,
    compute_row_hash,
    write_ref_sheet,
    write_sys_meta_sheet,
)
from app.services.tamc_color_scheme import TAMCColorScheme, get_color_scheme

logger = get_logger(__name__)


# ═══════════════════════════════════════════════════════════════════════════════
# Row Allocation
# ═══════════════════════════════════════════════════════════════════════════════


@dataclass
class RowAllocation:
    """Tracks person-to-row assignments for the export."""

    resident_rows: dict[str, int] = field(default_factory=dict)  # person_id -> row
    faculty_rows: dict[str, int] = field(default_factory=dict)
    extra_section_rows: dict[str, int] = field(default_factory=dict)
    last_resident_row: int = 0
    last_faculty_row: int = 0
    hidden_rows: set[int] = field(default_factory=set)


# ═══════════════════════════════════════════════════════════════════════════════
# Rotation Colors (from reference)
# ═══════════════════════════════════════════════════════════════════════════════

ROTATION_COLORS: dict[str, str] = {
    # Blue (FMIT family)
    "FMIT 2": "FF00B0F0",
    "FMIT2": "FF00B0F0",
    "FMIT2(R2 Role)": "FF00B0F0",
    "ITE": "FF00B0F0",
    # Yellow (outpatient/elective)
    "Peds EM": "FFFFFF00",
    "FMC": "FFFFFF00",
    "ELE": "FFFFFF00",
    "ELE (SM)": "FFFFFF00",
    "MSK-SEL": "FFFFFF00",
    "NBN": "FFFFFF00",
    "Neuro": "FFFFFF00",
    "Procedures": "FFFFFF00",
    "SM": "FFFFFF00",
    "POCUS": "FFFFFF00",
    "PROC": "FFFFFF00",
    "Gyn Clinic": "FFFFFF00",
    "Elect - Sports Med": "FFFFFF00",
    "Select Med (Neuro)": "FFFFFF00",
    "Peds Subspecialty": "FFFFFF00",
    "TAMC L and D": "FFFFFF00",
    # Red (inpatient/critical)
    "HILO": "FFFF0000",
    "Japan": "FFFF0000",
    "L and D night Float": "FFFF0000",
    "L and D Night Float": "FFFF0000",
    "Derm": "FFFF0000",
    "Peds Ward": "FFFF0000",
    "Cards": "FFFF0000",
    "IM": "FFFF0000",
    "Surg Exp": "FFFF0000",
    "Kapiolani L and D": "FFFF0000",
    "Kapiolani  L and D": "FFFF0000",
    # Black (night float)
    "NF": "FF000000",
    "Peds NF": "FF000000",
}

# Font colors for rotation labels (light gray on black backgrounds)
ROTATION_FONT_COLORS: dict[str, str] = {
    "NF": "FFE8E8E8",
    "Peds NF": "FFE8E8E8",
}

# ═══════════════════════════════════════════════════════════════════════════════
# Constants
# ═══════════════════════════════════════════════════════════════════════════════

BLACK_FILL = PatternFill("solid", fgColor="FF000000")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFFFF")
STAFF_CALL_FILL = PatternFill("solid", fgColor="FFFFF2CC")
THURS_FILL = PatternFill("solid", fgColor="FF99CCFF")
WEEKEND_FILL = PatternFill("solid", fgColor="FF9BC2E6")

NBSP = "\xa0"
NO_BORDER = Border()


class TAMCBlockExporter:
    """Build a pixel-perfect Block Template2 workbook from JSON schedule data."""

    # Column layout
    COL_ROTATION1 = 1
    COL_ROTATION2 = 2
    COL_TEMPLATE = 3
    COL_ROLE = 4
    COL_NAME = 5
    COL_SCHEDULE_START = 6
    COLS_PER_DAY = 2

    # Row bands (fixed positions matching reference)
    RESIDENT_START = 9
    RESIDENT_END = 28
    ROW_RES_TOTALS = 29
    ROW_FAC_HEADERS = 30
    FACULTY_START = 31
    FACULTY_END = 47

    # Full black separator rows (entire row is BLACK + NBSP)
    SEPARATOR_ROWS = {7, 8, 29, 48, 55, 66, 69}
    # Partial black rows (A:B = BLACK, rest has summary data)
    PARTIAL_BLACK_ROWS = {30}

    # Extra section row ranges
    TY_FLT_START = 49
    TY_FLT_END = 54
    USU_START = 56
    USU_END = 58
    MS_START = 61
    MS_END = 65
    CP_BHC_START = 67
    CP_BHC_END = 68
    APPT_HEADER_ROW = 71
    CALC_START_ROW = 72
    CALC_END_ROW = 85

    def __init__(
        self,
        color_scheme: TAMCColorScheme | None = None,
        block_config: dict[str, Any] | None = None,
    ) -> None:
        self.color_scheme = color_scheme or get_color_scheme()
        self.block_config = block_config or {}
        self.row_mappings: dict[str, int] = {}
        # Computed dynamically in export() — defaults to 62 (BJ) for 28-day blocks
        self.summary_col_start: int = 62

    # ══════════════════════════════════════════════════════════════════════
    # Public API
    # ══════════════════════════════════════════════════════════════════════

    def export(
        self,
        data: dict[str, Any],
        *,
        export_metadata: ExportMetadata | None = None,
        rotation_codes: list[str] | None = None,
        activity_codes: list[str] | None = None,
    ) -> bytes:
        """Generate a complete Block Template2 workbook from JSON data."""
        block_start = date.fromisoformat(data["block_start"])
        block_end = date.fromisoformat(data["block_end"])
        num_days = (block_end - block_start).days + 1
        max_col = self.COL_SCHEDULE_START + (num_days * self.COLS_PER_DAY) - 1

        # Summary columns start after the schedule grid (shifts right for >28-day blocks)
        self.summary_col_start = max_col + 1

        # Phase 1: Structure
        wb = self._create_workbook()
        ws = wb.active
        ws.title = "Block Template2"
        self._set_dimensions(ws, block_start, block_end)

        # Phase 2: Row allocation
        alloc = self._allocate_rows(data)
        self.row_mappings = {**alloc.resident_rows, **alloc.faculty_rows}

        # Phase 2b: Formatting
        call_data = data.get("call", {})
        call_rows = []
        if isinstance(call_data, dict):
            call_rows = call_data.get("nights", []) or []
        elif isinstance(call_data, list):
            call_rows = call_data

        self._write_headers(ws, block_start, block_end, call_rows)
        self._write_black_separators(ws, max_col)
        self._fill_black_cols_ab(ws)

        # Phase 3: Data injection
        residents = data.get("residents", []) or []
        faculty = data.get("faculty", []) or []

        self._write_rotation_labels(ws, residents, alloc)
        self._write_residents(ws, residents, block_start, alloc)
        self._write_faculty(ws, faculty, block_start, alloc)
        self._write_call_row(ws, call_rows, block_start, block_end)
        self._write_extra_sections(ws, alloc)

        # Phase 3b: Labels for rows 6
        self._write_label_row(ws, max_col)

        # Phase 4: Formulas and metadata
        self._write_resident_summaries(ws, block_start, block_end)
        self._write_faculty_summaries(ws, block_start, block_end)
        self._write_appointments_section(ws, block_start, block_end)
        self._add_conditional_formatting(ws, block_start, block_end)
        self._add_data_validation(ws, block_start, block_end)
        self._merge_header_cells(ws, block_start, block_end)
        self._write_anchor_sheet(wb, data, alloc)
        self._add_leave_formula_column(ws, block_start, block_end)

        # Hide unused rows
        for r in alloc.hidden_rows:
            ws.row_dimensions[r].hidden = True

        # Metadata sheets
        if export_metadata is not None:
            write_sys_meta_sheet(wb, export_metadata)
            if rotation_codes is not None and activity_codes is not None:
                write_ref_sheet(wb, rotation_codes, activity_codes)

        # Freeze panes
        ws.freeze_panes = "F9"

        # Final font/border pass
        self._apply_fonts_and_borders(ws, max_col)

        # Save
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ══════════════════════════════════════════════════════════════════════
    # Phase 1: Structure
    # ══════════════════════════════════════════════════════════════════════

    def _create_workbook(self) -> Workbook:
        wb = Workbook()
        return wb

    def _set_dimensions(self, ws, block_start: date, block_end: date) -> None:
        """Set column widths and row heights matching reference."""
        num_days = (block_end - block_start).days + 1
        max_col = self.COL_SCHEDULE_START + (num_days * self.COLS_PER_DAY) - 1

        # Column widths
        ws.column_dimensions["A"].width = 10.16
        ws.column_dimensions["B"].width = 12.16
        ws.column_dimensions["C"].width = 10.16
        ws.column_dimensions["D"].width = 11.83
        ws.column_dimensions["E"].width = 40.5
        for col in range(self.COL_SCHEDULE_START, max_col + 1):
            ws.column_dimensions[get_column_letter(col)].width = 10.16

        # Row heights
        ws.row_dimensions[1].height = 15  # Day full names
        ws.row_dimensions[2].height = 15  # Day abbreviations
        ws.row_dimensions[3].height = 20  # Dates
        ws.row_dimensions[4].height = 25  # Staff call
        ws.row_dimensions[5].height = 15  # Resident call
        ws.row_dimensions[6].height = 20  # Labels
        ws.row_dimensions[7].height = 20  # Black separator
        ws.row_dimensions[8].height = 21  # Summary header / black separator

        for row in range(self.RESIDENT_START, self.RESIDENT_END + 1):
            ws.row_dimensions[row].height = 21
        ws.row_dimensions[self.ROW_RES_TOTALS].height = 21
        ws.row_dimensions[self.ROW_FAC_HEADERS].height = 20

        for row in range(self.FACULTY_START, self.FACULTY_END + 1):
            ws.row_dimensions[row].height = 20
        for row in range(48, 70):
            ws.row_dimensions[row].height = 20
        ws.row_dimensions[70].height = 15  # White separator
        ws.row_dimensions[self.APPT_HEADER_ROW].height = 15

        for row in range(self.CALC_START_ROW, self.CALC_END_ROW + 1):
            ws.row_dimensions[row].height = 21

    # ══════════════════════════════════════════════════════════════════════
    # Phase 2: Formatting
    # ══════════════════════════════════════════════════════════════════════

    def _write_headers(
        self,
        ws,
        block_start: date,
        block_end: date,
        call_rows: list[dict[str, Any]],
    ) -> None:
        """Write header rows 1-5 with day names, abbreviations, dates, call."""
        num_days = (block_end - block_start).days + 1
        day_names = [
            "MONDAY",
            "TUESDAY",
            "WEDNESDAY",
            "THURSDAY",
            "FRIDAY",
            "SATURDAY",
            "SUNDAY",
        ]
        day_abbrevs = ["MON", "TUES", "WED", "THURS", "FRI", "SAT", "SUN"]

        # Black fill for A, B, E in rows 1-5
        for row in range(1, 6):
            ws.cell(row=row, column=1).fill = BLACK_FILL
            ws.cell(row=row, column=1).value = NBSP
            ws.cell(row=row, column=2).fill = BLACK_FILL
            ws.cell(row=row, column=2).value = NBSP
            ws.cell(row=row, column=5).fill = BLACK_FILL
            ws.cell(row=row, column=5).value = NBSP

        # Row 3, col E: "Date:"
        ws.cell(row=3, column=5).value = "Date:"
        ws.cell(row=3, column=5).fill = BLACK_FILL
        ws.cell(row=3, column=5).font = Font(
            name="Arial", size=16, bold=True, color="FFFFFFFF"
        )

        # Row 4: date range in C
        ws.cell(row=4, column=3).value = (
            f"{block_start.strftime('%d %b').lstrip('0')} - "
            f"{block_end.strftime('%d %b %Y').lstrip('0')}"
        )

        # Row 4: Staff Call label
        ws.cell(row=4, column=5).value = "Staff Call"
        ws.cell(row=4, column=5).fill = STAFF_CALL_FILL
        ws.cell(row=4, column=5).font = Font(name="Arial", size=20, bold=True)

        # Row 5: Resident Call label
        ws.cell(row=5, column=5).value = "Resident Call"
        ws.cell(row=5, column=5).fill = BLACK_FILL
        ws.cell(row=5, column=5).font = Font(
            name="Arial", size=14, bold=True, color="FFFFFFFF"
        )

        # Write day headers for each date
        for day_idx in range(num_days):
            d = block_start + timedelta(days=day_idx)
            am_col = self.COL_SCHEDULE_START + (day_idx * self.COLS_PER_DAY)
            dow = d.weekday()  # 0=Mon

            # Row 1: Full day name
            cell1 = ws.cell(row=1, column=am_col, value=day_names[dow])
            # Row 2: Day abbreviation
            cell2 = ws.cell(row=2, column=am_col, value=day_abbrevs[dow])
            # Row 3: Actual date (Excel serial date, displayed as M/D)
            cell3 = ws.cell(
                row=3, column=am_col, value=datetime(d.year, d.month, d.day)
            )
            cell3.number_format = "M/D"

            # Day-of-week coloring
            if dow == 3:  # Thursday
                cell1.fill = THURS_FILL
                cell2.fill = THURS_FILL
            elif dow >= 5:  # Saturday/Sunday
                cell1.fill = WEEKEND_FILL
                cell2.fill = WEEKEND_FILL
            else:
                cell1.fill = WHITE_FILL
                cell2.fill = WHITE_FILL

            # Staff call row 4: yellow fill for schedule cells
            ws.cell(row=4, column=am_col).fill = STAFF_CALL_FILL
            ws.cell(row=4, column=am_col + 1).fill = STAFF_CALL_FILL

    def _write_black_separators(self, ws, max_col: int) -> None:
        """Fill full black separator rows with BLACK + NBSP."""
        for row in self.SEPARATOR_ROWS:
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = BLACK_FILL
                cell.value = NBSP

    def _fill_black_cols_ab(self, ws) -> None:
        """Fill columns A and B with BLACK + NBSP for faculty and extra rows."""
        # Faculty rows
        for row in range(self.FACULTY_START, self.FACULTY_END + 1):
            ws.cell(row=row, column=1).fill = BLACK_FILL
            ws.cell(row=row, column=1).value = NBSP
            ws.cell(row=row, column=2).fill = BLACK_FILL
            ws.cell(row=row, column=2).value = NBSP

        # Extra section rows (TY/FLT, USU, MS, CP/BHC)
        for row in range(self.TY_FLT_START, self.TY_FLT_END + 1):
            ws.cell(row=row, column=1).value = NBSP
            ws.cell(row=row, column=2).value = NBSP
        for row in range(self.USU_START, self.USU_END + 1):
            ws.cell(row=row, column=1).value = NBSP
            ws.cell(row=row, column=2).value = NBSP
        for row in range(self.MS_START, self.MS_END + 1):
            ws.cell(row=row, column=1).value = NBSP
            ws.cell(row=row, column=2).value = NBSP
        for row in range(self.CP_BHC_START, self.CP_BHC_END + 1):
            ws.cell(row=row, column=1).value = NBSP
            ws.cell(row=row, column=2).value = NBSP

        # Partial black rows (summary/header)
        for row in self.PARTIAL_BLACK_ROWS:
            ws.cell(row=row, column=1).fill = BLACK_FILL
            ws.cell(row=row, column=1).value = NBSP
            ws.cell(row=row, column=2).fill = BLACK_FILL
            ws.cell(row=row, column=2).value = NBSP

        # Buffer rows 26-28 between residents and totals
        for row in range(26, 29):
            ws.cell(row=row, column=1).fill = BLACK_FILL
            ws.cell(row=row, column=1).value = NBSP
            ws.cell(row=row, column=2).fill = BLACK_FILL
            ws.cell(row=row, column=2).value = NBSP

    def _write_label_row(self, ws, max_col: int) -> None:
        """Write row 6 labels: TEMPLATE, ROLE, PROVIDER + BLACK schedule cols."""
        ws.cell(row=6, column=self.COL_TEMPLATE).value = "TEMPLATE"
        ws.cell(row=6, column=self.COL_ROLE).value = "ROLE"
        ws.cell(row=6, column=self.COL_NAME).value = "PROVIDER"

        # Black fill for schedule columns in row 6
        for col in range(self.COL_SCHEDULE_START, max_col + 1):
            ws.cell(row=6, column=col).fill = BLACK_FILL
            ws.cell(row=6, column=col).value = NBSP

    # ══════════════════════════════════════════════════════════════════════
    # Phase 3: Data Injection
    # ══════════════════════════════════════════════════════════════════════

    def _allocate_rows(self, data: dict[str, Any]) -> RowAllocation:
        """Assign rows to residents and faculty, matching reference layout."""
        alloc = RowAllocation()

        # Residents sorted by PGY DESC, then alpha
        residents = data.get("residents", []) or []

        def _pgy_sort_key(r: dict[str, Any]) -> tuple[int, str]:
            pgy_val = r.get("pgy")
            pgy_int = int(pgy_val) if pgy_val else 0
            return (-pgy_int, r.get("name", ""))

        residents_sorted = sorted(residents, key=_pgy_sort_key)

        current_row = self.RESIDENT_START
        for res in residents_sorted:
            pid = str(res.get("id", ""))
            if not pid:
                continue
            if current_row > self.RESIDENT_END:
                logger.warning("Resident band overflow at row %d", current_row)
                break
            alloc.resident_rows[pid] = current_row
            current_row += 1
        alloc.last_resident_row = current_row - 1

        # Hide unused resident slots
        for r in range(current_row, self.RESIDENT_END + 1):
            alloc.hidden_rows.add(r)

        # Faculty: core first (alpha), then adjunct (alpha)
        faculty = data.get("faculty", []) or []
        faculty_sorted = sorted(
            faculty,
            key=lambda f: (
                1 if f.get("faculty_role") == "adjunct" else 0,
                f.get("name", ""),
            ),
        )

        current_row = self.FACULTY_START
        for fac in faculty_sorted:
            pid = str(fac.get("id", ""))
            if not pid:
                continue
            if current_row > self.FACULTY_END:
                logger.warning("Faculty band overflow at row %d", current_row)
                break
            alloc.faculty_rows[pid] = current_row
            current_row += 1
        alloc.last_faculty_row = current_row - 1

        # Hide unused faculty slots
        for r in range(current_row, self.FACULTY_END + 1):
            alloc.hidden_rows.add(r)

        return alloc

    def _write_residents(
        self,
        ws,
        residents: list[dict[str, Any]],
        block_start: date,
        alloc: RowAllocation,
    ) -> None:
        """Write resident schedule codes into cells."""
        for person in residents:
            pid = str(person.get("id", ""))
            row = alloc.resident_rows.get(pid)
            if row is None:
                continue

            pgy = person.get("pgy")
            rotation1 = person.get("rotation1", "") or ""
            rotation2 = person.get("rotation2", "") or ""
            name = person.get("name", "")

            # Identity columns
            ws.cell(row=row, column=self.COL_TEMPLATE).value = f"R{pgy}" if pgy else ""
            ws.cell(row=row, column=self.COL_ROLE).value = f"PGY {pgy}" if pgy else ""
            ws.cell(row=row, column=self.COL_NAME).value = self._to_last_first(name)

            # Schedule columns
            self._write_schedule_days(ws, person, row, block_start)

    def _write_faculty(
        self,
        ws,
        faculty: list[dict[str, Any]],
        block_start: date,
        alloc: RowAllocation,
    ) -> None:
        """Write faculty schedule codes into cells."""
        for person in faculty:
            pid = str(person.get("id", ""))
            row = alloc.faculty_rows.get(pid)
            if row is None:
                continue

            name = person.get("name", "")
            faculty_role = person.get("faculty_role", "core")

            # Identity columns (no rotation for faculty, A/B are black)
            ws.cell(row=row, column=self.COL_TEMPLATE).value = (
                "ADJ" if faculty_role == "adjunct" else "C19"
            )
            ws.cell(row=row, column=self.COL_ROLE).value = "FAC"
            ws.cell(row=row, column=self.COL_NAME).value = self._to_last_first(name)

            # Schedule columns
            self._write_schedule_days(ws, person, row, block_start)

    def _write_schedule_days(
        self,
        ws,
        person: dict[str, Any],
        row: int,
        block_start: date,
    ) -> None:
        """Write AM/PM codes for each day and apply cell colors."""
        days = person.get("days", []) or []
        sorted_days = sorted(days, key=lambda d: str(d.get("date", "")))

        for day in sorted_days:
            day_date_str = day.get("date")
            if not day_date_str:
                continue
            day_date = date.fromisoformat(str(day_date_str))
            day_offset = (day_date - block_start).days
            am_col = self.COL_SCHEDULE_START + (day_offset * self.COLS_PER_DAY)
            pm_col = am_col + 1

            am_code = day.get("am", "") or ""
            pm_code = day.get("pm", "") or ""

            am_cell = ws.cell(row=row, column=am_col, value=am_code)
            pm_cell = ws.cell(row=row, column=pm_col, value=pm_code)

            if am_code:
                self._apply_cell_color(am_cell, am_code)
            if pm_code:
                self._apply_cell_color(pm_cell, pm_code)

    def _write_rotation_labels(
        self,
        ws,
        residents: list[dict[str, Any]],
        alloc: RowAllocation,
    ) -> None:
        """Write rotation labels in columns A and B with color coding."""
        overrides = self.block_config.get("rotation_overrides", {})

        for person in residents:
            pid = str(person.get("id", ""))
            row = alloc.resident_rows.get(pid)
            if row is None:
                continue

            name = person.get("name", "")
            rotation1 = person.get("rotation1", "") or ""
            rotation2 = person.get("rotation2", "") or ""

            # Use override name if available
            display_rot1 = overrides.get(name, rotation1)

            cell_a = ws.cell(row=row, column=self.COL_ROTATION1, value=display_rot1)
            fill, font = self._rotation_fill_and_font(display_rot1)
            if fill:
                cell_a.fill = fill
            if font:
                cell_a.font = font

            # Second rotation (mid-block switch)
            if rotation2:
                display_rot2 = rotation2
                cell_b = ws.cell(row=row, column=self.COL_ROTATION2, value=display_rot2)
                fill2, font2 = self._rotation_fill_and_font(display_rot2)
                if fill2:
                    cell_b.fill = fill2
                if font2:
                    cell_b.font = font2

    def _write_call_row(
        self,
        ws,
        call_rows: list[dict[str, Any]],
        block_start: date,
        block_end: date,
    ) -> None:
        """Fill call rows: faculty -> row 4, residents -> row 5."""
        for night in call_rows:
            night_date_str = night.get("date")
            if not night_date_str:
                continue
            night_date = date.fromisoformat(str(night_date_str))
            staff_name = night.get("staff", "")
            if not staff_name:
                continue

            person_type = night.get("person_type", "faculty")
            target_row = 5 if person_type == "resident" else 4

            day_offset = (night_date - block_start).days
            am_col = self.COL_SCHEDULE_START + (day_offset * self.COLS_PER_DAY)

            cell = ws.cell(row=target_row, column=am_col, value=staff_name)
            if target_row == 4:
                cell.fill = STAFF_CALL_FILL

    def _write_extra_sections(self, ws, alloc: RowAllocation) -> None:
        """Write TY/FLT, USU/IPAP, MS, CP/BHC placeholder rows from config."""
        # TY/FLT section (rows 49-54)
        ty_rotators = self.block_config.get("ty_rotators", [])
        for i, rotator in enumerate(ty_rotators):
            row = self.TY_FLT_START + i
            if row > self.TY_FLT_END:
                break
            slot = rotator.get("slot", "")
            name = rotator.get("name", "")
            rtype = rotator.get("type", "")
            ws.cell(row=row, column=self.COL_TEMPLATE).value = rtype or slot
            ws.cell(row=row, column=self.COL_ROLE).value = slot
            ws.cell(row=row, column=self.COL_NAME).value = name

        # USU/IPAP section (rows 56-58)
        usu_students = self.block_config.get("usu_students", [])
        for i, student in enumerate(usu_students):
            row = self.USU_START + i
            if row > self.USU_END:
                break
            ws.cell(row=row, column=self.COL_TEMPLATE).value = student.get(
                "role_code", "USU"
            )
            ws.cell(row=row, column=self.COL_ROLE).value = student.get("role", "")
            ws.cell(row=row, column=self.COL_NAME).value = student.get("name", "")

        # CP/BHC section (rows 67-68)
        support_staff = self.block_config.get("support_staff", [])
        for i, staff in enumerate(support_staff):
            row = self.CP_BHC_START + i
            if row > self.CP_BHC_END:
                break
            ws.cell(row=row, column=self.COL_TEMPLATE).value = staff.get(
                "role_code", ""
            )
            ws.cell(row=row, column=self.COL_ROLE).value = staff.get("role", "")
            ws.cell(row=row, column=self.COL_NAME).value = staff.get("name", "")

    # ══════════════════════════════════════════════════════════════════════
    # Phase 4: Formulas & Metadata
    # ══════════════════════════════════════════════════════════════════════

    def _write_resident_summaries(self, ws, block_start: date, block_end: date) -> None:
        """Write summary headers and formulas for resident rows."""
        sc = (
            self.summary_col_start
        )  # First summary column (shifts right for >28-day blocks)
        sched_end = get_column_letter(self._schedule_end_col(block_start, block_end))
        start_l = get_column_letter(self.COL_SCHEDULE_START)

        # Row 8: Resident summary headers
        headers = ["C", "CC", "CV", "(C+CC+CV)", "NF", "CC", "PC/OFF", "LV", "FMIT"]
        for idx, header in enumerate(headers):
            ws.cell(row=8, column=sc + idx).value = header

        # Column letters for the first three summary cols (used in total formula)
        c_let = get_column_letter(sc)  # C count
        cc_let = get_column_letter(sc + 1)  # CC count
        cv_let = get_column_letter(sc + 2)  # CV count

        # Rows 9-28: Resident formulas
        for row in range(self.RESIDENT_START, self.RESIDENT_END + 1):
            rng = f"{start_l}{row}:{sched_end}{row}"
            c_terms = '{"C","C-I","SM"}' if row == 9 else '{"C","CV","C-I","SM"}'
            ws.cell(
                row=row, column=sc
            ).value = f"=SUMPRODUCT(COUNTIF({rng}, {c_terms}))"
            ws.cell(row=row, column=sc + 1).value = f'=COUNTIF({rng}, "CC")'
            ws.cell(row=row, column=sc + 2).value = f'=COUNTIF({rng}, "CV")'
            ws.cell(
                row=row, column=sc + 3
            ).value = f"={c_let}{row}+{cc_let}{row}+{cv_let}{row}"
            ws.cell(row=row, column=sc + 4).value = f'=COUNTIF({rng}, "NF")'
            ws.cell(row=row, column=sc + 5).value = f'=COUNTIF({rng}, "CC")'
            ws.cell(
                row=row, column=sc + 6
            ).value = f'=SUMPRODUCT(COUNTIF({rng}, {{"PC","OFF","W"}}))'
            ws.cell(row=row, column=sc + 7).value = f'=COUNTIF({rng}, "LV")'
            ws.cell(row=row, column=sc + 8).value = f'=COUNTIF({rng}, "FMIT")'

        # Row 29: Totals
        for offset in range(9):
            col = sc + offset
            col_letter = get_column_letter(col)
            ws.cell(
                row=self.ROW_RES_TOTALS, column=col
            ).value = f"=SUM({col_letter}{self.RESIDENT_START}:{col_letter}{self.RESIDENT_END})"

    def _write_faculty_summaries(self, ws, block_start: date, block_end: date) -> None:
        """Write summary headers and formulas for faculty rows."""
        sc = self.summary_col_start
        sched_end = get_column_letter(self._schedule_end_col(block_start, block_end))
        start_l = get_column_letter(self.COL_SCHEDULE_START)

        # Row 30: Faculty summary headers
        fac_headers = [
            "C",
            "CC",
            "CV",
            "(C+CC+CV)",
            "AT",
            "ADM",
            "LV",
            "FMIT",
            "CALL",
        ]
        for idx, header in enumerate(fac_headers):
            ws.cell(row=self.ROW_FAC_HEADERS, column=sc + idx).value = header
        # SUN column (one past CALL)
        sun_col = sc + len(fac_headers)
        ws.cell(row=self.ROW_FAC_HEADERS, column=sun_col).value = "SUN"

        # Column letters for first three summary cols (total formula)
        c_let = get_column_letter(sc)
        cc_let = get_column_letter(sc + 1)
        cv_let = get_column_letter(sc + 2)

        # Compute Sunday column letters for SUN call count
        num_days = (block_end - block_start).days + 1
        sunday_col_letters: list[str] = []
        for day_offset in range(num_days):
            d = block_start + timedelta(days=day_offset)
            if d.weekday() == 6:  # Sunday
                am_col = self.COL_SCHEDULE_START + (day_offset * self.COLS_PER_DAY)
                sunday_col_letters.append(get_column_letter(am_col))
                sunday_col_letters.append(get_column_letter(am_col + 1))

        # Call row range for COUNTIF
        call_rng = f"{start_l}4:{sched_end}4"

        # Rows 31-47: Faculty formulas
        for row in range(self.FACULTY_START, self.FACULTY_END + 1):
            rng = f"{start_l}{row}:{sched_end}{row}"
            ws.cell(
                row=row, column=sc
            ).value = f'=SUMPRODUCT(COUNTIF({rng}, {{"C","SM"}}))'
            ws.cell(row=row, column=sc + 1).value = f'=COUNTIF({rng}, "CC")'
            ws.cell(row=row, column=sc + 2).value = f'=COUNTIF({rng}, "CV")'
            ws.cell(
                row=row, column=sc + 3
            ).value = f"={c_let}{row}+{cc_let}{row}+{cv_let}{row}"
            ws.cell(
                row=row, column=sc + 4
            ).value = f'=SUMPRODUCT(COUNTIF({rng}, {{"AT","PCAT","DO"}}))'
            ws.cell(
                row=row, column=sc + 5
            ).value = f'=SUMPRODUCT(COUNTIF({rng}, {{"GME","DFM","DOFM"}}))'
            ws.cell(row=row, column=sc + 6).value = f'=COUNTIF({rng}, "LV")'
            ws.cell(row=row, column=sc + 7).value = f'=COUNTIF({rng}, "FMIT")'

            # CALL count: match faculty last name against call row 4
            call_name = self._call_last_name_token(
                ws.cell(row=row, column=self.COL_NAME).value
            )
            ws.cell(row=row, column=sc + 8).value = (
                f'=COUNTIF({call_rng}, "{call_name}")' if call_name else None
            )

            # SUN column
            if call_name and sunday_col_letters:
                sun_terms = "+".join(
                    f'COUNTIF({c}4,"{call_name}")' for c in sunday_col_letters
                )
                ws.cell(row=row, column=sun_col).value = f"={sun_terms}"
            else:
                ws.cell(row=row, column=sun_col).value = 0

    def _write_appointments_section(
        self, ws, block_start: date, block_end: date
    ) -> None:
        """Write the appointments/daily staffing section (rows 71-85)."""
        num_days = (block_end - block_start).days + 1
        schedule_end = self.COL_SCHEDULE_START + (num_days * self.COLS_PER_DAY) - 1

        # Row 71: "Appointments:" header with BLACK bg
        ws.cell(row=self.APPT_HEADER_ROW, column=self.COL_NAME).value = "Appointments:"
        ws.cell(row=self.APPT_HEADER_ROW, column=self.COL_NAME).font = Font(
            name="Arial", size=11, bold=True
        )
        for col in range(1, 5):
            ws.cell(row=self.APPT_HEADER_ROW, column=col).fill = BLACK_FILL
            ws.cell(row=self.APPT_HEADER_ROW, column=col).value = NBSP

        labels = [
            "Screeners Needed",
            "Providers Virtual (CV)",
            "Interns in Clinic",
            "Residents in Clinic",
            "Attendings Needed Clinic",
            "Residents in PROC",
            "Residents in V Clinic",
            "Residents on HV",
            "Total Attendings Needed",
            "# Attendings Assigned",
            "# Primary Care Appts",
            "PR Count",
            "VAS Count",
            "PR/VAS Conflict",
        ]

        bold = Font(name="Arial", bold=True, size=10)
        for i, label in enumerate(labels):
            ws.cell(
                row=self.CALC_START_ROW + i, column=self.COL_NAME, value=label
            ).font = bold

        # Row range strings for formulas
        for col_idx in range(self.COL_SCHEDULE_START, schedule_end + 1):
            c = get_column_letter(col_idx)
            r_all = f"{c}${self.RESIDENT_START}:{c}${self.FACULTY_END}"
            r_res = f"{c}${self.RESIDENT_START}:{c}${self.RESIDENT_END}"
            r_fac = f"{c}${self.FACULTY_START}:{c}${self.FACULTY_END}"

            # Row 72: Screeners Needed
            ws.cell(row=72, column=col_idx).value = (
                f'=COUNTIF({r_all},"C")'
                f'+COUNTIF({r_all},"C30")'
                f'+COUNTIF({r_all},"C40")'
                f'+COUNTIF({r_all},"C60")'
                f'+COUNTIF({r_all},"CC")'
                f'+COUNTIF({r_all},"C-I")'
                f'+COUNTIF({r_all},"C-N")'
                f'+COUNTIF({r_all},"VAS")'
                f'+COUNTIF({r_all},"VASc")'
                f'+COUNTIF({r_all},"PR")'
                f'+COUNTIF({r_all},"OPR")'
                f'+COUNTIF({r_all},"SM")'
                f'+COUNTIF({r_all},"GAC")'
                f'+COUNTIF({r_all},"PE")'
                f'+COUNTIF({r_all},"C10")'
                f'+COUNTIF({r_all},"V1")'
                f'+COUNTIF({r_all},"V2")'
                f'+COUNTIF({r_all},"V3")'
                f'+COUNTIF({r_all},"PAP")*2'
                f'+COUNTIF({r_all},"CF2V")*0.5'
                f'+COUNTIF({r_all},"CV2F")*0.5'
            )

            # Row 73: Providers Virtual
            ws.cell(row=73, column=col_idx).value = (
                f'=COUNTIF({r_all},"CV")'
                f'+COUNTIF({r_all},"CV10")'
                f'+COUNTIF({r_all},"CF2V")*0.5'
                f'+COUNTIF({r_all},"CV2F")*0.5'
            )

            # Row 74: Interns in Clinic
            ws.cell(row=74, column=col_idx).value = (
                f'=COUNTIFS($D${self.RESIDENT_START}:$D${self.RESIDENT_END},"PGY 1",{r_res},"C")'
                f'+COUNTIFS($D${self.RESIDENT_START}:$D${self.RESIDENT_END},"PGY 1",{r_res},"C30")'
                f'+COUNTIFS($D${self.RESIDENT_START}:$D${self.RESIDENT_END},"PGY 1",{r_res},"C40")'
                f'+COUNTIFS($D${self.RESIDENT_START}:$D${self.RESIDENT_END},"PGY 1",{r_res},"C60")'
                f'+COUNTIFS($D${self.RESIDENT_START}:$D${self.RESIDENT_END},"PGY 1",{r_res},"CC")'
                f'+COUNTIFS($D${self.RESIDENT_START}:$D${self.RESIDENT_END},"PGY 1",{r_res},"C-I")'
            )

            # Row 75: Residents in Clinic (PGY 2/3)
            ws.cell(row=75, column=col_idx).value = (
                f'=COUNTIFS($D${self.RESIDENT_START}:$D${self.RESIDENT_END},"PGY 2",{r_res},"C")'
                f'+COUNTIFS($D${self.RESIDENT_START}:$D${self.RESIDENT_END},"PGY 3",{r_res},"C")'
                f'+COUNTIFS($D${self.RESIDENT_START}:$D${self.RESIDENT_END},"PGY 2",{r_res},"C30")'
                f'+COUNTIFS($D${self.RESIDENT_START}:$D${self.RESIDENT_END},"PGY 3",{r_res},"C30")'
                f'+COUNTIFS($D${self.RESIDENT_START}:$D${self.RESIDENT_END},"PGY 2",{r_res},"C40")'
                f'+COUNTIFS($D${self.RESIDENT_START}:$D${self.RESIDENT_END},"PGY 3",{r_res},"C40")'
                f'+COUNTIFS($D${self.RESIDENT_START}:$D${self.RESIDENT_END},"PGY 2",{r_res},"C60")'
                f'+COUNTIFS($D${self.RESIDENT_START}:$D${self.RESIDENT_END},"PGY 3",{r_res},"C60")'
                f'+COUNTIFS($D${self.RESIDENT_START}:$D${self.RESIDENT_END},"PGY 2",{r_res},"CC")'
                f'+COUNTIFS($D${self.RESIDENT_START}:$D${self.RESIDENT_END},"PGY 3",{r_res},"CC")'
                f'+COUNTIFS($D${self.RESIDENT_START}:$D${self.RESIDENT_END},"PGY 2",{r_res},"C-I")'
                f'+COUNTIFS($D${self.RESIDENT_START}:$D${self.RESIDENT_END},"PGY 3",{r_res},"C-I")'
                f'+COUNTIFS($D${self.RESIDENT_START}:$D${self.RESIDENT_END},"PGY 2",{r_res},"SM")'
                f'+COUNTIFS($D${self.RESIDENT_START}:$D${self.RESIDENT_END},"PGY 3",{r_res},"SM")'
                f'+COUNTIFS($D${self.RESIDENT_START}:$D${self.RESIDENT_END},"PGY 2",{r_res},"CV")'
                f'+COUNTIFS($D${self.RESIDENT_START}:$D${self.RESIDENT_END},"PGY 3",{r_res},"CV")'
            )

            # Row 76: Attendings Needed
            ws.cell(row=76, column=col_idx).value = f"={c}74*0.5+{c}75*0.25"

            # Row 77: Residents in PROC
            ws.cell(row=77, column=col_idx).value = (
                f'=COUNTIF({r_res},"VAS")+COUNTIF({r_res},"PR")'
                f'+COUNTIF({r_res},"OPR")+COUNTIF({r_res},"PROC")'
            )

            # Row 78: Residents in V Clinic
            ws.cell(row=78, column=col_idx).value = (
                f'=COUNTIF({r_res},"V1")'
                f'+COUNTIF({r_res},"V2")'
                f'+COUNTIF({r_res},"V3")'
                f'+COUNTIF({r_res},"CV")'
            )

            # Row 79: Residents on HV
            ws.cell(row=79, column=col_idx).value = f'=COUNTIF({r_res},"HV")'

            # Row 80: Total Attendings Needed
            ws.cell(row=80, column=col_idx).value = f"=ROUNDUP(SUM({c}76:{c}77),0)"

            # Row 81: # Attendings Assigned
            ws.cell(
                row=81, column=col_idx
            ).value = (
                f'=COUNTIF({r_fac},"AT")+COUNTIF({r_fac},"PCAT")+COUNTIF({r_fac},"DO")'
            )

            # Row 82: # Primary Care Appts (manual)
            ws.cell(row=82, column=col_idx).value = ""

            # Row 83: PR Count
            ws.cell(row=83, column=col_idx).value = f'=COUNTIF({r_res},"PR")'

            # Row 84: VAS Count
            ws.cell(row=84, column=col_idx).value = f'=COUNTIF({r_res},"VAS")'

            # Row 85: PR/VAS Conflict
            ws.cell(
                row=85, column=col_idx
            ).value = f'=IF(OR({c}83>1,{c}84>1,AND({c}83>=1,{c}84>=1)),"\u26a0","")'

    def _add_conditional_formatting(
        self, ws, block_start: date, block_end: date
    ) -> None:
        """Add CF rules for schedule codes based on color scheme."""
        if not self.color_scheme:
            return

        num_days = (block_end - block_start).days + 1
        schedule_end_col = self.COL_SCHEDULE_START + (num_days * self.COLS_PER_DAY) - 1

        grid_range = (
            f"{get_column_letter(self.COL_SCHEDULE_START)}{self.RESIDENT_START}:"
            f"{get_column_letter(schedule_end_col)}{self.CP_BHC_END}"
        )

        for code, bg_color in self.color_scheme._code_colors.items():
            fg_color = self.color_scheme._font_colors.get(code, "000000")
            rule = CellIsRule(
                operator="equal",
                formula=[f'"{code}"'],
                fill=PatternFill(start_color=bg_color, fill_type="solid"),
                font=Font(color=fg_color),
            )
            ws.conditional_formatting.add(grid_range, rule)

    def _add_data_validation(self, ws, block_start: date, block_end: date) -> None:
        """Add dropdown validation for rotation and activity columns."""
        rot_dv = DataValidation(
            type="list",
            formula1="ValidRotations",
            allow_blank=True,
            showErrorMessage=False,
        )
        ws.add_data_validation(rot_dv)

        act_dv = DataValidation(
            type="list",
            formula1="ValidActivities",
            allow_blank=True,
            showErrorMessage=False,
        )
        ws.add_data_validation(act_dv)

        num_days = (block_end - block_start).days + 1
        schedule_end_col = self.COL_SCHEDULE_START + (num_days * self.COLS_PER_DAY) - 1

        target_rows = sorted(set(self.row_mappings.values()))
        for row in target_rows:
            rot_dv.add(ws.cell(row=row, column=self.COL_ROTATION1))
            rot_dv.add(ws.cell(row=row, column=self.COL_ROTATION2))
            for col in range(self.COL_SCHEDULE_START, schedule_end_col + 1):
                act_dv.add(ws.cell(row=row, column=col))

    def _merge_header_cells(self, ws, block_start: date, block_end: date) -> None:
        """Merge AM/PM columns in header rows 1-5."""
        num_days = (block_end - block_start).days + 1
        schedule_end_col = self.COL_SCHEDULE_START + (num_days * self.COLS_PER_DAY) - 1

        for col in range(
            self.COL_SCHEDULE_START, schedule_end_col + 1, self.COLS_PER_DAY
        ):
            am_letter = get_column_letter(col)
            pm_letter = get_column_letter(col + 1)
            for row in range(1, 6):
                merge_range = f"{am_letter}{row}:{pm_letter}{row}"
                try:
                    ws.merge_cells(merge_range)
                except ValueError:
                    pass  # Already merged

    def _write_anchor_sheet(
        self, wb: Workbook, data: dict[str, Any], alloc: RowAllocation
    ) -> None:
        """Create veryHidden __ANCHORS__ sheet with UUIDs and row hashes."""
        if "__ANCHORS__" in wb.sheetnames:
            del wb["__ANCHORS__"]
        ws = wb.create_sheet("__ANCHORS__")
        ws.sheet_state = "veryHidden"

        ws.cell(row=1, column=1, value="person_id")
        ws.cell(row=1, column=2, value="block_assignment_id")
        ws.cell(row=1, column=3, value="row_hash")

        all_mappings = {**alloc.resident_rows, **alloc.faculty_rows}
        all_people = (data.get("residents", []) or []) + (data.get("faculty", []) or [])

        for person in all_people:
            person_id = person.get("id")
            if not person_id:
                continue

            pid_str = str(person_id)
            row = all_mappings.get(pid_str)
            if not row:
                continue

            rotation1 = person.get("rotation1")
            rotation2 = person.get("rotation2")
            days_codes = []
            for day in person.get("days", []):
                days_codes.append(day.get("am"))
                days_codes.append(day.get("pm"))

            try:
                pid_uuid = UUID(person_id)
            except (ValueError, AttributeError):
                pid_uuid = UUID(int=0)

            row_hash = compute_row_hash(pid_uuid, rotation1, rotation2, days_codes)

            ws.cell(row=row, column=1, value=str(person_id))
            ws.cell(row=row, column=2, value=str(person.get("block_assignment_id", "")))
            ws.cell(row=row, column=3, value=row_hash)

    def _add_leave_formula_column(self, ws, block_start: date, block_end: date) -> None:
        """Add LV Days column after the last summary column."""
        leave_col = self.summary_col_start + 9  # After 9 resident summary cols
        ws.cell(row=8, column=leave_col, value="LV Days")

        start_letter = get_column_letter(self.COL_SCHEDULE_START)
        actual_days = (block_end - block_start).days + 1
        end_letter = get_column_letter(
            self.COL_SCHEDULE_START + (actual_days * self.COLS_PER_DAY) - 1
        )

        for row in range(self.RESIDENT_START, self.CP_BHC_END + 1):
            ws.cell(
                row=row,
                column=leave_col,
                value=f'=COUNTIF({start_letter}{row}:{end_letter}{row}, "LV")/2',
            )

    # ══════════════════════════════════════════════════════════════════════
    # Final Presentation Pass
    # ══════════════════════════════════════════════════════════════════════

    def _apply_fonts_and_borders(self, ws, max_col: int) -> None:
        """Apply fonts by row zone and remove borders from body cells."""
        for row in ws.iter_rows(
            min_row=1,
            max_row=max(self.CALC_END_ROW, ws.max_row),
            min_col=1,
            max_col=max(max_col, ws.max_column),
        ):
            for cell in row:
                r = cell.row
                col = cell.column

                # Preserve existing font color from color scheme
                existing_color = (
                    cell.font.color if cell.font and cell.font.color else None
                )

                if r in (1, 2) and col >= self.COL_SCHEDULE_START:
                    cell.font = Font(name="Arial", size=18, color=existing_color)
                elif r == 4 and col >= self.COL_SCHEDULE_START:
                    cell.font = Font(name="Arial", size=20, color=existing_color)
                elif r == 5 and col >= self.COL_SCHEDULE_START:
                    cell.font = Font(name="Arial", size=14, color=existing_color)
                elif r == 6 and col <= 5:
                    cell.font = Font(
                        name="Arial", size=11, bold=True, color=existing_color
                    )
                elif r <= 5 and col <= 5:
                    cell.font = Font(
                        name="Arial", size=16, bold=True, color=existing_color
                    )
                elif r <= 8:
                    cell.font = Font(name="Arial", size=16, color=existing_color)
                else:
                    cell.font = Font(name="Arial", size=16, color=existing_color)

                # Remove borders from body cells
                if r >= 6:
                    cell.border = NO_BORDER

    # ══════════════════════════════════════════════════════════════════════
    # Utilities
    # ══════════════════════════════════════════════════════════════════════

    def _apply_cell_color(self, cell, code: str) -> None:
        """Apply background and font color based on schedule code."""
        if not self.color_scheme:
            return

        hex_color = self.color_scheme.get_code_color(code)
        if hex_color:
            cell.fill = PatternFill(start_color=hex_color, fill_type="solid")

        font_color = self.color_scheme.get_font_color(code)
        if font_color:
            cell.font = Font(color=font_color)
        elif hex_color:
            raw = hex_color[-6:]
            if raw == "FF0000":
                cell.font = Font(color="FFFFFFFF")

    def _rotation_fill_and_font(
        self, rotation_name: str
    ) -> tuple[PatternFill | None, Font | None]:
        """Get fill and font for a rotation label."""
        color = ROTATION_COLORS.get(rotation_name)
        if not color:
            return None, None

        fill = PatternFill("solid", fgColor=color)
        font_color = ROTATION_FONT_COLORS.get(rotation_name)
        if font_color:
            font = Font(name="Arial", size=16, color=font_color)
        elif color in ("FFFF0000", "FF000000"):
            font = Font(name="Arial", size=16, color="FFFFFFFF")
        else:
            font = Font(name="Arial", size=16)
        return fill, font

    def _call_last_name_token(self, raw_name: Any) -> str:
        """Extract upper-case last name token for COUNTIF matching."""
        if not raw_name:
            return ""
        text = str(raw_name).replace("*", "").strip()
        if not text:
            return ""
        if "," in text:
            last = text.split(",", 1)[0]
        else:
            last = text.split()[-1]
        return " ".join(last.split()).upper()

    def _to_last_first(self, name: str) -> str:
        """Ensure name is in 'Last, First' format."""
        if not name:
            return ""
        if "," in name:
            return name  # Already in Last, First format
        parts = name.split()
        if len(parts) >= 2:
            return f"{parts[-1]}, {' '.join(parts[:-1])}"
        return name

    def _schedule_end_col(self, block_start: date, block_end: date) -> int:
        num_days = (block_end - block_start).days + 1
        return self.COL_SCHEDULE_START + (num_days * self.COLS_PER_DAY) - 1
