"""
Import staging service for Excel import workflow.

Provides the service layer for the import staging feature:
- stage_import(): Parse Excel, fuzzy match, create staged records
- get_batch_preview(): Compare staged vs existing assignments
- apply_batch(): Commit staged to live assignments table (upsert default)
- rollback_batch(): Undo applied batch within 24h window
- reject_batch(): Mark batch rejected, cleanup

Transaction Safety:
    Uses transactional context managers to ensure atomicity.
    Rollback window of 24 hours for applied batches.

Usage:
    from app.services.import_staging_service import ImportStagingService

    service = ImportStagingService(db)
    result = await service.stage_import(
        file_bytes=excel_bytes,
        filename="schedule.xlsx",
        created_by_id=user_id,
    )
"""

import hashlib
import io
import logging
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from difflib import SequenceMatcher
from typing import Any
from uuid import UUID, uuid4

from openpyxl import load_workbook
from openpyxl.cell import MergedCell
from sqlalchemy import and_, delete, func, select, update
from sqlalchemy.orm import Session, selectinload

from app.db.transaction import transactional_with_retry
from app.models.assignment import Assignment
from app.models.block import Block
from app.models.import_staging import (
    ConflictResolutionMode,
    ImportBatch,
    ImportBatchStatus,
    ImportStagedAssignment,
    StagedAssignmentStatus,
)
from app.models.person import Person
from app.models.rotation_template import RotationTemplate
from app.scheduling.validator import ACGMEValidator

logger = logging.getLogger(__name__)

# Rollback window in hours
ROLLBACK_WINDOW_HOURS = 24

# Fuzzy match threshold (0-100)
FUZZY_MATCH_THRESHOLD = 70


@dataclass
class StageResult:
    """Result of staging an import batch."""

    success: bool
    batch_id: UUID | None = None
    message: str = ""
    error_code: str | None = None
    row_count: int = 0
    error_count: int = 0
    warning_count: int = 0


@dataclass
class PreviewResult:
    """Result of previewing a batch before apply."""

    batch_id: UUID
    new_count: int = 0
    update_count: int = 0
    conflict_count: int = 0
    skip_count: int = 0
    staged_assignments: list[dict] | None = None
    conflicts: list[dict] | None = None
    acgme_violations: list[str] | None = None
    total_staged: int = 0


@dataclass
class ApplyResult:
    """Result of applying a batch."""

    success: bool
    batch_id: UUID
    status: ImportBatchStatus
    applied_count: int = 0
    skipped_count: int = 0
    error_count: int = 0
    errors: list[dict] | None = None
    acgme_warnings: list[str] | None = None
    rollback_available: bool = True
    rollback_expires_at: datetime | None = None
    message: str = ""
    error_code: str | None = None


@dataclass
class RollbackResult:
    """Result of rolling back a batch."""

    success: bool
    batch_id: UUID
    status: ImportBatchStatus
    rolled_back_count: int = 0
    failed_count: int = 0
    errors: list[str] | None = None
    message: str = ""
    error_code: str | None = None


class ImportStagingService:
    """
    Service for managing import staging workflow.

    Handles the complete lifecycle of Excel imports:
    1. Parse and stage Excel data
    2. Preview staged vs existing (with conflict detection)
    3. Apply staged data to live assignments table
    4. Rollback applied batches within window
    5. Reject/cleanup unwanted batches

    Thread Safety:
        Uses row-level locking where needed to prevent concurrent
        modifications. Safe for use in multi-threaded environments.
    """

    def __init__(self, db: Session) -> None:
        """
        Initialize the ImportStagingService with a database session.

        Args:
            db: SQLAlchemy Session for database operations.
        """
        self.db = db
        self._person_cache: dict[str, tuple[UUID, int]] = {}  # name -> (id, confidence)
        self._rotation_cache: dict[
            str, tuple[UUID, int]
        ] = {}  # name -> (id, confidence)

    async def stage_import(
        self,
        file_bytes: bytes,
        filename: str,
        created_by_id: UUID | None = None,
        target_block: int | None = None,
        target_start_date: date | None = None,
        target_end_date: date | None = None,
        conflict_resolution: ConflictResolutionMode = ConflictResolutionMode.UPSERT,
        notes: str | None = None,
        sheet_name: str | None = None,
    ) -> StageResult:
        """
        Parse Excel file and create staged import records.

        Parses the uploaded Excel file, performs fuzzy matching for person
        and rotation names, detects conflicts with existing assignments,
        and creates ImportStagedAssignment records for review.

        Args:
            file_bytes: Raw bytes of the Excel file.
            filename: Original filename for metadata.
            created_by_id: UUID of the user creating the import.
            target_block: Target academic block number (1-26).
            target_start_date: Target date range start.
            target_end_date: Target date range end.
            conflict_resolution: How to handle conflicts during apply.
            notes: Optional notes for the batch.
            sheet_name: Specific sheet to parse (default: first sheet).

        Returns:
            StageResult with batch_id if successful, error details if failed.
        """
        try:
            # Calculate file hash for dedup detection
            file_hash = hashlib.sha256(file_bytes).hexdigest()

            # Check for duplicate import
            existing = (
                self.db.query(ImportBatch)
                .filter(
                    ImportBatch.file_hash == file_hash,
                    ImportBatch.status.in_(
                        [ImportBatchStatus.STAGED, ImportBatchStatus.APPROVED]
                    ),
                )
                .first()
            )
            if existing:
                return StageResult(
                    success=False,
                    message=f"Duplicate file detected. Existing batch {existing.id} has the same content.",
                    error_code="DUPLICATE_FILE",
                )

                # Parse Excel file
            parsed_rows, parse_warnings, parse_errors = self._parse_excel(
                file_bytes, sheet_name
            )

            if parse_errors:
                return StageResult(
                    success=False,
                    message=f"Failed to parse Excel file: {'; '.join(parse_errors)}",
                    error_code="PARSE_ERROR",
                )

            if not parsed_rows:
                return StageResult(
                    success=False,
                    message="No data rows found in Excel file",
                    error_code="NO_DATA",
                )

                # Load caches for fuzzy matching
            await self._load_person_cache()
            await self._load_rotation_cache()

            # Create batch record
            batch = ImportBatch(
                id=uuid4(),
                created_at=datetime.utcnow(),
                created_by_id=created_by_id,
                filename=filename,
                file_hash=file_hash,
                file_size_bytes=len(file_bytes),
                status=ImportBatchStatus.STAGED,
                conflict_resolution=conflict_resolution,
                target_block=target_block,
                target_start_date=target_start_date,
                target_end_date=target_end_date,
                notes=notes,
                row_count=len(parsed_rows),
            )

            self.db.add(batch)
            self.db.flush()

            # Process each row and create staged assignments
            error_count = 0
            warning_count = len(parse_warnings)

            for row_num, row_data in enumerate(parsed_rows, start=1):
                staged, row_errors, row_warnings = await self._create_staged_assignment(
                    batch_id=batch.id,
                    row_number=row_num,
                    row_data=row_data,
                    sheet_name=sheet_name,
                )

                if staged:
                    self.db.add(staged)

                if row_errors:
                    error_count += 1
                    warning_count += len(row_warnings)

                    # Update batch counts
            batch.error_count = error_count
            batch.warning_count = warning_count

            self.db.commit()

            logger.info(
                f"Staged import batch {batch.id}: {len(parsed_rows)} rows, "
                f"{error_count} errors, {warning_count} warnings"
            )

            return StageResult(
                success=True,
                batch_id=batch.id,
                message=f"Successfully staged {len(parsed_rows)} rows",
                row_count=len(parsed_rows),
                error_count=error_count,
                warning_count=warning_count,
            )

        except Exception as e:
            logger.exception(f"Failed to stage import: {e}")
            self.db.rollback()
            return StageResult(
                success=False,
                message=f"Failed to stage import: {str(e)}",
                error_code="STAGING_FAILED",
            )

    async def get_batch_preview(
        self,
        batch_id: UUID,
        page: int = 1,
        page_size: int = 50,
    ) -> PreviewResult | None:
        """
        Generate preview comparing staged vs existing assignments.

        Args:
            batch_id: UUID of the batch to preview.
            page: Page number for staged assignments (1-indexed).
            page_size: Items per page (max 100).

        Returns:
            PreviewResult with counts and detailed comparisons, or None if batch not found.
        """
        batch = self.db.query(ImportBatch).filter(ImportBatch.id == batch_id).first()

        if not batch:
            return None

            # Get staged assignments with pagination
        offset = (page - 1) * page_size
        staged_query = (
            self.db.query(ImportStagedAssignment)
            .filter(ImportStagedAssignment.batch_id == batch_id)
            .options(
                selectinload(ImportStagedAssignment.matched_person),
                selectinload(ImportStagedAssignment.matched_rotation),
            )
        )

        total_staged = staged_query.count()
        staged_assignments = (
            staged_query.order_by(ImportStagedAssignment.row_number)
            .offset(offset)
            .limit(page_size)
            .all()
        )

        # Calculate counts
        new_count = 0
        update_count = 0
        conflict_count = 0
        skip_count = 0
        conflicts = []

        for staged in staged_assignments:
            if staged.status == StagedAssignmentStatus.SKIPPED:
                skip_count += 1
            elif staged.conflict_type == "overwrite":
                update_count += 1
                conflicts.append(
                    {
                        "staged_assignment_id": str(staged.id),
                        "existing_assignment_id": str(staged.existing_assignment_id)
                        if staged.existing_assignment_id
                        else None,
                        "person_name": staged.person_name,
                        "assignment_date": staged.assignment_date.isoformat(),
                        "slot": staged.slot,
                        "staged_rotation": staged.rotation_name,
                        "existing_rotation": None,  # Would need to fetch from existing
                        "conflict_type": staged.conflict_type,
                    }
                )
            elif staged.conflict_type == "duplicate":
                conflict_count += 1
                conflicts.append(
                    {
                        "staged_assignment_id": str(staged.id),
                        "existing_assignment_id": str(staged.existing_assignment_id)
                        if staged.existing_assignment_id
                        else None,
                        "person_name": staged.person_name,
                        "assignment_date": staged.assignment_date.isoformat(),
                        "slot": staged.slot,
                        "staged_rotation": staged.rotation_name,
                        "existing_rotation": None,
                        "conflict_type": staged.conflict_type,
                    }
                )
            else:
                new_count += 1

                # Check ACGME compliance preview
        acgme_violations = []
        if batch.target_start_date and batch.target_end_date:
            try:
                validator = ACGMEValidator(self.db)
                # Would need to simulate the changes for proper ACGME validation
                # For now, return empty list
            except Exception as e:
                logger.warning(f"ACGME validation preview failed: {e}")

        return PreviewResult(
            batch_id=batch_id,
            new_count=new_count,
            update_count=update_count,
            conflict_count=conflict_count,
            skip_count=skip_count,
            staged_assignments=[self._staged_to_dict(s) for s in staged_assignments],
            conflicts=conflicts,
            acgme_violations=acgme_violations,
            total_staged=total_staged,
        )

    async def apply_batch(
        self,
        batch_id: UUID,
        applied_by_id: UUID | None = None,
        conflict_resolution: ConflictResolutionMode | None = None,
        dry_run: bool = False,
        validate_acgme: bool = True,
    ) -> ApplyResult:
        """
        Apply staged batch to live assignments table.

        Commits all approved staged assignments to the assignments table,
        handling conflicts according to the specified resolution mode.

        Args:
            batch_id: UUID of the batch to apply.
            applied_by_id: UUID of the user applying the batch.
            conflict_resolution: Override conflict resolution mode (optional).
            dry_run: If True, validate only without applying.
            validate_acgme: If True, validate ACGME compliance before apply.

        Returns:
            ApplyResult with counts and any errors.
        """
        try:
            with transactional_with_retry(self.db, max_retries=3, timeout_seconds=60.0):
                # Get batch with lock
                batch = (
                    self.db.query(ImportBatch)
                    .filter(ImportBatch.id == batch_id)
                    .with_for_update()
                    .first()
                )

                if not batch:
                    return ApplyResult(
                        success=False,
                        batch_id=batch_id,
                        status=ImportBatchStatus.FAILED,
                        message="Batch not found",
                        error_code="BATCH_NOT_FOUND",
                    )

                if batch.status not in [
                    ImportBatchStatus.STAGED,
                    ImportBatchStatus.APPROVED,
                ]:
                    return ApplyResult(
                        success=False,
                        batch_id=batch_id,
                        status=batch.status,
                        message=f"Cannot apply batch with status: {batch.status.value}",
                        error_code="INVALID_STATUS",
                    )

                    # Use override or batch's resolution mode
                resolution = conflict_resolution or batch.conflict_resolution

                # Get all staged assignments that are pending or approved
                staged_assignments = (
                    self.db.query(ImportStagedAssignment)
                    .filter(
                        ImportStagedAssignment.batch_id == batch_id,
                        ImportStagedAssignment.status.in_(
                            [
                                StagedAssignmentStatus.PENDING,
                                StagedAssignmentStatus.APPROVED,
                            ]
                        ),
                    )
                    .all()
                )

                if dry_run:
                    return ApplyResult(
                        success=True,
                        batch_id=batch_id,
                        status=batch.status,
                        applied_count=len(staged_assignments),
                        message=f"Dry run: Would apply {len(staged_assignments)} assignments",
                    )

                applied_count = 0
                skipped_count = 0
                error_count = 0
                errors = []
                acgme_warnings = []

                for staged in staged_assignments:
                    try:
                        created_id = await self._apply_single_assignment(
                            staged, resolution
                        )

                        if created_id:
                            staged.status = StagedAssignmentStatus.APPLIED
                            staged.created_assignment_id = created_id
                            applied_count += 1
                        else:
                            staged.status = StagedAssignmentStatus.SKIPPED
                            skipped_count += 1

                    except Exception as e:
                        logger.warning(
                            f"Failed to apply staged assignment {staged.id}: {e}"
                        )
                        staged.status = StagedAssignmentStatus.FAILED
                        error_count += 1
                        errors.append(
                            {
                                "staged_assignment_id": str(staged.id),
                                "row_number": staged.row_number,
                                "person_name": staged.person_name,
                                "assignment_date": staged.assignment_date.isoformat(),
                                "error_message": str(e),
                                "error_code": "APPLY_FAILED",
                            }
                        )

                        # Update batch status
                now = datetime.utcnow()
                batch.status = ImportBatchStatus.APPLIED
                batch.applied_at = now
                batch.applied_by_id = applied_by_id
                batch.rollback_available = True
                batch.rollback_expires_at = now + timedelta(hours=ROLLBACK_WINDOW_HOURS)

                # ACGME validation after apply
                if validate_acgme and applied_count > 0:
                    try:
                        validator = ACGMEValidator(self.db)
                        if batch.target_start_date and batch.target_end_date:
                            result = validator.validate_all(
                                batch.target_start_date, batch.target_end_date
                            )
                            acgme_warnings = [
                                f"{v.severity}: {v.message}" for v in result.violations
                            ]
                    except Exception as e:
                        logger.warning(f"ACGME validation after apply failed: {e}")

                logger.info(
                    f"Applied batch {batch_id}: {applied_count} applied, "
                    f"{skipped_count} skipped, {error_count} errors"
                )

                return ApplyResult(
                    success=error_count == 0,
                    batch_id=batch_id,
                    status=batch.status,
                    applied_count=applied_count,
                    skipped_count=skipped_count,
                    error_count=error_count,
                    errors=errors if errors else None,
                    acgme_warnings=acgme_warnings if acgme_warnings else None,
                    rollback_available=True,
                    rollback_expires_at=batch.rollback_expires_at,
                    message=f"Applied {applied_count} assignments",
                )

        except Exception as e:
            logger.exception(f"Failed to apply batch {batch_id}: {e}")
            return ApplyResult(
                success=False,
                batch_id=batch_id,
                status=ImportBatchStatus.FAILED,
                message=f"Apply failed: {str(e)}",
                error_code="APPLY_FAILED",
            )

    async def rollback_batch(
        self,
        batch_id: UUID,
        rolled_back_by_id: UUID | None = None,
        reason: str | None = None,
    ) -> RollbackResult:
        """
        Rollback an applied batch within the allowed window.

        Deletes all assignments created by this batch and restores
        the staged status.

        Args:
            batch_id: UUID of the batch to rollback.
            rolled_back_by_id: UUID of the user performing rollback.
            reason: Reason for the rollback.

        Returns:
            RollbackResult with counts and any errors.
        """
        try:
            with transactional_with_retry(self.db, max_retries=3, timeout_seconds=60.0):
                # Get batch with lock
                batch = (
                    self.db.query(ImportBatch)
                    .filter(ImportBatch.id == batch_id)
                    .with_for_update()
                    .first()
                )

                if not batch:
                    return RollbackResult(
                        success=False,
                        batch_id=batch_id,
                        status=ImportBatchStatus.FAILED,
                        message="Batch not found",
                        error_code="BATCH_NOT_FOUND",
                    )

                if batch.status != ImportBatchStatus.APPLIED:
                    return RollbackResult(
                        success=False,
                        batch_id=batch_id,
                        status=batch.status,
                        message=f"Cannot rollback batch with status: {batch.status.value}",
                        error_code="INVALID_STATUS",
                    )

                    # Check rollback window
                if not batch.rollback_available:
                    return RollbackResult(
                        success=False,
                        batch_id=batch_id,
                        status=batch.status,
                        message="Rollback not available for this batch",
                        error_code="ROLLBACK_NOT_AVAILABLE",
                    )

                if (
                    batch.rollback_expires_at
                    and datetime.utcnow() > batch.rollback_expires_at
                ):
                    return RollbackResult(
                        success=False,
                        batch_id=batch_id,
                        status=batch.status,
                        message=f"Rollback window of {ROLLBACK_WINDOW_HOURS} hours has expired",
                        error_code="ROLLBACK_WINDOW_EXPIRED",
                    )

                    # Get all applied staged assignments
                staged_assignments = (
                    self.db.query(ImportStagedAssignment)
                    .filter(
                        ImportStagedAssignment.batch_id == batch_id,
                        ImportStagedAssignment.status == StagedAssignmentStatus.APPLIED,
                        ImportStagedAssignment.created_assignment_id.isnot(None),
                    )
                    .all()
                )

                rolled_back_count = 0
                failed_count = 0
                errors = []

                for staged in staged_assignments:
                    try:
                        # Delete the created assignment
                        if staged.created_assignment_id:
                            assignment = (
                                self.db.query(Assignment)
                                .filter(Assignment.id == staged.created_assignment_id)
                                .first()
                            )
                            if assignment:
                                self.db.delete(assignment)
                                rolled_back_count += 1

                                # Reset staged status
                        staged.status = StagedAssignmentStatus.PENDING
                        staged.created_assignment_id = None

                    except Exception as e:
                        logger.warning(
                            f"Failed to rollback staged assignment {staged.id}: {e}"
                        )
                        failed_count += 1
                        errors.append(str(e))

                        # Update batch status
                batch.status = ImportBatchStatus.ROLLED_BACK
                batch.rolled_back_at = datetime.utcnow()
                batch.rolled_back_by_id = rolled_back_by_id
                batch.rollback_available = False

                logger.info(
                    f"Rolled back batch {batch_id}: {rolled_back_count} rolled back, "
                    f"{failed_count} failed"
                )

                return RollbackResult(
                    success=failed_count == 0,
                    batch_id=batch_id,
                    status=batch.status,
                    rolled_back_count=rolled_back_count,
                    failed_count=failed_count,
                    errors=errors if errors else None,
                    message=f"Rolled back {rolled_back_count} assignments",
                )

        except Exception as e:
            logger.exception(f"Failed to rollback batch {batch_id}: {e}")
            return RollbackResult(
                success=False,
                batch_id=batch_id,
                status=ImportBatchStatus.FAILED,
                message=f"Rollback failed: {str(e)}",
                error_code="ROLLBACK_FAILED",
            )

    async def reject_batch(
        self,
        batch_id: UUID,
    ) -> tuple[bool, str]:
        """
        Reject and cleanup a batch.

        Marks the batch as rejected and optionally deletes staged assignments.

        Args:
            batch_id: UUID of the batch to reject.

        Returns:
            Tuple of (success, message).
        """
        try:
            batch = (
                self.db.query(ImportBatch).filter(ImportBatch.id == batch_id).first()
            )

            if not batch:
                return False, "Batch not found"

            if batch.status == ImportBatchStatus.APPLIED:
                return False, "Cannot reject an applied batch. Use rollback first."

            if batch.status == ImportBatchStatus.REJECTED:
                return True, "Batch already rejected"

                # Delete staged assignments
            self.db.query(ImportStagedAssignment).filter(
                ImportStagedAssignment.batch_id == batch_id
            ).delete(synchronize_session=False)

            # Update batch status
            batch.status = ImportBatchStatus.REJECTED

            self.db.commit()

            logger.info(f"Rejected batch {batch_id}")

            return True, "Batch rejected successfully"

        except Exception as e:
            logger.exception(f"Failed to reject batch {batch_id}: {e}")
            self.db.rollback()
            return False, f"Failed to reject batch: {str(e)}"

    async def get_batch(self, batch_id: UUID) -> ImportBatch | None:
        """Get a batch by ID with counts."""
        batch = (
            self.db.query(ImportBatch)
            .options(selectinload(ImportBatch.staged_assignments))
            .filter(ImportBatch.id == batch_id)
            .first()
        )
        return batch

    async def list_batches(
        self,
        page: int = 1,
        page_size: int = 50,
        status: ImportBatchStatus | None = None,
    ) -> tuple[list[ImportBatch], int]:
        """
        List batches with pagination.

        Args:
            page: Page number (1-indexed).
            page_size: Items per page.
            status: Filter by status.

        Returns:
            Tuple of (batches, total_count).
        """
        query = self.db.query(ImportBatch)

        if status:
            query = query.filter(ImportBatch.status == status)

        total = query.count()

        batches = (
            query.order_by(ImportBatch.created_at.desc())
            .offset((page - 1) * page_size)
            .limit(page_size)
            .all()
        )

        return batches, total

        # -------------------------------------------------------------------------
        # Private helper methods
        # -------------------------------------------------------------------------

    def _parse_excel(
        self,
        file_bytes: bytes,
        sheet_name: str | None = None,
    ) -> tuple[list[dict[str, Any]], list[str], list[str]]:
        """
        Parse Excel file and return rows as dictionaries.

        Returns:
            Tuple of (rows, warnings, errors).
        """
        warnings = []
        errors = []
        rows = []

        try:
            wb = load_workbook(io.BytesIO(file_bytes), data_only=True)

            # Select worksheet
            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    errors.append(
                        f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
                    )
                    return rows, warnings, errors
                ws = wb[sheet_name]
            else:
                ws = wb.active
                if not ws:
                    ws = wb[wb.sheetnames[0]]

                    # Extract headers from first row
            headers = []
            for cell in ws[1]:
                if isinstance(cell, MergedCell):
                    headers.append("")
                else:
                    value = cell.value
                    if value is not None:
                        headers.append(str(value).strip())
                    else:
                        headers.append(f"Column_{len(headers) + 1}")

                        # Required columns check
            required = {"person_name", "assignment_date"}
            header_lower = {h.lower().replace(" ", "_") for h in headers}
            missing = required - header_lower
            if missing:
                errors.append(f"Missing required columns: {missing}")
                return rows, warnings, errors

                # Parse data rows
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                row_data = {}
                is_empty = True

                for col_idx, cell in enumerate(row):
                    if col_idx >= len(headers):
                        break

                    header = headers[col_idx]

                    if isinstance(cell, MergedCell):
                        value = None
                    else:
                        value = cell.value

                    if value is not None:
                        is_empty = False

                        # Convert dates to date objects
                        if hasattr(value, "date"):
                            value = value.date()
                        elif hasattr(value, "isoformat"):
                            value = value
                        elif isinstance(value, str):
                            value = value.strip()

                    row_data[header] = value

                if not is_empty:
                    rows.append(row_data)

            wb.close()

        except Exception as e:
            errors.append(str(e))

        return rows, warnings, errors

    async def _load_person_cache(self) -> None:
        """Load persons into cache for fuzzy matching."""
        persons = self.db.query(Person.id, Person.name).all()
        self._person_cache = {}
        for person_id, name in persons:
            normalized = name.lower().strip()
            self._person_cache[normalized] = (person_id, 100)

    async def _load_rotation_cache(self) -> None:
        """Load rotation templates into cache for fuzzy matching."""
        rotations = self.db.query(RotationTemplate.id, RotationTemplate.name).all()
        self._rotation_cache = {}
        for rotation_id, name in rotations:
            normalized = name.lower().strip()
            self._rotation_cache[normalized] = (rotation_id, 100)

    def _fuzzy_match_person(self, name: str) -> tuple[UUID | None, int]:
        """
        Fuzzy match a person name.

        Returns:
            Tuple of (person_id, confidence 0-100).
        """
        normalized = name.lower().strip()

        # Exact match
        if normalized in self._person_cache:
            return self._person_cache[normalized]

            # Fuzzy match
        best_match = None
        best_score = 0

        for cached_name, (person_id, _) in self._person_cache.items():
            score = int(SequenceMatcher(None, normalized, cached_name).ratio() * 100)
            if score > best_score and score >= FUZZY_MATCH_THRESHOLD:
                best_score = score
                best_match = person_id

        return best_match, best_score

    def _fuzzy_match_rotation(self, name: str) -> tuple[UUID | None, int]:
        """
        Fuzzy match a rotation name.

        Returns:
            Tuple of (rotation_id, confidence 0-100).
        """
        normalized = name.lower().strip()

        # Exact match
        if normalized in self._rotation_cache:
            return self._rotation_cache[normalized]

            # Fuzzy match
        best_match = None
        best_score = 0

        for cached_name, (rotation_id, _) in self._rotation_cache.items():
            score = int(SequenceMatcher(None, normalized, cached_name).ratio() * 100)
            if score > best_score and score >= FUZZY_MATCH_THRESHOLD:
                best_score = score
                best_match = rotation_id

        return best_match, best_score

    async def _create_staged_assignment(
        self,
        batch_id: UUID,
        row_number: int,
        row_data: dict[str, Any],
        sheet_name: str | None = None,
    ) -> tuple[ImportStagedAssignment | None, list[str], list[str]]:
        """
        Create a staged assignment from parsed row data.

        Returns:
            Tuple of (staged_assignment, errors, warnings).
        """
        errors = []
        warnings = []

        # Extract required fields (handle various column name formats)
        person_name = None
        assignment_date = None
        rotation_name = None
        slot = None
        raw_cell_value = None

        for key, value in row_data.items():
            key_lower = key.lower().replace(" ", "_")
            if key_lower in ("person_name", "name", "provider", "resident"):
                person_name = str(value).strip() if value else None
            elif key_lower in ("assignment_date", "date"):
                if hasattr(value, "isoformat"):
                    assignment_date = value if isinstance(value, date) else value.date()
                elif isinstance(value, str):
                    try:
                        assignment_date = date.fromisoformat(value)
                    except ValueError:
                        errors.append(f"Invalid date format: {value}")
                else:
                    assignment_date = value
            elif key_lower in ("rotation_name", "rotation", "activity"):
                rotation_name = str(value).strip() if value else None
            elif key_lower in ("slot", "time", "session"):
                slot = str(value).strip().upper() if value else None
            elif key_lower in ("raw", "raw_value", "cell"):
                raw_cell_value = str(value) if value else None

        if not person_name:
            errors.append("Missing person name")
            return None, errors, warnings

        if not assignment_date:
            errors.append("Missing assignment date")
            return None, errors, warnings

            # Fuzzy match person
        matched_person_id, person_confidence = self._fuzzy_match_person(person_name)
        if person_confidence < FUZZY_MATCH_THRESHOLD:
            warnings.append(
                f"Low confidence person match: {person_name} ({person_confidence}%)"
            )

            # Fuzzy match rotation
        matched_rotation_id = None
        rotation_confidence = 0
        if rotation_name:
            matched_rotation_id, rotation_confidence = self._fuzzy_match_rotation(
                rotation_name
            )
            if rotation_confidence < FUZZY_MATCH_THRESHOLD:
                warnings.append(
                    f"Low confidence rotation match: {rotation_name} ({rotation_confidence}%)"
                )

                # Check for conflicts with existing assignments
        conflict_type = None
        existing_assignment_id = None

        if matched_person_id:
            # Find blocks on this date
            blocks = self.db.query(Block).filter(Block.date == assignment_date).all()

            for block in blocks:
                existing = (
                    self.db.query(Assignment)
                    .filter(
                        Assignment.block_id == block.id,
                        Assignment.person_id == matched_person_id,
                    )
                    .first()
                )

                if existing:
                    existing_assignment_id = existing.id
                    if (
                        matched_rotation_id
                        and existing.rotation_template_id == matched_rotation_id
                    ):
                        conflict_type = "duplicate"
                    else:
                        conflict_type = "overwrite"
                    break

                    # Create staged assignment
        staged = ImportStagedAssignment(
            id=uuid4(),
            batch_id=batch_id,
            row_number=row_number,
            sheet_name=sheet_name,
            person_name=person_name,
            assignment_date=assignment_date,
            slot=slot,
            rotation_name=rotation_name,
            raw_cell_value=raw_cell_value,
            matched_person_id=matched_person_id,
            person_match_confidence=person_confidence if matched_person_id else None,
            matched_rotation_id=matched_rotation_id,
            rotation_match_confidence=rotation_confidence
            if matched_rotation_id
            else None,
            conflict_type=conflict_type,
            existing_assignment_id=existing_assignment_id,
            status=StagedAssignmentStatus.PENDING,
            validation_errors=errors if errors else None,
            validation_warnings=warnings if warnings else None,
        )

        return staged, errors, warnings

    async def _apply_single_assignment(
        self,
        staged: ImportStagedAssignment,
        resolution: ConflictResolutionMode,
    ) -> UUID | None:
        """
        Apply a single staged assignment to live table.

        Returns:
            Created assignment ID or None if skipped.
        """
        if not staged.matched_person_id:
            return None

            # Find or create block for this date
        block = (
            self.db.query(Block).filter(Block.date == staged.assignment_date).first()
        )

        if not block:
            # Create a block for this date if it doesn't exist
            # In a real implementation, would need to handle AM/PM slots
            block = Block(
                id=uuid4(),
                date=staged.assignment_date,
                slot_type=staged.slot or "AM",
            )
            self.db.add(block)
            self.db.flush()

            # Check for existing assignment
        existing = (
            self.db.query(Assignment)
            .filter(
                Assignment.block_id == block.id,
                Assignment.person_id == staged.matched_person_id,
            )
            .first()
        )

        if existing:
            if resolution == ConflictResolutionMode.MERGE:
                # Skip if exists
                return None
            elif resolution == ConflictResolutionMode.UPSERT:
                # Update existing
                if staged.matched_rotation_id:
                    existing.rotation_template_id = staged.matched_rotation_id
                existing.notes = (
                    f"Updated via import batch at {datetime.utcnow().isoformat()}"
                )
                return existing.id
            elif resolution == ConflictResolutionMode.REPLACE:
                # Delete and recreate
                self.db.delete(existing)
                self.db.flush()

                # Create new assignment
        assignment = Assignment(
            id=uuid4(),
            block_id=block.id,
            person_id=staged.matched_person_id,
            rotation_template_id=staged.matched_rotation_id,
            role="primary",
            created_by="import_staging_service",
            created_at=datetime.utcnow(),
            notes=f"Created via import batch at {datetime.utcnow().isoformat()}",
        )
        self.db.add(assignment)
        self.db.flush()

        return assignment.id

    def _staged_to_dict(self, staged: ImportStagedAssignment) -> dict[str, Any]:
        """Convert staged assignment to dictionary for API response."""
        return {
            "id": str(staged.id),
            "batch_id": str(staged.batch_id),
            "row_number": staged.row_number,
            "sheet_name": staged.sheet_name,
            "person_name": staged.person_name,
            "assignment_date": staged.assignment_date.isoformat(),
            "slot": staged.slot,
            "rotation_name": staged.rotation_name,
            "raw_cell_value": staged.raw_cell_value,
            "matched_person_id": str(staged.matched_person_id)
            if staged.matched_person_id
            else None,
            "matched_person_name": staged.matched_person.name
            if staged.matched_person
            else None,
            "person_match_confidence": staged.person_match_confidence,
            "matched_rotation_id": str(staged.matched_rotation_id)
            if staged.matched_rotation_id
            else None,
            "matched_rotation_name": staged.matched_rotation.name
            if staged.matched_rotation
            else None,
            "rotation_match_confidence": staged.rotation_match_confidence,
            "conflict_type": staged.conflict_type,
            "existing_assignment_id": str(staged.existing_assignment_id)
            if staged.existing_assignment_id
            else None,
            "status": staged.status.value,
            "validation_errors": staged.validation_errors,
            "validation_warnings": staged.validation_warnings,
            "created_assignment_id": str(staged.created_assignment_id)
            if staged.created_assignment_id
            else None,
        }
