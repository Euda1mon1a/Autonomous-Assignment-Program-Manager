"""Half-day schedule Excel staging and diff service."""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any
from uuid import UUID, uuid4

from openpyxl import load_workbook
from openpyxl.cell import MergedCell
from sqlalchemy import and_, select
from sqlalchemy.orm import Session, selectinload

from app.core.logging import get_logger
from app.db.transaction import transactional
from app.models.activity import Activity, ActivityCategory
from app.models.half_day_assignment import HalfDayAssignment
from app.models.import_staging import (
    ImportBatch,
    ImportBatchStatus,
    ImportStagedAssignment,
    StagedAssignmentStatus,
)
from app.models.person import Person
from app.models.schedule_draft import (
    DraftAssignmentChangeType,
    DraftSourceType,
    ScheduleDraft,
)
from app.models.schedule_override import ScheduleOverride
from app.schemas.half_day_import import (
    HalfDayDiffEntry,
    HalfDayDiffMetrics,
    HalfDayDiffType,
)
from app.services.schedule_draft_service import ScheduleDraftService
from app.utils.academic_blocks import get_block_dates, get_block_number_for_date

logger = get_logger(__name__)

DATA_START_ROW = 9
DATA_END_ROW = 69
NAME_COL = 5  # Column E
DATE_ROW = 3  # Row with actual date values
SCHEDULE_START_COL = 6  # Column F

# Normalize Excel cell values to canonical activity codes
SPECIAL_CODE_MAP = {
    "LV": {"AM": "LV-AM", "PM": "LV-PM"},
    "LVAM": {"AM": "LV-AM", "PM": "LV-AM"},
    "LVPM": {"AM": "LV-PM", "PM": "LV-PM"},
    "PC": {"AM": "PC", "PM": "PC"},
    "OFF": {"AM": "OFF", "PM": "OFF"},
    "W": {"AM": "W", "PM": "W"},
}


@dataclass
class ParsedSlot:
    person_name: str
    person_id: UUID | None
    assignment_date: date
    time_of_day: str
    raw_value: str | None
    excel_code: str | None
    warnings: list[str]
    row_number: int | None = None


@dataclass
class DraftResult:
    """Result of creating a schedule draft from staged diffs."""

    success: bool
    batch_id: UUID
    draft_id: UUID | None = None
    message: str = ""
    error_code: str | None = None
    total_selected: int = 0
    added: int = 0
    modified: int = 0
    removed: int = 0
    skipped: int = 0
    failed: int = 0
    failed_ids: list[UUID] = field(default_factory=list)


class HalfDayImportService:
    """Stage Excel schedule and compute diffs vs current schedule."""

    def __init__(self, db: Session) -> None:
        self.db = db
        self._activity_map: dict[str, str] = {}
        self._person_map: dict[str, UUID] = {}

    def stage_block_template2(
        self,
        file_bytes: bytes,
        block_number: int,
        academic_year: int,
        created_by_id: UUID | None = None,
        notes: str | None = None,
        filename: str | None = None,
    ) -> tuple[ImportBatch, HalfDayDiffMetrics, list[str]]:
        """Parse Block Template2 Excel, stage diffs, and return metrics."""
        warnings: list[str] = []

        block_dates = get_block_dates(block_number, academic_year)
        start_date = block_dates.start_date
        end_date = block_dates.end_date

        parsed_slots, parse_warnings = self._parse_block_template2(
            file_bytes, start_date, end_date
        )
        warnings.extend(parse_warnings)

        if not parsed_slots:
            raise ValueError("No schedule slots parsed from Excel")

        self._load_activity_map()
        self._load_person_map()

        # Attach person IDs + normalize codes
        normalized_slots: list[ParsedSlot] = []
        for slot in parsed_slots:
            slot_warnings: list[str] = []
            normalized_name = self._normalize_person_name(slot.person_name)
            person_id = self._person_map.get(normalized_name)
            if not person_id:
                slot_warnings.append("Person not found in database")

            excel_code = self._normalize_excel_code(slot.raw_value, slot.time_of_day)
            if excel_code and excel_code not in self._activity_map:
                slot_warnings.append(f"Unknown activity code '{excel_code}'")
            normalized_slots.append(
                ParsedSlot(
                    person_name=slot.person_name,
                    person_id=person_id,
                    assignment_date=slot.assignment_date,
                    time_of_day=slot.time_of_day,
                    raw_value=slot.raw_value,
                    excel_code=excel_code,
                    warnings=slot_warnings,
                    row_number=slot.row_number,
                )
            )

            # Load current schedule (include overrides)
        current_assignments = self._get_current_schedule(
            start_date=start_date,
            end_date=end_date,
            include_overrides=True,
        )
        current_map: dict[tuple[UUID, date, str], HalfDayAssignment] = {}
        for assignment in current_assignments:
            if assignment.person_id:
                current_map[
                    (assignment.person_id, assignment.date, assignment.time_of_day)
                ] = assignment

                # Compute diffs and stage
        diff_entries: list[HalfDayDiffEntry] = []
        metrics = HalfDayDiffMetrics()
        metrics.total_slots = sum(1 for slot in normalized_slots if slot.person_id)

        batch = ImportBatch(
            id=uuid4(),
            created_at=datetime.utcnow(),
            created_by_id=created_by_id,
            filename=filename or "block_template2.xlsx",
            status=ImportBatchStatus.STAGED,
            target_block=block_number,
            target_start_date=start_date,
            target_end_date=end_date,
            notes=notes or "half-day schedule import (Block Template2)",
            row_count=len(parsed_slots),
        )
        self.db.add(batch)
        self.db.flush()

        for slot in normalized_slots:
            if not slot.person_id:
                continue

            key = (slot.person_id, slot.assignment_date, slot.time_of_day)
            existing = current_map.get(key)
            existing_code = (
                self._normalize_activity_code(existing.activity) if existing else None
            )
            excel_code = slot.excel_code

            diff_type = None
            if excel_code is None and existing_code is None:
                continue
            if excel_code is None and existing_code is not None:
                diff_type = HalfDayDiffType.REMOVED
            elif excel_code is not None and existing_code is None:
                diff_type = HalfDayDiffType.ADDED
            elif excel_code != existing_code:
                diff_type = HalfDayDiffType.MODIFIED

            if not diff_type:
                continue

            diff_entries.append(
                HalfDayDiffEntry(
                    person_id=slot.person_id,
                    person_name=slot.person_name,
                    assignment_date=slot.assignment_date,
                    time_of_day=slot.time_of_day,
                    diff_type=diff_type,
                    excel_value=excel_code,
                    current_value=existing_code,
                    warnings=slot.warnings,
                )
            )

            # Track metrics
            metrics.changed_slots += 1
            if diff_type == HalfDayDiffType.ADDED:
                metrics.added += 1
            elif diff_type == HalfDayDiffType.REMOVED:
                metrics.removed += 1
            else:
                metrics.modified += 1

            if excel_code:
                metrics.by_activity[excel_code] = (
                    metrics.by_activity.get(excel_code, 0) + 1
                )
            elif existing_code:
                metrics.by_activity[existing_code] = (
                    metrics.by_activity.get(existing_code, 0) + 1
                )

                # Stage diff row
            staged = ImportStagedAssignment(
                id=uuid4(),
                batch_id=batch.id,
                row_number=slot.row_number,
                person_name=slot.person_name,
                assignment_date=slot.assignment_date,
                slot=slot.time_of_day,
                rotation_name=excel_code,
                raw_cell_value=slot.raw_value,
                matched_person_id=slot.person_id,
                person_match_confidence=100,
                conflict_type=diff_type.value,
                existing_assignment_id=existing.id if existing else None,
                validation_warnings={
                    "existing_code": existing_code,
                    "excel_code": excel_code,
                    "diff_type": diff_type.value,
                },
            )
            self.db.add(staged)

            # Finalize metrics
        if metrics.total_slots > 0:
            metrics.percent_changed = round(
                (metrics.changed_slots / metrics.total_slots) * 100, 2
            )
        metrics.manual_half_days = metrics.changed_slots
        metrics.manual_hours = float(metrics.manual_half_days * 4)

        self.db.commit()
        return batch, metrics, warnings

    def preview_batch(
        self,
        batch_id: UUID,
        page: int = 1,
        page_size: int = 50,
    ) -> tuple[HalfDayDiffMetrics, list[HalfDayDiffEntry], int]:
        """Preview a staged half-day batch."""
        batch = self.db.query(ImportBatch).filter(ImportBatch.id == batch_id).first()
        if not batch:
            raise ValueError("Import batch not found")

        if page < 1:
            page = 1
        if page_size < 1:
            page_size = 50

        total_query = self.db.query(ImportStagedAssignment).filter(
            ImportStagedAssignment.batch_id == batch_id
        )
        total_diffs = total_query.count()

        staged_rows = (
            total_query.order_by(ImportStagedAssignment.row_number)
            .offset((page - 1) * page_size)
            .limit(page_size)
            .all()
        )

        metrics = HalfDayDiffMetrics()
        metrics.total_slots = int(batch.row_count or 0)
        for row in total_query.all():
            metrics.changed_slots += 1
            if row.conflict_type == HalfDayDiffType.ADDED.value:
                metrics.added += 1
            elif row.conflict_type == HalfDayDiffType.REMOVED.value:
                metrics.removed += 1
            else:
                metrics.modified += 1

            code = row.rotation_name or None
            if code:
                metrics.by_activity[code] = metrics.by_activity.get(code, 0) + 1

        if metrics.total_slots > 0:
            metrics.percent_changed = round(
                (metrics.changed_slots / metrics.total_slots) * 100, 2
            )
        metrics.manual_half_days = metrics.changed_slots
        metrics.manual_hours = float(metrics.manual_half_days * 4)

        diffs: list[HalfDayDiffEntry] = []
        for row in staged_rows:
            current_value = None
            warnings = []
            if row.validation_warnings:
                current_value = row.validation_warnings.get("existing_code")
            diffs.append(
                HalfDayDiffEntry(
                    person_id=row.matched_person_id,
                    person_name=row.person_name,
                    assignment_date=row.assignment_date,
                    time_of_day=row.slot or "AM",
                    diff_type=HalfDayDiffType(row.conflict_type),
                    excel_value=row.rotation_name,
                    current_value=current_value,
                    warnings=warnings,
                )
            )

        return metrics, diffs, total_diffs

    def create_draft_from_batch(
        self,
        batch_id: UUID,
        staged_ids: list[UUID] | None = None,
        created_by_id: UUID | None = None,
        notes: str | None = None,
    ) -> DraftResult:
        """Create a schedule draft from staged half-day diffs."""
        batch = self.db.query(ImportBatch).filter(ImportBatch.id == batch_id).first()
        if not batch:
            return DraftResult(
                success=False,
                batch_id=batch_id,
                message="Import batch not found",
                error_code="BATCH_NOT_FOUND",
            )
        if batch.status not in (ImportBatchStatus.STAGED, ImportBatchStatus.APPROVED):
            return DraftResult(
                success=False,
                batch_id=batch_id,
                message="Import batch status must be staged or approved",
                error_code="INVALID_BATCH_STATUS",
            )

        staged_query = self.db.query(ImportStagedAssignment).filter(
            ImportStagedAssignment.batch_id == batch_id
        )

        staged_id_set = set(staged_ids or [])
        if staged_ids:
            selected_rows = staged_query.filter(
                ImportStagedAssignment.id.in_(staged_id_set)
            ).all()
            skipped_rows = staged_query.filter(
                ~ImportStagedAssignment.id.in_(staged_id_set)
            ).all()
        else:
            selected_rows = staged_query.all()
            skipped_rows = []

        if not selected_rows:
            return DraftResult(
                success=False,
                batch_id=batch_id,
                message="No staged diffs selected",
                error_code="NO_SELECTION",
            )

        eligible_rows = [row for row in selected_rows if row.matched_person_id]
        if not eligible_rows:
            return DraftResult(
                success=False,
                batch_id=batch_id,
                message="Selected diffs have no matched people",
                error_code="NO_MATCHED_PEOPLE",
            )

        # Determine draft scope from batch or selected rows
        min_date = min(row.assignment_date for row in eligible_rows)
        max_date = max(row.assignment_date for row in eligible_rows)
        start_date = batch.target_start_date or min_date
        end_date = batch.target_end_date or max_date
        block_number = batch.target_block
        if block_number is None:
            block_number, _ = get_block_number_for_date(start_date)

        draft_service = ScheduleDraftService(self.db)
        draft_result = None
        added = modified = removed = failed = 0
        failed_ids: list[UUID] = []
        total_selected = len(eligible_rows)
        skipped = len(skipped_rows)
        error_code = None

        try:
            with transactional(self.db):
                draft_result = draft_service.create_draft_sync(
                    source_type=DraftSourceType.IMPORT,
                    start_date=start_date,
                    end_date=end_date,
                    block_number=block_number,
                    created_by_id=created_by_id,
                    notes=notes or f"Excel import batch {batch_id}",
                    commit=False,
                )

                if not draft_result.success or not draft_result.draft_id:
                    error_code = draft_result.error_code or "CREATE_FAILED"
                    raise ValueError(draft_result.message or "Failed to create draft")

                for row in eligible_rows:
                    diff_type = row.conflict_type or ""
                    if diff_type == HalfDayDiffType.ADDED.value:
                        change_type = DraftAssignmentChangeType.ADD
                        added += 1
                    elif diff_type == HalfDayDiffType.REMOVED.value:
                        change_type = DraftAssignmentChangeType.DELETE
                        removed += 1
                    elif diff_type == HalfDayDiffType.MODIFIED.value:
                        change_type = DraftAssignmentChangeType.MODIFY
                        modified += 1
                    else:
                        failed += 1
                        failed_ids.append(row.id)
                        continue

                    activity_code = (
                        row.rotation_name
                        if change_type != DraftAssignmentChangeType.DELETE
                        else None
                    )
                    existing_id = (
                        row.existing_assignment_id
                        if change_type != DraftAssignmentChangeType.ADD
                        else None
                    )
                    if change_type != DraftAssignmentChangeType.ADD and not existing_id:
                        failed += 1
                        failed_ids.append(row.id)
                        continue
                    draft_assignment_id = draft_service.add_assignment_to_draft_sync(
                        draft_id=draft_result.draft_id,
                        person_id=row.matched_person_id,
                        assignment_date=row.assignment_date,
                        time_of_day=row.slot,  # None→"ALL" handled by draft service
                        activity_code=activity_code,
                        rotation_id=row.matched_rotation_id,
                        change_type=change_type,
                        existing_assignment_id=existing_id,
                        commit=False,
                    )
                    if not draft_assignment_id:
                        failed += 1
                        failed_ids.append(row.id)
                        continue
                    row.status = StagedAssignmentStatus.APPROVED

                if failed_ids:
                    error_code = "ROW_FAILURE"
                    raise ValueError(
                        f"Draft creation failed for {len(failed_ids)} staged rows"
                    )

            message = f"Draft {draft_result.draft_id} created from {total_selected} staged diffs"
            return DraftResult(
                success=True,
                batch_id=batch_id,
                draft_id=draft_result.draft_id,
                message=message,
                total_selected=total_selected,
                added=added,
                modified=modified,
                removed=removed,
                skipped=skipped,
                failed=failed,
                failed_ids=failed_ids,
            )
        except Exception as exc:
            error_message = str(exc) or "Draft creation failed"
            return DraftResult(
                success=False,
                batch_id=batch_id,
                draft_id=None,
                message=error_message,
                error_code=error_code or "DRAFT_CREATE_FAILED",
                total_selected=total_selected,
                added=added,
                modified=modified,
                removed=removed,
                skipped=skipped,
                failed=failed or len(failed_ids),
                failed_ids=failed_ids,
            )

    def _get_current_schedule(
        self,
        start_date: date,
        end_date: date,
        include_overrides: bool = True,
    ) -> list[HalfDayAssignment]:
        """Load current half-day assignments for a date range."""
        assignments = (
            self.db.query(HalfDayAssignment)
            .options(
                selectinload(HalfDayAssignment.person),
                selectinload(HalfDayAssignment.activity),
            )
            .filter(
                HalfDayAssignment.date >= start_date,
                HalfDayAssignment.date <= end_date,
            )
            .order_by(HalfDayAssignment.date, HalfDayAssignment.time_of_day)
            .all()
        )

        if include_overrides:
            assignments = self._apply_overrides(assignments, start_date, end_date)

        return assignments

    def _apply_overrides(
        self,
        assignments: list[HalfDayAssignment],
        start_date: date,
        end_date: date,
    ) -> list[HalfDayAssignment]:
        if not assignments:
            return assignments

        overrides = (
            self.db.query(ScheduleOverride)
            .filter(
                ScheduleOverride.is_active.is_(True),
                ScheduleOverride.effective_date >= start_date,
                ScheduleOverride.effective_date <= end_date,
            )
            .all()
        )
        if not overrides:
            return assignments

        replacement_ids = {
            o.replacement_person_id
            for o in overrides
            if o.override_type == "coverage" and o.replacement_person_id
        }
        replacement_people = {}
        if replacement_ids:
            people = self.db.query(Person).filter(Person.id.in_(replacement_ids)).all()
            replacement_people = {p.id: p for p in people}

        overrides_by_assignment = {o.half_day_assignment_id: o for o in overrides}
        effective: list[HalfDayAssignment] = []

        for assignment in assignments:
            override = overrides_by_assignment.get(assignment.id)
            if not override:
                effective.append(assignment)
                continue

            if override.override_type == "cancellation":
                continue

            if override.override_type == "gap":
                clone = HalfDayAssignment(
                    id=assignment.id,
                    person_id=assignment.person_id,
                    date=assignment.date,
                    time_of_day=assignment.time_of_day,
                    activity_id=None,
                    counts_toward_fmc_capacity=assignment.counts_toward_fmc_capacity,
                    source=assignment.source,
                    block_assignment_id=assignment.block_assignment_id,
                    is_override=assignment.is_override,
                    override_reason=assignment.override_reason,
                    overridden_by=assignment.overridden_by,
                    overridden_at=assignment.overridden_at,
                    created_at=assignment.created_at,
                    updated_at=assignment.updated_at,
                )
                clone.activity = Activity(
                    name="GAP",
                    code="gap",
                    display_abbreviation="GAP",
                    activity_category=ActivityCategory.ADMINISTRATIVE.value,
                )
                clone.person = assignment.person
                clone.is_gap = True
                effective.append(clone)
                continue

            replacement = replacement_people.get(override.replacement_person_id)
            if not replacement:
                effective.append(assignment)
                continue

            clone = HalfDayAssignment(
                id=assignment.id,
                person_id=replacement.id,
                date=assignment.date,
                time_of_day=assignment.time_of_day,
                activity_id=assignment.activity_id,
                counts_toward_fmc_capacity=assignment.counts_toward_fmc_capacity,
                source=assignment.source,
                block_assignment_id=assignment.block_assignment_id,
                is_override=assignment.is_override,
                override_reason=assignment.override_reason,
                overridden_by=assignment.overridden_by,
                overridden_at=assignment.overridden_at,
                created_at=assignment.created_at,
                updated_at=assignment.updated_at,
            )
            clone.activity = assignment.activity
            clone.person = replacement
            effective.append(clone)

        return effective

    def _parse_block_template2(
        self, file_bytes: bytes, start_date: date, end_date: date
    ) -> tuple[list[ParsedSlot], list[str]]:
        """Parse Block Template2 Excel into slots."""
        warnings: list[str] = []
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active

        # Build date columns from row 3
        date_cols: list[tuple[int, date]] = []
        col = SCHEDULE_START_COL
        while True:
            cell = ws.cell(row=DATE_ROW, column=col)
            if isinstance(cell, MergedCell):
                col += 1
                continue
            val = cell.value
            if isinstance(val, datetime):
                date_cols.append((col, val.date()))
            elif isinstance(val, date):
                date_cols.append((col, val))
            else:
                if col >= SCHEDULE_START_COL + 56:
                    break
            col += 2
            if col > 200:
                break

        if not date_cols:
            raise ValueError("Could not find date columns in row 3")

            # Verify date range if possible
        parsed_start = date_cols[0][1]
        parsed_end = date_cols[-1][1]
        if parsed_start != start_date or parsed_end != end_date:
            warnings.append(
                f"Excel date range {parsed_start} to {parsed_end} "
                f"does not match block {start_date} to {end_date}"
            )

        slots: list[ParsedSlot] = []
        for row_idx in range(DATA_START_ROW, DATA_END_ROW + 1):
            name_cell = ws.cell(row=row_idx, column=NAME_COL)
            name_val = name_cell.value
            if not name_val:
                continue
            person_name = str(name_val).replace("*", "").strip()

            for col, slot_date in date_cols:
                am_val = ws.cell(row=row_idx, column=col).value
                pm_val = ws.cell(row=row_idx, column=col + 1).value

                am_raw = self._clean_cell_value(am_val)
                pm_raw = self._clean_cell_value(pm_val)

                slots.append(
                    ParsedSlot(
                        person_name=person_name,
                        person_id=None,
                        assignment_date=slot_date,
                        time_of_day="AM",
                        raw_value=am_raw,
                        excel_code=None,
                        warnings=[],
                        row_number=row_idx,
                    )
                )
                slots.append(
                    ParsedSlot(
                        person_name=person_name,
                        person_id=None,
                        assignment_date=slot_date,
                        time_of_day="PM",
                        raw_value=pm_raw,
                        excel_code=None,
                        warnings=[],
                        row_number=row_idx,
                    )
                )

        wb.close()
        return slots, warnings

    def _load_activity_map(self) -> None:
        self._activity_map = {}
        activities = self.db.query(Activity).all()
        for activity in activities:
            if activity.code:
                self._activity_map[self._normalize_token(activity.code)] = activity.code
            if activity.display_abbreviation:
                self._activity_map[
                    self._normalize_token(activity.display_abbreviation)
                ] = activity.code

    def _load_person_map(self) -> None:
        self._person_map = {}
        people = self.db.query(Person).all()
        for person in people:
            normalized = self._normalize_person_name(person.name)
            self._person_map[normalized] = person.id

    def _normalize_excel_code(
        self, raw_value: str | None, time_of_day: str
    ) -> str | None:
        if raw_value is None:
            return None
        token = self._normalize_token(raw_value)
        if not token:
            return None

        if token in SPECIAL_CODE_MAP:
            return SPECIAL_CODE_MAP[token].get(time_of_day, token)

            # Strip AM/PM suffixes
        if token.endswith("-AM") or token.endswith("-PM"):
            token = token[:-3]

        return self._activity_map.get(token, token)

    def _normalize_activity_code(self, activity: Activity | None) -> str | None:
        if not activity:
            return None
        return activity.code or None

    def _normalize_person_name(self, name: str) -> str:
        if not name:
            return ""
        name = name.replace("*", "").strip()
        if "," in name:
            last, first = [part.strip() for part in name.split(",", 1)]
            name = f"{first} {last}".strip()
        name = re.sub(r"\s+", " ", name)
        return name.lower()

    def _normalize_token(self, value: str) -> str:
        return re.sub(r"\s+", "", (value or "").strip().upper())

    def _clean_cell_value(self, value: Any) -> str | None:
        if value is None:
            return None
        if isinstance(value, str):
            cleaned = value.strip()
            return cleaned or None
        return str(value).strip()
