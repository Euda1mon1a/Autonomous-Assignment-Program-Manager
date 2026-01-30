"""
block Format Exporter.

Exports schedule data to block xlsx format by filling in a template.
Preserves all formatting, colors, formulas, and merged cells from the template.
"""

import json
from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path
from typing import Any
from uuid import UUID

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.logging import get_logger
from app.models.assignment import Assignment
from app.models.block import Block
from app.models.person import Person
from app.models.rotation_template import RotationTemplate

logger = get_logger(__name__)

# Default paths - __file__ is at app/services/block_format_exporter.py
# .parent.parent.parent gets to backend root (/app in container)
DEFAULT_TEMPLATE_PATH = (
    Path(__file__).parent.parent.parent / "data" / "local" / "Block10_TEMPLATE.xlsx"
)
DEFAULT_MAPPING_PATH = (
    Path(__file__).parent.parent.parent
    / "data"
    / "local"
    / "block_template_mapping.json"
)


class BlockFormatExporter:
    """
    Exports schedule data in block xlsx format.

    Uses a template xlsx file and fills in daily assignments
    while preserving all formatting.
    """

    def __init__(
        self,
        session: AsyncSession,
        template_path: Path | str | None = None,
        mapping_path: Path | str | None = None,
    ) -> None:
        self.session = session
        self.template_path = (
            Path(template_path) if template_path else DEFAULT_TEMPLATE_PATH
        )
        self.mapping_path = Path(mapping_path) if mapping_path else DEFAULT_MAPPING_PATH
        self._mapping: dict[str, Any] | None = None
        self._date_to_cols: dict[date, tuple[int, int]] | None = None

    def _load_mapping(self) -> dict[str, Any]:
        """Load the template mapping JSON."""
        if self._mapping is None:
            with open(self.mapping_path) as f:
                self._mapping = json.load(f)
        return self._mapping

    def _build_date_column_map(self) -> dict[date, tuple[int, int]]:
        """Build a mapping from date to (am_col, pm_col)."""
        if self._date_to_cols is not None:
            return self._date_to_cols

        mapping = self._load_mapping()
        self._date_to_cols = {}

        for key, col_info in mapping.get("column_mapping", {}).items():
            date_str = col_info.get("date")
            if date_str:
                # Parse date string like "2025-03-12"
                d = datetime.strptime(date_str, "%Y-%m-%d").date()
                self._date_to_cols[d] = (col_info["am_col"], col_info["pm_col"])

        return self._date_to_cols

    async def _get_assignments(
        self,
        start_date: date,
        end_date: date,
    ) -> list[Assignment]:
        """Fetch assignments for the date range."""
        stmt = (
            select(Assignment)
            .join(Block, Assignment.block_id == Block.id)
            .options(
                selectinload(Assignment.person),
                selectinload(Assignment.rotation_template),
                selectinload(Assignment.block),
            )
            .where(Block.date >= start_date)
            .where(Block.date <= end_date)
            .order_by(Block.date, Block.time_of_day)
        )
        result = await self.session.execute(stmt)
        return list(result.scalars().all())

    async def _get_residents_in_order(self) -> list[Person]:
        """Get residents ordered by PGY level (3, 2, 1) then name."""
        stmt = (
            select(Person)
            .where(Person.person_type == "resident")
            .where(Person.is_active == True)  # noqa: E712
            .order_by(Person.pgy_level.desc(), Person.name)
        )
        result = await self.session.execute(stmt)
        return list(result.scalars().all())

    def _find_resident_row(
        self,
        ws: Worksheet,
        resident_name: str,
        mapping: dict[str, Any],
    ) -> int | None:
        """Find the row number for a resident by name in column E."""
        data_start = mapping["structure"]["data_start_row"]
        data_end = mapping["structure"]["data_end_row"]

        # Normalize name for comparison
        search_name = resident_name.lower().strip()

        for row in range(data_start, data_end + 1):
            cell_value = ws.cell(row, 5).value  # Column E = Provider name
            if cell_value:
                # Handle both "Last, First" and with asterisk suffix
                cell_name = str(cell_value).replace("*", "").lower().strip()
                if (
                    cell_name == search_name
                    or search_name in cell_name
                    or cell_name in search_name
                ):
                    return row

        return None

    def _get_assignment_abbrev(self, assignment: Assignment) -> str:
        """Get the display abbreviation for an assignment."""
        if assignment.rotation_template:
            abbrev = assignment.rotation_template.abbreviation
            # Map DB abbreviations to display abbreviations
            display_map = {
                "NF-PM": "NF",
                "OFF-AM": "OFF",
                "LEC-PM": "LEC",
                "W-AM": "W",
                "W-PM": "W",
                "FMIT-R": "FMIT",
                "SM-AM": "SM",
                "PR-AM": "PR",
                "IM-INT": "IM",
                "GYN-CLIN": "GYN",
                "SURG-EXP": "SURG",
                "PEDS-W": "PedW",
                "KAPI-LD": "KAP",
                # Night float variants
                "LDNF": "L&D",
                "PNF": "PedNF",
                "NEURO-NF": "NEURO",  # First half; mid-block switches to NF
                "NF-ENDO": "NF",  # Night float with MS:Endo secondary
                # Other mappings
                "FMC": "C",  # Family Medicine Clinic → Clinic
                "C": "C",
            }
            return display_map.get(abbrev, abbrev)
        return ""

    async def export(
        self,
        start_date: date,
        end_date: date,
        output_path: Path | str | None = None,
    ) -> bytes:
        """
        Export schedule to block format xlsx.

        Args:
            start_date: First day of the block
            end_date: Last day of the block
            output_path: Optional path to save the file

        Returns:
            Bytes of the xlsx file
        """
        logger.info(f"Exporting block format for {start_date} to {end_date}")

        # Load template and mapping
        wb = load_workbook(self.template_path)
        ws = wb.active
        mapping = self._load_mapping()
        date_cols = self._build_date_column_map()

        # Get assignments
        assignments = await self._get_assignments(start_date, end_date)
        logger.info(f"Found {len(assignments)} assignments")

        # Build lookup: (person_id, date, time_of_day) -> assignment
        assignment_lookup: dict[tuple[UUID, date, str], Assignment] = {}
        for a in assignments:
            if a.person and a.block:
                key = (a.person.id, a.block.date, a.block.time_of_day)
                assignment_lookup[key] = a

                # Build row lookup: person_name -> row number
                # Template uses "Last, First" format; normalize to match DB "First Last"
        row_lookup: dict[str, int] = {}
        data_start = mapping["structure"]["data_start_row"]
        data_end = mapping["structure"]["data_end_row"]

        def normalize_template_name(name: str) -> str:
            """Normalize template name 'Last, First' to 'last, first'."""
            return name.replace("*", "").strip().lower()

        def get_last_name(name: str) -> str:
            """Extract last name from 'Last, First' or 'First Last' format."""
            name = name.replace("*", "").strip().lower()
            if ", " in name:
                return name.split(", ")[0]  # "last, first" -> "last"
            parts = name.split()
            return parts[-1] if parts else name  # "first last" -> "last"

        def normalize_db_name(name: str) -> str:
            """Convert DB 'First Last' to template format 'last, first'."""
            name = name.strip().lower()
            parts = name.split()
            if len(parts) >= 2:
                return f"{parts[-1]}, {' '.join(parts[:-1])}"  # "first last" -> "last, first"
            return name

            # Build lookups: full name and last-name-only fallback

        last_name_lookup: dict[str, int] = {}
        for row in range(data_start, data_end + 1):
            cell_value = ws.cell(row, 5).value  # Column E
            if cell_value:
                normalized = normalize_template_name(str(cell_value))
                row_lookup[normalized] = row
                # Also index by last name for fallback matching
                last_name = get_last_name(str(cell_value))
                last_name_lookup[last_name] = row

                # Build date column map relative to start_date
                # Template mapping has hardcoded dates; we map by day offset instead
        date_col_map: dict[date, tuple[int, int]] = {}
        col_mapping = mapping.get("column_mapping", {})
        template_dates = sorted(
            [
                datetime.strptime(v["date"], "%Y-%m-%d").date()
                for v in col_mapping.values()
                if v.get("date")
            ]
        )
        if template_dates:
            template_start = template_dates[0]
            for key, col_info in col_mapping.items():
                if col_info.get("date"):
                    template_date = datetime.strptime(
                        col_info["date"], "%Y-%m-%d"
                    ).date()
                    day_offset = (template_date - template_start).days
                    actual_date = start_date + timedelta(days=day_offset)
                    date_col_map[actual_date] = (col_info["am_col"], col_info["pm_col"])

                    # Fill in assignments
        filled_count = 0
        for a in assignments:
            if not a.person or not a.block:
                continue

                # Find row by normalized name (convert DB "First Last" to template "last, first")
            person_name = normalize_db_name(a.person.name)
            row = row_lookup.get(person_name)
            if not row:
                # Try last-name-only fallback (handles nickname mismatches)
                db_last_name = get_last_name(a.person.name)
                row = last_name_lookup.get(db_last_name)
            if not row:
                # Try partial match as last resort
                for name, r in row_lookup.items():
                    if person_name in name or name in person_name:
                        row = r
                        break

            if not row:
                logger.debug(f"No row found for {a.person.name}")
                continue

                # Find column using relative date mapping
            block_date = a.block.date
            if block_date not in date_col_map:
                logger.debug(f"No column for date {block_date}")
                continue

            am_col, pm_col = date_col_map[block_date]
            col = am_col if a.block.time_of_day == "AM" else pm_col

            # Get abbreviation
            abbrev = self._get_assignment_abbrev(a)
            if abbrev:
                ws.cell(row, col).value = abbrev
                filled_count += 1

        logger.info(f"Filled {filled_count} cells")

        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        xlsx_bytes = output.read()

        # Optionally save to file
        if output_path:
            output_path = Path(output_path)
            output_path.write_bytes(xlsx_bytes)
            logger.info(f"Saved to {output_path}")

        return xlsx_bytes


async def export_block_format(
    session: AsyncSession,
    start_date: date,
    end_date: date,
    template_path: Path | str | None = None,
    output_path: Path | str | None = None,
) -> bytes:
    """
    Convenience function to export schedule in block format.

    Args:
        session: Database session
        start_date: First day of the block
        end_date: Last day of the block
        template_path: Path to template xlsx (uses default if None)
        output_path: Optional path to save the file

    Returns:
        Bytes of the xlsx file
    """
    exporter = BlockFormatExporter(session, template_path)
    return await exporter.export(start_date, end_date, output_path)
