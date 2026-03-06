"""Canonical schedule export service.

Exports schedules from half_day_assignments (descriptive truth) into
Block Template2 XLSX using a fixed, formatted template.

Pipeline:
  half_day_assignments -> HalfDayJSONExporter -> JSONToXlsxConverter -> xlsx
"""

from __future__ import annotations

import io
import json
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Protection
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.logging import get_logger
from app.models.academic_block import AcademicBlock
from app.models.activity import Activity
from app.models.rotation_template import RotationTemplate
from app.services.excel_metadata import (
    ExportMetadata,
    write_baseline_sheet,
    write_ref_sheet,
    write_sys_meta_sheet,
)
from app.services.export.tamc_block_exporter import TAMCBlockExporter
from app.services.half_day_json_exporter import HalfDayJSONExporter
from app.services.json_to_xlsx_converter import JSONToXlsxConverter
from app.utils.academic_blocks import get_block_dates

logger = get_logger(__name__)

PHANTOM_FILL = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
LOCKED = Protection(locked=True)


class CanonicalScheduleExportService:
    """Export a block schedule in the official Block Template2 format."""

    def __init__(self, db: Session) -> None:
        self.db = db

    def export_block_xlsx(
        self,
        block_number: int,
        academic_year: int,
        include_faculty: bool = True,
        include_overrides: bool = True,
        include_qa_sheet: bool = True,  # noqa: ARG002 — kept for API compat
        preserve_template_identity_fields: bool = True,  # noqa: ARG002
        presentation_profile: str = "tamc_handjam_v2",  # noqa: ARG002
        output_path: Path | str | None = None,
    ) -> bytes:
        """Export a block schedule to XLSX using TAMCBlockExporter.

        Note: ``include_qa_sheet``, ``preserve_template_identity_fields``, and
        ``presentation_profile`` are accepted for API compatibility but have no
        effect — TAMCBlockExporter always produces the full-fidelity template.
        """
        block_dates = get_block_dates(block_number, academic_year)

        data = self._export_json_data(
            block_dates.start_date,
            block_dates.end_date,
            include_faculty=include_faculty,
            include_overrides=include_overrides,
        )

        # Query ref codes for Phase 1 __REF__ sheet (single-save pattern)
        rotation_codes = [
            r[0]
            for r in self.db.query(RotationTemplate.abbreviation).distinct().all()
            if r[0]
        ]
        activity_codes = [
            a[0] for a in self.db.query(Activity.code).distinct().all() if a[0]
        ]

        meta = ExportMetadata(
            academic_year=academic_year,
            block_number=block_number,
            export_timestamp=datetime.now(UTC).isoformat(),
        )

        exporter = TAMCBlockExporter(
            block_config=self._load_block_config(block_number),
        )
        xlsx_bytes = exporter.export(
            data,
            export_metadata=meta,
            rotation_codes=rotation_codes,
            activity_codes=activity_codes,
        )

        # Optional xlwings finishing pass
        if self._should_use_xlwings():
            xlsx_bytes = self._apply_xlwings_finishing(xlsx_bytes)

        if output_path:
            Path(output_path).write_bytes(xlsx_bytes)

        return xlsx_bytes

    def export_year_xlsx(
        self,
        academic_year: int,
        include_faculty: bool = True,
        include_overrides: bool = True,
        output_path: Path | str | None = None,
    ) -> bytes:
        """Export all 14 blocks for an academic year into a single workbook."""
        wb = Workbook()
        wb.remove(wb.active)  # type: ignore

        stmt = (
            select(AcademicBlock)
            .where(AcademicBlock.academic_year == academic_year)
            .order_by(AcademicBlock.block_number)
        )
        blocks = self.db.execute(stmt).scalars().all()

        if not blocks:
            raise ValueError(f"No academic blocks found for year {academic_year}")

        block_map = {}
        all_faculty_by_name: dict[str, dict[str, Any]] = {}
        template_path = self._template_path()
        structure_path = self._structure_path()
        effective_preserve = structure_path is not None and template_path is not None

        converter = JSONToXlsxConverter(
            template_path=template_path,
            structure_xml_path=structure_path,
            use_block_template2=True,
            apply_colors=True,
            strict_row_mapping=structure_path is not None,
            include_qa_sheet=False,  # Single QA sheet for 14 blocks is messy
            preserve_template_identity_fields=effective_preserve,
            presentation_profile="tamc_handjam_v2",
        )

        for block in blocks:
            logger.info(f"Exporting Block {block.block_number} for yearly workbook")
            data = self._export_json_data(
                block.start_date,
                block.end_date,
                include_faculty=include_faculty,
                include_overrides=include_overrides,
            )

            # Accumulate faculty across all blocks for YTD_SUMMARY
            for f in data.get("faculty", []):
                name = f.get("name", "")
                if name:
                    all_faculty_by_name[name] = f

            # Convert to temporary workbook to extract sheet
            temp_bytes = converter.convert_from_json(data)
            temp_wb = load_workbook(io.BytesIO(temp_bytes))
            temp_ws = temp_wb["Block Template2"]

            # Create sheet in master workbook
            sheet_title = f"Block {block.block_number}"
            ws = wb.create_sheet(title=sheet_title)

            # Copy values and styles
            self._copy_worksheet(temp_ws, ws)

            # Apply phantom columns for stub blocks (0 and 13)
            self._apply_phantom_columns(ws, block)

            # Hide unused faculty rows for cleaner presentation
            # Base template has 50 faculty rows starting at row 31
            active_faculty_count = len(data.get("faculty", []))
            last_populated_row = 30 + active_faculty_count
            for row_idx in range(last_populated_row + 1, 81):
                ws.row_dimensions[row_idx].hidden = True

            # Collect baseline cell data for hand-jam tracking
            baseline_cells = self._collect_baseline_data(ws, data)
            if baseline_cells:
                write_baseline_sheet(wb, sheet_title, baseline_cells)

            block_map[sheet_title] = str(block.id)
            temp_wb.close()

        # Build YTD Summary sheet using union of all faculty across all blocks
        all_faculty = sorted(
            all_faculty_by_name.values(), key=lambda f: f.get("name", "")
        )
        summary_ws = wb.create_sheet(title="YTD_SUMMARY", index=0)
        self._build_ytd_summary_sheet(summary_ws, blocks, all_faculty)

        # Ensure YTD_SUMMARY is active
        wb.active = summary_ws

        # Phase 1 metadata — single-save (write directly to wb before save)
        rotation_codes = [
            r[0]
            for r in self.db.query(RotationTemplate.abbreviation).distinct().all()
            if r[0]
        ]
        activity_codes = [
            a[0] for a in self.db.query(Activity.code).distinct().all() if a[0]
        ]
        meta = ExportMetadata(
            academic_year=academic_year,
            export_timestamp=datetime.now(UTC).isoformat(),
            block_map=block_map,
        )
        write_sys_meta_sheet(wb, meta)
        write_ref_sheet(wb, rotation_codes, activity_codes)

        # Save to bytes (single save — no double load/save cycle)
        buffer = io.BytesIO()
        wb.save(buffer)
        xlsx_bytes = buffer.getvalue()

        # Skip xlwings for year export — sheets are named "Block {n}" not
        # "Block Template2", so the finishing pass would fail on sheet lookup.
        # TODO: Multi-sheet xlwings support for annual workbooks.

        if output_path:
            Path(output_path).write_bytes(xlsx_bytes)

        return xlsx_bytes

    def _should_use_xlwings(self) -> bool:
        """Check if xlwings finishing pass should be applied."""
        from app.core.config import get_settings
        from app.services.xlwings_availability import is_xlwings_available

        return get_settings().XLWINGS_ENABLED and is_xlwings_available()

    def _apply_xlwings_finishing(self, xlsx_bytes: bytes) -> bytes:
        """Run xlwings finishing pass on openpyxl-generated bytes.

        On any error, logs a warning and returns the original openpyxl output
        (graceful degradation — the schedule must always export).
        """
        import tempfile

        from app.services.tamc_color_scheme import get_color_scheme
        from app.services.xlwings_bridge import XlwingsBridge

        tmp_path = None
        try:
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                tmp.write(xlsx_bytes)
                tmp_path = Path(tmp.name)

            from app.core.config import get_settings

            settings = get_settings()
            bridge = XlwingsBridge(
                color_scheme=get_color_scheme(),
                visible=settings.XLWINGS_VISIBLE,
            )
            bridge.apply_finishing_pass(tmp_path)
            return tmp_path.read_bytes()
        except Exception:
            logger.warning(
                "xlwings finishing pass failed — returning openpyxl output",
                exc_info=True,
            )
            return xlsx_bytes
        finally:
            if tmp_path is not None:
                tmp_path.unlink(missing_ok=True)

    def _copy_worksheet(self, source, target) -> None:
        """Deep copy worksheet content and styles between workbooks."""
        from copy import copy

        for row in source.iter_rows():
            for cell in row:
                new_cell = target.cell(
                    row=cell.row, column=cell.column, value=cell.value
                )
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)

                if cell.comment:
                    new_cell.comment = copy(cell.comment)

        # Copy column dimensions
        for col, dim in source.column_dimensions.items():
            target.column_dimensions[col] = copy(dim)

        # Copy row dimensions (preserves hidden state from converter)
        for row_idx, dim in source.row_dimensions.items():
            target.row_dimensions[row_idx] = copy(dim)

        # Copy merged cells
        for merged_range in source.merged_cells.ranges:
            target.merge_cells(str(merged_range))

        # Copy data validation
        for dv in source.data_validations.dataValidation:
            target.add_data_validation(copy(dv))

        # Copy conditional formatting
        for rule_range, rules in source.conditional_formatting._cf_rules.items():
            for rule in rules:
                target.conditional_formatting.add(rule_range, copy(rule))

    def _apply_phantom_columns(self, ws, block: AcademicBlock) -> None:
        """Gray out and lock columns for days that don't exist in stub blocks."""
        # BT2_COL_SCHEDULE_START = 6
        # COLS_PER_DAY = 2
        actual_days = (block.end_date - block.start_date).days + 1
        if actual_days >= 28:
            return

        # Data starts at row 9, ends at 69
        for day_index in range(actual_days, 28):
            for slot in range(2):  # AM, PM
                col = 6 + (day_index * 2) + slot
                for row in range(9, 70):
                    cell = ws.cell(row=row, column=col)
                    cell.fill = PHANTOM_FILL
                    cell.protection = LOCKED
                    cell.value = None

        ws.protection.sheet = True
        # No password — visual lock only (prevents accidental edits)

    def _build_ytd_summary_sheet(
        self, ws, blocks: list[AcademicBlock], faculty: list[dict[str, Any]]
    ) -> None:
        """Build YTD_SUMMARY sheet using dynamic cross-sheet SUMIF formulas."""
        headers = [
            "Faculty Name",
            "YTD Clinic (C+SM)",
            "YTD CC",
            "YTD CV",
            "YTD Total Clinic",
            "YTD AT (AT+PCAT+DO)",
            "YTD Admin",
            "YTD Leave",
            "YTD FMIT Weeks",
            "YTD Call Nights",
        ]

        # Write headers
        from openpyxl.styles import Font

        bold_font = Font(bold=True)
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = bold_font

        block_names = [f"'Block {b.block_number}'" for b in blocks]

        # Write faculty rows
        for row_idx, f in enumerate(faculty, start=2):
            name = f.get("name", "")
            ws.cell(row=row_idx, column=1, value=name)

            if not name:
                continue

            # Helper to generate cross-sheet SUMIF
            def cross_sheet_sumif(col_letter: str, current_row: int = row_idx) -> str:
                # e.g., SUMIF('Block 0'!$E$31:$E$80, $A2, 'Block 0'!BJ$31:BJ$80) + ...
                terms = [
                    f"SUMIF({b}!$E$31:$E$80, $A{current_row}, {b}!{col_letter}$31:{col_letter}$80)"
                    for b in block_names
                ]
                return f"=({'+'.join(terms)})"

            # B: YTD Clinic (C+SM) -> BJ
            ws.cell(row=row_idx, column=2, value=cross_sheet_sumif("BJ"))
            # C: YTD CC -> BK
            ws.cell(row=row_idx, column=3, value=cross_sheet_sumif("BK"))
            # D: YTD CV -> BL
            ws.cell(row=row_idx, column=4, value=cross_sheet_sumif("BL"))
            # E: YTD Total Clinic -> B+C+D
            ws.cell(row=row_idx, column=5, value=f"=B{row_idx}+C{row_idx}+D{row_idx}")
            # F: YTD AT (AT+PCAT+DO) -> BN
            ws.cell(row=row_idx, column=6, value=cross_sheet_sumif("BN"))
            # G: YTD Admin -> BO
            ws.cell(row=row_idx, column=7, value=cross_sheet_sumif("BO"))
            # H: YTD Leave -> BP
            ws.cell(row=row_idx, column=8, value=cross_sheet_sumif("BP"))
            # I: YTD FMIT Weeks -> BQ / 14
            ws.cell(row=row_idx, column=9, value=f"{cross_sheet_sumif('BQ')}/14")
            # J: YTD Call Nights -> BR
            ws.cell(row=row_idx, column=10, value=cross_sheet_sumif("BR"))

    def _collect_baseline_data(self, ws, data: dict[str, Any]) -> list[dict[str, Any]]:
        """Extract cell data from a block sheet for baseline fingerprinting.

        Scans the data region (rows 9-69, cols F onward) and records every
        non-empty cell with its reference, value, and source from the JSON data.
        """
        from openpyxl.utils import get_column_letter

        cells: list[dict[str, Any]] = []

        # BT2 layout: data starts row 9, name col E (5), schedule cols start F (6)
        # 28 days × 2 slots (AM/PM) = 56 columns of schedule data
        DATA_START_ROW = 9
        DATA_END_ROW = 69  # Max possible (residents + faculty)
        SCHEDULE_START_COL = 6
        SCHEDULE_END_COL = SCHEDULE_START_COL + 56  # 28 days × 2 slots

        # Build person-name -> source lookup from JSON data
        source_by_person: dict[str, str] = {}
        for person_list_key in ("residents", "faculty"):
            for person in data.get(person_list_key, []):
                name = person.get("name", "")
                source_by_person[name] = person.get("source", "solver")

        for row in range(DATA_START_ROW, DATA_END_ROW + 1):
            person_name = ws.cell(row=row, column=5).value  # Column E = name
            if not person_name:
                continue

            # Compute row hash from all schedule cells in this row
            day_values = []
            for col in range(SCHEDULE_START_COL, SCHEDULE_END_COL):
                day_values.append(ws.cell(row=row, column=col).value)

            rotation1 = ws.cell(row=row, column=3).value  # Column C
            rotation2 = ws.cell(row=row, column=4).value  # Column D
            row_hash = ""  # Will be populated if person_id available

            person_source = source_by_person.get(str(person_name), "solver")

            for col in range(SCHEDULE_START_COL, SCHEDULE_END_COL):
                val = ws.cell(row=row, column=col).value
                if val is not None and str(val).strip():
                    col_letter = get_column_letter(col)
                    cells.append(
                        {
                            "cell_ref": f"{col_letter}{row}",
                            "value": str(val),
                            "row_hash": row_hash,
                            "source": person_source,
                        }
                    )

        return cells

    # Faculty excluded from export (placeholders and removed staff)
    _EXCLUDED_FACULTY_NAMES: set[str] = {"Kate Bohringer"}

    def _export_json_data(
        self,
        start_date: date,
        end_date: date,
        include_faculty: bool,
        include_overrides: bool,
    ) -> dict[str, Any]:
        from datetime import timedelta

        from app.models.person import Person

        exporter = HalfDayJSONExporter(self.db)
        data = exporter.export(
            block_start=start_date,
            block_end=end_date,
            include_faculty=include_faculty,
            include_call=True,
            include_overrides=include_overrides,
        )

        # Filter out excluded faculty (e.g. departed faculty not yet deleted)
        if data.get("faculty"):
            data["faculty"] = [
                f
                for f in data["faculty"]
                if f.get("name", "") not in self._EXCLUDED_FACULTY_NAMES
            ]

        # Ensure adjunct faculty appear even without assignments (blank rows)
        if include_faculty:
            existing_names = {f.get("name") for f in data.get("faculty", [])}
            adjunct_people = (
                self.db.query(Person)
                .filter(Person.type == "faculty", Person.faculty_role == "adjunct")
                .all()
            )
            for p in adjunct_people:
                if p.name not in existing_names:
                    days = []
                    current = start_date
                    while current <= end_date:
                        days.append(
                            {
                                "date": current.isoformat(),
                                "weekday": current.strftime("%a"),
                                "am": None,
                                "pm": None,
                            }
                        )
                        current = current + timedelta(days=1)
                    data.setdefault("faculty", []).append(
                        {
                            "id": str(p.id),
                            "name": p.name,
                            "pgy": None,
                            "faculty_role": "adjunct",
                            "rotation1": "",
                            "rotation2": "",
                            "days": days,
                        }
                    )

        return data

    def _inject_metadata(
        self,
        xlsx_bytes: bytes,
        academic_year: int,
        block_number: int | None = None,
        output_path: Path | str | None = None,
        block_map: dict[str, str] | None = None,
    ) -> bytes:
        """Add __SYS_META__ and __REF__ sheets to exported workbook."""
        wb = load_workbook(io.BytesIO(xlsx_bytes))

        rotation_codes = [
            r[0]
            for r in self.db.query(RotationTemplate.abbreviation).distinct().all()
            if r[0]
        ]
        activity_codes = [
            a[0] for a in self.db.query(Activity.code).distinct().all() if a[0]
        ]

        meta = ExportMetadata(
            academic_year=academic_year,
            block_number=block_number,
            export_timestamp=datetime.now(UTC).isoformat(),
            block_map=block_map,
        )
        write_sys_meta_sheet(wb, meta)
        write_ref_sheet(wb, rotation_codes, activity_codes)

        buf = io.BytesIO()
        wb.save(buf)
        final_bytes = buf.getvalue()

        if output_path:
            with open(output_path, "wb") as f:
                f.write(final_bytes)

        return final_bytes

    def _load_block_config(self, block_number: int) -> dict:
        """Load block-specific config from block_config.json."""
        config_path = Path(__file__).parent / "export" / "block_config.json"
        if not config_path.exists():
            return {}
        with open(config_path) as f:
            return json.load(f).get("blocks", {}).get(str(block_number), {})

    def _data_dir(self) -> Path:
        # backend/app/services -> backend
        return Path(__file__).resolve().parents[2] / "data"

    def _template_path(self) -> Path | None:
        template = self._data_dir() / "BlockTemplate2_Official.xlsx"
        if not template.exists():
            logger.warning(
                "Canonical template missing: %s — export will generate BT2 layout "
                "programmatically (no pre-formatted styling).",
                template,
            )
            return None
        return template

    def _structure_path(self) -> Path | None:
        structure = self._data_dir() / "BlockTemplate2_Structure.xml"
        if not structure.exists():
            logger.info(
                "Structure XML not found at %s — using dynamic UUID-based row mappings.",
                structure,
            )
            return None
        return structure
