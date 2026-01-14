"""
Block Schedule Sheet Parser.

Parses TRIPLER-format block schedule xlsx files.
Extracts resident rotation assignments for import into the scheduler.

Expected XLSX structure:
    Row 1: Block number + day names (THURS, FRI, SAT, ...)
    Row 3: Dates (12-Mar, 13-Mar, ...)
    Row 4: Date range (12Mar-08Apr) + Staff Call assignments
    Row 6: Headers (TEMPLATE, ROLE, PROVIDER, ...)
    Row 7+: Data rows

Data row format:
    Col A: Primary rotation template (Hilo, NF, FMC, etc.)
    Col B: Secondary assignment (optional)
    Col C: Role (R1, R2, R3 for residents)
    Col D: PGY level (PGY 1, PGY 2, PGY 3)
    Col E: Name (Last, First format, may have * suffix)
    Col F+: Daily activities (W, LV, LEC, etc.)
"""

import io
import re
from dataclasses import dataclass
from pathlib import Path
from typing import BinaryIO

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.core.logging import get_logger

logger = get_logger(__name__)


@dataclass
class BlockAssignmentRow:
    """Parsed block assignment from xlsx."""

    rotation_template: str
    secondary_rotation: str | None
    person_name: str
    pgy_level: int
    block_number: int
    role: str  # R1, R2, R3

    def to_dict(self) -> dict:
        """Convert to dictionary for import pipeline."""
        return {
            "rotation": self.rotation_template,
            "secondary_rotation": self.secondary_rotation,
            "name": self.person_name,
            "pgy_level": self.pgy_level,
            "block": self.block_number,
            "role": self.role,
        }


class BlockScheduleParser:
    """
    Parses TRIPLER-format block schedule xlsx.

    Extracts resident rotation assignments and feeds them into
    the existing BlockAssignmentImportService pipeline.
    """

    # Rotation name mappings (xlsx name/abbreviation -> DB abbreviation)
    # Maps xlsx values to actual database rotation template abbreviations
    ROTATION_MAPPINGS: dict[str, str] = {
        # Night Float variations
        "NF": "NF-PM",
        "Night Float": "NF-PM",
        "NEURO": "NEURO",
        "Neurology": "NEURO",
        # Sports Medicine
        "SM": "SM-AM",
        "Sports Medicine": "SM-AM",
        # FMIT
        "FMIT 2": "FMIT-R",
        "FMIT 1": "FMIT-R",
        "FMIT": "FMIT-R",
        # L&D Night Float
        "L and D night float": "LDNF",
        "L and D NF": "LDNF",
        "L&D Night Float": "LDNF",
        "L&D NF": "LDNF",
        # Surgical Experience
        "Surg Exp": "SURG-EXP",
        "Surgical Experience": "SURG-EXP",
        # GYN Clinic
        "Gyn Clinic": "GYN-CLIN",
        "GYN Clinic": "GYN-CLIN",
        "Gynecology Clinic": "GYN-CLIN",
        # Pediatrics
        "Peds Ward": "PEDS-W",
        "Pediatrics Ward": "PEDS-W",
        "Peds NF": "PNF",
        "Pediatrics Night Float": "PNF",
        # Kapiolani
        "Kapiolani L and D": "KAPI-LD",
        "Kapiolani L&D": "KAPI-LD",
        # Procedures
        "PROC": "PR-AM",
        "Procedures": "PR-AM",
        # Internal Medicine
        "IM": "IM-INT",
        "Internal Medicine": "IM-INT",
        # Endocrinology
        "MS: Endo": "ENDO",
        "Endocrinology": "ENDO",
        # Hilo
        "Hilo": "HILO",
        # FMC
        "FMC": "FMC",
        "Family Medicine Clinic": "FMC",
        # POCUS
        "POCUS": "POCUS",
        "Point of Care Ultrasound": "POCUS",
    }

    # Role patterns that identify residents
    RESIDENT_ROLES = {"R1", "R2", "R3"}

    def __init__(self) -> None:
        self._block_number: int | None = None
        self._header_row: int | None = None
        self._column_map: dict[str, int] = {}

    def parse_file(self, file_path: str | Path) -> list[BlockAssignmentRow]:
        """
        Parse a block schedule xlsx file.

        Args:
            file_path: Path to the xlsx file

        Returns:
            List of BlockAssignmentRow objects
        """
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        return self._parse_worksheet(ws)

    def parse_bytes(self, file_bytes: bytes) -> list[BlockAssignmentRow]:
        """
        Parse block schedule from bytes (for API uploads).

        Args:
            file_bytes: Excel file as bytes

        Returns:
            List of BlockAssignmentRow objects
        """
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active
        return self._parse_worksheet(ws)

    def parse_file_handle(self, file_handle: BinaryIO) -> list[BlockAssignmentRow]:
        """
        Parse block schedule from file handle.

        Args:
            file_handle: Open file handle

        Returns:
            List of BlockAssignmentRow objects
        """
        wb = load_workbook(file_handle, data_only=True)
        ws = wb.active
        return self._parse_worksheet(ws)

    def _parse_worksheet(self, ws: Worksheet) -> list[BlockAssignmentRow]:
        """
        Parse a worksheet and extract resident assignments.

        Args:
            ws: openpyxl Worksheet

        Returns:
            List of BlockAssignmentRow objects
        """
        # Step 1: Find block number from row 1
        self._block_number = self._find_block_number(ws)
        if self._block_number is None:
            logger.warning("Could not find block number in worksheet")
            # Try to extract from filename or default
            self._block_number = 0

        logger.info(f"Parsing block {self._block_number}")

        # Step 2: Find header row
        self._header_row = self._find_header_row(ws)
        if self._header_row is None:
            raise ValueError("Could not find header row (TEMPLATE/ROLE/PROVIDER)")

        logger.info(f"Found header row at {self._header_row}")

        # Step 3: Build column map
        self._build_column_map(ws)

        # Step 4: Extract resident rows
        assignments = self._extract_resident_rows(ws)

        logger.info(f"Extracted {len(assignments)} resident assignments")
        return assignments

    def _find_block_number(self, ws: Worksheet) -> int | None:
        """
        Find block number from row 1.

        Scans first row for a standalone integer (the block number).
        """
        for cell in ws[1]:
            if cell.value is not None:
                # Try to parse as integer
                try:
                    val = int(cell.value)
                    if 1 <= val <= 13:  # Valid block numbers
                        return val
                except (ValueError, TypeError):
                    # Check if it's a string that's just a number
                    if isinstance(cell.value, str):
                        stripped = cell.value.strip()
                        if stripped.isdigit():
                            val = int(stripped)
                            if 1 <= val <= 13:
                                return val
        return None

    def _find_header_row(self, ws: Worksheet) -> int | None:
        """
        Find the header row containing TEMPLATE, ROLE, PROVIDER.

        Scans first 20 rows looking for these keywords.
        """
        keywords = {"TEMPLATE", "ROLE", "PROVIDER", "PGY"}

        for row_idx in range(1, min(21, ws.max_row + 1)):
            row_values = []
            for cell in ws[row_idx]:
                if cell.value is not None:
                    val = str(cell.value).strip().upper()
                    row_values.append(val)

            # Check if this row contains header keywords
            found = sum(1 for v in row_values if v in keywords)
            if found >= 2:  # At least 2 keywords found
                return row_idx

        return None

    def _build_column_map(self, ws: Worksheet) -> None:
        """
        Build a map of column names to indices.

        Based on header row, identify which columns contain what data.
        """
        if self._header_row is None:
            return

        self._column_map = {}
        for idx, cell in enumerate(ws[self._header_row]):
            if cell.value is not None:
                key = str(cell.value).strip().upper()
                self._column_map[key] = idx

        logger.debug(f"Column map: {self._column_map}")

    def _extract_resident_rows(self, ws: Worksheet) -> list[BlockAssignmentRow]:
        """
        Extract all resident assignment rows.

        Filters to rows where Role column = R1, R2, or R3.
        """
        if self._header_row is None:
            return []

        assignments = []
        start_row = self._header_row + 1

        for row_idx in range(start_row, ws.max_row + 1):
            row_data = [cell.value for cell in ws[row_idx]]

            # Skip empty rows
            if not any(row_data):
                continue

            # Extract fields based on position (columns A-E)
            # Col A: Template, Col B: Secondary, Col C: Role, Col D: PGY, Col E: Name
            template = self._get_cell_value(row_data, 0)
            secondary = self._get_cell_value(row_data, 1)
            role = self._get_cell_value(row_data, 2)
            pgy_str = self._get_cell_value(row_data, 3)
            name = self._get_cell_value(row_data, 4)

            # Filter to residents only
            if role not in self.RESIDENT_ROLES:
                continue

            # Skip rows without a name
            if not name:
                continue

            # Skip rows without a rotation template
            if not template:
                continue

            # Parse PGY level
            pgy_level = self._parse_pgy_level(pgy_str)

            # Normalize the rotation name
            normalized_rotation = self._normalize_rotation(template)
            normalized_secondary = (
                self._normalize_rotation(secondary) if secondary else None
            )

            # Normalize the person name
            normalized_name = self._normalize_name(name)

            assignment = BlockAssignmentRow(
                rotation_template=normalized_rotation,
                secondary_rotation=normalized_secondary,
                person_name=normalized_name,
                pgy_level=pgy_level,
                block_number=self._block_number or 0,
                role=role,
            )
            assignments.append(assignment)

            logger.debug(f"Extracted: {assignment}")

        return assignments

    def _get_cell_value(self, row_data: list, col_idx: int) -> str | None:
        """Safely get cell value as string."""
        if col_idx >= len(row_data):
            return None
        val = row_data[col_idx]
        if val is None:
            return None
        return str(val).strip()

    def _parse_pgy_level(self, pgy_str: str | None) -> int:
        """
        Parse PGY level from string like "PGY 3" or "PGY3".

        Returns 0 if cannot parse.
        """
        if not pgy_str:
            return 0

        # Match "PGY 1", "PGY1", "PGY-1", etc.
        match = re.search(r"PGY[- ]?(\d+)", pgy_str.upper())
        if match:
            return int(match.group(1))
        return 0

    def _normalize_rotation(self, rotation: str | None) -> str:
        """
        Normalize rotation name to match DB templates.

        Applies known mappings and cleans up formatting.
        """
        if not rotation:
            return ""

        # Strip whitespace
        rotation = rotation.strip()

        # Check if there's a known mapping
        if rotation in self.ROTATION_MAPPINGS:
            return self.ROTATION_MAPPINGS[rotation]

        # Check case-insensitive
        for key, value in self.ROTATION_MAPPINGS.items():
            if key.upper() == rotation.upper():
                return value

        # Return as-is if no mapping found
        return rotation

    def _normalize_name(self, name: str | None) -> str:
        """
        Normalize person name.

        - Strips asterisks (*) used to mark chief residents
        - Handles "Last, First" format
        - Cleans up whitespace
        """
        if not name:
            return ""

        # Strip asterisks
        name = name.replace("*", "").strip()

        # Clean up multiple spaces
        name = re.sub(r"\s+", " ", name)

        return name

    def to_import_format(
        self, assignments: list[BlockAssignmentRow]
    ) -> list[dict[str, str]]:
        """
        Convert assignments to the format expected by BlockAssignmentImportService.

        Returns list of dicts with:
        - resident_name: str
        - rotation_name: str
        - block_number: str
        """
        return [
            {
                "resident_name": a.person_name,
                "rotation_name": a.rotation_template,
                "block_number": str(a.block_number),
            }
            for a in assignments
        ]


def parse_block_schedule(file_path: str | Path) -> list[dict]:
    """
    Convenience function to parse a block schedule file.

    Args:
        file_path: Path to xlsx file

    Returns:
        List of assignment dicts ready for import
    """
    parser = BlockScheduleParser()
    assignments = parser.parse_file(file_path)
    return parser.to_import_format(assignments)
