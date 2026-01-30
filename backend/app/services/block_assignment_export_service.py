"""
Block Assignment Export Service.

Handles:
- Export to CSV format
- Export to Excel format (via openpyxl if available)
- Filtering by academic year, block, rotation, resident
- Grouping options
"""

import csv
import io
from datetime import datetime
from uuid import UUID

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import joinedload

from app.core.logging import get_logger
from app.models.block_assignment import BlockAssignment
from app.models.person import Person
from app.models.rotation_template import RotationTemplate
from app.schemas.block_assignment_import import (
    BlockAssignmentExportRequest,
    ExportFormat,
)

logger = get_logger(__name__)


class BlockAssignmentExportService:
    """
    Service for exporting block assignments to CSV/Excel.

    Key features:
    - Multiple export formats (CSV, XLSX)
    - Flexible filtering
    - Grouping options
    - Admin-only (full names exported, no PERSEC anonymization)
    """

    def __init__(self, session: AsyncSession) -> None:
        self.session = session

    async def export(
        self, request: BlockAssignmentExportRequest
    ) -> tuple[bytes, str, str]:
        """
        Export block assignments based on request options.

        Args:
            request: Export request with filters and options

        Returns:
            Tuple of (file_bytes, filename, content_type)
        """
        # Query assignments with filters
        assignments = await self._query_assignments(request)

        # Generate export
        if request.format == ExportFormat.XLSX:
            file_bytes = self._generate_xlsx(assignments, request)
            filename = self._generate_filename(request, "xlsx")
            content_type = (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            file_bytes = self._generate_csv(assignments, request)
            filename = self._generate_filename(request, "csv")
            content_type = "text/csv"

        logger.info(
            f"Export generated: {len(assignments)} assignments, format={request.format.value}"
        )

        return file_bytes, filename, content_type

    async def _query_assignments(
        self, request: BlockAssignmentExportRequest
    ) -> list[BlockAssignment]:
        """Query assignments with filters."""
        query = (
            select(BlockAssignment)
            .options(
                joinedload(BlockAssignment.resident),
                joinedload(BlockAssignment.rotation_template),
            )
            .where(BlockAssignment.academic_year == request.academic_year)
        )

        # Apply filters
        if request.block_numbers:
            query = query.where(BlockAssignment.block_number.in_(request.block_numbers))

        if request.rotation_ids:
            query = query.where(
                BlockAssignment.rotation_template_id.in_(request.rotation_ids)
            )

        if request.resident_ids:
            query = query.where(BlockAssignment.resident_id.in_(request.resident_ids))

            # Order by block number, then resident name
        query = query.order_by(
            BlockAssignment.block_number,
            BlockAssignment.resident_id,
        )

        result = await self.session.execute(query)
        return list(result.scalars().unique().all())

    def _generate_csv(
        self,
        assignments: list[BlockAssignment],
        request: BlockAssignmentExportRequest,
    ) -> bytes:
        """Generate CSV export."""
        output = io.StringIO()
        writer = csv.writer(output)

        # Build header
        header = ["block_number", "rotation_abbrev", "resident_name"]
        if request.include_pgy_level:
            header.append("pgy_level")
        if request.include_leave_status:
            header.extend(["has_leave", "leave_days"])

        writer.writerow(header)

        # Write data rows
        for assignment in assignments:
            row = [
                assignment.block_number,
                (
                    assignment.rotation_template.abbreviation
                    if assignment.rotation_template
                    else ""
                ),
                assignment.resident.name if assignment.resident else "",
            ]

            if request.include_pgy_level:
                pgy = assignment.resident.pgy_level if assignment.resident else None
                row.append(pgy if pgy else "")

            if request.include_leave_status:
                row.extend([assignment.has_leave, assignment.leave_days])

            writer.writerow(row)

        return output.getvalue().encode("utf-8")

    def _generate_xlsx(
        self,
        assignments: list[BlockAssignment],
        request: BlockAssignmentExportRequest,
    ) -> bytes:
        """Generate Excel export with formatting."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            # Fallback to CSV if openpyxl not available
            logger.warning("openpyxl not available, falling back to CSV")
            return self._generate_csv(assignments, request)

        wb = Workbook()
        ws = wb.active
        ws.title = f"Block Assignments {request.academic_year}"

        # Define styles
        header_font = Font(bold=True)
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_font_white = Font(bold=True, color="FFFFFF")

        # Activity type color mapping
        activity_colors = {
            "clinic": "90EE90",  # Light green
            "inpatient": "FFD700",  # Gold
            "outpatient": "87CEEB",  # Sky blue
            "procedures": "DDA0DD",  # Plum
            "call": "FFA07A",  # Light salmon
            "education": "F0E68C",  # Khaki
            "off": "D3D3D3",  # Light gray
            "conference": "E6E6FA",  # Lavender
        }

        # Build header
        headers = ["Block", "Rotation", "Resident"]
        if request.include_pgy_level:
            headers.append("PGY")
        if request.include_leave_status:
            headers.extend(["Leave?", "Leave Days"])

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill

            # Write data rows
        for row_num, assignment in enumerate(assignments, 2):
            ws.cell(row=row_num, column=1, value=assignment.block_number)

            rotation_abbrev = (
                assignment.rotation_template.abbreviation
                if assignment.rotation_template
                else ""
            )
            ws.cell(row=row_num, column=2, value=rotation_abbrev)

            resident_name = assignment.resident.name if assignment.resident else ""
            ws.cell(row=row_num, column=3, value=resident_name)

            col_offset = 4
            if request.include_pgy_level:
                pgy = assignment.resident.pgy_level if assignment.resident else None
                ws.cell(row=row_num, column=col_offset, value=pgy if pgy else "")
                col_offset += 1

            if request.include_leave_status:
                ws.cell(row=row_num, column=col_offset, value=assignment.has_leave)
                ws.cell(row=row_num, column=col_offset + 1, value=assignment.leave_days)

                # Apply rotation type coloring to rotation cell
            if assignment.rotation_template:
                rotation_type = assignment.rotation_template.rotation_type
                if rotation_type and rotation_type.lower() in activity_colors:
                    color = activity_colors[rotation_type.lower()]
                    ws.cell(row=row_num, column=2).fill = PatternFill(
                        start_color=color, end_color=color, fill_type="solid"
                    )

                    # Auto-size columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, AttributeError):
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

            # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    def _generate_filename(
        self, request: BlockAssignmentExportRequest, extension: str
    ) -> str:
        """Generate export filename."""
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")

        # Build filename parts
        parts = [f"block_assignments_{request.academic_year}"]

        if request.block_numbers:
            if len(request.block_numbers) == 1:
                parts.append(f"block{request.block_numbers[0]}")
            else:
                parts.append(f"{len(request.block_numbers)}blocks")

        parts.append(timestamp)

        return "_".join(parts) + f".{extension}"

        # Factory function for dependency injection


def get_block_assignment_export_service(
    session: AsyncSession,
) -> BlockAssignmentExportService:
    """Get BlockAssignmentExportService instance."""
    return BlockAssignmentExportService(session)
