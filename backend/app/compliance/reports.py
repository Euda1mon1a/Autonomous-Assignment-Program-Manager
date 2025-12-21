"""
Compliance Report Generator.

Generates comprehensive ACGME compliance reports including:
- Work hour violation summaries
- Supervision ratio analysis
- Leave utilization metrics
- Schedule coverage analysis
- Trend analysis over time
- Export to PDF and Excel formats
"""

import io
from collections import defaultdict
from datetime import date, datetime, timedelta
from typing import Any
from uuid import UUID

import pandas as pd
import plotly.graph_objects as go
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from sqlalchemy import func, select
from sqlalchemy.orm import Session, joinedload

from app.models.absence import Absence
from app.models.assignment import Assignment
from app.models.block import Block
from app.models.person import Person
from app.scheduling.validator import ACGMEValidator


class ComplianceReportData:
    """Data container for compliance report metrics."""

    def __init__(self):
        """Initialize compliance report data structure."""
        self.acgme_violations: list[dict[str, Any]] = []
        self.work_hour_summary: dict[str, Any] = {}
        self.supervision_summary: dict[str, Any] = {}
        self.leave_utilization: dict[str, Any] = {}
        self.coverage_metrics: dict[str, Any] = {}
        self.trend_data: dict[str, list[dict[str, Any]]] = {}
        self.resident_summaries: list[dict[str, Any]] = []
        self.period_start: date | None = None
        self.period_end: date | None = None
        self.generated_at: datetime = datetime.utcnow()


class ComplianceReportGenerator:
    """
    Generates comprehensive ACGME compliance reports.

    This generator creates detailed compliance reports for medical residency
    programs, ensuring ACGME (Accreditation Council for Graduate Medical
    Education) requirements are met.

    Features:
        - ACGME compliance validation (80-hour rule, 1-in-7 rule, supervision)
        - Work hour violation summaries by resident and time period
        - Supervision ratio analysis and trends
        - Leave utilization reports (vacation, medical, deployment)
        - Schedule coverage analysis
        - Trend analysis over time with visualizations
        - Export to PDF (detailed reports) and Excel (data analysis)

    Example:
        >>> generator = ComplianceReportGenerator(db)
        >>> report_data = generator.generate_compliance_data(
        ...     start_date=date(2025, 1, 1),
        ...     end_date=date(2025, 3, 31)
        ... )
        >>> pdf_bytes = generator.export_to_pdf(report_data)
        >>> excel_bytes = generator.export_to_excel(report_data)
    """

    # Constants
    MAX_WEEKLY_HOURS = 80
    HOURS_PER_HALF_DAY = 6
    ROLLING_WINDOW_WEEKS = 4

    def __init__(self, db: Session):
        """
        Initialize compliance report generator.

        Args:
            db: SQLAlchemy database session for querying data
        """
        self.db = db
        self.validator = ACGMEValidator(db)

    def generate_compliance_data(
        self,
        start_date: date,
        end_date: date,
        resident_ids: list[UUID] | None = None,
        pgy_levels: list[int] | None = None,
        include_violations_only: bool = False,
    ) -> ComplianceReportData:
        """
        Generate comprehensive compliance report data.

        This method queries the database and compiles all compliance metrics
        for the specified time period and filters.

        Args:
            start_date: Start date of reporting period
            end_date: End date of reporting period (inclusive)
            resident_ids: Optional list of specific resident IDs to include
            pgy_levels: Optional list of PGY levels to filter (e.g., [1, 2, 3])
            include_violations_only: If True, only include residents with violations

        Returns:
            ComplianceReportData: Complete compliance metrics for the period

        Process:
            1. Query residents based on filters
            2. For each resident:
               - Calculate work hours (by week and rolling 4-week windows)
               - Check ACGME violations (80-hour, 1-in-7, supervision)
               - Analyze leave utilization
            3. Calculate supervision ratios by block
            4. Compute coverage metrics
            5. Generate trend data for time series analysis
            6. Compile summary statistics

        Example:
            >>> data = generator.generate_compliance_data(
            ...     start_date=date(2025, 1, 1),
            ...     end_date=date(2025, 1, 31),
            ...     pgy_levels=[1],  # Only PGY-1 residents
            ...     include_violations_only=True
            ... )
            >>> print(f"Violations found: {len(data.acgme_violations)}")
        """
        report_data = ComplianceReportData()
        report_data.period_start = start_date
        report_data.period_end = end_date

        # Build resident query with filters
        resident_query = self.db.query(Person).filter(Person.type == "resident")

        if resident_ids:
            resident_query = resident_query.filter(Person.id.in_(resident_ids))

        if pgy_levels:
            resident_query = resident_query.filter(Person.pgy_level.in_(pgy_levels))

        residents = resident_query.order_by(Person.pgy_level, Person.name).all()

        # Get all assignments for the period
        assignments = (
            self.db.query(Assignment)
            .options(
                joinedload(Assignment.block),
                joinedload(Assignment.person),
                joinedload(Assignment.rotation_template),
            )
            .join(Block)
            .filter(Block.date >= start_date, Block.date <= end_date)
            .all()
        )

        # Get all absences for the period
        absences = (
            self.db.query(Absence)
            .filter(Absence.start_date <= end_date, Absence.end_date >= start_date)
            .all()
        )

        # Process each resident
        for resident in residents:
            resident_summary = self._analyze_resident(
                resident, assignments, absences, start_date, end_date
            )

            # Filter by violations if requested
            if include_violations_only and not resident_summary["has_violations"]:
                continue

            report_data.resident_summaries.append(resident_summary)

            # Collect violations
            report_data.acgme_violations.extend(resident_summary["violations"])

        # Generate aggregate metrics
        report_data.work_hour_summary = self._calculate_work_hour_summary(
            report_data.resident_summaries
        )
        report_data.supervision_summary = self._calculate_supervision_summary(
            assignments, start_date, end_date
        )
        report_data.leave_utilization = self._calculate_leave_utilization(
            absences, residents, start_date, end_date
        )
        report_data.coverage_metrics = self._calculate_coverage_metrics(
            assignments, start_date, end_date
        )
        report_data.trend_data = self._calculate_trend_data(
            assignments, absences, start_date, end_date
        )

        return report_data

    def _analyze_resident(
        self,
        resident: Person,
        all_assignments: list[Assignment],
        all_absences: list[Absence],
        start_date: date,
        end_date: date,
    ) -> dict[str, Any]:
        """
        Analyze a single resident's compliance metrics.

        Args:
            resident: Resident person object
            all_assignments: All assignments in the period
            all_absences: All absences in the period
            start_date: Period start date
            end_date: Period end date

        Returns:
            dict: Resident summary with metrics and violations
        """
        # Filter assignments for this resident
        resident_assignments = [
            a for a in all_assignments if a.person_id == resident.id
        ]

        # Filter absences for this resident
        resident_absences = [
            a for a in all_absences if a.person_id == resident.id
        ]

        # Calculate work hours by week
        hours_by_week = self._calculate_hours_by_week(
            resident_assignments, start_date, end_date
        )

        # Calculate rolling 4-week averages
        rolling_averages = self._calculate_rolling_averages(resident_assignments)

        # Check for violations
        violations = []

        # 80-hour violations
        violations.extend(
            self.validator._check_80_hour_rule(resident, resident_assignments)
        )

        # 1-in-7 violations
        violations.extend(
            self.validator._check_1_in_7_rule(resident, resident_assignments)
        )

        # Calculate absence statistics
        total_absence_days = sum(a.duration_days for a in resident_absences)
        absence_by_type = defaultdict(int)
        for absence in resident_absences:
            absence_by_type[absence.absence_type] += absence.duration_days

        # Determine max weekly hours
        max_weekly_hours = max(hours_by_week.values()) if hours_by_week else 0
        avg_weekly_hours = (
            sum(hours_by_week.values()) / len(hours_by_week) if hours_by_week else 0
        )

        return {
            "resident_id": str(resident.id),
            "resident_name": resident.name,
            "pgy_level": resident.pgy_level,
            "total_assignments": len(resident_assignments),
            "total_hours": len(resident_assignments) * self.HOURS_PER_HALF_DAY,
            "avg_weekly_hours": avg_weekly_hours,
            "max_weekly_hours": max_weekly_hours,
            "rolling_averages": rolling_averages,
            "hours_by_week": hours_by_week,
            "total_absence_days": total_absence_days,
            "absence_by_type": dict(absence_by_type),
            "violations": [
                {
                    "type": v.type,
                    "severity": v.severity,
                    "message": v.message,
                    "details": v.details,
                }
                for v in violations
            ],
            "has_violations": len(violations) > 0,
            "violation_count": len(violations),
        }

    def _calculate_hours_by_week(
        self, assignments: list[Assignment], start_date: date, end_date: date
    ) -> dict[str, float]:
        """
        Calculate work hours grouped by week.

        Args:
            assignments: List of assignments
            start_date: Period start
            end_date: Period end

        Returns:
            dict: Week start date (ISO format) -> hours worked
        """
        hours_by_week = defaultdict(float)

        for assignment in assignments:
            if not assignment.block:
                continue

            block_date = assignment.block.date
            if block_date < start_date or block_date > end_date:
                continue

            # Get Monday of the week (ISO week start)
            week_start = block_date - timedelta(days=block_date.weekday())
            week_key = week_start.isoformat()

            hours_by_week[week_key] += self.HOURS_PER_HALF_DAY

        return dict(hours_by_week)

    def _calculate_rolling_averages(
        self, assignments: list[Assignment]
    ) -> list[dict[str, Any]]:
        """
        Calculate rolling 4-week average work hours.

        Args:
            assignments: List of assignments

        Returns:
            list: Rolling window data with window_start, window_end, avg_weekly_hours
        """
        if not assignments:
            return []

        # Get hours by date
        hours_by_date = defaultdict(float)
        for assignment in assignments:
            if assignment.block:
                hours_by_date[assignment.block.date] += self.HOURS_PER_HALF_DAY

        if not hours_by_date:
            return []

        dates = sorted(hours_by_date.keys())
        rolling_data = []
        window_days = self.ROLLING_WINDOW_WEEKS * 7

        # Calculate rolling averages
        for i in range(len(dates)):
            window_start = dates[i]
            window_end = window_start + timedelta(days=window_days - 1)

            # Sum hours in window
            total_hours = sum(
                hours
                for d, hours in hours_by_date.items()
                if window_start <= d <= window_end
            )

            avg_weekly = total_hours / self.ROLLING_WINDOW_WEEKS

            rolling_data.append(
                {
                    "window_start": window_start.isoformat(),
                    "window_end": window_end.isoformat(),
                    "total_hours": total_hours,
                    "avg_weekly_hours": round(avg_weekly, 1),
                    "exceeds_limit": avg_weekly > self.MAX_WEEKLY_HOURS,
                }
            )

        return rolling_data

    def _calculate_work_hour_summary(
        self, resident_summaries: list[dict[str, Any]]
    ) -> dict[str, Any]:
        """
        Calculate aggregate work hour summary statistics.

        Args:
            resident_summaries: List of resident summary dicts

        Returns:
            dict: Aggregate work hour statistics
        """
        if not resident_summaries:
            return {
                "total_residents": 0,
                "residents_with_violations": 0,
                "total_violations": 0,
                "avg_weekly_hours": 0,
                "max_weekly_hours": 0,
                "compliance_rate": 100.0,
            }

        total_residents = len(resident_summaries)
        residents_with_violations = sum(
            1 for r in resident_summaries if r["has_violations"]
        )
        total_violations = sum(r["violation_count"] for r in resident_summaries)

        avg_weekly_hours = (
            sum(r["avg_weekly_hours"] for r in resident_summaries) / total_residents
        )
        max_weekly_hours = max(r["max_weekly_hours"] for r in resident_summaries)

        compliance_rate = (
            (total_residents - residents_with_violations) / total_residents * 100
        )

        return {
            "total_residents": total_residents,
            "residents_with_violations": residents_with_violations,
            "total_violations": total_violations,
            "avg_weekly_hours": round(avg_weekly_hours, 1),
            "max_weekly_hours": round(max_weekly_hours, 1),
            "compliance_rate": round(compliance_rate, 1),
        }

    def _calculate_supervision_summary(
        self, assignments: list[Assignment], start_date: date, end_date: date
    ) -> dict[str, Any]:
        """
        Calculate supervision ratio summary.

        Args:
            assignments: All assignments
            start_date: Period start
            end_date: Period end

        Returns:
            dict: Supervision ratio statistics
        """
        # Group by block
        assignments_by_block = defaultdict(list)
        for assignment in assignments:
            if assignment.block:
                assignments_by_block[assignment.block_id].append(assignment)

        total_blocks = len(assignments_by_block)
        blocks_with_violations = 0
        supervision_violations = []

        for block_id, block_assignments in assignments_by_block.items():
            residents = []
            faculty = []

            for assignment in block_assignments:
                person = assignment.person
                if not person:
                    continue

                if person.type == "resident":
                    residents.append(person)
                elif person.type == "faculty":
                    faculty.append(person)

            if not residents:
                continue

            # Calculate required faculty
            pgy1_count = sum(1 for r in residents if r.pgy_level == 1)
            other_count = len(residents) - pgy1_count

            # 1:2 for PGY-1, 1:4 for others
            required = max(1, (pgy1_count + 1) // 2 + (other_count + 3) // 4)

            if len(faculty) < required:
                blocks_with_violations += 1
                block = (
                    self.db.query(Block).filter(Block.id == block_id).first()
                )
                supervision_violations.append(
                    {
                        "block_id": str(block_id),
                        "block_date": block.date.isoformat() if block else None,
                        "residents": len(residents),
                        "pgy1_count": pgy1_count,
                        "faculty": len(faculty),
                        "required_faculty": required,
                        "deficit": required - len(faculty),
                    }
                )

        compliance_rate = (
            (total_blocks - blocks_with_violations) / total_blocks * 100
            if total_blocks > 0
            else 100.0
        )

        return {
            "total_blocks": total_blocks,
            "blocks_with_violations": blocks_with_violations,
            "supervision_violations": supervision_violations,
            "compliance_rate": round(compliance_rate, 1),
        }

    def _calculate_leave_utilization(
        self,
        absences: list[Absence],
        residents: list[Person],
        start_date: date,
        end_date: date,
    ) -> dict[str, Any]:
        """
        Calculate leave utilization statistics.

        Args:
            absences: All absences
            residents: All residents
            start_date: Period start
            end_date: Period end

        Returns:
            dict: Leave utilization metrics
        """
        total_absence_days = 0
        absence_by_type = defaultdict(int)
        absence_by_person = defaultdict(int)

        for absence in absences:
            # Calculate overlapping days within period
            overlap_start = max(absence.start_date, start_date)
            overlap_end = min(absence.end_date, end_date)

            if overlap_start <= overlap_end:
                days = (overlap_end - overlap_start).days + 1
                total_absence_days += days
                absence_by_type[absence.absence_type] += days
                absence_by_person[str(absence.person_id)] += days

        # Calculate average per resident
        avg_absence_days = (
            total_absence_days / len(residents) if residents else 0
        )

        # Period length in days
        period_days = (end_date - start_date).days + 1

        # Absence rate as percentage
        total_possible_days = len(residents) * period_days
        absence_rate = (
            total_absence_days / total_possible_days * 100
            if total_possible_days > 0
            else 0
        )

        return {
            "total_absence_days": total_absence_days,
            "avg_absence_days_per_resident": round(avg_absence_days, 1),
            "absence_rate_percent": round(absence_rate, 1),
            "absence_by_type": dict(absence_by_type),
            "most_common_type": (
                max(absence_by_type.items(), key=lambda x: x[1])[0]
                if absence_by_type
                else None
            ),
        }

    def _calculate_coverage_metrics(
        self, assignments: list[Assignment], start_date: date, end_date: date
    ) -> dict[str, Any]:
        """
        Calculate schedule coverage metrics.

        Args:
            assignments: All assignments
            start_date: Period start
            end_date: Period end

        Returns:
            dict: Coverage statistics
        """
        # Get total workday blocks (exclude weekends)
        total_workday_blocks = (
            self.db.query(Block)
            .filter(
                Block.date >= start_date,
                Block.date <= end_date,
                Block.is_weekend == False,  # noqa: E712
            )
            .count()
        )

        # Count assigned blocks
        assigned_blocks = len({a.block_id for a in assignments})

        # Coverage rate
        coverage_rate = (
            assigned_blocks / total_workday_blocks * 100
            if total_workday_blocks > 0
            else 0
        )

        # Group by role
        assignments_by_role = defaultdict(int)
        for assignment in assignments:
            assignments_by_role[assignment.role] += 1

        return {
            "total_workday_blocks": total_workday_blocks,
            "assigned_blocks": assigned_blocks,
            "unassigned_blocks": total_workday_blocks - assigned_blocks,
            "coverage_rate_percent": round(coverage_rate, 1),
            "assignments_by_role": dict(assignments_by_role),
        }

    def _calculate_trend_data(
        self,
        assignments: list[Assignment],
        absences: list[Absence],
        start_date: date,
        end_date: date,
    ) -> dict[str, list[dict[str, Any]]]:
        """
        Calculate trend data for time series analysis.

        Args:
            assignments: All assignments
            absences: All absences
            start_date: Period start
            end_date: Period end

        Returns:
            dict: Trend data by category (weekly_hours, coverage, absences)
        """
        trend_data = {
            "weekly_hours": [],
            "coverage": [],
            "absences": [],
        }

        # Calculate weekly trends
        current = start_date
        while current <= end_date:
            week_end = min(current + timedelta(days=6), end_date)

            # Hours this week
            week_assignments = [
                a
                for a in assignments
                if a.block and current <= a.block.date <= week_end
            ]
            total_hours = len(week_assignments) * self.HOURS_PER_HALF_DAY

            # Coverage this week
            week_blocks = (
                self.db.query(Block)
                .filter(
                    Block.date >= current,
                    Block.date <= week_end,
                    Block.is_weekend == False,  # noqa: E712
                )
                .count()
            )
            assigned = len({a.block_id for a in week_assignments})
            coverage_rate = assigned / week_blocks * 100 if week_blocks > 0 else 0

            # Absences this week
            week_absences = [
                a
                for a in absences
                if a.start_date <= week_end and a.end_date >= current
            ]
            absence_days = sum(
                min(a.end_date, week_end).toordinal()
                - max(a.start_date, current).toordinal()
                + 1
                for a in week_absences
            )

            trend_data["weekly_hours"].append(
                {
                    "week_start": current.isoformat(),
                    "week_end": week_end.isoformat(),
                    "total_hours": total_hours,
                }
            )

            trend_data["coverage"].append(
                {
                    "week_start": current.isoformat(),
                    "week_end": week_end.isoformat(),
                    "coverage_rate": round(coverage_rate, 1),
                }
            )

            trend_data["absences"].append(
                {
                    "week_start": current.isoformat(),
                    "week_end": week_end.isoformat(),
                    "absence_days": absence_days,
                }
            )

            current = week_end + timedelta(days=1)

        return trend_data

    def export_to_pdf(
        self,
        report_data: ComplianceReportData,
        include_charts: bool = True,
        include_details: bool = True,
    ) -> bytes:
        """
        Export compliance report to PDF format.

        Creates a comprehensive PDF report with summary statistics,
        detailed violations, and optional trend charts.

        Args:
            report_data: Compiled compliance data
            include_charts: Whether to include trend charts
            include_details: Whether to include detailed violation listings

        Returns:
            bytes: PDF file content

        PDF Structure:
            1. Cover page with period and summary
            2. Executive summary (compliance rates, key metrics)
            3. ACGME violations section (if any)
            4. Work hour analysis
            5. Supervision ratio analysis
            6. Leave utilization summary
            7. Coverage metrics
            8. Trend charts (if enabled)
            9. Detailed resident summaries (if enabled)

        Example:
            >>> pdf_bytes = generator.export_to_pdf(report_data)
            >>> with open("compliance_report.pdf", "wb") as f:
            ...     f.write(pdf_bytes)
        """
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        story = []
        styles = getSampleStyleSheet()

        # Custom styles
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=24,
            textColor=colors.HexColor("#1f4788"),
            spaceAfter=30,
        )

        heading_style = ParagraphStyle(
            "CustomHeading",
            parent=styles["Heading2"],
            fontSize=16,
            textColor=colors.HexColor("#1f4788"),
            spaceAfter=12,
        )

        # Title
        story.append(Paragraph("ACGME Compliance Report", title_style))
        story.append(Spacer(1, 0.2 * inch))

        # Period
        period_text = (
            f"<b>Reporting Period:</b> {report_data.period_start} to "
            f"{report_data.period_end}"
        )
        story.append(Paragraph(period_text, styles["Normal"]))

        # Generated timestamp
        gen_text = f"<b>Generated:</b> {report_data.generated_at.strftime('%Y-%m-%d %H:%M UTC')}"
        story.append(Paragraph(gen_text, styles["Normal"]))
        story.append(Spacer(1, 0.3 * inch))

        # Executive Summary
        story.append(Paragraph("Executive Summary", heading_style))

        summary_data = [
            ["Metric", "Value"],
            [
                "Total Residents",
                str(report_data.work_hour_summary.get("total_residents", 0)),
            ],
            [
                "Compliance Rate",
                f"{report_data.work_hour_summary.get('compliance_rate', 100)}%",
            ],
            [
                "Total Violations",
                str(report_data.work_hour_summary.get("total_violations", 0)),
            ],
            [
                "Avg Weekly Hours",
                f"{report_data.work_hour_summary.get('avg_weekly_hours', 0)} hrs",
            ],
            [
                "Coverage Rate",
                f"{report_data.coverage_metrics.get('coverage_rate_percent', 0)}%",
            ],
        ]

        summary_table = Table(summary_data, colWidths=[3 * inch, 2 * inch])
        summary_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 12),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )
        story.append(summary_table)
        story.append(Spacer(1, 0.3 * inch))

        # ACGME Violations
        if report_data.acgme_violations:
            story.append(PageBreak())
            story.append(Paragraph("ACGME Violations", heading_style))

            for violation in report_data.acgme_violations[:20]:  # Limit to 20
                viol_text = (
                    f"<b>{violation['type']}</b> ({violation['severity']}): "
                    f"{violation['message']}"
                )
                story.append(Paragraph(viol_text, styles["Normal"]))
                story.append(Spacer(1, 0.1 * inch))

        # Work Hour Summary
        story.append(PageBreak())
        story.append(Paragraph("Work Hour Analysis", heading_style))

        wh_text = f"""
        The work hour analysis covers {report_data.work_hour_summary.get('total_residents', 0)} residents
        over the reporting period. The average weekly hours worked was
        {report_data.work_hour_summary.get('avg_weekly_hours', 0)} hours, with a maximum of
        {report_data.work_hour_summary.get('max_weekly_hours', 0)} hours in any given week.
        """
        story.append(Paragraph(wh_text, styles["Normal"]))
        story.append(Spacer(1, 0.2 * inch))

        # Supervision Summary
        story.append(Paragraph("Supervision Ratio Analysis", heading_style))

        sup_text = f"""
        Analyzed {report_data.supervision_summary.get('total_blocks', 0)} clinical blocks.
        Supervision compliance rate: {report_data.supervision_summary.get('compliance_rate', 100)}%.
        Violations found in {report_data.supervision_summary.get('blocks_with_violations', 0)} blocks.
        """
        story.append(Paragraph(sup_text, styles["Normal"]))
        story.append(Spacer(1, 0.2 * inch))

        # Leave Utilization
        story.append(Paragraph("Leave Utilization", heading_style))

        leave_text = f"""
        Total absence days: {report_data.leave_utilization.get('total_absence_days', 0)}.
        Average per resident: {report_data.leave_utilization.get('avg_absence_days_per_resident', 0)} days.
        Absence rate: {report_data.leave_utilization.get('absence_rate_percent', 0)}%.
        """
        story.append(Paragraph(leave_text, styles["Normal"]))
        story.append(Spacer(1, 0.2 * inch))

        # Coverage Metrics
        story.append(Paragraph("Schedule Coverage", heading_style))

        cov_text = f"""
        Total workday blocks: {report_data.coverage_metrics.get('total_workday_blocks', 0)}.
        Assigned blocks: {report_data.coverage_metrics.get('assigned_blocks', 0)}.
        Coverage rate: {report_data.coverage_metrics.get('coverage_rate_percent', 0)}%.
        """
        story.append(Paragraph(cov_text, styles["Normal"]))

        # Detailed Resident Summaries (if enabled)
        if include_details and report_data.resident_summaries:
            story.append(PageBreak())
            story.append(Paragraph("Detailed Resident Summaries", heading_style))

            for resident in report_data.resident_summaries[:10]:  # Limit to 10
                res_text = f"""
                <b>{resident['resident_name']}</b> (PGY-{resident['pgy_level']})<br/>
                Total Hours: {resident['total_hours']} | Avg Weekly: {resident['avg_weekly_hours']:.1f}<br/>
                Violations: {resident['violation_count']}<br/>
                Absence Days: {resident['total_absence_days']}
                """
                story.append(Paragraph(res_text, styles["Normal"]))
                story.append(Spacer(1, 0.15 * inch))

        # Build PDF
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()

    def export_to_excel(
        self, report_data: ComplianceReportData, include_charts: bool = True
    ) -> bytes:
        """
        Export compliance report to Excel format.

        Creates a multi-sheet Excel workbook with detailed data tables
        suitable for further analysis.

        Args:
            report_data: Compiled compliance data
            include_charts: Whether to include embedded charts

        Returns:
            bytes: Excel file content

        Excel Structure:
            - Summary sheet: Key metrics and statistics
            - Violations sheet: All ACGME violations
            - Residents sheet: Per-resident statistics
            - Supervision sheet: Supervision ratio violations
            - Leave sheet: Leave utilization by type and person
            - Coverage sheet: Coverage metrics by week
            - Trends sheet: Time series data

        Example:
            >>> excel_bytes = generator.export_to_excel(report_data)
            >>> with open("compliance_report.xlsx", "wb") as f:
            ...     f.write(excel_bytes)
        """
        wb = Workbook()
        wb.remove(wb.active)

        # Styling
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(
            start_color="1F4788", end_color="1F4788", fill_type="solid"
        )
        header_font_white = Font(bold=True, size=11, color="FFFFFF")

        # Summary Sheet
        ws_summary = wb.create_sheet("Summary")
        ws_summary.append(["ACGME Compliance Report Summary"])
        ws_summary.append([])
        ws_summary.append(
            [
                "Period:",
                f"{report_data.period_start} to {report_data.period_end}",
            ]
        )
        ws_summary.append(
            ["Generated:", report_data.generated_at.strftime("%Y-%m-%d %H:%M UTC")]
        )
        ws_summary.append([])

        ws_summary.append(["Metric", "Value"])
        ws_summary["A6"].font = header_font_white
        ws_summary["B6"].font = header_font_white
        ws_summary["A6"].fill = header_fill
        ws_summary["B6"].fill = header_fill

        ws_summary.append(
            [
                "Total Residents",
                report_data.work_hour_summary.get("total_residents", 0),
            ]
        )
        ws_summary.append(
            [
                "Compliance Rate (%)",
                report_data.work_hour_summary.get("compliance_rate", 100),
            ]
        )
        ws_summary.append(
            [
                "Total Violations",
                report_data.work_hour_summary.get("total_violations", 0),
            ]
        )
        ws_summary.append(
            [
                "Avg Weekly Hours",
                report_data.work_hour_summary.get("avg_weekly_hours", 0),
            ]
        )
        ws_summary.append(
            [
                "Max Weekly Hours",
                report_data.work_hour_summary.get("max_weekly_hours", 0),
            ]
        )
        ws_summary.append(
            [
                "Coverage Rate (%)",
                report_data.coverage_metrics.get("coverage_rate_percent", 0),
            ]
        )

        ws_summary.column_dimensions["A"].width = 25
        ws_summary.column_dimensions["B"].width = 20

        # Violations Sheet
        ws_violations = wb.create_sheet("Violations")
        ws_violations.append(["Type", "Severity", "Message", "Details"])

        for cell in ["A1", "B1", "C1", "D1"]:
            ws_violations[cell].font = header_font_white
            ws_violations[cell].fill = header_fill

        for violation in report_data.acgme_violations:
            ws_violations.append(
                [
                    violation["type"],
                    violation["severity"],
                    violation["message"],
                    str(violation.get("details", "")),
                ]
            )

        ws_violations.column_dimensions["A"].width = 20
        ws_violations.column_dimensions["B"].width = 12
        ws_violations.column_dimensions["C"].width = 50
        ws_violations.column_dimensions["D"].width = 30

        # Residents Sheet
        ws_residents = wb.create_sheet("Residents")
        ws_residents.append(
            [
                "Name",
                "PGY",
                "Total Hours",
                "Avg Weekly",
                "Max Weekly",
                "Violations",
                "Absence Days",
            ]
        )

        for cell in ["A1", "B1", "C1", "D1", "E1", "F1", "G1"]:
            ws_residents[cell].font = header_font_white
            ws_residents[cell].fill = header_fill

        for resident in report_data.resident_summaries:
            ws_residents.append(
                [
                    resident["resident_name"],
                    resident["pgy_level"],
                    resident["total_hours"],
                    round(resident["avg_weekly_hours"], 1),
                    round(resident["max_weekly_hours"], 1),
                    resident["violation_count"],
                    resident["total_absence_days"],
                ]
            )

        ws_residents.column_dimensions["A"].width = 25
        for col in ["B", "C", "D", "E", "F", "G"]:
            ws_residents.column_dimensions[col].width = 12

        # Supervision Sheet
        ws_supervision = wb.create_sheet("Supervision")
        ws_supervision.append(
            ["Block Date", "Residents", "PGY-1", "Faculty", "Required", "Deficit"]
        )

        for cell in ["A1", "B1", "C1", "D1", "E1", "F1"]:
            ws_supervision[cell].font = header_font_white
            ws_supervision[cell].fill = header_fill

        for violation in report_data.supervision_summary.get(
            "supervision_violations", []
        ):
            ws_supervision.append(
                [
                    violation.get("block_date", ""),
                    violation["residents"],
                    violation["pgy1_count"],
                    violation["faculty"],
                    violation["required_faculty"],
                    violation["deficit"],
                ]
            )

        for col in ["A", "B", "C", "D", "E", "F"]:
            ws_supervision.column_dimensions[col].width = 15

        # Leave Sheet
        ws_leave = wb.create_sheet("Leave Utilization")
        ws_leave.append(["Absence Type", "Total Days"])

        for cell in ["A1", "B1"]:
            ws_leave[cell].font = header_font_white
            ws_leave[cell].fill = header_fill

        for absence_type, days in report_data.leave_utilization.get(
            "absence_by_type", {}
        ).items():
            ws_leave.append([absence_type, days])

        ws_leave.column_dimensions["A"].width = 20
        ws_leave.column_dimensions["B"].width = 12

        # Trends Sheet
        ws_trends = wb.create_sheet("Trends")
        ws_trends.append(["Week Start", "Week End", "Total Hours", "Coverage %", "Absence Days"])

        for cell in ["A1", "B1", "C1", "D1", "E1"]:
            ws_trends[cell].font = header_font_white
            ws_trends[cell].fill = header_fill

        # Combine trend data
        hours_trends = report_data.trend_data.get("weekly_hours", [])
        coverage_trends = report_data.trend_data.get("coverage", [])
        absence_trends = report_data.trend_data.get("absences", [])

        for i in range(len(hours_trends)):
            ws_trends.append(
                [
                    hours_trends[i]["week_start"],
                    hours_trends[i]["week_end"],
                    hours_trends[i]["total_hours"],
                    coverage_trends[i]["coverage_rate"] if i < len(coverage_trends) else 0,
                    absence_trends[i]["absence_days"] if i < len(absence_trends) else 0,
                ]
            )

        for col in ["A", "B", "C", "D", "E"]:
            ws_trends.column_dimensions[col].width = 15

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def generate_trend_chart(
        self, report_data: ComplianceReportData, chart_type: str = "weekly_hours"
    ) -> bytes:
        """
        Generate a trend chart using Plotly.

        Args:
            report_data: Compliance report data
            chart_type: Type of chart (weekly_hours, coverage, absences)

        Returns:
            bytes: PNG image of the chart
        """
        trend_data = report_data.trend_data.get(chart_type, [])

        if not trend_data:
            # Return empty image
            return b""

        # Extract data based on chart type
        if chart_type == "weekly_hours":
            x = [d["week_start"] for d in trend_data]
            y = [d["total_hours"] for d in trend_data]
            title = "Weekly Work Hours Trend"
            y_label = "Total Hours"
        elif chart_type == "coverage":
            x = [d["week_start"] for d in trend_data]
            y = [d["coverage_rate"] for d in trend_data]
            title = "Schedule Coverage Trend"
            y_label = "Coverage Rate (%)"
        elif chart_type == "absences":
            x = [d["week_start"] for d in trend_data]
            y = [d["absence_days"] for d in trend_data]
            title = "Weekly Absence Days Trend"
            y_label = "Absence Days"
        else:
            return b""

        # Create Plotly figure
        fig = go.Figure()
        fig.add_trace(go.Scatter(x=x, y=y, mode="lines+markers", name=chart_type))

        fig.update_layout(
            title=title,
            xaxis_title="Week",
            yaxis_title=y_label,
            template="plotly_white",
        )

        # Convert to PNG bytes
        img_bytes = fig.to_image(format="png")
        return img_bytes
