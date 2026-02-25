"""Celery tasks for Excel import processing."""

import base64
import io
import logging
from typing import cast
from uuid import UUID, uuid4

from celery import shared_task
from openpyxl import load_workbook

from app.db.session import task_session_scope
from app.models.import_staging import (
    ImportBatch,
    ImportBatchStatus,
    ConflictResolutionMode,
)
from app.services.excel_metadata import read_sys_meta
from app.services.longitudinal_validator import run_longitudinal_validation

logger = logging.getLogger(__name__)


@shared_task(bind=True, name="app.tasks.import_tasks.process_yearly_workbook")
def process_yearly_workbook(
    self,
    file_bytes: bytes,
    academic_year: int,
    user_id: str,
    conflict_resolution: str = "upsert",
) -> str:
    """
    Parse a 14-sheet master workbook into staged assignments.

    Creates a parent batch for the academic year and child batches for each block.
    """
    with task_session_scope() as db:
        try:
            decoded_file_bytes = base64.b64decode(file_bytes)
            wb = load_workbook(io.BytesIO(decoded_file_bytes), data_only=True)
            meta = read_sys_meta(wb)

            if not meta or not meta.block_map:
                raise ValueError(
                    "Workbook is not a valid year-level master schedule (missing block_map)"
                )

            if meta.academic_year != academic_year:
                raise ValueError(
                    f"Workbook is for AY {meta.academic_year}, expected {academic_year}"
                )

            # 1. Create Parent Batch
            parent_batch = ImportBatch(
                id=uuid4(),
                academic_year=academic_year,
                status=ImportBatchStatus.STAGED,
                created_by_id=UUID(user_id),
                filename="master_schedule.xlsx",
                notes=f"Yearly master workbook for AY {academic_year}",
                conflict_resolution=ConflictResolutionMode(conflict_resolution),
            )
            db.add(parent_batch)
            db.flush()

            from app.services.half_day_import_service import HalfDayImportService
            from app.services.block_generation_service import BlockGenerationService

            import_service = HalfDayImportService(db)
            block_gen_service = BlockGenerationService(db)

            # Ensure all academic blocks exist for this year
            academic_blocks = block_gen_service.ensure_academic_blocks_exist(
                academic_year
            )
            blocks_by_uuid = {str(b.id): b for b in academic_blocks}

            block_map = meta.block_map

            for sheet_name in wb.sheetnames:
                if sheet_name.startswith("__"):
                    continue

                block_uuid = block_map.get(sheet_name)
                if not block_uuid:
                    logger.warning(
                        f"Sheet '{sheet_name}' not found in block_map, skipping"
                    )
                    continue

                # 2. Resolve block by UUID from metadata and ensure daily blocks exist
                try:
                    normalized_block_uuid = str(UUID(block_uuid))
                except ValueError:
                    logger.warning(
                        f"Invalid block UUID '{block_uuid}' for sheet '{sheet_name}', skipping"
                    )
                    continue

                target_block_obj = blocks_by_uuid.get(normalized_block_uuid)
                if not target_block_obj:
                    logger.warning(
                        f"Academic block {normalized_block_uuid} not found, skipping sheet '{sheet_name}'"
                    )
                    continue
                block_num = int(target_block_obj.block_number)

                # Ensure half-day blocks exist for this block's date range
                block_gen_service.generate_daily_blocks(
                    target_block_obj.start_date, target_block_obj.end_date, block_num
                )

                # Create child batch for this block
                child_batch = ImportBatch(
                    id=uuid4(),
                    parent_batch_id=parent_batch.id,
                    target_block=block_num,
                    status=ImportBatchStatus.STAGED,
                    created_by_id=UUID(user_id),
                    filename=f"{sheet_name}.xlsx",
                    notes="Staged from master workbook",
                    conflict_resolution=ConflictResolutionMode(conflict_resolution),
                )
                db.add(child_batch)
                db.flush()

                # 3. Parse and stage assignments for this sheet using HalfDayImportService
                import_service.stage_block_sheet(
                    wb=wb,
                    sheet_name=sheet_name,
                    block_number=block_num,
                    academic_year=academic_year,
                    batch_id=cast(UUID, child_batch.id),
                )

            # 4. Run Longitudinal Validation (Cross-Block ACGME)
            run_longitudinal_validation(db, cast(UUID, parent_batch.id))

            return str(parent_batch.id)

        except Exception as e:
            logger.exception(f"Yearly workbook processing failed: {e}")
            raise
