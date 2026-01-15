#!/usr/bin/env python3
"""
Validate xlsx export against ROSETTA xlsx ground truth.

This is the final validation checkpoint in the central dogma pipeline:
    DB → XML (validated) → xlsx → compare to ROSETTA.xlsx

Usage:
    cd backend
    python scripts/validate_xlsx_vs_rosetta.py
"""

import sys
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

# Add backend to path
sys.path.insert(0, str(Path(__file__).parent.parent))


# ═══════════════════════════════════════════════════════════════════════════════
# ROSETTA XLSX VALIDATION
# ═══════════════════════════════════════════════════════════════════════════════

# Path to ROSETTA xlsx ground truth
ROSETTA_XLSX_PATH = (
    Path(__file__).parent.parent
    / "docs"
    / "scheduling"
    / "Block10_ROSETTA_CORRECT.xlsx"
)

# ROSETTA validation format structure
ROSETTA_SHEET_NAME = "Block10_CORRECT"
GENERATED_SHEET_NAME = "Block10_GENERATED"
COL_RESIDENT_NAME = 1  # Resident name column
COL_PGY = 2  # PGY level
COL_ROTATION1 = 3  # First rotation
COL_ROTATION2 = 4  # Second rotation (mid-block)
COL_NOTES = 5  # Notes
COL_SCHEDULE_START = 6  # First schedule column
COL_SCHEDULE_END = 61  # Last schedule column (28 days × 2)


def compare_xlsx(
    generated_path: Path | str,
    rosetta_path: Path | str,
) -> list[dict]:
    """
    Compare generated xlsx to ROSETTA xlsx cell-by-cell.

    Compares schedule columns (6-61) for each resident.
    Residents are matched by name (column 1).

    Args:
        generated_path: Path to generated xlsx
        rosetta_path: Path to ROSETTA ground truth xlsx

    Returns:
        List of mismatch dicts with row, col, expected, actual
    """
    gen_wb = load_workbook(generated_path, data_only=True)
    ros_wb = load_workbook(rosetta_path, data_only=True)

    if GENERATED_SHEET_NAME not in gen_wb.sheetnames:
        # Try first sheet
        gen_sheet = gen_wb.active
    else:
        gen_sheet = gen_wb[GENERATED_SHEET_NAME]

    if ROSETTA_SHEET_NAME not in ros_wb.sheetnames:
        raise ValueError(f"Sheet '{ROSETTA_SHEET_NAME}' not found in ROSETTA xlsx")

    ros_sheet = ros_wb[ROSETTA_SHEET_NAME]

    mismatches = []

    # Build lookup of ROSETTA rows by resident name
    rosetta_by_name = {}
    for row in range(2, ros_sheet.max_row + 1):  # Skip header row
        name = ros_sheet.cell(row=row, column=COL_RESIDENT_NAME).value
        if name:
            rosetta_by_name[name.strip()] = row

    # Compare each generated resident against ROSETTA
    for gen_row in range(2, gen_sheet.max_row + 1):  # Skip header row
        gen_name = gen_sheet.cell(row=gen_row, column=COL_RESIDENT_NAME).value
        if not gen_name:
            continue

        gen_name = gen_name.strip()
        ros_row = rosetta_by_name.get(gen_name)

        if ros_row is None:
            mismatches.append(
                {
                    "row": gen_row,
                    "col": 1,
                    "provider": gen_name,
                    "expected": "(not in ROSETTA)",
                    "actual": gen_name,
                }
            )
            continue

        # Compare schedule columns
        for col in range(COL_SCHEDULE_START, COL_SCHEDULE_END + 1):
            gen_val = gen_sheet.cell(row=gen_row, column=col).value
            ros_val = ros_sheet.cell(row=ros_row, column=col).value

            # Normalize for comparison
            gen_str = str(gen_val or "").strip().upper()
            ros_str = str(ros_val or "").strip().upper()

            if gen_str != ros_str:
                mismatches.append(
                    {
                        "row": gen_row,
                        "col": col,
                        "provider": gen_name,
                        "expected": ros_val,
                        "actual": gen_val,
                    }
                )

    return mismatches


def get_mismatch_summary(mismatches: list[dict]) -> dict:
    """Summarize mismatches by provider and time of day."""
    if not mismatches:
        return {"total": 0, "by_provider": {}, "by_time": {"AM": 0, "PM": 0}}

    by_provider = {}
    by_time = {"AM": 0, "PM": 0}

    for m in mismatches:
        # Count by provider
        provider = m.get("provider", "unknown")
        by_provider[provider] = by_provider.get(provider, 0) + 1

        # Count by time (even col = AM, odd col = PM)
        col = m.get("col", 0)
        time = "AM" if col % 2 == 0 else "PM"
        by_time[time] += 1

    return {
        "total": len(mismatches),
        "by_provider": by_provider,
        "by_time": by_time,
        "sample_mismatches": mismatches[:10],
    }


def main():
    """Run xlsx validation against ROSETTA."""
    print("=" * 70)
    print("ROSETTA XLSX VALIDATION")
    print("=" * 70)
    print()

    # Check ROSETTA exists
    if not ROSETTA_XLSX_PATH.exists():
        print(f"ERROR: ROSETTA xlsx not found at {ROSETTA_XLSX_PATH}")
        sys.exit(1)

    print(f"ROSETTA xlsx: {ROSETTA_XLSX_PATH}")
    print()

    # Generate xlsx from XML
    print("Generating xlsx from XML...")

    from app.services.schedule_xml_exporter import ScheduleXMLExporter
    from app.services.xml_to_xlsx_converter import XMLToXlsxConverter

    # Block 10 dates
    BLOCK_10_START = date(2026, 3, 12)
    BLOCK_10_END = date(2026, 4, 8)

    # Block 10 residents (same as XML validation)
    BLOCK_10_RESIDENTS = [
        {
            "name": "Byrnes, Katherine",
            "pgy": 1,
            "rotation1": "Peds NF",
            "rotation2": "Peds Ward",
        },
        {"name": "Monsivais, Joshua", "pgy": 1, "rotation1": "IM", "rotation2": None},
        {"name": "Sawyer, Tessa", "pgy": 1, "rotation1": "FMC", "rotation2": None},
        {"name": "Sloss, Meleighe", "pgy": 1, "rotation1": "PROC", "rotation2": None},
        {"name": "Travis, Colin", "pgy": 1, "rotation1": "KAP", "rotation2": None},
        {
            "name": "Wilhelm, Clara",
            "pgy": 1,
            "rotation1": "Peds Ward",
            "rotation2": "Peds NF",
        },
        {"name": "Chen, Michael", "pgy": 2, "rotation1": "FMC", "rotation2": None},
        {"name": "Headid, Ronald", "pgy": 2, "rotation1": "LDNF", "rotation2": None},
        {"name": "You, Jae", "pgy": 3, "rotation1": "NEURO", "rotation2": "NF"},
    ]

    # Generate XML
    exporter = ScheduleXMLExporter(BLOCK_10_START, BLOCK_10_END)
    xml_string = exporter.export(BLOCK_10_RESIDENTS)
    print(f"Generated XML: {len(xml_string)} chars")

    # Convert to xlsx
    output_path = Path(__file__).parent.parent / "Block10_GENERATED.xlsx"
    converter = XMLToXlsxConverter()
    xlsx_bytes = converter.convert_from_string(xml_string, output_path)
    print(f"Generated xlsx: {len(xlsx_bytes)} bytes")
    print(f"Saved to: {output_path}")
    print()

    # Compare to ROSETTA
    print("Comparing to ROSETTA xlsx...")
    mismatches = compare_xlsx(output_path, ROSETTA_XLSX_PATH)

    if mismatches:
        summary = get_mismatch_summary(mismatches)
        print()
        print("=" * 70)
        print("VALIDATION FAILED")
        print("=" * 70)
        print(f"Total mismatches: {summary['total']}")
        print()
        print("By provider:")
        for name, count in sorted(summary["by_provider"].items()):
            print(f"  {name}: {count}")
        print()
        print("By time of day:")
        print(f"  AM: {summary['by_time']['AM']}")
        print(f"  PM: {summary['by_time']['PM']}")
        print()
        print("Sample mismatches:")
        for m in summary["sample_mismatches"]:
            print(
                f"  Row {m['row']}, Col {m['col']} ({m['provider']}): "
                f"expected '{m['expected']}', got '{m['actual']}'"
            )
        print()
        print("=" * 70)
        print("FIX THESE DISCREPANCIES BEFORE PROCEEDING")
        print("=" * 70)
        sys.exit(1)
    else:
        print()
        print("=" * 70)
        print("VALIDATION PASSED")
        print("=" * 70)
        print("Generated xlsx matches ROSETTA xlsx!")
        print("Central dogma pipeline validated end-to-end.")
        print("=" * 70)


if __name__ == "__main__":
    main()
