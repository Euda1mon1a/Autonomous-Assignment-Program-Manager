"""Generate BlockTemplate2_Official.xlsx ("Fat Template") for canonical export.

Creates a pre-formatted Excel workbook with oversized row bands:
  - Rows 1-3: Header (dates, day names, AM/PM labels — filled at export time)
  - Row 4: Staff Call (faculty on-call)
  - Row 5: Resident Call
  - Rows 6-8: Section dividers / blank
  - Rows 9-30: Resident band (22 slots, PGY 3→2→1)
  - Rows 31-80: Faculty band (50 slots)
  - Rows 81+: (unused)

Summary formula columns BJ-BR are NOT pre-filled here because they depend
on actual data layout at export time. The converter writes those.

Usage:
    python -m scripts.generate_fat_template
    # or
    python scripts/generate_fat_template.py
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ── Constants matching xml_to_xlsx_converter.py ──
BT2_COL_ROTATION1 = 1
BT2_COL_ROTATION2 = 2
BT2_COL_TEMPLATE = 3
BT2_COL_ROLE = 4
BT2_COL_NAME = 5
BT2_COL_SCHEDULE_START = 6
COLS_PER_DAY = 2
TOTAL_DAYS = 28
TOTAL_SCHEDULE_COLS = TOTAL_DAYS * COLS_PER_DAY  # 56

# Row bands
ROW_HEADER_START = 1
ROW_HEADER_END = 3
ROW_STAFF_CALL = 4
ROW_RESIDENT_CALL = 5
ROW_DIVIDER_START = 6
ROW_DIVIDER_END = 8
ROW_RESIDENT_START = 9
ROW_RESIDENT_END = 30
ROW_FACULTY_START = 31
ROW_FACULTY_END = 80

# Colors
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=9)
CALL_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
DIVIDER_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
RESIDENT_ALT_FILL = PatternFill(
    start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
)
FACULTY_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="C0C0C0"),
    right=Side(style="thin", color="C0C0C0"),
    top=Side(style="thin", color="C0C0C0"),
    bottom=Side(style="thin", color="C0C0C0"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)
BODY_FONT = Font(size=8)
NAME_FONT = Font(size=8, bold=True)


def generate() -> Path:
    """Generate the Fat Template and return the output path."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Block Template2"

    last_col = BT2_COL_SCHEDULE_START + TOTAL_SCHEDULE_COLS - 1  # col 61

    # ── Column widths ──
    ws.column_dimensions[get_column_letter(BT2_COL_ROTATION1)].width = 10  # A
    ws.column_dimensions[get_column_letter(BT2_COL_ROTATION2)].width = 10  # B
    ws.column_dimensions[get_column_letter(BT2_COL_TEMPLATE)].width = 5  # C
    ws.column_dimensions[get_column_letter(BT2_COL_ROLE)].width = 6  # D
    ws.column_dimensions[get_column_letter(BT2_COL_NAME)].width = 18  # E
    for c in range(BT2_COL_SCHEDULE_START, last_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 4.5

    # ── Header rows 1-3 ──
    # Row 1: Identity column headers
    headers = {
        BT2_COL_ROTATION1: "Rot 1",
        BT2_COL_ROTATION2: "Rot 2",
        BT2_COL_TEMPLATE: "Tmpl",
        BT2_COL_ROLE: "Role",
        BT2_COL_NAME: "Name",
    }
    for col, label in headers.items():
        cell = ws.cell(row=1, column=col, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # Row 1 schedule area: placeholder "Day 1 AM", "Day 1 PM", etc.
    # (Overwritten at export time by _fill_header_row with actual dates)
    for day in range(TOTAL_DAYS):
        am_col = BT2_COL_SCHEDULE_START + (day * 2)
        pm_col = am_col + 1
        for col in (am_col, pm_col):
            cell = ws.cell(row=1, column=col)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER
            cell.border = THIN_BORDER

    # Rows 2-3: Sub-headers (day names, AM/PM — filled at export time)
    for row in range(2, ROW_HEADER_END + 1):
        for col in range(1, last_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER
            cell.border = THIN_BORDER

    # ── Call rows (4-5) ──
    for row in (ROW_STAFF_CALL, ROW_RESIDENT_CALL):
        label = "Staff Call" if row == ROW_STAFF_CALL else "Res Call"
        ws.cell(row=row, column=BT2_COL_NAME, value=label).font = NAME_FONT
        for col in range(1, last_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = CALL_FILL
            cell.border = THIN_BORDER
            cell.font = BODY_FONT
            cell.alignment = CENTER

    # ── Divider rows (6-8) ──
    for row in range(ROW_DIVIDER_START, ROW_DIVIDER_END + 1):
        for col in range(1, last_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = DIVIDER_FILL
            cell.border = THIN_BORDER

    # ── Resident band (9-30) ──
    for row in range(ROW_RESIDENT_START, ROW_RESIDENT_END + 1):
        use_alt = (row - ROW_RESIDENT_START) % 2 == 1
        for col in range(1, last_col + 1):
            cell = ws.cell(row=row, column=col)
            if use_alt:
                cell.fill = RESIDENT_ALT_FILL
            cell.border = THIN_BORDER
            cell.font = BODY_FONT
            if col == BT2_COL_NAME:
                cell.alignment = LEFT
                cell.font = NAME_FONT
            else:
                cell.alignment = CENTER

    # ── Faculty band (31-80) ──
    for row in range(ROW_FACULTY_START, ROW_FACULTY_END + 1):
        for col in range(1, last_col + 1):
            cell = ws.cell(row=row, column=col)
            if (row - ROW_FACULTY_START) % 2 == 0:
                cell.fill = FACULTY_FILL
            cell.border = THIN_BORDER
            cell.font = BODY_FONT
            if col == BT2_COL_NAME:
                cell.alignment = LEFT
                cell.font = NAME_FONT
            else:
                cell.alignment = CENTER

    # ── Freeze panes: freeze at F4 (scroll schedule independently) ──
    ws.freeze_panes = "F4"

    # ── Print settings ──
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # ── Save ──
    output = (
        Path(__file__).resolve().parents[1] / "data" / "BlockTemplate2_Official.xlsx"
    )
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    print(f"Fat Template generated: {output}")
    print(
        f"  Resident band: rows {ROW_RESIDENT_START}-{ROW_RESIDENT_END} "
        f"({ROW_RESIDENT_END - ROW_RESIDENT_START + 1} slots)"
    )
    print(
        f"  Faculty band:  rows {ROW_FACULTY_START}-{ROW_FACULTY_END} "
        f"({ROW_FACULTY_END - ROW_FACULTY_START + 1} slots)"
    )
    return output


if __name__ == "__main__":
    generate()
