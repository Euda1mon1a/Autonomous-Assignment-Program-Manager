#!/usr/bin/env python3
"""
Validate ROSETTA_COMPLETE.xml against ROSETTA_COMPLETE.xlsx.

This script compares XML and XLSX files cell-by-cell to ensure
they contain the same data. Does NOT validate against scheduling rules.
"""

import sys
from datetime import date, timedelta
from pathlib import Path
from xml.etree import ElementTree

from openpyxl import load_workbook

# Add backend to path
sys.path.insert(0, str(Path(__file__).parent.parent))

ROSETTA_XML = (
    Path(__file__).parent.parent.parent
    / "docs"
    / "scheduling"
    / "Block10_ROSETTA_COMPLETE.xml"
)

ROSETTA_XLSX = (
    Path(__file__).parent.parent.parent
    / "docs"
    / "scheduling"
    / "Block10_ROSETTA_COMPLETE.xlsx"
)

# Block 10 dates
BLOCK_START = date(2026, 3, 12)
BLOCK_END = date(2026, 4, 8)

# XLSX column mapping (Block Template2 format)
# Row 3 = Headers: Rotation 1, Rotation 2, Template, Role, Provider, dates...
# Data rows start at row 9 (after header, call rows, and empty rows)
COL_ROTATION1 = 1  # "Hilo", "NF", "FMC", etc.
COL_ROTATION2 = 2  # Second-half rotation (if mid-block switch)
COL_TEMPLATE = 3  # "R3", "R2", "R1", "C19", "ADJ", etc.
COL_ROLE = 4  # "PGY 3", "PGY 2", "PGY 1", "FAC", etc.
COL_NAME = 5  # Provider name ("Connolly, Laura", etc.)
COL_SCHEDULE_START = 6  # First schedule column (Mar 12 AM)
DATA_ROW_START = 9  # First data row (after headers and call rows)


def parse_xml() -> tuple[dict, dict, dict]:
    """
    Parse ROSETTA XML into structured data.

    Returns:
        Tuple of (residents, faculty, call_schedule)
    """
    tree = ElementTree.parse(ROSETTA_XML)
    root = tree.getroot()

    # Parse call schedule
    call_schedule = {}
    call_elem = root.find("call")
    if call_elem is not None:
        for night in call_elem.findall("night"):
            call_schedule[night.get("date")] = night.get("faculty")

    # Parse residents
    residents = {}
    for elem in root.findall("resident"):
        name = elem.get("name")
        residents[name] = {
            "pgy": elem.get("pgy"),
            "rotation1": elem.get("rotation1", ""),
            "rotation2": elem.get("rotation2", ""),
            "days": {
                day.get("date"): {"am": day.get("am"), "pm": day.get("pm")}
                for day in elem.findall("day")
            },
        }

    # Parse faculty
    faculty = {}
    for elem in root.findall("faculty"):
        name = elem.get("name")
        faculty[name] = {
            "role": elem.get("role", ""),
            "notes": elem.get("notes", ""),
            "days": {
                day.get("date"): {"am": day.get("am"), "pm": day.get("pm")}
                for day in elem.findall("day")
            },
        }

    return residents, faculty, call_schedule


def parse_xlsx() -> tuple[dict, dict, dict]:
    """
    Parse ROSETTA XLSX (Block Template2 format) into structured data.

    Returns:
        Tuple of (residents, faculty, call_schedule)
    """
    wb = load_workbook(ROSETTA_XLSX, data_only=True)
    sheet = wb.active

    residents = {}
    faculty = {}
    call_schedule = {}

    # Build date-to-column mapping
    date_to_cols = {}
    current = BLOCK_START
    col = COL_SCHEDULE_START
    while current <= BLOCK_END:
        date_to_cols[current.isoformat()] = col
        current += timedelta(days=1)
        col += 2  # AM, PM

    # Parse call row (row 4)
    for date_str, start_col in date_to_cols.items():
        call_val = sheet.cell(row=4, column=start_col).value
        if call_val and call_val not in ("Staff Call", "CALL"):
            call_schedule[date_str] = call_val

    # Parse data rows (starting at row 9)
    for row in range(DATA_ROW_START, sheet.max_row + 1):
        name = sheet.cell(row=row, column=COL_NAME).value
        if not name or str(name).strip() == "":
            continue

        role = sheet.cell(row=row, column=COL_ROLE).value or ""
        rotation1 = sheet.cell(row=row, column=COL_ROTATION1).value or ""
        rotation2 = sheet.cell(row=row, column=COL_ROTATION2).value or ""

        # Read schedule
        days = {}
        for date_str, start_col in date_to_cols.items():
            am = sheet.cell(row=row, column=start_col).value or ""
            pm = sheet.cell(row=row, column=start_col + 1).value or ""
            days[date_str] = {"am": str(am), "pm": str(pm)}

        # Determine PGY level from role
        pgy = ""
        if "PGY 1" in str(role) or "PGY1" in str(role):
            pgy = "1"
        elif "PGY 2" in str(role) or "PGY2" in str(role):
            pgy = "2"
        elif "PGY 3" in str(role) or "PGY3" in str(role):
            pgy = "3"

        entry = {
            "pgy": pgy,
            "rotation1": str(rotation1),
            "rotation2": str(rotation2),
            "role": str(role),
            "days": days,
        }

        # Determine if resident or faculty based on role
        if pgy:
            residents[name] = entry
        elif "FAC" in str(role):
            faculty[name] = entry
        # Skip other roles (ADJ, SPEC, etc. that aren't in XML)

    return residents, faculty, call_schedule


def compare_schedules(xml_data: dict, xlsx_data: dict, label: str) -> list[dict]:
    """
    Compare XML and XLSX schedule data.

    Returns list of mismatches.
    """
    mismatches = []

    # Check for names in XML not in XLSX
    for name in xml_data:
        if name not in xlsx_data:
            mismatches.append(
                {
                    "name": name,
                    "issue": "In XML but not in XLSX",
                    "type": "missing",
                }
            )

    # Check for names in XLSX not in XML
    for name in xlsx_data:
        if name not in xml_data:
            mismatches.append(
                {
                    "name": name,
                    "issue": "In XLSX but not in XML",
                    "type": "missing",
                }
            )

    # Compare day-by-day for matching names
    for name in xml_data:
        if name not in xlsx_data:
            continue

        xml_entry = xml_data[name]
        xlsx_entry = xlsx_data[name]

        # Compare each day
        for date_str in xml_entry["days"]:
            xml_day = xml_entry["days"].get(date_str, {})
            xlsx_day = xlsx_entry["days"].get(date_str, {})

            xml_am = xml_day.get("am", "")
            xml_pm = xml_day.get("pm", "")
            xlsx_am = xlsx_day.get("am", "")
            xlsx_pm = xlsx_day.get("pm", "")

            if xml_am != xlsx_am:
                mismatches.append(
                    {
                        "name": name,
                        "date": date_str,
                        "slot": "AM",
                        "xml": xml_am,
                        "xlsx": xlsx_am,
                        "type": "value",
                    }
                )

            if xml_pm != xlsx_pm:
                mismatches.append(
                    {
                        "name": name,
                        "date": date_str,
                        "slot": "PM",
                        "xml": xml_pm,
                        "xlsx": xlsx_pm,
                        "type": "value",
                    }
                )

    return mismatches


def main():
    """Run XML ↔ XLSX validation."""
    print("=" * 70)
    print("ROSETTA XML ↔ XLSX VALIDATION")
    print("=" * 70)
    print()

    # Check files exist
    if not ROSETTA_XML.exists():
        print(f"ERROR: XML file not found: {ROSETTA_XML}")
        sys.exit(1)
    if not ROSETTA_XLSX.exists():
        print(f"ERROR: XLSX file not found: {ROSETTA_XLSX}")
        sys.exit(1)

    # Parse both files
    print("Parsing XML...")
    xml_residents, xml_faculty, xml_call = parse_xml()
    print(f"  Residents: {len(xml_residents)}")
    print(f"  Faculty: {len(xml_faculty)}")
    print(f"  Call nights: {len(xml_call)}")
    print()

    print("Parsing XLSX...")
    xlsx_residents, xlsx_faculty, xlsx_call = parse_xlsx()
    print(f"  Residents: {len(xlsx_residents)}")
    print(f"  Faculty: {len(xlsx_faculty)}")
    print(f"  Call nights: {len(xlsx_call)}")
    print()

    # Compare residents
    print("=" * 70)
    print("RESIDENT COMPARISON")
    print("=" * 70)
    resident_mismatches = compare_schedules(xml_residents, xlsx_residents, "Resident")

    if resident_mismatches:
        print(f"\nFound {len(resident_mismatches)} resident mismatches:")
        by_name = {}
        for m in resident_mismatches:
            by_name.setdefault(m["name"], []).append(m)

        for name in sorted(by_name.keys()):
            items = by_name[name]
            print(f"\n{name}:")
            for m in items[:10]:
                if m["type"] == "missing":
                    print(f"  {m['issue']}")
                else:
                    print(
                        f"  {m['date']} {m['slot']}: XML={m['xml']!r} vs XLSX={m['xlsx']!r}"
                    )
            if len(items) > 10:
                print(f"  ... and {len(items) - 10} more")
    else:
        print("\nAll residents match!")

    # Compare faculty
    print()
    print("=" * 70)
    print("FACULTY COMPARISON")
    print("=" * 70)
    faculty_mismatches = compare_schedules(xml_faculty, xlsx_faculty, "Faculty")

    if faculty_mismatches:
        print(f"\nFound {len(faculty_mismatches)} faculty mismatches:")
        by_name = {}
        for m in faculty_mismatches:
            by_name.setdefault(m["name"], []).append(m)

        for name in sorted(by_name.keys()):
            items = by_name[name]
            print(f"\n{name}:")
            for m in items[:10]:
                if m["type"] == "missing":
                    print(f"  {m['issue']}")
                else:
                    print(
                        f"  {m['date']} {m['slot']}: XML={m['xml']!r} vs XLSX={m['xlsx']!r}"
                    )
            if len(items) > 10:
                print(f"  ... and {len(items) - 10} more")
    else:
        print("\nAll faculty match!")

    # Summary
    print()
    print("=" * 70)
    print("SUMMARY")
    print("=" * 70)
    total = len(resident_mismatches) + len(faculty_mismatches)
    print(f"\nResident mismatches: {len(resident_mismatches)}")
    print(f"Faculty mismatches:  {len(faculty_mismatches)}")
    print(f"TOTAL:               {total}")

    if total == 0:
        print("\n✓ XML and XLSX are in sync!")
    else:
        print("\n✗ XML and XLSX have differences")

    print("=" * 70)


if __name__ == "__main__":
    main()
