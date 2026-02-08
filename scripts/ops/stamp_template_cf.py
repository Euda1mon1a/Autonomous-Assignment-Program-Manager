#!/usr/bin/env python3
"""Stamp TAMC conditional formatting rules into an Excel template.

One-time setup: run this to make the template self-coloring.
Any schedule code written into the grid will auto-color via CF rules.

Usage:
    # Stamp the canonical template (default):
    ./backend/venv/bin/python scripts/ops/stamp_template_cf.py

    # Stamp a specific workbook/sheet (e.g., the handjam):
    ./backend/venv/bin/python scripts/ops/stamp_template_cf.py \
        --workbook ~/Downloads/handjam.xlsx --sheet "Block 10"
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

try:
    from defusedxml import ElementTree
except ImportError:
    from xml.etree import ElementTree  # type: ignore[no-redef]

from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill

REPO_ROOT = Path(__file__).resolve().parents[2]
COLOR_XML = REPO_ROOT / "docs" / "scheduling" / "TAMC_Color_Scheme_Reference.xml"
DEFAULT_TEMPLATE = REPO_ROOT / "backend" / "data" / "BlockTemplate2_Official.xlsx"

SCHEDULE_RANGE = "F9:BI45"
ROTATION_RANGE = "A9:B45"


def _strip_alpha(hex_color: str) -> str:
    """Remove FF alpha prefix: 'FFFFC000' -> 'FFC000'."""
    if hex_color.startswith("FF") and len(hex_color) == 8:
        return hex_color[2:]
    return hex_color


def _parse_color_scheme(xml_path: Path):
    """Parse TAMC XML into (code->fill, code->font, rotation->fill) dicts."""
    tree = ElementTree.parse(str(xml_path))
    root = tree.getroot()

    code_fills: dict[str, str] = {}
    for group in root.findall(".//fill_colors/color_group"):
        hex_c = _strip_alpha(group.get("hex_color", ""))
        for el in group.findall("code"):
            code = el.get("value", "")
            if code:
                code_fills[code] = hex_c

    code_fonts: dict[str, str] = {}
    for fg in root.findall(".//font_colors/font_color_group"):
        name = fg.get("name", "")
        if name in ("theme_dark", "black"):
            continue  # default black — no need for CF font rule
        hex_c = fg.get("hex_color", "")
        if hex_c.startswith("theme:"):
            hex_c = fg.get("actual_rgb", "FF000000")
        hex_c = _strip_alpha(hex_c)
        for el in fg.findall("code"):
            code = el.get("value", "")
            if code:
                code_fonts[code] = hex_c

    rotation_fills: dict[str, str] = {}
    for rt in root.findall(".//rotation_column_colors/rotation_type"):
        hex_c = _strip_alpha(rt.get("hex_color", ""))
        for rot_name in rt.get("rotations", "").split(","):
            rot_name = rot_name.strip()
            if rot_name:
                rotation_fills[rot_name] = hex_c

    return code_fills, code_fonts, rotation_fills


def stamp_cf_rules(wb_path: Path, sheet_name: str | None = None) -> int:
    """Add conditional formatting rules to a workbook. Returns rule count."""
    wb = load_workbook(wb_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    code_fills, code_fonts, rotation_fills = _parse_color_scheme(COLOR_XML)
    count = 0

    # Schedule grid: one CellIsRule per code
    for code, fill_hex in code_fills.items():
        fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
        font_hex = code_fonts.get(code, "000000")
        font = Font(color=font_hex)
        ws.conditional_formatting.add(
            SCHEDULE_RANGE,
            CellIsRule(operator="equal", formula=[f'"{code}"'], fill=fill, font=font),
        )
        count += 1

    # Rotation columns: one rule per rotation name
    for rot_name, fill_hex in rotation_fills.items():
        fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
        font_hex = "FFFFFF" if fill_hex in ("000000", "FF0000", "7030A0") else "000000"
        font = Font(color=font_hex)
        ws.conditional_formatting.add(
            ROTATION_RANGE,
            CellIsRule(operator="equal", formula=[f'"{rot_name}"'], fill=fill, font=font),
        )
        count += 1

    wb.save(wb_path)
    wb.close()
    print(f"Stamped {count} CF rules into {wb_path.name} [{ws.title}]")
    return count


def main() -> int:
    parser = argparse.ArgumentParser(description="Stamp TAMC CF rules into Excel template")
    parser.add_argument(
        "--workbook", type=str, default=str(DEFAULT_TEMPLATE),
        help="Path to workbook (default: BlockTemplate2_Official.xlsx)",
    )
    parser.add_argument("--sheet", type=str, default=None, help="Sheet name (default: active)")
    args = parser.parse_args()

    wb_path = Path(args.workbook).expanduser()
    if not wb_path.exists():
        print(f"ERROR: {wb_path} not found")
        return 1

    stamp_cf_rules(wb_path, args.sheet)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
