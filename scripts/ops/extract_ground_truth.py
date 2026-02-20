"""Extract ground truth from handjam workbook → structured JSON.

Reads every schedule cell from block tabs that have full data (named residents),
producing labeled examples: (person, date, am/pm, display_code, rotation, block).

Handles two column layouts automatically:
  - Blocks 2-6: names in col C (3), schedule cols 4-59
  - Blocks 7+:  names in col E (5), schedule cols 6-61

Usage:
    python scripts/ops/extract_ground_truth.py \
        --workbook ~/Downloads/"Current AY 25-26 pulled 20 FEB 2026.xlsx" \
        --output /tmp/ground_truth.json
"""

from __future__ import annotations

import argparse
import json
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

# Row constants (consistent across all block tabs)
ROW_DATES = 3
ROW_STAFF_CALL = 4
ROW_RES_CALL = 5
COLS_PER_DAY = 2

NBSP = "\xa0"


def _detect_layout(ws) -> dict:
    """Auto-detect column layout by finding where dates start.

    Checks row 3 first (standard), then row 2 (blocks 3-4 variant where
    dates are in row 2 and staff call is in row 3).

    Returns dict with col_name, col_schedule_start, col_schedule_end,
    col_template, col_role, col_rotation1, col_rotation2, and date_row.
    """
    # Find the first date cell — try row 3, then row 2
    first_date_col = None
    date_row = ROW_DATES
    for check_row in [3, 2]:
        for col in range(1, 70):
            v = ws.cell(row=check_row, column=col).value
            if isinstance(v, (datetime, date)):
                first_date_col = col
                date_row = check_row
                break
        if first_date_col is not None:
            break

    if first_date_col is None:
        return None

    # Schedule starts at the first date column
    col_schedule_start = first_date_col
    # 28 days * 2 cols/day = 56 columns
    col_schedule_end = col_schedule_start + 28 * COLS_PER_DAY - 1

    # Name column is one before schedule start
    col_name = col_schedule_start - 1

    # Identity columns are before the name column
    # Layout A (blocks 7+): rot1=1, rot2=2, template=3, role=4, name=5, sched=6
    # Layout B (blocks 2-6): template=1, role=2, name=3, sched=4
    # Determine staff_call and resident_call rows based on date_row
    if date_row == 3:
        staff_call_row = 4
        res_call_row = 5
    else:
        # Dates in row 2 → staff call in row 3, resident call in row 4
        staff_call_row = 3
        res_call_row = 4

    if col_name >= 5:
        # Layout A: 5 identity columns
        return {
            "col_rotation1": col_name - 4,  # 1
            "col_rotation2": col_name - 3,  # 2
            "col_template": col_name - 2,   # 3
            "col_role": col_name - 1,       # 4
            "col_name": col_name,           # 5
            "col_schedule_start": col_schedule_start,  # 6
            "col_schedule_end": col_schedule_end,      # 61
            "date_row": date_row,
            "staff_call_row": staff_call_row,
            "res_call_row": res_call_row,
        }
    else:
        # Layout B: 3 identity columns (no rotation columns in sheet)
        return {
            "col_rotation1": None,
            "col_rotation2": None,
            "col_template": col_name - 2 if col_name >= 3 else None,  # 1
            "col_role": col_name - 1 if col_name >= 2 else None,      # 2
            "col_name": col_name,           # 3
            "col_schedule_start": col_schedule_start,  # 4
            "col_schedule_end": col_schedule_end,      # 59
            "date_row": date_row,
            "staff_call_row": staff_call_row,
            "res_call_row": res_call_row,
        }


def _detect_person_rows(ws, col_name: int, col_template: int | None, col_role: int | None):
    """Scan for resident and faculty rows dynamically.

    Returns (resident_rows, faculty_rows) as lists of row numbers.
    """
    resident_markers = {"R1", "R2", "R3"}
    faculty_markers = {"C19", "C17"}
    role_resident = {"PGY 1", "PGY 2", "PGY 3", "PGY1", "PGY2", "PGY3"}
    role_faculty = {"FAC", "NP", "MD"}

    resident_rows = []
    faculty_rows = []

    for row_idx in range(1, min(ws.max_row + 1, 120)):
        name_val = ws.cell(row=row_idx, column=col_name).value
        if not name_val or not str(name_val).strip() or str(name_val).strip() == NBSP:
            continue
        name_str = str(name_val).strip()
        # Must look like a name (has comma for "Last, First")
        if "," not in name_str:
            continue
        # Skip short all-caps (schedule codes leaked into name col)
        if len(name_str) <= 4 and name_str.upper() == name_str:
            continue

        # Determine type from template/role columns
        template_val = ""
        if col_template:
            tv = ws.cell(row=row_idx, column=col_template).value
            template_val = str(tv).strip().upper() if tv and str(tv).strip() != NBSP else ""

        role_val = ""
        if col_role:
            rv = ws.cell(row=row_idx, column=col_role).value
            role_val = str(rv).strip().upper() if rv and str(rv).strip() != NBSP else ""

        if template_val in faculty_markers or role_val in role_faculty:
            faculty_rows.append(row_idx)
        elif template_val in resident_markers or role_val in role_resident:
            resident_rows.append(row_idx)
        else:
            # Default: resident (conservative)
            resident_rows.append(row_idx)

    return resident_rows, faculty_rows


def extract_workbook(wb_path: Path) -> dict:
    """Extract ground truth from all viable block tabs."""
    wb = load_workbook(wb_path, read_only=True, data_only=True)

    result = {
        "source": str(wb_path.name),
        "blocks": [],
    }

    block_tabs = [s for s in wb.sheetnames if s.startswith("Block ") and s != "Block Schedule"]

    for tab_name in block_tabs:
        ws = wb[tab_name]

        # Auto-detect column layout
        layout = _detect_layout(ws)
        if layout is None:
            continue

        col_name = layout["col_name"]
        col_schedule_start = layout["col_schedule_start"]
        col_schedule_end = layout["col_schedule_end"]
        col_template = layout.get("col_template")
        col_role = layout.get("col_role")
        col_rotation1 = layout.get("col_rotation1")
        col_rotation2 = layout.get("col_rotation2")
        date_row = layout.get("date_row", ROW_DATES)
        staff_call_row = layout.get("staff_call_row", ROW_STAFF_CALL)
        res_call_row = layout.get("res_call_row", ROW_RES_CALL)

        # Detect person rows dynamically
        resident_rows, faculty_rows = _detect_person_rows(ws, col_name, col_template, col_role)
        if not resident_rows and not faculty_rows:
            continue

        # Extract dates from detected date row
        dates = []
        for col in range(col_schedule_start, col_schedule_end + 1, COLS_PER_DAY):
            v = ws.cell(row=date_row, column=col).value
            if isinstance(v, datetime):
                dates.append(v.date())
            elif isinstance(v, date):
                dates.append(v)
            else:
                dates.append(None)

        block_data = {
            "tab_name": tab_name,
            "layout": "A" if col_name >= 5 else "B",
            "dates": [d.isoformat() if d else None for d in dates],
            "num_days": len([d for d in dates if d]),
            "staff_call": [],
            "resident_call": [],
            "people": [],
        }

        # Staff call row
        for i, d in enumerate(dates):
            col = col_schedule_start + i * COLS_PER_DAY
            v = ws.cell(row=staff_call_row, column=col).value
            if v and v != NBSP:
                block_data["staff_call"].append({
                    "date": d.isoformat() if d else None,
                    "day_index": i,
                    "staff": str(v).strip(),
                })

        # Resident call row
        for i, d in enumerate(dates):
            col = col_schedule_start + i * COLS_PER_DAY
            v = ws.cell(row=res_call_row, column=col).value
            if v and v != NBSP:
                block_data["resident_call"].append({
                    "date": d.isoformat() if d else None,
                    "day_index": i,
                    "note": str(v).strip(),
                })

        # Extract people from detected rows
        for person_rows, person_type in [
            (resident_rows, "resident"),
            (faculty_rows, "faculty"),
        ]:
            for row in person_rows:
                name_raw = ws.cell(row=row, column=col_name).value
                if not name_raw or name_raw == NBSP or not str(name_raw).strip():
                    continue

                name = str(name_raw).strip()

                rot1 = None
                rot2 = None
                template = None
                role = None

                if col_rotation1:
                    rot1_raw = ws.cell(row=row, column=col_rotation1).value
                    rot1 = str(rot1_raw).strip() if rot1_raw and rot1_raw != NBSP else None
                if col_rotation2:
                    rot2_raw = ws.cell(row=row, column=col_rotation2).value
                    rot2 = str(rot2_raw).strip() if rot2_raw and rot2_raw != NBSP else None
                if col_template:
                    template_raw = ws.cell(row=row, column=col_template).value
                    template = str(template_raw).strip() if template_raw and template_raw != NBSP else None
                if col_role:
                    role_raw = ws.cell(row=row, column=col_role).value
                    role = str(role_raw).strip() if role_raw and role_raw != NBSP else None

                person = {
                    "name": name,
                    "row": row,
                    "type": person_type,
                    "rotation1": rot1,
                    "rotation2": rot2,
                    "template": template,
                    "role": role,
                    "days": [],
                }

                for i, d in enumerate(dates):
                    am_col = col_schedule_start + i * COLS_PER_DAY
                    pm_col = am_col + 1

                    am_raw = ws.cell(row=row, column=am_col).value
                    pm_raw = ws.cell(row=row, column=pm_col).value

                    am = str(am_raw).strip() if am_raw and am_raw != NBSP else None
                    pm = str(pm_raw).strip() if pm_raw and pm_raw != NBSP else None

                    person["days"].append({
                        "date": d.isoformat() if d else None,
                        "day_index": i,
                        "am": am,
                        "pm": pm,
                    })

                person["filled_cells"] = sum(
                    1 for day in person["days"]
                    if day["am"] or day["pm"]
                )
                block_data["people"].append(person)

        result["blocks"].append(block_data)

    wb.close()

    # Summary
    total_people = sum(len(b["people"]) for b in result["blocks"])
    total_cells = sum(
        p["filled_cells"]
        for b in result["blocks"]
        for p in b["people"]
    )
    result["summary"] = {
        "blocks_extracted": len(result["blocks"]),
        "block_names": [b["tab_name"] for b in result["blocks"]],
        "total_people": total_people,
        "total_filled_cells": total_cells,
    }

    return result


def main():
    parser = argparse.ArgumentParser(description="Extract ground truth from handjam workbook")
    parser.add_argument("--workbook", required=True, help="Path to handjam .xlsx")
    parser.add_argument("--output", default="/tmp/ground_truth.json", help="Output JSON path")
    args = parser.parse_args()

    wb_path = Path(args.workbook).expanduser()
    if not wb_path.exists():
        print(f"ERROR: {wb_path} not found")
        return

    print(f"Reading {wb_path.name}...")
    data = extract_workbook(wb_path)

    out = Path(args.output)
    out.write_text(json.dumps(data, indent=2))
    print(f"\nWrote {out} ({out.stat().st_size:,} bytes)")
    print(f"\nSummary:")
    for k, v in data["summary"].items():
        print(f"  {k}: {v}")

    # Quick code frequency table
    code_counts: dict[str, int] = {}
    for block in data["blocks"]:
        for person in block["people"]:
            for day in person["days"]:
                for half in ("am", "pm"):
                    code = day.get(half)
                    if code:
                        code_counts[code] = code_counts.get(code, 0) + 1

    print(f"\n  Unique display codes: {len(code_counts)}")
    print(f"  Top 20:")
    for code, cnt in sorted(code_counts.items(), key=lambda x: -x[1])[:20]:
        print(f"    {code:12s} {cnt:5d}")


if __name__ == "__main__":
    main()
