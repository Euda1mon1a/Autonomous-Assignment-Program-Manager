#!/usr/bin/env python3
"""Audit parity between DB half-day assignments and Block Template2 XLSX export.

This script compares three views for the same block:
1. Raw DB truth (no overrides)
2. Effective export truth (overrides applied)
3. Rendered XLSX cells

It is intended to answer "DB issue or export issue?" quickly and deterministically.
"""

from __future__ import annotations

import argparse
import os
import sys
from collections import Counter
from datetime import date
from pathlib import Path
from typing import Any

from defusedxml import ElementTree
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[2]
BACKEND_DIR = REPO_ROOT / "backend"
sys.path.insert(0, str(BACKEND_DIR))


def _load_env() -> None:
    env_path = REPO_ROOT / ".env"
    if not env_path.exists():
        return

    for raw in env_path.read_text().splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        existing = os.environ.get(key)
        if key == "CORS_ORIGINS":
            if not existing or '"' not in existing:
                os.environ[key] = value
            continue
        if not existing:
            os.environ[key] = value

    if not os.environ.get("DATABASE_URL") and os.environ.get("DB_PASSWORD"):
        os.environ["DATABASE_URL"] = (
            f"postgresql://scheduler:{os.environ['DB_PASSWORD']}@localhost:5432/"
            "residency_scheduler"
        )


_load_env()

from app.db.session import SessionLocal  # noqa: E402
from app.services.canonical_schedule_export_service import (  # noqa: E402
    CanonicalScheduleExportService,
)
from app.services.half_day_json_exporter import HalfDayJSONExporter  # noqa: E402
from app.services.json_to_xlsx_converter import JSONToXlsxConverter  # noqa: E402
from app.utils.academic_blocks import get_block_dates  # noqa: E402


SectionCounts = dict[str, Counter[str]]
PersonCodeCounts = dict[str, Counter[str]]


def _norm_code(value: Any) -> str:
    if value is None:
        return ""
    code = str(value).strip()
    if not code:
        return ""
    return code.upper()


def _name_key(name: str) -> str:
    return " ".join(name.replace(",", " ").split()).lower()


def _count_codes_from_people(people: list[dict[str, Any]]) -> tuple[Counter[str], PersonCodeCounts]:
    code_counts: Counter[str] = Counter()
    per_person: PersonCodeCounts = {}

    for person in people:
        person_name = str(person.get("name", "") or "").strip()
        key = _name_key(person_name)
        if not key:
            continue

        person_counter: Counter[str] = Counter()
        for day in person.get("days", []) or []:
            am = _norm_code(day.get("am", ""))
            pm = _norm_code(day.get("pm", ""))
            if am:
                person_counter[am] += 1
                code_counts[am] += 1
            if pm:
                person_counter[pm] += 1
                code_counts[pm] += 1

        per_person[key] = person_counter

    return code_counts, per_person


def _count_codes_from_payload(payload: dict[str, Any]) -> tuple[SectionCounts, dict[str, PersonCodeCounts]]:
    resident_codes, resident_people = _count_codes_from_people(payload.get("residents", []) or [])
    faculty_codes, faculty_people = _count_codes_from_people(payload.get("faculty", []) or [])
    combined_codes = resident_codes + faculty_codes

    return (
        {
            "residents": resident_codes,
            "faculty": faculty_codes,
            "combined": combined_codes,
        },
        {
            "residents": resident_people,
            "faculty": faculty_people,
        },
    )


def _load_structure_rows(structure_path: Path) -> dict[str, list[tuple[str, int]]]:
    tree = ElementTree.parse(str(structure_path))
    root = tree.getroot()

    residents: list[tuple[str, int]] = []
    for node in root.findall("./residents/resident"):
        name = (node.get("name") or "").strip()
        row = int(node.get("row") or "0")
        if name and row > 0:
            residents.append((name, row))

    faculty: list[tuple[str, int]] = []
    for node in root.findall("./faculty/person"):
        name = (node.get("name") or "").strip()
        row = int(node.get("row") or "0")
        if name and row > 0:
            faculty.append((name, row))

    return {"residents": residents, "faculty": faculty}


def _count_codes_from_sheet(
    xlsx_path: Path,
    block_start: date,
    block_end: date,
    rows_by_section: dict[str, list[tuple[str, int]]],
    active_person_keys: dict[str, set[str]] | None = None,
) -> tuple[SectionCounts, dict[str, PersonCodeCounts], SectionCounts]:
    wb = load_workbook(xlsx_path, data_only=False)
    ws = wb["Block Template2"] if "Block Template2" in wb.sheetnames else wb.active

    total_days = (block_end - block_start).days + 1
    schedule_start_col = 6
    schedule_end_col = schedule_start_col + (total_days * 2) - 1

    section_counts: SectionCounts = {"residents": Counter(), "faculty": Counter(), "combined": Counter()}
    per_section_people: dict[str, PersonCodeCounts] = {"residents": {}, "faculty": {}}
    inactive_counts: SectionCounts = {"residents": Counter(), "faculty": Counter(), "combined": Counter()}

    for section in ("residents", "faculty"):
        for expected_name, row in rows_by_section[section]:
            person_key = _name_key(expected_name)
            person_counter: Counter[str] = Counter()

            for col in range(schedule_start_col, schedule_end_col + 1):
                code = _norm_code(ws.cell(row=row, column=col).value)
                if not code:
                    continue
                person_counter[code] += 1

            include_in_active = True
            if active_person_keys is not None:
                include_in_active = person_key in active_person_keys.get(section, set())

            if include_in_active:
                per_section_people[section][person_key] = person_counter
                section_counts[section].update(person_counter)
            else:
                inactive_counts[section].update(person_counter)

    section_counts["combined"] = section_counts["residents"] + section_counts["faculty"]
    inactive_counts["combined"] = inactive_counts["residents"] + inactive_counts["faculty"]
    return section_counts, per_section_people, inactive_counts


def _counter_diff(expected: Counter[str], actual: Counter[str]) -> list[tuple[str, int, int, int]]:
    diffs: list[tuple[str, int, int, int]] = []
    for code in sorted(set(expected) | set(actual)):
        exp = expected.get(code, 0)
        act = actual.get(code, 0)
        if exp != act:
            diffs.append((code, exp, act, act - exp))
    return diffs


def _focus_count(counter: Counter[str], focus_codes: list[str]) -> int:
    return sum(counter.get(code, 0) for code in focus_codes)


def _print_counter_snapshot(label: str, counts: SectionCounts, focus_codes: list[str]) -> None:
    print(f"\n{label}:")
    for section in ("residents", "faculty", "combined"):
        total = sum(counts[section].values())
        focus_total = _focus_count(counts[section], focus_codes)
        print(f"  {section:9s} total_nonempty={total:4d} focus({','.join(focus_codes)})={focus_total:4d}")


def _print_focus_people(
    title: str,
    expected_people: PersonCodeCounts,
    actual_people: PersonCodeCounts,
    focus_codes: list[str],
    limit: int = 20,
) -> None:
    print(f"\n{title} (focus codes: {', '.join(focus_codes)}):")
    rows: list[tuple[str, int, int, int]] = []
    for key in sorted(set(expected_people) | set(actual_people)):
        expected = _focus_count(expected_people.get(key, Counter()), focus_codes)
        actual = _focus_count(actual_people.get(key, Counter()), focus_codes)
        if expected == 0 and actual == 0:
            continue
        rows.append((key, expected, actual, actual - expected))

    if not rows:
        print("  (no non-zero focus codes)")
        return

    rows.sort(key=lambda r: (abs(r[3]), r[0]), reverse=True)
    print("  person_key                               expected  xlsx  delta")
    for key, expected, actual, delta in rows[:limit]:
        print(f"  {key:38s} {expected:8d} {actual:5d} {delta:6d}")


def _resolve_output_path(block_number: int, academic_year: int, requested: str | None) -> Path:
    if requested:
        return Path(requested).resolve()
    return Path(f"/tmp/block{block_number}_ay{academic_year}_parity.xlsx")


def main() -> int:
    parser = argparse.ArgumentParser(description="Audit DB vs XLSX export parity for a block")
    parser.add_argument("--block", type=int, required=True, help="Block number (1-13)")
    parser.add_argument("--academic-year", type=int, required=True, help="Academic year start, e.g. 2025")
    parser.add_argument("--xlsx", type=str, default=None, help="Use existing XLSX file instead of exporting")
    parser.add_argument(
        "--output",
        type=str,
        default=None,
        help="Output XLSX path when --xlsx is not provided",
    )
    parser.add_argument(
        "--focus-codes",
        type=str,
        default="C",
        help="Comma-separated codes to highlight (default: C)",
    )
    parser.add_argument(
        "--no-fail",
        action="store_true",
        help="Always exit 0 even when parity mismatches are detected",
    )
    args = parser.parse_args()

    focus_codes = [c.strip().upper() for c in args.focus_codes.split(",") if c.strip()]
    if not focus_codes:
        focus_codes = ["C"]

    block_dates = get_block_dates(args.block, args.academic_year)
    block_start = block_dates.start_date
    block_end = block_dates.end_date

    session = SessionLocal()
    try:
        exporter = HalfDayJSONExporter(session)
        raw_payload = exporter.export(
            block_start=block_start,
            block_end=block_end,
            include_faculty=True,
            include_call=False,
            include_overrides=False,
        )
        effective_payload = exporter.export(
            block_start=block_start,
            block_end=block_end,
            include_faculty=True,
            include_call=False,
            include_overrides=True,
        )

        raw_counts, _ = _count_codes_from_payload(raw_payload)
        effective_counts, effective_people = _count_codes_from_payload(effective_payload)

        service = CanonicalScheduleExportService(session)
        template_path = service._template_path()
        structure_path = service._structure_path()
        conversion_stats: dict[str, Any] = {}

        if args.xlsx:
            xlsx_path = Path(args.xlsx).resolve()
            if not xlsx_path.exists():
                raise FileNotFoundError(f"XLSX not found: {xlsx_path}")
        else:
            xlsx_path = _resolve_output_path(args.block, args.academic_year, args.output)
            converter = JSONToXlsxConverter(
                template_path=template_path,
                structure_xml_path=structure_path,
                use_block_template2=True,
                apply_colors=True,
                strict_row_mapping=True,
            )
            converter.convert_from_json(effective_payload, output_path=xlsx_path)
            conversion_stats = converter.last_conversion_stats

        rows_by_section = _load_structure_rows(structure_path)
        active_keys = {
            "residents": set(effective_people["residents"].keys()),
            "faculty": set(effective_people["faculty"].keys()),
        }
        sheet_counts, sheet_people, inactive_sheet_counts = _count_codes_from_sheet(
            xlsx_path=xlsx_path,
            block_start=block_start,
            block_end=block_end,
            rows_by_section=rows_by_section,
            active_person_keys=active_keys,
        )

        print(f"Block {args.block} AY{args.academic_year} parity audit")
        print(f"Date range: {block_start} -> {block_end}")
        print(f"Workbook: {xlsx_path}")

        _print_counter_snapshot("RAW (no overrides)", raw_counts, focus_codes)
        _print_counter_snapshot("EFFECTIVE (overrides applied)", effective_counts, focus_codes)
        _print_counter_snapshot("XLSX (rendered cells)", sheet_counts, focus_codes)
        if sum(inactive_sheet_counts["combined"].values()) > 0:
            _print_counter_snapshot(
                "XLSX (non-exported mapped rows, informational)",
                inactive_sheet_counts,
                focus_codes,
            )

        mismatch_found = False
        for section in ("residents", "faculty", "combined"):
            diffs = _counter_diff(effective_counts[section], sheet_counts[section])
            if diffs:
                mismatch_found = True
                print(f"\nMISMATCH {section.upper()} (expected=effective, actual=xlsx):")
                print("  code   expected  xlsx  delta")
                for code, expected, actual, delta in diffs:
                    print(f"  {code:5s} {expected:8d} {actual:5d} {delta:6d}")

        _print_focus_people(
            "Faculty per-person parity",
            effective_people["faculty"],
            sheet_people["faculty"],
            focus_codes,
        )

        if conversion_stats:
            print("\nConverter diagnostics:")
            print(
                "  residents: "
                f"input={conversion_stats['residents']['people_input']}, "
                f"written={conversion_stats['residents']['people_written']}, "
                f"unmapped={len(conversion_stats['residents']['unmapped_names'])}"
            )
            print(
                "  faculty:   "
                f"input={conversion_stats['faculty']['people_input']}, "
                f"written={conversion_stats['faculty']['people_written']}, "
                f"unmapped={len(conversion_stats['faculty']['unmapped_names'])}"
            )

        status = "PASS" if not mismatch_found else "FAIL"
        print(f"\nSTATUS: {status}")

        if mismatch_found and not args.no_fail:
            return 1
        return 0
    finally:
        session.close()


if __name__ == "__main__":
    raise SystemExit(main())
