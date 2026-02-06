#!/usr/bin/env python3
"""Generate values-only Excel payloads from canonical block export.

This is a pragmatic fallback when presentation parity is hard:
- Keep your handjam workbook formatting untouched.
- Copy/paste (or apply) only assignment values.

Default ranges target Block Template2:
- Call row: F4:BI4
- Schedule grid: F9:BI43
"""

from __future__ import annotations

import argparse
import csv
import os
import shutil
import subprocess
import sys
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import range_boundaries

REPO_ROOT = Path(__file__).resolve().parents[2]
BACKEND_DIR = REPO_ROOT / "backend"
sys.path.insert(0, str(BACKEND_DIR))


def _load_env() -> None:
    env_path = REPO_ROOT / ".env"
    if not env_path.exists():
        return

    for raw in env_path.read_text().splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        existing = os.environ.get(key)
        if key == "CORS_ORIGINS":
            if not existing or '"' not in existing:
                os.environ[key] = value
            continue
        if not existing:
            os.environ[key] = value

    if not os.environ.get("DATABASE_URL") and os.environ.get("DB_PASSWORD"):
        os.environ["DATABASE_URL"] = (
            f"postgresql://scheduler:{os.environ['DB_PASSWORD']}@localhost:5432/"
            "residency_scheduler"
        )


_load_env()


def _load_backend_export_service():
    try:
        from app.db.session import SessionLocal  # noqa: E402
        from app.services.canonical_schedule_export_service import (  # noqa: E402
            CanonicalScheduleExportService,
        )
    except ImportError as exc:  # pragma: no cover - environment guard
        if "NotRequired" in str(exc):
            raise SystemExit(
                "Python 3.11+ runtime required for backend imports. "
                "Run with ./backend/venv/bin/python."
            ) from exc
        raise
    return SessionLocal, CanonicalScheduleExportService


def _extract_matrix(ws, cell_range: str) -> list[list[Any]]:
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    rows: list[list[Any]] = []
    for row in range(min_row, max_row + 1):
        values: list[Any] = []
        for col in range(min_col, max_col + 1):
            values.append(ws.cell(row=row, column=col).value)
        rows.append(values)
    return rows


def _apply_matrix(ws, start_range: str, matrix: list[list[Any]]) -> int:
    min_col, min_row, max_col, max_row = range_boundaries(start_range)
    expected_rows = (max_row - min_row) + 1
    expected_cols = (max_col - min_col) + 1
    if len(matrix) != expected_rows:
        raise ValueError(
            f"Row mismatch for {start_range}: expected {expected_rows}, got {len(matrix)}"
        )
    if matrix and any(len(row) != expected_cols for row in matrix):
        raise ValueError(
            f"Column mismatch for {start_range}: expected {expected_cols} columns"
        )

    merged_lookup: dict[tuple[int, int], tuple[int, int]] = {}
    for merged_range in ws.merged_cells.ranges:
        if merged_range.max_col < min_col or merged_range.min_col > max_col:
            continue
        if merged_range.max_row < min_row or merged_range.min_row > max_row:
            continue
        top_left = (merged_range.min_row, merged_range.min_col)
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                if min_row <= row <= max_row and min_col <= col <= max_col:
                    merged_lookup[(row, col)] = top_left

    written = 0
    merged_target_values: dict[tuple[int, int], Any] = {}
    for r_idx, row_values in enumerate(matrix):
        row = min_row + r_idx
        for c_idx, value in enumerate(row_values):
            col = min_col + c_idx
            target = merged_lookup.get((row, col))
            if target is not None:
                existing = merged_target_values.get(target)
                if target not in merged_target_values:
                    merged_target_values[target] = value
                elif existing in (None, "") and value not in (None, ""):
                    # Keep non-empty source value when multiple cells map to one merged target.
                    merged_target_values[target] = value
            else:
                ws.cell(row=row, column=col).value = value

    for (target_row, target_col), value in merged_target_values.items():
        ws.cell(row=target_row, column=target_col).value = value

    for r_idx, row_values in enumerate(matrix):
        row = min_row + r_idx
        for c_idx, _ in enumerate(row_values):
            col = min_col + c_idx
            cell = ws.cell(row=row, column=col)
            value = cell.value
            if value not in (None, ""):
                written += 1
    return written


def _matrix_to_tsv(matrix: list[list[Any]]) -> str:
    lines: list[str] = []
    for row in matrix:
        line = "\t".join("" if value is None else str(value) for value in row)
        lines.append(line)
    return "\n".join(lines) + "\n"


def _write_tsv(path: Path, matrix: list[list[Any]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle, delimiter="\t", lineterminator="\n")
        for row in matrix:
            writer.writerow(["" if value is None else str(value) for value in row])


def _copy_to_clipboard(text: str) -> bool:
    if shutil.which("pbcopy") is None:
        return False
    subprocess.run(["pbcopy"], input=text.encode("utf-8"), check=True)
    return True


def _resolve_sheet(workbook, requested: str | None):
    if requested and requested in workbook.sheetnames:
        return workbook[requested]
    if "Block Template2" in workbook.sheetnames:
        return workbook["Block Template2"]
    return workbook.active


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Create values-only payloads or apply values into a formatted workbook."
    )
    parser.add_argument("--block", type=int, required=True, help="Block number (1-13)")
    parser.add_argument(
        "--academic-year", type=int, required=True, help="Academic year start, e.g. 2025"
    )
    parser.add_argument(
        "--source-xlsx",
        type=str,
        default=None,
        help="Use an existing source workbook instead of exporting from DB",
    )
    parser.add_argument(
        "--source-sheet",
        type=str,
        default="Block Template2",
        help="Source sheet name (default: Block Template2)",
    )
    parser.add_argument(
        "--target-workbook",
        type=str,
        default=None,
        help="If provided, apply values into this workbook (preserves formatting)",
    )
    parser.add_argument(
        "--target-sheet",
        type=str,
        default="Block Template2",
        help="Target sheet name (default: Block Template2)",
    )
    parser.add_argument(
        "--output",
        type=str,
        default=None,
        help=(
            "Output path. For target mode: patched workbook path. "
            "For payload mode: payload workbook path."
        ),
    )
    parser.add_argument(
        "--grid-range",
        type=str,
        default="F9:BI43",
        help="Grid range for schedule values (default: F9:BI43)",
    )
    parser.add_argument(
        "--call-range",
        type=str,
        default="F4:BI4",
        help="Call range (default: F4:BI4)",
    )
    parser.add_argument(
        "--skip-call",
        action="store_true",
        help="Do not include call row values",
    )
    parser.add_argument(
        "--grid-tsv",
        type=str,
        default=None,
        help="Optional path to write grid TSV (tab-separated)",
    )
    parser.add_argument(
        "--pbcopy-grid",
        action="store_true",
        help="Copy grid TSV payload to macOS clipboard",
    )
    args = parser.parse_args()

    source_path: Path
    temp_export: Path | None = None

    if args.source_xlsx:
        source_path = Path(args.source_xlsx).expanduser().resolve()
        if not source_path.exists():
            raise FileNotFoundError(f"Source workbook not found: {source_path}")
    else:
        SessionLocal, CanonicalScheduleExportService = _load_backend_export_service()
        with NamedTemporaryFile(
            prefix=f"block{args.block}_ay{args.academic_year}_source_",
            suffix=".xlsx",
            delete=False,
        ) as handle:
            temp_export = Path(handle.name)
        source_path = temp_export

        session = SessionLocal()
        try:
            service = CanonicalScheduleExportService(session)
            service.export_block_xlsx(
                block_number=args.block,
                academic_year=args.academic_year,
                include_faculty=True,
                include_qa_sheet=False,
                preserve_template_identity_fields=True,
                presentation_profile="tamc_handjam_v2",
                output_path=source_path,
            )
        finally:
            session.close()

    source_wb = load_workbook(source_path, data_only=False)
    source_ws = _resolve_sheet(source_wb, args.source_sheet)

    grid_matrix = _extract_matrix(source_ws, args.grid_range)
    call_matrix = _extract_matrix(source_ws, args.call_range) if not args.skip_call else []

    if args.grid_tsv:
        grid_tsv_path = Path(args.grid_tsv).expanduser().resolve()
        _write_tsv(grid_tsv_path, grid_matrix)
    else:
        grid_tsv_path = None

    if args.pbcopy_grid:
        copied = _copy_to_clipboard(_matrix_to_tsv(grid_matrix))
        if not copied:
            print("Warning: pbcopy not found, clipboard copy skipped", file=sys.stderr)

    if args.target_workbook:
        target_input = Path(args.target_workbook).expanduser().resolve()
        if not target_input.exists():
            raise FileNotFoundError(f"Target workbook not found: {target_input}")

        output_path = (
            Path(args.output).expanduser().resolve()
            if args.output
            else target_input.with_name(f"{target_input.stem}_values_applied.xlsx")
        )

        target_wb = load_workbook(target_input)
        target_ws = _resolve_sheet(target_wb, args.target_sheet)
        written_grid = _apply_matrix(target_ws, args.grid_range, grid_matrix)
        written_call = 0
        if not args.skip_call:
            written_call = _apply_matrix(target_ws, args.call_range, call_matrix)

        target_wb.save(output_path)
        print("Mode: apply")
        print(f"Source workbook: {source_path}")
        print(f"Target workbook: {target_input}")
        print(f"Output workbook: {output_path}")
        print(f"Grid range {args.grid_range}: {written_grid} non-empty values written")
        if not args.skip_call:
            print(f"Call range {args.call_range}: {written_call} non-empty values written")
    else:
        output_path = (
            Path(args.output).expanduser().resolve()
            if args.output
            else Path(f"/tmp/block{args.block}_ay{args.academic_year}_paste_payload.xlsx")
        )

        payload_wb = Workbook()
        grid_ws = payload_wb.active
        grid_ws.title = "Grid_Paste"
        for r_idx, row_values in enumerate(grid_matrix, start=1):
            for c_idx, value in enumerate(row_values, start=1):
                grid_ws.cell(row=r_idx, column=c_idx).value = value

        if not args.skip_call:
            call_ws = payload_wb.create_sheet("Call_Paste")
            for r_idx, row_values in enumerate(call_matrix, start=1):
                for c_idx, value in enumerate(row_values, start=1):
                    call_ws.cell(row=r_idx, column=c_idx).value = value

        payload_wb.save(output_path)
        print("Mode: payload")
        print(f"Source workbook: {source_path}")
        print(f"Output payload workbook: {output_path}")
        print(
            "Copy Grid_Paste!A1:BD35 into target sheet range "
            f"{args.grid_range} using Paste Special -> Values."
        )
        if not args.skip_call:
            print(
                "Copy Call_Paste!A1:BD1 into target sheet range "
                f"{args.call_range} using Paste Special -> Values."
            )

    if grid_tsv_path:
        print(f"Grid TSV: {grid_tsv_path}")

    if temp_export and temp_export.exists():
        temp_export.unlink()

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
