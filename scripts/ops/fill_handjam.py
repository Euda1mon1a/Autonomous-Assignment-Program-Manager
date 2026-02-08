#!/usr/bin/env python3
"""Fill Handjam Workbook -- Write DB schedule data into the human-maintained Excel workbook.

Design philosophy:
    - The handjam workbook owns ALL presentation (borders, widths, fonts, summary formulas,
      header colors, freeze panes, column sizes)
    - Colors come from conditional formatting rules stamped into the template
      (run stamp_template_cf.py first, or colors auto-apply if CF rules exist)
    - This script ONLY writes string values into cells -- zero color logic

Pipeline:
    Handjam workbook (formatting + CF rules) + DB (data) -> fill_handjam -> filled workbook

Usage:
    # Fill Block 10 into the handjam workbook:
    python scripts/ops/fill_handjam.py --block 10 --academic-year 2026 \
        --workbook ~/Downloads/TAMC_Schedule.xlsx

    # Preview name->row mapping only (no DB query, no writes):
    python scripts/ops/fill_handjam.py --block 10 --academic-year 2026 \
        --workbook ~/Downloads/TAMC_Schedule.xlsx --preview

    # Use pre-exported JSON (no DB connection needed):
    python scripts/ops/fill_handjam.py --block 10 --academic-year 2026 \
        --workbook ~/Downloads/TAMC_Schedule.xlsx --json /tmp/block10_data.json

    # Stamp CF rules onto the handjam first (one-time per sheet):
    python scripts/ops/stamp_template_cf.py \
        --workbook ~/Downloads/TAMC_Schedule.xlsx --sheet "Block 10"
"""

from __future__ import annotations

import argparse
import os
import sys
from datetime import date
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[2]
BACKEND_DIR = REPO_ROOT / "backend"
sys.path.insert(0, str(BACKEND_DIR))


def _load_env() -> None:
    """Load .env file for DATABASE_URL etc."""
    env_path = REPO_ROOT / ".env"
    if not env_path.exists():
        return
    for raw in env_path.read_text().splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        value = value.strip().strip('"').strip("'")
        key = key.strip()
        existing = os.environ.get(key)
        if key == "CORS_ORIGINS":
            if not existing or '"' not in existing:
                os.environ[key] = value
            continue
        if not existing:
            os.environ[key] = value
    if not os.environ.get("DATABASE_URL") and os.environ.get("DB_PASSWORD"):
        os.environ["DATABASE_URL"] = (
            f"postgresql://scheduler:{os.environ['DB_PASSWORD']}@localhost:5432/"
            "residency_scheduler"
        )


_load_env()

from openpyxl import load_workbook  # noqa: E402
from openpyxl.styles import PatternFill  # noqa: E402

# ---------------------------------------------------------------
# Block Template2 layout constants
# ---------------------------------------------------------------
COL_ROTATION1 = 1
COL_ROTATION2 = 2
COL_NAME = 5
COL_SCHEDULE_START = 6
COLS_PER_DAY = 2
MAX_DAYS = 28
COL_SCHEDULE_END = COL_SCHEDULE_START + (MAX_DAYS * COLS_PER_DAY) - 1

ROW_STAFF_CALL = 4
ROW_RESIDENT_START = 9
ROW_RESIDENT_END = 25
ROW_FACULTY_START = 31
ROW_FACULTY_END = 45


# ---------------------------------------------------------------
# Name matching
# ---------------------------------------------------------------

def _normalize_name(name: str) -> str:
    return name.strip().rstrip("*").strip().lower()


def _extract_last_name(name: str) -> str:
    name = _normalize_name(name)
    if "," in name:
        return name.split(",")[0].strip()
    parts = name.split()
    return parts[-1] if parts else name


def _discover_name_rows(ws, row_start: int, row_end: int) -> dict[str, int]:
    """Read names from column E -> last_name -> row mapping."""
    mapping: dict[str, int] = {}
    for row in range(row_start, row_end + 1):
        val = ws.cell(row=row, column=COL_NAME).value
        if val and isinstance(val, str) and val.strip():
            last = _extract_last_name(val)
            if last in mapping:
                for existing_name, existing_row in list(mapping.items()):
                    if existing_name == last:
                        full = _normalize_name(
                            ws.cell(row=existing_row, column=COL_NAME).value or ""
                        )
                        mapping[full] = existing_row
                        del mapping[last]
                        break
                mapping[_normalize_name(val)] = row
            else:
                mapping[last] = row
    return mapping


def _match_name(db_name: str, wb_names: dict[str, int]) -> int | None:
    last = _extract_last_name(db_name)
    if last in wb_names:
        return wb_names[last]
    norm = _normalize_name(db_name)
    if norm in wb_names:
        return wb_names[norm]
    parts = _normalize_name(db_name).split()
    if len(parts) >= 2:
        converted = f"{parts[-1]}, {' '.join(parts[:-1])}"
        if converted in wb_names:
            return wb_names[converted]
    return None


# ---------------------------------------------------------------
# Grid clearing
# ---------------------------------------------------------------

def _clear_grid(ws, row_start: int, row_end: int,
                col_start: int = COL_SCHEDULE_START,
                col_end: int = COL_SCHEDULE_END) -> int:
    """Clear grid cells (values + explicit fills so CF rules can take effect)."""
    count = 0
    no_fill = PatternFill(fill_type=None)
    for row in range(row_start, row_end + 1):
        for col in range(col_start, col_end + 1):
            cell = ws.cell(row=row, column=col)
            cell.value = None
            cell.fill = no_fill
            count += 1
    return count


# ---------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------

def _load_schedule_data(
    block_number: int,
    academic_year: int,
    json_path: Path | None = None,
) -> dict:
    import json as json_mod

    if json_path:
        print(f"\nLoading from JSON: {json_path}")
        with open(json_path) as f:
            return json_mod.load(f)

    from app.db.session import SessionLocal
    from app.services.half_day_json_exporter import HalfDayJSONExporter
    from app.utils.academic_blocks import get_block_dates

    block_dates = get_block_dates(block_number, academic_year)
    print(f"\nQuerying DB for Block {block_number} AY{academic_year}...")
    session = SessionLocal()
    try:
        exporter = HalfDayJSONExporter(session)
        return exporter.export(
            block_start=block_dates.start_date,
            block_end=block_dates.end_date,
            include_faculty=True,
            include_call=True,
            include_overrides=True,
        )
    finally:
        session.close()


# ---------------------------------------------------------------
# Main fill logic
# ---------------------------------------------------------------

def fill_handjam(
    workbook_path: Path,
    block_number: int,
    academic_year: int,
    output_path: Path,
    sheet_name: str | None = None,
    preview: bool = False,
    json_path: Path | None = None,
) -> dict:
    """Fill handjam workbook with DB schedule data (values only)."""
    wb = load_workbook(workbook_path)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        ws = wb[sheet_name]
    else:
        target = f"Block {block_number}"
        if target in wb.sheetnames:
            ws = wb[target]
        else:
            raise ValueError(
                f"Sheet '{target}' not found. Available: {wb.sheetnames}. "
                "Use --sheet to specify."
            )

    resident_names = _discover_name_rows(ws, ROW_RESIDENT_START, ROW_RESIDENT_END)
    faculty_names = _discover_name_rows(ws, ROW_FACULTY_START, ROW_FACULTY_END)

    print(f"\nWorkbook: {workbook_path.name}")
    print(f"Sheet:    {ws.title}")
    print(f"\nResidents ({len(resident_names)}):")
    for name, row in sorted(resident_names.items(), key=lambda x: x[1]):
        print(f"  Row {row:2d}: {name}")
    print(f"\nFaculty ({len(faculty_names)}):")
    for name, row in sorted(faculty_names.items(), key=lambda x: x[1]):
        print(f"  Row {row:2d}: {name}")

    if preview:
        print("\n[PREVIEW] No DB query, no changes written.")
        wb.close()
        return {"matched": 0, "unmatched": 0, "cells_written": 0}

    data = _load_schedule_data(block_number, academic_year, json_path)
    block_start = date.fromisoformat(data["block_start"])
    block_end = date.fromisoformat(data["block_end"])
    num_days = (block_end - block_start).days + 1
    print(f"\nBlock {block_number}: {block_start} to {block_end} ({num_days} days)")

    # Clear existing grid (values + fills) so CF rules take effect
    _clear_grid(ws, ROW_RESIDENT_START, ROW_RESIDENT_END)
    _clear_grid(ws, ROW_FACULTY_START, ROW_FACULTY_END)
    _clear_grid(ws, ROW_STAFF_CALL, ROW_STAFF_CALL)

    all_wb_names = {**resident_names, **faculty_names}
    stats = {"matched": 0, "unmatched": 0, "cells_written": 0}
    unmatched: list[str] = []

    for person in data.get("residents", []) + data.get("faculty", []):
        db_name = person.get("name", "")
        row = _match_name(db_name, all_wb_names)

        if row is None:
            unmatched.append(db_name)
            stats["unmatched"] += 1
            continue

        stats["matched"] += 1

        # Write rotation names (values only -- CF handles colors)
        ws.cell(row=row, column=COL_ROTATION1).value = person.get("rotation1") or None
        ws.cell(row=row, column=COL_ROTATION2).value = person.get("rotation2") or None

        for day_idx, day in enumerate(person.get("days", [])):
            if day_idx >= MAX_DAYS:
                break

            am_code = day.get("am", "")
            pm_code = day.get("pm", "")

            day_date = date.fromisoformat(day["date"])
            is_weekend = day_date.weekday() in (5, 6)
            if not am_code and is_weekend:
                am_code = "W"
            if not pm_code and is_weekend:
                pm_code = "W"

            am_col = COL_SCHEDULE_START + (day_idx * COLS_PER_DAY)
            pm_col = am_col + 1

            if am_code:
                ws.cell(row=row, column=am_col).value = am_code
                stats["cells_written"] += 1
            if pm_code:
                ws.cell(row=row, column=pm_col).value = pm_code
                stats["cells_written"] += 1

    # Call row
    call_written = 0
    for night in data.get("call", {}).get("nights", []):
        night_date = date.fromisoformat(night["date"])
        day_offset = (night_date - block_start).days
        if 0 <= day_offset < MAX_DAYS:
            am_col = COL_SCHEDULE_START + (day_offset * COLS_PER_DAY)
            staff_name = night.get("staff", "")
            if staff_name:
                if "," in staff_name:
                    last_name = staff_name.split(",")[0].strip()
                else:
                    parts = staff_name.split()
                    last_name = parts[-1] if parts else staff_name
                ws.cell(row=ROW_STAFF_CALL, column=am_col).value = last_name
                call_written += 1

    wb.save(output_path)
    wb.close()

    print(f"\n{'='*50}")
    print("RESULTS")
    print(f"{'='*50}")
    print(f"  Matched:       {stats['matched']} people")
    print(f"  Unmatched:     {stats['unmatched']} people")
    if unmatched:
        for name in unmatched:
            print(f"    - {name}")
    print(f"  Cells written: {stats['cells_written']}")
    print(f"  Call nights:   {call_written}")
    print(f"  Output:        {output_path}")

    return stats


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Fill handjam workbook with DB schedule data (values only)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s --block 10 --academic-year 2026 --workbook ~/Downloads/TAMC_Schedule.xlsx
  %(prog)s --block 10 --academic-year 2026 --workbook ~/Downloads/TAMC_Schedule.xlsx --preview
  %(prog)s --block 10 --academic-year 2026 --workbook ~/Downloads/TAMC_Schedule.xlsx --json /tmp/block10.json
        """,
    )
    parser.add_argument("--block", type=int, required=True, help="Block number (1-13)")
    parser.add_argument("--academic-year", type=int, required=True, help="Starting year (e.g., 2025 for AY25-26)")
    parser.add_argument("--workbook", type=str, required=True, help="Path to handjam workbook (.xlsx)")
    parser.add_argument("--output", type=str, default=None, help="Output path (default: Block{N}_FILLED.xlsx)")
    parser.add_argument("--sheet", type=str, default=None, help="Sheet name (default: 'Block {N}')")
    parser.add_argument("--json", type=str, default=None, help="Path to pre-exported JSON (skip DB query)")
    parser.add_argument("--preview", action="store_true", help="Preview name->row mapping only, no writes")

    args = parser.parse_args()

    workbook_path = Path(args.workbook).expanduser()
    if not workbook_path.exists():
        print(f"ERROR: Workbook not found: {workbook_path}")
        return 1

    output_path = (
        Path(args.output).expanduser() if args.output
        else workbook_path.parent / f"Block{args.block}_FILLED.xlsx"
    )

    json_path = Path(args.json).expanduser() if args.json else None

    fill_handjam(
        workbook_path=workbook_path,
        block_number=args.block,
        academic_year=args.academic_year,
        output_path=output_path,
        sheet_name=args.sheet,
        preview=args.preview,
        json_path=json_path,
    )

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
