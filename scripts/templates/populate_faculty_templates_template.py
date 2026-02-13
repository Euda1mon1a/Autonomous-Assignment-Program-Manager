"""Populate faculty weekly templates from Excel data template.

Analyzes multiple weeks of Excel data per faculty member to extract
recurring weekly patterns and populates the faculty_weekly_templates table.

Logic:
- For each (faculty, day_of_week, time_of_day), finds most common activity
- Filters out rotating activities (CALL, PCAT, DO) — handled by preload system
- Filters out one-time events (RETREAT) — not part of weekly pattern
- Uses week_number=NULL for default patterns (applies to all weeks)

Usage:
  1. Copy to /tmp/populate_faculty_templates_blockNN.py
  2. Fill in configuration
  3. Run: backend/.venv/bin/python /tmp/populate_faculty_templates_blockNN.py

Reference implementation: scripts/data/block11_import/ (local only, gitignored)
"""

from datetime import date, timedelta
from collections import Counter

import openpyxl
import psycopg2

CONN = "dbname=residency_scheduler user=scheduler host=localhost"

EXCEL_PATH = "/path/to/block_import.xlsx"
SHEET_NAME = "Block NN Import"

BLOCK_CONFIG = {
    "start_date": date(2026, 1, 1),
    "end_date": date(2026, 1, 28),
}

FACULTY_ROWS = range(31, 41)
NAME_COL = 5
DATA_COL_START = 6

# Activity codes to skip when determining default pattern
# These are rotating/one-time and handled separately
SKIP_CODES = {"CALL", "PCAT", "DO", "RETREAT", "retreat"}

# CODE_MAP from Excel display codes to DB codes
CODE_MAP = {
    "C": "fm_clinic",
    "SM": "sm_clinic",
    # Add block-specific mappings
}

# NAME_MAP from Excel to DB names
NAME_MAP = {
    # "Last, First": "First Last",
}


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[SHEET_NAME]

    start = BLOCK_CONFIG["start_date"]
    num_days = (BLOCK_CONFIG["end_date"] - start).days + 1
    slots = []
    for i in range(num_days):
        d = start + timedelta(days=i)
        slots.append((d, "AM"))
        slots.append((d, "PM"))

    conn = psycopg2.connect(CONN)
    cur = conn.cursor()

    # Build lookups
    cur.execute("SELECT id, name FROM people")
    people = {name: pid for pid, name in cur.fetchall()}

    cur.execute("SELECT id, code FROM activities")
    activities = {code: aid for aid, code in cur.fetchall()}

    for row in FACULTY_ROWS:
        excel_name = ws.cell(row=row, column=NAME_COL).value
        if not excel_name or not str(excel_name).strip():
            continue
        excel_name = str(excel_name).strip()
        db_name = NAME_MAP.get(excel_name)
        if not db_name:
            continue
        person_id = people.get(db_name)
        if not person_id:
            continue

        # Collect activities by (day_of_week, time_of_day)
        pattern = {}  # {(dow, tod): Counter}
        for idx, (d, tod) in enumerate(slots):
            col = DATA_COL_START + idx
            val = ws.cell(row=row, column=col).value
            if not val:
                continue
            code = str(val).strip()
            if code in SKIP_CODES:
                continue
            db_code = CODE_MAP.get(code, code.lower())

            # Python weekday: Mon=0..Sun=6 → DB: Sun=0, Mon=1..Sat=6
            db_dow = (d.weekday() + 1) % 7
            key = (db_dow, tod)
            pattern.setdefault(key, Counter())[db_code] += 1

        # Update or create templates
        for (dow, tod), counter in pattern.items():
            best_code = counter.most_common(1)[0][0]
            activity_id = activities.get(best_code)
            if not activity_id:
                print(f"  SKIP: Activity '{best_code}' not in DB for {db_name}")
                continue

            # Try update existing
            cur.execute(
                """UPDATE faculty_weekly_templates
                SET activity_id = %s
                WHERE person_id = %s AND day_of_week = %s AND time_of_day = %s
                  AND week_number IS NULL""",
                (activity_id, person_id, dow, tod),
            )
            if cur.rowcount == 0:
                # Insert new
                cur.execute(
                    """INSERT INTO faculty_weekly_templates
                    (id, person_id, day_of_week, time_of_day, activity_id,
                     week_number, created_at, updated_at)
                    VALUES (gen_random_uuid(), %s, %s, %s, %s, NULL, NOW(), NOW())""",
                    (person_id, dow, tod, activity_id),
                )

        print(f"  {db_name}: {len(pattern)} slots updated")

    conn.commit()
    cur.close()
    conn.close()
    print("Done.")


if __name__ == "__main__":
    main()
