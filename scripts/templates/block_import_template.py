"""Block schedule import template.

Template for importing block schedule data from an Excel handjam into the database.
Copy this file and fill in the block-specific configuration to import a new block.

Usage:
  1. Copy to /tmp/import_blockNN.py
  2. Fill in BLOCK_CONFIG, NAME_MAP, CODE_MAP
  3. Run: backend/.venv/bin/python /tmp/import_blockNN.py

Prerequisite:
  - Database backup: pg_dump -U scheduler -d residency_scheduler -Fc -f /tmp/blockNN_pre_import.dump
  - Excel file accessible at EXCEL_PATH

Reference implementation: scripts/data/block11_import/ (local only, gitignored)
"""

import uuid
from datetime import date, timedelta, datetime
from collections import Counter

import openpyxl
import psycopg2

# ── Configuration ────────────────────────────────────────────────────────────

CONN = "dbname=residency_scheduler user=scheduler host=localhost"

EXCEL_PATH = "/path/to/block_import.xlsx"
SHEET_NAME = "Block NN Import"

BLOCK_CONFIG = {
    "block_number": 0,       # e.g., 11
    "academic_year": 2025,   # AY start year
    "start_date": date(2026, 1, 1),  # First day of block
    "end_date": date(2026, 1, 28),   # Last day of block
}

# Excel row ranges (1-indexed)
RESIDENT_ROWS = range(9, 26)   # Adjust per block
FACULTY_ROWS = range(31, 41)   # Adjust per block
NAME_COL = 5                    # Column E = names (Last, First format)
DATA_COL_START = 6              # Column F = first half-day slot
DATA_COL_END = 61               # Column BI = last half-day slot (28 days x 2)

# ── Name Mapping ─────────────────────────────────────────────────────────────
# Maps "Last, First" from Excel to "First Last" in DB
# Fill in for each block's roster

NAME_MAP = {
    # "Last, First": "First Last",
    # "Last, First*": "First Last",  # asterisk = chief resident
}

# ── Activity Code Mapping ────────────────────────────────────────────────────
# Maps Excel display codes to DB activity codes
# Most are lowercase or match directly; exceptions listed here

CODE_MAP = {
    "C": "fm_clinic",
    "SM": "sm_clinic",
    "retreat": "RETREAT",
    "L&D": "KAP-LD",       # Default; override per person if TAMC vs KAP
    "OFF": "off",
    "W": "W",
    "LV": "LV",
    "TDY": "TDY",
    "DEP": "DEP",
    "SLV": "SLV",
    # Add block-specific codes here
}

# Leave-type activity codes (for absence sync)
LEAVE_CODES = {"LV", "LV-AM", "LV-PM", "TDY", "DEP", "SLV"}

# Maps leave codes to absence_type
LEAVE_TYPE_MAP = {
    "LV": "vacation",
    "LV-AM": "vacation",
    "LV-PM": "vacation",
    "SLV": "vacation",
    "TDY": "tdy",
    "DEP": "deployment",
}


# ── Step 1: Parse Excel ──────────────────────────────────────────────────────

def parse_excel():
    """Parse Excel handjam and return list of (name, date, time_of_day, code)."""
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[SHEET_NAME]

    start = BLOCK_CONFIG["start_date"]
    num_days = (BLOCK_CONFIG["end_date"] - start).days + 1

    slots = []
    for i in range(num_days):
        d = start + timedelta(days=i)
        slots.append((d, "AM"))
        slots.append((d, "PM"))

    records = []
    for row in list(RESIDENT_ROWS) + list(FACULTY_ROWS):
        excel_name = ws.cell(row=row, column=NAME_COL).value
        if not excel_name or not str(excel_name).strip():
            continue
        excel_name = str(excel_name).strip()

        db_name = NAME_MAP.get(excel_name)
        if not db_name:
            print(f"  WARNING: No name mapping for '{excel_name}' (row {row})")
            continue

        for idx, (d, tod) in enumerate(slots):
            col = DATA_COL_START + idx
            val = ws.cell(row=row, column=col).value
            code = str(val).strip() if val else None
            if code:
                db_code = CODE_MAP.get(code, code.lower())
                records.append((db_name, d, tod, db_code))

    return records


# ── Step 2: Load HDAs ────────────────────────────────────────────────────────

def load_hdas(records):
    """Upsert half-day assignments from parsed Excel records."""
    conn = psycopg2.connect(CONN)
    cur = conn.cursor()

    # Build person lookup
    cur.execute("SELECT id, name FROM people")
    people = {name: pid for pid, name in cur.fetchall()}

    # Build activity lookup
    cur.execute("SELECT id, code FROM activities")
    activities = {code: aid for aid, code in cur.fetchall()}

    created, updated, skipped = 0, 0, 0
    for name, d, tod, code in records:
        person_id = people.get(name)
        activity_id = activities.get(code)

        if not person_id:
            print(f"  SKIP: Person '{name}' not in DB")
            skipped += 1
            continue
        if not activity_id:
            print(f"  SKIP: Activity '{code}' not in DB")
            skipped += 1
            continue

        # Try update first
        cur.execute(
            """UPDATE half_day_assignments
            SET activity_id = %s, updated_at = NOW()
            WHERE person_id = %s AND date = %s AND time_of_day = %s""",
            (activity_id, person_id, d, tod),
        )
        if cur.rowcount > 0:
            updated += cur.rowcount
        else:
            # Insert new
            cur.execute(
                """INSERT INTO half_day_assignments
                (id, person_id, date, time_of_day, activity_id, source, created_at, updated_at)
                VALUES (%s, %s, %s, %s, %s, %s, NOW(), NOW())""",
                (str(uuid.uuid4()), person_id, d, tod, activity_id, "manual"),
            )
            created += 1

    conn.commit()
    cur.close()
    conn.close()
    print(f"  HDAs: {updated} updated, {created} created, {skipped} skipped")


# ── Step 3: Sync Absences ────────────────────────────────────────────────────

def sync_absences(records):
    """Extract leave-type activities and sync to absences table."""
    conn = psycopg2.connect(CONN)
    cur = conn.cursor()

    # Build person lookup
    cur.execute("SELECT id, name FROM people")
    people = {name: pid for pid, name in cur.fetchall()}

    # Group leave records by person
    leave_by_person = {}
    for name, d, tod, code in records:
        if code.upper() in LEAVE_CODES:
            leave_by_person.setdefault(name, []).append((d, tod, code))

    # Delete existing block-specific absences
    block_num = BLOCK_CONFIG["block_number"]
    cur.execute(
        "DELETE FROM absences WHERE notes LIKE %s",
        (f"Block {block_num}%",),
    )
    print(f"  Deleted {cur.rowcount} existing Block {block_num} absences")

    # Consolidate consecutive dates into ranges per person/type
    for name, entries in leave_by_person.items():
        person_id = people.get(name)
        if not person_id:
            continue

        # Group by absence_type
        by_type = {}
        for d, tod, code in entries:
            atype = LEAVE_TYPE_MAP.get(code.upper(), "vacation")
            by_type.setdefault(atype, set()).add(d)

        for atype, date_set in by_type.items():
            # Find contiguous date ranges
            sorted_dates = sorted(date_set)
            ranges = []
            range_start = sorted_dates[0]
            range_end = sorted_dates[0]
            for d in sorted_dates[1:]:
                if (d - range_end).days <= 1:
                    range_end = d
                else:
                    ranges.append((range_start, range_end))
                    range_start = d
                    range_end = d
            ranges.append((range_start, range_end))

            for start_d, end_d in ranges:
                cur.execute(
                    """INSERT INTO absences
                    (id, person_id, start_date, end_date, absence_type,
                     is_blocking, is_away_from_program, return_date_tentative,
                     status, notes, created_at)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
                    (
                        str(uuid.uuid4()),
                        person_id,
                        start_d,
                        end_d,
                        atype,
                        atype in ("tdy", "deployment"),
                        True,
                        False,
                        "approved",
                        f"Block {block_num} - synced from Excel handjam.",
                        datetime.utcnow(),
                    ),
                )
                print(f"  Created: {name} {atype} {start_d} - {end_d}")

    conn.commit()
    cur.close()
    conn.close()


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    print("=== Parsing Excel ===")
    records = parse_excel()
    print(f"  Parsed {len(records)} records")

    print("\n=== Loading HDAs ===")
    load_hdas(records)

    print("\n=== Syncing Absences ===")
    sync_absences(records)

    print("\nDone.")


if __name__ == "__main__":
    main()
