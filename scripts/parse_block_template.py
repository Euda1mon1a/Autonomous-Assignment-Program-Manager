#!/usr/bin/env python3
"""
Parse Block_Template2.xlsx and extract structure to XML.

This script reads an Excel template and generates an XML file that captures
the row/column layout for data population.
"""
from pathlib import Path
from xml.dom import minidom
from xml.etree.ElementTree import Element, SubElement, tostring

from openpyxl import load_workbook


def parse_template(xlsx_path: str, output_xml: str, sheet_name: str = "Block Template2"):
    """Parse Excel template and generate XML structure file."""
    wb = load_workbook(xlsx_path)
    sheet = wb[sheet_name]

    root = Element("block_template")
    root.set("name", sheet_name)
    root.set("source", Path(xlsx_path).name)

    # Layout section (hardcoded based on known structure)
    layout = SubElement(root, "layout")
    SubElement(layout, "call_row").set("row", "4")
    SubElement(layout, "resident_call_row").set("row", "5")
    SubElement(layout, "schedule_col_start").text = "6"
    SubElement(layout, "cols_per_day").text = "2"

    # Find residents and faculty by scanning provider column (E = col 5)
    residents = SubElement(root, "residents")
    faculty = SubElement(root, "faculty")

    resident_count = 0
    faculty_count = 0

    for row in range(1, 100):
        name = sheet.cell(row=row, column=5).value  # Column E = Provider
        role = sheet.cell(row=row, column=4).value  # Column D = Role
        template = sheet.cell(row=row, column=3).value  # Column C = Template
        rotation1 = sheet.cell(row=row, column=1).value  # Column A = First rotation
        rotation2 = sheet.cell(row=row, column=2).value  # Column B = Second rotation

        # Clean up non-breaking spaces
        if name and isinstance(name, str):
            name = name.replace("\xa0", "").strip()

        if name and isinstance(name, str) and "," in name:
            role_str = str(role).strip() if role else ""

            if "PGY" in role_str:
                # Resident
                pgy = role_str.replace("PGY ", "").strip()
                res = SubElement(residents, "resident")
                res.set("row", str(row))
                res.set("name", name)
                res.set("pgy", pgy)
                if rotation1:
                    res.set("rotation1", str(rotation1).replace("\xa0", "").strip())
                if rotation2:
                    res.set("rotation2", str(rotation2).replace("\xa0", "").strip())
                resident_count += 1
            elif role_str in ("FAC", "PSY"):
                # Faculty
                fac = SubElement(faculty, "person")
                fac.set("row", str(row))
                fac.set("name", name)
                if template:
                    fac.set("template", str(template).replace("\xa0", "").strip())
                faculty_count += 1

    # Pretty print
    xml_str = minidom.parseString(tostring(root)).toprettyxml(indent="  ")

    # Write output
    with open(output_xml, "w") as f:
        f.write(xml_str)

    print(f"Generated {output_xml}")
    print(f"  Residents: {resident_count}")
    print(f"  Faculty: {faculty_count}")

    return xml_str


if __name__ == "__main__":
    import sys

    if len(sys.argv) >= 2:
        xlsx_path = sys.argv[1]
    else:
        xlsx_path = "Block10_Template2_FILLED.xlsx"

    output_xml = "docs/scheduling/BlockTemplate2_Structure.xml"

    parse_template(xlsx_path, output_xml)
