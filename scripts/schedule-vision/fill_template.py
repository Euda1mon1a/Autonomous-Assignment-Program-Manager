"""Fill BlockTemplate2 with display codes using rules + lookup table.

Strategy (hybrid-lookup mode):
  1. Build a lookup table from ground truth training data:
     (rotation, db_code, weekday/weekend) → most common display code
  2. Apply deterministic rules (NF expansion, weekend collapse, clinic mapping)
  3. Where lookup disagrees with rules and has ≥2 training examples, use lookup
  4. For unseen combos, fall back to rules

This approach captures coordinator patterns the rules miss (sub-clinic types,
faculty patterns, rotation-specific overrides) while keeping rules as a
reliable fallback for edge cases.

Usage:
    python fill_template.py --truth data/features_all.json
    python fill_template.py --mode rules-only
    python fill_template.py --mode lookup-only --truth data/features_all.json
"""

from __future__ import annotations

import argparse
import json
from collections import Counter
from datetime import date, datetime, timedelta
from pathlib import Path
from xml.etree import ElementTree

import joblib
import numpy as np
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

# Template layout constants
COL_ROTATION1 = 1
COL_ROTATION2 = 2
COL_TEMPLATE = 3
COL_ROLE = 4
COL_NAME = 5
COL_SCHEDULE_START = 6
COLS_PER_DAY = 2
MAX_DAYS = 28
ROW_DATES = 3
ROW_STAFF_CALL = 4
ROW_RES_CALL = 5


# ── Rule-based transform ──────────────────────────────────

NF_MAP = [
    ("LDNF", "L&D"), ("PEDSNF", "PedsNF"), ("PEDNF", "PedsNF"),
    ("PNF", "PedsNF"), ("PEDSW", "PedsNF"), ("NF/ENDO", "NF"), ("NF", "NF"),
]
WEEKEND_OVERRIDES = {"FMIT": "FMIT", "IMW": "IMW", "KAP": "KAP"}
CLINIC_MAP = {"GYN": "GYN", "SM": "SM"}
FACULTY_COLLAPSE = {"GME", "CV", "DFM", "SM", "LEC", "ADV", "AT", "DO"}
CODE_NORM = {"IM": "IMW", "LDNF": "L&D", "PedNF": "PedsNF"}


def _nf_display(rot: str, rot2: str) -> str | None:
    combined = f"{rot} {rot2}".upper()
    for key, display in NF_MAP:
        if key in combined:
            return display
    if "NEURO" in combined:
        return "NF"
    return None


def rule_transform(
    db_code: str,
    rotation: str,
    rotation2: str,
    is_weekend: bool,
    is_faculty: bool,
) -> str:
    """Apply deterministic display rules to a DB code."""
    code = db_code.strip()
    rot_u = (rotation or "").upper()
    rot2_u = (rotation2 or "").upper()

    if code in CODE_NORM:
        code = CODE_NORM[code]

    # NF code → rotation-specific night float name
    if code == "NF":
        specific = _nf_display(rot_u, rot2_u)
        if specific and specific != "NF":
            return specific

    # OFF on night float rotation → night float name
    if code == "OFF":
        nf = _nf_display(rot_u, rot2_u)
        if nf:
            return nf

    # Weekend "W" on NF/special rotations → rotation-specific code
    if is_weekend and code == "W":
        nf = _nf_display(rot_u, rot2_u)
        if nf:
            return nf
        for rk, disp in WEEKEND_OVERRIDES.items():
            if rk in rot_u:
                return disp

    # Faculty weekend collapse: admin codes → W
    if is_faculty and is_weekend and code in FACULTY_COLLAPSE:
        return "W"

    # Clinic code → rotation-specific clinic
    if code == "C":
        for rk, cc in CLINIC_MAP.items():
            if rk.upper() in rot_u or rk.upper() in rot2_u:
                return cc

    return code


# ── Name matching ──────────────────────────────────────────

def _last_name(name: str) -> str:
    name = name.strip()
    if "," in name:
        return name.split(",")[0].strip().lower()
    parts = name.split()
    return parts[-1].lower() if parts else name.lower()


def load_name_map(structure_path: str) -> dict[str, int]:
    tree = ElementTree.parse(structure_path)
    root = tree.getroot()
    mapping = {}
    for tag in ("resident", "person"):
        for elem in root.iter(tag):
            name = elem.get("name", "")
            row = elem.get("row")
            if name and row and name not in ("Unassigned", "Unassigned-NF"):
                mapping[name] = int(row)
    return mapping


def resolve_row(db_name: str, name_map: dict[str, int]) -> int | None:
    if db_name in name_map:
        return name_map[db_name]
    db_last = _last_name(db_name)
    for xml_name, row in name_map.items():
        xml_last = _last_name(xml_name)
        if db_last == xml_last:
            return row
    return None


# ── ML prediction with confidence ─────────────────────────

def _encode_single(feat: dict, encoders: dict, cat_fields: list[str]) -> np.ndarray:
    row = []
    row.append(feat.get("pgy_level") or 0)
    row.append(feat.get("block_number") or 0)
    val = feat.get("day_of_week")
    row.append(val if val is not None else -1)
    row.append(1 if feat.get("is_weekend") else 0)
    row.append(feat.get("day_index") or 0)
    row.append(1 if feat.get("half") == "pm" else 0)
    row.append(1 if feat.get("font_bold") else 0)
    row.append(1 if feat.get("db_matches_template") else 0)
    row.append(feat.get("week_in_block") or 0)
    for field in cat_fields:
        val = feat.get(field) or "__NONE__"
        enc = encoders[field]
        if val in enc.classes_:
            row.append(enc.transform([val])[0])
        else:
            row.append(len(enc.classes_))
    return np.array(row, dtype=np.float32)


def predict_cell_with_confidence(
    rf, encoders, label_enc, cat_fields,
    db_code: str, person_type: str,
    rot1: str, rot2: str, pgy: int,
    block_number: int, day_index: int,
    day_date: date, half: str,
    other_half_code: str,
) -> tuple[str, float]:
    """Predict display code + confidence for a single cell."""
    if not db_code:
        return "", 0.0

    feat = {
        "pgy_level": pgy,
        "block_number": block_number,
        "day_of_week": day_date.weekday(),
        "is_weekend": day_date.weekday() in (5, 6),
        "day_index": day_index,
        "half": half,
        "font_bold": False,
        "db_matches_template": False,
        "week_in_block": day_index // 7,
        "rotation1": rot1,
        "rotation2": rot2,
        "template": "",
        "role": "",
        "db_code": db_code,
        "template_expected": "",
        "db_other_half": other_half_code,
        "fill_rgb": "",
        "font_rgb": "",
        "person_type": person_type,
    }

    row = _encode_single(feat, encoders, cat_fields)
    proba = rf.predict_proba([row])[0]
    pred_idx = proba.argmax()
    confidence = proba[pred_idx]
    pred_code = label_enc.inverse_transform([pred_idx])[0]
    return pred_code, confidence


# ── Ground truth and lookup table ─────────────────────────

def load_ground_truth(
    truth_path: str, block_number: int
) -> dict[tuple[str, int, str], str]:
    """Load ground truth from features JSON.

    Returns {(last_name, day_index, half): truth_code}.
    """
    features = json.loads(Path(truth_path).read_text())
    truth = {}
    for f in features:
        if f.get("block_number") != block_number:
            continue
        if not f.get("truth_code"):
            continue
        name = f.get("person", "")
        last = _last_name(name)
        day_idx = f.get("day_index")
        half = f.get("half")
        if day_idx is not None and half:
            truth[(last, day_idx, half)] = f["truth_code"]
    return truth


def build_lookup_tables(
    truth_path: str,
) -> list[dict[tuple, tuple[str, int]]]:
    """Build hierarchical lookup tables from training data.

    Returns a list of tables from most specific to broadest:
      0: (rotation, db_code, weekday/weekend, person_type, half)
      1: (rotation, db_code, weekday/weekend, person_type)
      2: (rotation, db_code, weekday/weekend)
      3: (rotation, db_code)
      4: (db_code,)

    Each table maps key → (most_common_truth_code, count).
    """
    features = json.loads(Path(truth_path).read_text())

    def _wd(f):
        return "wkend" if f.get("is_weekend") else "wkday"

    key_fns = [
        lambda f: (f.get("rotation1") or "?", f.get("db_code"), _wd(f),
                   f.get("person_type", "resident"), f.get("half")),
        lambda f: (f.get("rotation1") or "?", f.get("db_code"), _wd(f),
                   f.get("person_type", "resident")),
        lambda f: (f.get("rotation1") or "?", f.get("db_code"), _wd(f)),
        lambda f: (f.get("rotation1") or "?", f.get("db_code")),
        lambda f: (f.get("db_code"),),
    ]

    tables = []
    for kf in key_fns:
        counts: dict[tuple, Counter] = {}
        for f in features:
            truth = f.get("truth_code")
            db_code = f.get("db_code")
            if not truth or not db_code:
                continue
            key = kf(f)
            if key not in counts:
                counts[key] = Counter()
            counts[key][truth] += 1
        resolved = {k: c.most_common(1)[0] for k, c in counts.items()}
        tables.append(resolved)

    return tables


def lookup_display(
    tables: list[dict[tuple, tuple[str, int]]],
    rot: str, db_code: str, is_weekend: bool, person_type: str,
    half: str = "am",
    min_count: int = 1,
) -> str | None:
    """Hierarchical lookup: try most specific key first, broaden on miss."""
    wd = "wkend" if is_weekend else "wkday"
    rot_val = rot or "?"

    keys = [
        (rot_val, db_code, wd, person_type, half),
        (rot_val, db_code, wd, person_type),
        (rot_val, db_code, wd),
        (rot_val, db_code),
        (db_code,),
    ]

    for table, key in zip(tables, keys):
        if key in table:
            code, count = table[key]
            if count >= min_count:
                return code

    return None


# ── Template filling ──────────────────────────────────────

def _write_cell(ws, row, col, value):
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        cell.value = value


def fill_template(
    db_json_path: str,
    model_path: str,
    template_path: str,
    structure_path: str,
    output_path: str,
    block_number: int = 10,
    preserve_identity: bool = True,
    mode: str = "hybrid",
    ml_threshold: float = 0.6,
    truth_path: str | None = None,
):
    """Fill the template with display codes.

    Modes:
        hybrid: Rules first, ML overlay when confident
        ml-only: Pure ML predictions
        rules-only: Only rule-based transforms
    """
    data = json.loads(Path(db_json_path).read_text())
    block_start = date.fromisoformat(data["block_start"])
    block_end = date.fromisoformat(data["block_end"])

    # Load ML model (only for ML modes)
    rf = encoders = label_enc = cat_fields = None
    if mode in ("ml-only", "hybrid"):
        model_data = joblib.load(model_path)
        rf = model_data["rf"]
        encoders = model_data["encoders"]
        label_enc = encoders["__label__"]
        cat_fields = [
            "rotation1", "rotation2", "template", "role", "db_code",
            "template_expected", "db_other_half", "fill_rgb", "font_rgb",
            "person_type",
        ]

    # Load ground truth for comparison + lookup table
    truth = {}
    lookup = None
    if truth_path:
        truth = load_ground_truth(truth_path, block_number)
        print(f"  Ground truth: {len(truth)} cells for block {block_number}")
        if mode in ("lookup-only", "hybrid-lookup"):
            lookup = build_lookup_tables(truth_path)
            print(f"  Lookup tables: {len(lookup)} levels, "
                  f"{sum(len(t) for t in lookup)} total entries")

    # Load template
    wb = load_workbook(template_path)
    ws = wb.active
    name_map = load_name_map(structure_path)

    print(f"  Block {block_number}: {block_start} to {block_end}")
    print(f"  Mode: {mode}" + (f" (ML threshold={ml_threshold})" if mode == "hybrid" else ""))

    # ── 1. Update dates ──
    for day_idx in range(MAX_DAYS):
        day_date = block_start + timedelta(days=day_idx)
        col = COL_SCHEDULE_START + day_idx * COLS_PER_DAY
        cell = ws.cell(row=ROW_DATES, column=col)
        if not isinstance(cell, MergedCell):
            cell.value = datetime(day_date.year, day_date.month, day_date.day)

    _write_cell(ws, 1, 3, str(block_number))
    range_str = (
        f"{block_start.strftime('%d%b').lstrip('0')}-"
        f"{block_end.strftime('%d%b').lstrip('0')}"
    )
    _write_cell(ws, ROW_STAFF_CALL, 3, range_str)

    # ── 2. Fill schedule ──
    matched = 0
    skipped = 0
    stats = Counter()  # rule_kept, ml_override, ml_only, rules_only
    truth_stats = Counter()  # correct, wrong, missing

    faculty_names = {p["name"] for p in data.get("faculty", [])}

    for person in data.get("residents", []) + data.get("faculty", []):
        name = person["name"]
        row = resolve_row(name, name_map)
        if row is None:
            print(f"  SKIP: No row for '{name}'")
            skipped += 1
            continue

        matched += 1
        is_faculty = name in faculty_names
        person_type = "faculty" if is_faculty else "resident"
        rot1 = person.get("rotation1") or ""
        rot2 = person.get("rotation2") or ""
        pgy = person.get("pgy") or 0

        if not preserve_identity:
            _write_cell(ws, row, COL_ROTATION1, rot1 or None)
            _write_cell(ws, row, COL_ROTATION2, rot2 or None)

        person_last = _last_name(name)

        for i, day in enumerate(person.get("days", [])):
            if i >= MAX_DAYS:
                break

            day_date = date.fromisoformat(day["date"])
            is_weekend = day_date.weekday() in (5, 6)
            raw_am = day.get("am") or ""
            raw_pm = day.get("pm") or ""

            for half, raw_code, other_code in [
                ("am", raw_am, raw_pm),
                ("pm", raw_pm, raw_am),
            ]:
                if not raw_code:
                    # Weekend empty → W
                    display = "W" if is_weekend else ""
                else:
                    display = _resolve_display(
                        mode, ml_threshold,
                        raw_code, rot1, rot2, is_weekend, is_faculty,
                        rf, encoders, label_enc, cat_fields,
                        person_type, pgy, block_number, i, day_date,
                        half, other_code, stats, lookup=lookup,
                    )

                col = COL_SCHEDULE_START + i * COLS_PER_DAY
                if half == "pm":
                    col += 1
                _write_cell(ws, row, col, display or None)

                # Ground truth comparison
                if truth:
                    gt = truth.get((person_last, i, half))
                    if gt:
                        if display == gt:
                            truth_stats["correct"] += 1
                        else:
                            truth_stats["wrong"] += 1
                    else:
                        truth_stats["no_truth"] += 1

    # ── 3. Fill call row ──
    call_written = 0
    for night in data.get("call", {}).get("nights", []):
        night_date = date.fromisoformat(night["date"])
        offset = (night_date - block_start).days
        if 0 <= offset < MAX_DAYS:
            staff = night.get("staff", "")
            if staff:
                last = staff.split()[-1] if " " in staff else staff
                col = COL_SCHEDULE_START + offset * COLS_PER_DAY
                _write_cell(ws, ROW_STAFF_CALL, col, last)
                call_written += 1

    # ── 4. Save ──
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    wb.close()

    # ── 5. Report ──
    total = sum(stats.values())
    print(f"\n  Results ({mode}):")
    print(f"    People: {matched} matched, {skipped} skipped")
    print(f"    Call nights: {call_written}")
    if mode == "hybrid-lookup":
        print(f"    Rule changed:      {stats['rule_changed']:5d} ({_pct(stats['rule_changed'], total)})")
        print(f"    Lookup supplement: {stats['lookup_supplement']:5d} ({_pct(stats['lookup_supplement'], total)})")
        print(f"    Unchanged:         {stats['unchanged']:5d} ({_pct(stats['unchanged'], total)})")
    elif mode == "lookup-only":
        print(f"    Lookup hit:   {stats['lookup_hit']:5d} ({_pct(stats['lookup_hit'], total)})")
        print(f"    Lookup miss:  {stats['lookup_miss']:5d} ({_pct(stats['lookup_miss'], total)})")
    elif mode == "hybrid":
        print(f"    Rules kept:   {stats['rule_kept']:5d} ({_pct(stats['rule_kept'], total)})")
        print(f"    ML override:  {stats['ml_override']:5d} ({_pct(stats['ml_override'], total)})")
        print(f"    ML same:      {stats['ml_agrees']:5d} ({_pct(stats['ml_agrees'], total)})")
    elif mode == "ml-only":
        print(f"    ML cells:     {stats['ml_only']:5d}")
    else:
        print(f"    Rule cells:   {stats['rules_only']:5d}")

    if truth:
        tc = truth_stats["correct"]
        tw = truth_stats["wrong"]
        tt = tc + tw
        print(f"\n  Ground truth comparison:")
        print(f"    Matched:  {tc}/{tt} ({_pct(tc, tt)})")
        print(f"    Wrong:    {tw}/{tt}")

    print(f"\n    Output: {out}")
    return out


def _pct(n: int, total: int) -> str:
    return f"{100*n/total:.1f}%" if total else "0%"


def _resolve_display(
    mode, ml_threshold,
    raw_code, rot1, rot2, is_weekend, is_faculty,
    rf, encoders, label_enc, cat_fields,
    person_type, pgy, block_number, day_index, day_date,
    half, other_code, stats, lookup=None,
) -> str:
    """Resolve the display code using the selected strategy."""

    if mode == "rules-only":
        code = rule_transform(raw_code, rot1, rot2, is_weekend, is_faculty)
        stats["rules_only"] += 1
        return code

    if mode == "ml-only":
        pred, _ = predict_cell_with_confidence(
            rf, encoders, label_enc, cat_fields,
            raw_code, person_type, rot1, rot2, pgy,
            block_number, day_index, day_date, half, other_code,
        )
        stats["ml_only"] += 1
        return pred

    if mode == "lookup-only":
        if lookup:
            code = lookup_display(
                lookup, rot1, raw_code, is_weekend, person_type, half=half,
            )
            if code:
                stats["lookup_hit"] += 1
                return code
        # Fallback to rules
        stats["lookup_miss"] += 1
        return rule_transform(raw_code, rot1, rot2, is_weekend, is_faculty)

    if mode == "hybrid-lookup":
        # Strategy: rules first for definitive changes, lookup for the rest
        rule_code = rule_transform(raw_code, rot1, rot2, is_weekend, is_faculty)

        if rule_code != raw_code:
            # Rule made a definitive change (NF expansion, clinic mapping, etc.)
            # Trust the rule — it's pattern-matched and reliable
            stats["rule_changed"] += 1
            return rule_code

        # Rule didn't change the code — try lookup for data-driven patterns
        if lookup:
            lookup_code = lookup_display(
                lookup, rot1, raw_code, is_weekend, person_type,
                half=half, min_count=2,
            )
            if lookup_code and lookup_code != raw_code:
                stats["lookup_supplement"] += 1
                return lookup_code

        stats["unchanged"] += 1
        return rule_code

    # Default hybrid (rules + ML confidence) — kept for comparison
    rule_code = rule_transform(raw_code, rot1, rot2, is_weekend, is_faculty)
    ml_code, confidence = predict_cell_with_confidence(
        rf, encoders, label_enc, cat_fields,
        raw_code, person_type, rot1, rot2, pgy,
        block_number, day_index, day_date, half, other_code,
    )

    if ml_code == rule_code:
        stats["ml_agrees"] += 1
        return rule_code
    elif confidence >= ml_threshold:
        stats["ml_override"] += 1
        return ml_code
    else:
        stats["rule_kept"] += 1
        return rule_code


def main():
    parser = argparse.ArgumentParser(
        description="Fill BlockTemplate2 with hybrid rule + ML display codes",
    )
    parser.add_argument(
        "--db-json", default="/tmp/block10_data.json",
    )
    parser.add_argument(
        "--model", default="data/model_all.pkl",
    )
    parser.add_argument(
        "--template",
        default=str(
            Path.home()
            / "Autonomous-Assignment-Program-Manager"
            / "backend/data/BlockTemplate2_Official.xlsx"
        ),
    )
    parser.add_argument(
        "--structure",
        default=str(
            Path.home()
            / "Autonomous-Assignment-Program-Manager"
            / "backend/data/BlockTemplate2_Structure.xml"
        ),
    )
    parser.add_argument("--output", default="data/Block10_ML.xlsx")
    parser.add_argument("--block", type=int, default=10)
    parser.add_argument(
        "--mode",
        choices=["lookup-only", "hybrid-lookup", "rules-only", "hybrid", "ml-only"],
        default="lookup-only",
    )
    parser.add_argument(
        "--threshold", type=float, default=0.6,
        help="ML confidence threshold for hybrid mode (default: 0.6)",
    )
    parser.add_argument(
        "--truth", default=None,
        help="Ground truth features JSON for accuracy comparison",
    )
    parser.add_argument(
        "--overwrite-identity", action="store_true",
    )
    args = parser.parse_args()

    fill_template(
        args.db_json, args.model, args.template,
        args.structure, args.output, args.block,
        preserve_identity=not args.overwrite_identity,
        mode=args.mode, ml_threshold=args.threshold,
        truth_path=args.truth,
    )


if __name__ == "__main__":
    main()
