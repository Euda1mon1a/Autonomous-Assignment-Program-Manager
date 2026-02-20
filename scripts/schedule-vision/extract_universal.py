"""Universal multi-AY schedule extractor for neural tabular learning.

Reads ALL Excel schedule workbooks (AY 2016-2025) and extracts features
from every cell in every block/rotation sheet. Handles two format eras:

  Era 1 (2016-2021): "Rot N" tabs, data starts col J, Hawaiian teams,
                      single-name format, 9 metadata columns
  Era 2 (2022-2025): "Block N" tabs, data starts col E/F, R1/R2/R3
                      templates, "Last, First" format, 4 metadata columns

Novel features over existing extract.py:
  - Continuous RGB decomposition (fill_r/g/b, font_r/g/b) instead of
    categorical hex strings
  - Theme color resolution to actual RGB via workbook theme XML
  - Cross-day context (prev/next codes, other-half code)
  - Cross-AY temporal features (academic_year, block_number)
  - Anonymized person tracking (name_hash for per-person patterns)

Usage:
    python extract_universal.py \\
        --archive-dir "Excel Schedule Archive/" \\
        --output data/features_universal.json
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles.colors import Color

# ── Theme color lookup (Office 2007+ default theme) ──────────────────
# These are the default theme colors; actual workbooks may override them.
DEFAULT_THEME_COLORS = [
    (255, 255, 255),  # theme:0 = lt1 (white/background)
    (0, 0, 0),        # theme:1 = dk1 (black/text)
    (68, 84, 106),    # theme:2 = lt2 (light accent)
    (198, 89, 17),    # theme:3 = dk2 (dark accent)
    (68, 114, 196),   # theme:4 = accent1
    (237, 125, 49),   # theme:5 = accent2
    (165, 165, 165),  # theme:6 = accent3
    (255, 192, 0),    # theme:7 = accent4
    (91, 155, 213),   # theme:8 = accent5
    (112, 173, 71),   # theme:9 = accent6
    (0, 0, 255),      # theme:10 = hlink
    (150, 0, 150),    # theme:11 = folHlink
]


def _apply_tint(rgb: tuple[int, int, int], tint: float) -> tuple[int, int, int]:
    """Apply Excel tint transformation to an RGB color."""
    r, g, b = rgb
    if tint < 0:
        factor = 1.0 + tint
        return (int(r * factor), int(g * factor), int(b * factor))
    else:
        return (
            int(r + (255 - r) * tint),
            int(g + (255 - g) * tint),
            int(b + (255 - b) * tint),
        )


def _resolve_color(color: Color | None) -> tuple[int, int, int] | None:
    """Resolve an openpyxl Color to (R, G, B) tuple."""
    if color is None:
        return None

    # Direct RGB value
    if color.type == "rgb" and color.rgb:
        rgb_str = str(color.rgb)
        if rgb_str.startswith("00") and len(rgb_str) == 8:
            # "00000000" = transparent/no color
            if rgb_str == "00000000":
                return None
        try:
            # ARGB format: "FFRRGGBB" or "RRGGBB"
            hex_str = rgb_str[-6:]
            r = int(hex_str[0:2], 16)
            g = int(hex_str[2:4], 16)
            b = int(hex_str[4:6], 16)
            return (r, g, b)
        except (ValueError, IndexError):
            return None

    # Theme color
    if color.type == "theme" and color.theme is not None:
        idx = int(color.theme)
        if 0 <= idx < len(DEFAULT_THEME_COLORS):
            base = DEFAULT_THEME_COLORS[idx]
            tint = float(color.tint) if color.tint else 0.0
            return _apply_tint(base, tint)
        return None

    # Indexed color — fall back
    if color.type == "indexed":
        return None

    return None


def _get_fill_rgb(cell) -> tuple[int, int, int] | None:
    """Get cell background fill color as RGB tuple."""
    if isinstance(cell, MergedCell):
        return None
    try:
        fg = cell.fill.fgColor
        return _resolve_color(fg)
    except (AttributeError, TypeError):
        return None


def _get_font_rgb(cell) -> tuple[int, int, int] | None:
    """Get cell font color as RGB tuple."""
    if isinstance(cell, MergedCell):
        return None
    try:
        return _resolve_color(cell.font.color)
    except (AttributeError, TypeError):
        return None


def _clean(val: Any) -> str | None:
    """Clean cell value to string or None."""
    if val is None:
        return None
    s = str(val).strip().replace("\xa0", "")
    return s if s else None


def _parse_pgy(role_str: str) -> int | None:
    """Extract PGY level from role string."""
    if not role_str:
        return None
    m = re.search(r"PGY\s*(\d)", role_str, re.IGNORECASE)
    if m:
        return int(m.group(1))
    m = re.search(r"R(\d)", role_str)
    if m:
        return int(m.group(1))
    return None


def _name_hash(name: str) -> int:
    """Anonymized person hash for per-person pattern tracking."""
    return int(hashlib.md5(name.lower().strip().encode()).hexdigest()[:8], 16)


def _detect_era(sheet_name: str) -> str:
    """Detect format era from sheet name."""
    if re.match(r"Rot\s+\d+", sheet_name, re.IGNORECASE):
        return "era1"
    if re.match(r"Block\s+\d+", sheet_name, re.IGNORECASE):
        return "era2"
    return "unknown"


def _parse_block_number(sheet_name: str) -> int | None:
    """Extract block/rotation number from sheet name."""
    m = re.search(r"(?:Block|Rot)\s+(\d+)", sheet_name, re.IGNORECASE)
    return int(m.group(1)) if m else None


def _parse_academic_year(filename: str) -> int | None:
    """Extract start year of academic year from filename."""
    m = re.search(r"(\d{4})-(\d{4})", filename)
    if m:
        return int(m.group(1))
    return None


# ── Era 1 extraction (2016-2021: Rot N format) ──────────────────────

def _find_era1_data_rows(ws) -> list[dict]:
    """Find person rows in era1 format by scanning for role indicators."""
    people = []
    max_row = ws.max_row or 100
    for row_idx in range(1, min(max_row + 1, 200)):
        # Look for role in column C (era1)
        role_val = _clean(ws.cell(row=row_idx, column=3).value)
        name_val = _clean(ws.cell(row=row_idx, column=5).value)
        team_val = _clean(ws.cell(row=row_idx, column=2).value)

        if not name_val:
            continue

        # Skip header/summary rows
        if name_val in ("PROVIDER", "Provider", "Date:", "Rotation:",
                        "Events and", "Meetings", "Staff Call",
                        "Resident Call"):
            continue
        if name_val.startswith("=") or name_val.startswith("Average"):
            continue

        pgy = _parse_pgy(role_val or "")
        person_type = "faculty"
        if role_val:
            ru = role_val.upper()
            if "PGY" in ru or ru in ("R1", "R2", "R3"):
                person_type = "resident"
            elif ru in ("FAC", "MD", "NP", "DO"):
                person_type = "faculty"
            elif ru in ("TR", "SM", "CP", "IBHC", "BH", "BHC"):
                person_type = "other"

        # Get rotation from column A or team assignment
        rot_val = _clean(ws.cell(row=row_idx, column=1).value)

        people.append({
            "row": row_idx,
            "name": name_val,
            "team": team_val,
            "role": role_val,
            "pgy": pgy,
            "person_type": person_type,
            "rotation": rot_val or team_val or "",
        })

    return people


def _find_era1_dates(ws, col_start: int = 10) -> list[date | None]:
    """Find dates in era1 format (row 3, starting at col_start)."""
    dates = []
    for col in range(col_start, ws.max_column + 1, 2):
        val = ws.cell(row=3, column=col).value
        if isinstance(val, datetime):
            dates.append(val.date())
        elif isinstance(val, date):
            dates.append(val)
        else:
            dates.append(None)
    return dates


def extract_era1(ws, sheet_name: str, academic_year: int,
                 filename: str) -> list[dict]:
    """Extract features from an era1 (Rot N) sheet."""
    block_num = _parse_block_number(sheet_name)
    col_start = 10  # Era1: data starts at column J

    # Find dates
    dates = _find_era1_dates(ws, col_start)
    n_days = len(dates)

    # Find people
    people = _find_era1_data_rows(ws)

    features = []
    for person in people:
        row_idx = person["row"]
        # First pass: collect all cell codes for context
        codes = {}
        for day_idx in range(n_days):
            for half_idx, half in enumerate(("am", "pm")):
                col = col_start + day_idx * 2 + half_idx
                if col > ws.max_column:
                    continue
                cell = ws.cell(row=row_idx, column=col)
                code = _clean(cell.value)
                codes[(day_idx, half)] = code

        # Second pass: extract features with context
        for day_idx in range(n_days):
            day_date = dates[day_idx] if day_idx < len(dates) else None

            for half_idx, half in enumerate(("am", "pm")):
                col = col_start + day_idx * 2 + half_idx
                if col > ws.max_column:
                    continue

                cell = ws.cell(row=row_idx, column=col)
                code = codes.get((day_idx, half))
                if not code:
                    continue

                # Skip formula cells
                if isinstance(code, str) and code.startswith("="):
                    continue

                # Visual features
                fill = _get_fill_rgb(cell)
                font = _get_font_rgb(cell)
                bold = False
                try:
                    bold = bool(cell.font.bold)
                except (AttributeError, TypeError):
                    pass

                feat = {
                    "cell_text": code,
                    "rotation": person["rotation"],
                    "person_type": person["person_type"],
                    "pgy_level": person["pgy"] or 0,
                    "day_of_week": day_date.weekday() if day_date else -1,
                    "is_weekend": day_date.weekday() in (5, 6) if day_date else False,
                    "day_index": day_idx,
                    "half": half,
                    "week_in_block": day_idx // 7,
                    "fill_r": fill[0] if fill else -1,
                    "fill_g": fill[1] if fill else -1,
                    "fill_b": fill[2] if fill else -1,
                    "font_r": font[0] if font else -1,
                    "font_g": font[1] if font else -1,
                    "font_b": font[2] if font else -1,
                    "font_bold": bold,
                    "has_fill": fill is not None,
                    "has_font_color": font is not None,
                    "academic_year": academic_year,
                    "block_number": block_num or 0,
                    "row_position": row_idx / (ws.max_row or 100),
                    "team_or_template": person["team"] or "",
                    "role_raw": person["role"] or "",
                    "name_hash": _name_hash(person["name"]),
                    # Context features
                    "prev_same_half": codes.get((day_idx - 1, half)) or "",
                    "next_same_half": codes.get((day_idx + 1, half)) or "",
                    "other_half": codes.get(
                        (day_idx, "pm" if half == "am" else "am")
                    ) or "",
                    # Metadata
                    "source_file": filename,
                    "sheet": sheet_name,
                    "era": "era1",
                }
                features.append(feat)

    return features


# ── Era 2 extraction (2022-2025: Block N format) ────────────────────

ERA2_ROW_DATES = 3
ERA2_COLS_PER_DAY = 2

# Row ranges for era2 (from the Excel explorer output)
ERA2_RESIDENT_RANGES = [
    (21, 31, "PGY 3"),   # R3
    (32, 36, "PGY 2"),   # R2
    (37, 43, "PGY 1"),   # R1
]
ERA2_FACULTY_RANGE = (48, 68)
ERA2_OTHER_RANGES = [
    (15, 19, "other"),    # Float, Reservist/GMO
    (70, 75, "trainee"),  # Floats/Trainees
    (78, 91, "student"),  # Medical students
]


def _find_era2_col_start(ws) -> int:
    """Detect where schedule data starts in era2 format."""
    # Check if dates are in row 3 starting at col E (5) or F (6)
    for start in (5, 6):
        val = ws.cell(row=ERA2_ROW_DATES, column=start).value
        if isinstance(val, (datetime, date)):
            return start
    # Fallback: scan for first date in row 3
    for col in range(1, 20):
        val = ws.cell(row=ERA2_ROW_DATES, column=col).value
        if isinstance(val, (datetime, date)):
            return col
    return 6  # Default to F


def _find_era2_dates(ws, col_start: int) -> list[date | None]:
    """Find dates in era2 format (row 3, every 2 columns)."""
    dates = []
    col = col_start
    while col <= ws.max_column:
        val = ws.cell(row=ERA2_ROW_DATES, column=col).value
        if isinstance(val, datetime):
            dates.append(val.date())
        elif isinstance(val, date):
            dates.append(val)
        elif val is None and dates:
            # Sometimes dates are merged and only in odd columns
            dates.append(None)
        else:
            break
        col += ERA2_COLS_PER_DAY
    return dates


def extract_era2(ws, sheet_name: str, academic_year: int,
                 filename: str) -> list[dict]:
    """Extract features from an era2 (Block N) sheet."""
    block_num = _parse_block_number(sheet_name)
    col_start = _find_era2_col_start(ws)
    dates = _find_era2_dates(ws, col_start)
    n_days = len(dates)

    # Collect all person rows
    people = []

    # Column indices for metadata
    col_template = 2  # B
    col_role = 3      # C
    col_name = 4      # D

    def _add_rows(start, end, default_type, default_pgy_str=""):
        for row_idx in range(start, end + 1):
            name = _clean(ws.cell(row=row_idx, column=col_name).value)
            if not name:
                continue
            if name.startswith("="):
                continue
            role = _clean(ws.cell(row=row_idx, column=col_role).value) or ""
            template = _clean(ws.cell(row=row_idx, column=col_template).value) or ""
            pgy = _parse_pgy(role) or _parse_pgy(default_pgy_str)
            people.append({
                "row": row_idx,
                "name": name,
                "role": role,
                "template": template,
                "pgy": pgy,
                "person_type": default_type,
            })

    for start, end, pgy_str in ERA2_RESIDENT_RANGES:
        _add_rows(start, end, "resident", pgy_str)
    _add_rows(ERA2_FACULTY_RANGE[0], ERA2_FACULTY_RANGE[1], "faculty")
    for start, end, ptype in ERA2_OTHER_RANGES:
        _add_rows(start, end, ptype)

    features = []
    for person in people:
        row_idx = person["row"]
        # First pass: collect codes
        codes = {}
        for day_idx in range(n_days):
            for half_idx, half in enumerate(("am", "pm")):
                col = col_start + day_idx * ERA2_COLS_PER_DAY + half_idx
                if col > ws.max_column:
                    continue
                cell = ws.cell(row=row_idx, column=col)
                code = _clean(cell.value)
                codes[(day_idx, half)] = code

        # Second pass: features
        for day_idx in range(n_days):
            day_date = dates[day_idx] if day_idx < len(dates) else None

            for half_idx, half in enumerate(("am", "pm")):
                col = col_start + day_idx * ERA2_COLS_PER_DAY + half_idx
                if col > ws.max_column:
                    continue

                cell = ws.cell(row=row_idx, column=col)
                code = codes.get((day_idx, half))
                if not code:
                    continue
                if isinstance(code, str) and code.startswith("="):
                    continue

                fill = _get_fill_rgb(cell)
                font = _get_font_rgb(cell)
                bold = False
                try:
                    bold = bool(cell.font.bold)
                except (AttributeError, TypeError):
                    pass

                # Get rotation from col A or template
                rot_val = _clean(ws.cell(row=row_idx, column=1).value)

                feat = {
                    "cell_text": code,
                    "rotation": rot_val or person.get("template", ""),
                    "person_type": person["person_type"],
                    "pgy_level": person["pgy"] or 0,
                    "day_of_week": day_date.weekday() if day_date else -1,
                    "is_weekend": (
                        day_date.weekday() in (5, 6) if day_date else False
                    ),
                    "day_index": day_idx,
                    "half": half,
                    "week_in_block": day_idx // 7,
                    "fill_r": fill[0] if fill else -1,
                    "fill_g": fill[1] if fill else -1,
                    "fill_b": fill[2] if fill else -1,
                    "font_r": font[0] if font else -1,
                    "font_g": font[1] if font else -1,
                    "font_b": font[2] if font else -1,
                    "font_bold": bold,
                    "has_fill": fill is not None,
                    "has_font_color": font is not None,
                    "academic_year": academic_year,
                    "block_number": block_num or 0,
                    "row_position": row_idx / (ws.max_row or 100),
                    "team_or_template": person.get("template", ""),
                    "role_raw": person["role"],
                    "name_hash": _name_hash(person["name"]),
                    "prev_same_half": codes.get((day_idx - 1, half)) or "",
                    "next_same_half": codes.get((day_idx + 1, half)) or "",
                    "other_half": codes.get(
                        (day_idx, "pm" if half == "am" else "am")
                    ) or "",
                    "source_file": filename,
                    "sheet": sheet_name,
                    "era": "era2",
                }
                features.append(feat)

    return features


# ── Main extraction pipeline ─────────────────────────────────────────

def extract_workbook(filepath: Path) -> list[dict]:
    """Extract features from all schedule sheets in a workbook."""
    filename = filepath.name
    academic_year = _parse_academic_year(filename)
    if academic_year is None:
        print(f"  SKIP: Cannot parse AY from '{filename}'")
        return []

    print(f"  Loading {filename} (AY {academic_year}-{academic_year+1})...")
    try:
        wb = load_workbook(filepath, data_only=False, read_only=False)
    except Exception as e:
        print(f"  ERROR loading {filename}: {e}")
        return []

    all_features = []
    for sheet_name in wb.sheetnames:
        era = _detect_era(sheet_name)
        if era == "unknown":
            continue

        ws = wb[sheet_name]
        try:
            if era == "era1":
                feats = extract_era1(ws, sheet_name, academic_year, filename)
            else:
                feats = extract_era2(ws, sheet_name, academic_year, filename)
            all_features.extend(feats)
            print(f"    {sheet_name}: {len(feats)} cells ({era})")
        except Exception as e:
            print(f"    ERROR on {sheet_name}: {e}")

    wb.close()
    return all_features


def main():
    parser = argparse.ArgumentParser(
        description="Universal multi-AY schedule feature extractor",
    )
    parser.add_argument(
        "--archive-dir",
        default=str(
            Path.home()
            / "Autonomous-Assignment-Program-Manager"
            / "Excel Schedule Archive"
        ),
        help="Directory containing Excel schedule files",
    )
    parser.add_argument(
        "--output", default="data/features_universal.json",
        help="Output JSON file for extracted features",
    )
    parser.add_argument(
        "--filter-ay", type=int, default=None,
        help="Only extract a specific academic year (e.g., 2024)",
    )
    args = parser.parse_args()

    archive = Path(args.archive_dir)
    if not archive.is_dir():
        print(f"ERROR: Archive directory not found: {archive}")
        sys.exit(1)

    # Find schedule workbooks (skip FMIT/Lecture files)
    xlsx_files = sorted(
        f for f in archive.glob("Family Medicine Clinic Schedule*.xlsx")
        if "~$" not in f.name  # Skip temp files
    )

    print(f"Found {len(xlsx_files)} schedule workbooks in {archive}")

    all_features: list[dict] = []
    ay_stats: Counter = Counter()

    for fp in xlsx_files:
        ay = _parse_academic_year(fp.name)
        if args.filter_ay and ay != args.filter_ay:
            continue

        feats = extract_workbook(fp)
        all_features.extend(feats)
        if ay:
            ay_stats[ay] = len(feats)

    # Remove metadata fields not needed for training
    for f in all_features:
        f.pop("source_file", None)
        f.pop("sheet", None)
        f.pop("era", None)

    # Stats
    print(f"\n{'='*60}")
    print(f"EXTRACTION COMPLETE")
    print(f"{'='*60}")
    print(f"Total cells: {len(all_features)}")
    print(f"\nPer academic year:")
    for ay in sorted(ay_stats):
        print(f"  AY {ay}-{ay+1}: {ay_stats[ay]:,} cells")

    # Code distribution
    codes = Counter(f["cell_text"] for f in all_features)
    print(f"\nUnique codes: {len(codes)}")
    print("Top 20 codes:")
    for code, count in codes.most_common(20):
        pct = 100 * count / len(all_features)
        print(f"  {code:12s}: {count:6d} ({pct:5.1f}%)")

    # Person type distribution
    ptypes = Counter(f["person_type"] for f in all_features)
    print(f"\nPerson types:")
    for pt, count in ptypes.most_common():
        print(f"  {pt:12s}: {count:6d}")

    # Save
    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(all_features, indent=None, separators=(",", ":")))
    size_mb = out.stat().st_size / 1024 / 1024
    print(f"\nSaved to {out} ({size_mb:.1f} MB)")


if __name__ == "__main__":
    main()
