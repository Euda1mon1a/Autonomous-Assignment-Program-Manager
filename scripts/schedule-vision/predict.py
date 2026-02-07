"""Apply trained model to predict display codes for a schedule block.

Takes a DB export JSON (same format as HalfDayJSONExporter output),
applies the trained Random Forest model, and writes predicted display
codes alongside the original DB codes.

Usage:
    python predict.py \
        --db-json /tmp/block11_data.json \
        --model data/model_all.pkl \
        --output data/block11_predicted.json
"""

from __future__ import annotations

import argparse
import json
from datetime import date
from pathlib import Path

import joblib
import numpy as np


def _last_name(name: str) -> str:
    name = name.strip()
    if "," in name:
        return name.split(",")[0].strip().lower()
    parts = name.split()
    return parts[-1].lower() if parts else name.lower()


def predict_schedule(
    db_json_path: str,
    model_path: str,
    workbook_path: str | None = None,
    output_path: str | None = None,
) -> dict:
    """Predict display codes for a schedule block.

    Args:
        db_json_path: Path to DB export JSON.
        model_path: Path to trained model (.pkl).
        workbook_path: Optional handjam workbook for visual features.
        output_path: Optional output path for predicted JSON.

    Returns:
        Modified data dict with predictions added to each day.
    """
    model_data = joblib.load(model_path)
    rf = model_data["rf"]
    encoders = model_data["encoders"]
    label_enc = encoders["__label__"]

    cat_fields = [
        "rotation1", "rotation2", "template", "role", "db_code",
        "template_expected", "db_other_half", "fill_rgb", "font_rgb",
        "person_type",
    ]

    data = json.loads(Path(db_json_path).read_text())
    block_start = date.fromisoformat(data["block_start"])

    # Parse templates for lookup if workbook provided
    templates: dict = {}
    if workbook_path:
        try:
            from extract import parse_templates
            templates = parse_templates(Path(workbook_path).expanduser())
        except Exception:
            pass

    # If workbook provided, extract visual features
    visual_features: dict = {}
    if workbook_path:
        try:
            from extract import extract_block, _parse_block_number

            wb_path = Path(workbook_path).expanduser()
            # Determine block tab name
            block_num = None
            for person in data.get("residents", []):
                if person.get("days"):
                    first_date = date.fromisoformat(person["days"][0]["date"])
                    # Rough block number from date
                    break

            # Try common tab names
            from openpyxl import load_workbook
            wb = load_workbook(wb_path, read_only=True)
            for sheet in wb.sheetnames:
                if "Block" in sheet and "Deprecated" not in sheet:
                    bn = _parse_block_number(sheet)
                    if bn is not None:
                        # Check if dates overlap
                        features = extract_block(wb_path, sheet)
                        for f in features:
                            if f.get("date") and f.get("date") == data["block_start"]:
                                visual_features = {
                                    (f["person"], f["day_index"], f["half"]): f
                                    for f in features
                                }
                                break
                        if visual_features:
                            break
            wb.close()
        except Exception as e:
            print(f"  Warning: Could not extract visual features: {e}")

    total_predicted = 0
    total_changed = 0

    for person_type, key in [("resident", "residents"), ("faculty", "faculty")]:
        for person in data.get(key, []):
            name = person["name"]
            rot1 = person.get("rotation1", "")
            rot2 = person.get("rotation2", "")
            pgy = person.get("pgy") or 0

            for i, day in enumerate(person.get("days", [])):
                day_date = date.fromisoformat(day["date"])
                is_weekend = day_date.weekday() in (5, 6)

                for half in ["am", "pm"]:
                    db_code = day.get(half, "")
                    if not db_code:
                        continue

                    # Build feature dict
                    vf = visual_features.get((name, i, half), {})

                    # Template lookup
                    tpl_expected = ""
                    if templates:
                        from extract import lookup_template_code
                        tpl_expected = lookup_template_code(
                            rot1, pgy, data.get("block_number"),
                            day_date.weekday(), half, templates
                        ) or ""

                    # DB other-half code
                    other_half = "pm" if half == "am" else "am"
                    db_other_half = day.get(other_half, "")

                    feat = {
                        "pgy_level": pgy,
                        "block_number": data.get("block_number", 0),
                        "day_of_week": day_date.weekday(),
                        "is_weekend": is_weekend,
                        "day_index": i,
                        "half": half,
                        "font_bold": vf.get("font_bold", False),
                        "db_matches_template": db_code == tpl_expected if tpl_expected else False,
                        "week_in_block": i // 7,
                        "rotation1": rot1,
                        "rotation2": rot2,
                        "template": vf.get("template", ""),
                        "role": vf.get("role", ""),
                        "db_code": db_code,
                        "template_expected": tpl_expected,
                        "db_other_half": db_other_half,
                        "fill_rgb": vf.get("fill_rgb", ""),
                        "font_rgb": vf.get("font_rgb", ""),
                        "person_type": person_type,
                    }

                    # Encode
                    row = _encode_single(feat, encoders, cat_fields)
                    pred_idx = rf.predict([row])[0]
                    pred_code = label_enc.inverse_transform([pred_idx])[0]

                    # Store prediction
                    day[f"{half}_predicted"] = pred_code
                    total_predicted += 1
                    if pred_code != db_code:
                        total_changed += 1

    print(f"\n  Predicted {total_predicted} cells")
    print(f"  Changed {total_changed} codes ({100*total_changed/total_predicted:.1f}%)")
    print(f"  Kept {total_predicted - total_changed} codes unchanged")

    if output_path:
        out = Path(output_path)
        out.parent.mkdir(parents=True, exist_ok=True)
        out.write_text(json.dumps(data, indent=2))
        print(f"  Wrote predictions to {out}")

    return data


def _encode_single(feat: dict, encoders: dict, cat_fields: list[str]) -> np.ndarray:
    """Encode a single feature dict for prediction."""
    row = []

    # Numeric features (same order as learn.py NUM_FIELDS)
    row.append(feat.get("pgy_level") or 0)
    row.append(feat.get("block_number") or 0)
    val = feat.get("day_of_week")
    row.append(val if val is not None else -1)
    row.append(1 if feat.get("is_weekend") else 0)
    row.append(feat.get("day_index") or 0)
    row.append(1 if feat.get("half") == "pm" else 0)
    row.append(1 if feat.get("font_bold") else 0)
    row.append(1 if feat.get("db_matches_template") else 0)
    row.append(feat.get("week_in_block") or 0)

    # Categorical features
    for field in cat_fields:
        val = feat.get(field) or "__NONE__"
        enc = encoders[field]
        if val in enc.classes_:
            row.append(enc.transform([val])[0])
        else:
            row.append(len(enc.classes_))

    return np.array(row, dtype=np.float32)


def main():
    parser = argparse.ArgumentParser(
        description="Predict display codes using trained model",
    )
    parser.add_argument("--db-json", required=True,
                        help="DB export JSON (HalfDayJSONExporter format)")
    parser.add_argument("--model", default="data/model_all.pkl",
                        help="Trained model from learn.py")
    parser.add_argument("--workbook", default=None,
                        help="Optional handjam workbook for visual features")
    parser.add_argument("--output", default=None,
                        help="Output JSON with predictions")
    args = parser.parse_args()

    if not args.output:
        stem = Path(args.db_json).stem
        args.output = f"data/{stem}_predicted.json"

    predict_schedule(args.db_json, args.model, args.workbook, args.output)


if __name__ == "__main__":
    main()
