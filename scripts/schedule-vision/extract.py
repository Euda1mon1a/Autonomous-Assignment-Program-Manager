"""Extract visual features from handjam workbook for ML training.

Reads each schedule cell's text AND visual properties (fill color, font color)
along with contextual features (rotation, PGY level, block, day of week).
Aligns with DB export to create labeled training examples.

Supports multi-block extraction for larger training sets.

Usage:
    # Single block
    python extract.py \
        --workbook ~/Downloads/"Current AY 25-26 pulled 29 JAN 2026.xlsx" \
        --db-json /tmp/block10_data.json \
        --block "Block 10"

    # Multiple blocks (auto-discovers matching DB JSON files)
    python extract.py \
        --workbook ~/Downloads/"Current AY 25-26 pulled 29 JAN 2026.xlsx" \
        --blocks "Block 7" "Block 8" "Block 9" "Block 10" \
        --db-dir /tmp/
"""

from __future__ import annotations

import argparse
import json
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles.colors import Color

# ── Template lookup ──────────────────────────────────────────
# Maps (rotation1_upper, pgy) → template row name on Templates sheet.
# When pgy=None, matches any PGY level.
ROTATION_TO_TEMPLATE: dict[tuple[str, int | None], str] = {
    # PGY 3
    ("GERI", 3): "Geriatrics 7/20",
    ("FMC", 3): "FMC 4/21",
    ("FMIT", 3): "FMIT PGY-3",
    ("NF", 3): "Night Float * / Cards 7/26",
    ("NICU", 3): "NICU/NF 7/26",
    ("EM", 3): "Peds ER 7/20",
    ("HILO", 3): "ICU (Hilo)",
    ("PSYCH", 3): "Psychiatry*9/24",
    ("DERM", 3): "Derm (2 weeks)",
    ("JAPAN", 3): "Okinawa, Japan",
    ("MS", 3): "Med Select",
    ("NF/ENDO", 3): "Night Float * / Cards 7/26",
    ("NEURO/NF", 3): "Night Float * / Cards 7/26",
    # PGY 2
    ("CARD", 2): "Cardiology 08/22",
    ("PM", 2): "CM/PM & RSH 4/21",
    ("ER", 2): "Emergency Room R2 7/20",
    ("FMC", 2): "FMC*10/18",
    ("FMIT", 2): "FMIT 2A*12/15",
    ("GYN", 2): "GYN*3/21",
    ("TAMC-LD", 2): "L&D Pit-Boss Nights*06/16",
    ("LDNF", 2): "L&D Pit-Boss Nights*06/16",
    ("PEDS-S", 2): "Peds-SPC*4/21",
    ("PEDS-C", 2): "Peds Clinic (PedC)",
    ("PROC", 2): "Procedures 4/21",
    ("SURG", 2): "Surg Exp 7/24",
    ("SM", 2): "Sports Medicine R2 4/21",
    ("MSK-SEL", 2): "MSK Select",
    ("NF", 2): "NF/NICU 7/24",
    ("ELEC", 2): "Elective, Generic 1",
    ("PSYCH", 2): "Psychiatry*9/24",
    # PGY 1
    ("ER", 1): "Emergency Room",
    ("FMIT", 1): "FMIT 1",
    ("ICU", 1): "ICU 1/20",
    ("KAP-LD", 1): "KAP L&D",
    ("TAMC-LD", 1): "L&D 1  7/16",
    ("IMW", 1): "Medicine Ward",
    ("NBN", 1): "Newborn Nursery 8/12",
    ("PEDSW", 1): "Peds Ward (PedW) - Day schedule",
    ("PEDS-C", 1): "Peds Clinic (PedC)",
    ("PNF", 1): "NF FMIT",
    ("GSW", 1): "General Surgery Ward",
    ("SM", 1): "MSK Select",
    # PGY-independent fallbacks
    ("HILO", None): "ICU (Hilo)",
    ("JAPAN", None): "Okinawa, Japan",
    ("PSYCH", None): "Psychiatry*9/24",
    ("FMIT", None): "FMIT 1",
    ("NF", None): "Night Float * / Cards 7/26",
    ("IMW", None): "Medicine Ward",
    ("KAP-LD", None): "KAP L&D",
    ("NBN", None): "Newborn Nursery 8/12",
    ("PEDSW", None): "Peds Ward (PedW) - Day schedule",
    ("GYN", None): "GYN*3/21",
    ("PROC", None): "Procedures 4/21",
    ("SURG", None): "Surg Exp 7/24",
    ("SM", None): "Sports Medicine R2 4/21",
    ("FMC", None): "FMC*10/18",
    ("EM", None): "Emergency Room R2 7/20",
    ("GERI", None): "Geriatrics 7/20",
    ("NICU", None): "NICU/NF 7/26",
    ("LDNF", None): "L&D Pit-Boss Nights*06/16",
    ("TAMC-LD", None): "L&D 1  7/16",
    ("ELEC", None): "Elective, Generic 1",
    ("PEDS-S", None): "Peds-SPC*4/21",
    ("MSK-SEL", None): "MSK Select",
    ("PNF", None): "NF FMIT",
    ("NF/ENDO", None): "Night Float * / Cards 7/26",
    ("NEURO/NF", None): "Night Float * / Cards 7/26",
}

# Columns on Templates sheet → (day_of_week, half)
_TEMPLATE_COL_MAP = {
    2: (3, "am"), 3: (3, "pm"),     # Thu
    4: (4, "am"), 5: (4, "pm"),     # Fri
    6: (5, "am"), 7: (5, "pm"),     # Sat
    8: (6, "am"), 9: (6, "pm"),     # Sun
    10: (0, "am"), 11: (0, "pm"),   # Mon
    12: (1, "am"), 13: (1, "pm"),   # Tue
    14: (2, "am"), 15: (2, "pm"),   # Wed
    16: (3, "am"), 17: (3, "pm"),   # Thu (2nd)
    18: (4, "am"), 19: (4, "pm"),   # Fri (2nd)
}

_TEMPLATE_CODES = {"R1", "R2", "R3", "C19", "Ca9", "SPEC", "ADJ", "FAC"}


def parse_templates(wb_path: Path) -> dict[str, dict[tuple[int, str], str]]:
    """Parse Templates sheet → {template_name: {(dow, half): code}}.

    The Templates sheet contains weekly patterns for each rotation.
    """
    wb = load_workbook(wb_path, data_only=True)
    if "Templates" not in wb.sheetnames:
        wb.close()
        return {}
    ws = wb["Templates"]

    templates: dict[str, dict[tuple[int, str], str]] = {}

    for row in range(4, 80):
        name = ws.cell(row=row, column=1).value
        if not name or name == NBSP:
            continue
        name_str = str(name).strip()
        # Skip section headers
        if name_str.startswith("PGY") or "ARCHIVED" in name_str or "Elective Opp" in name_str:
            continue

        pattern: dict[tuple[int, str], str] = {}
        for col, (dow, half) in _TEMPLATE_COL_MAP.items():
            val = ws.cell(row=row, column=col).value
            if val and str(val).strip() and val != NBSP:
                key = (dow, half)
                # For Thu/Fri (appear twice), prefer later (standard week) version
                if key not in pattern or col >= 16:
                    pattern[key] = str(val).strip()

        if pattern:
            templates[name_str] = pattern

    wb.close()
    return templates


def lookup_template_code(
    rotation1: str | None,
    pgy: int | None,
    block_num: int | None,
    day_of_week: int | None,
    half: str | None,
    templates: dict[str, dict[tuple[int, str], str]],
) -> str | None:
    """Look up expected display code from Templates sheet.

    Returns the code or None if no match.
    """
    if not rotation1 or day_of_week is None or not half:
        return None

    rot_upper = rotation1.upper().strip()

    # Special case: FMC PGY1 has block-dependent template
    if rot_upper == "FMC" and pgy == 1 and block_num is not None:
        if block_num <= 6:
            tname = "FMC for Block 1-6 (ver Aug23)"
        else:
            tname = "FMC for Block 7-13 (ver Aug23)"
        tpl = templates.get(tname)
        if tpl:
            return tpl.get((day_of_week, half))

    # Try (rotation, pgy) first, then (rotation, None) fallback
    tname = ROTATION_TO_TEMPLATE.get((rot_upper, pgy))
    if not tname:
        tname = ROTATION_TO_TEMPLATE.get((rot_upper, None))
    if not tname:
        return None

    tpl = templates.get(tname)
    if not tpl:
        return None

    return tpl.get((day_of_week, half))


# Layout constants (same as extract_ground_truth.py)
ROW_DATES = 3
ROW_RESIDENTS_START = 9
ROW_RESIDENTS_END = 25
ROW_FACULTY_START = 31
ROW_FACULTY_END = 47
COL_ROTATION1 = 1
COL_ROTATION2 = 2
COL_TEMPLATE = 3
COL_ROLE = 4
COL_NAME = 5
COL_SCHEDULE_START = 6
COLS_PER_DAY = 2
MAX_DAYS = 28
NBSP = "\xa0"


def _clean(val) -> str | None:
    if val is None or val == NBSP:
        return None
    s = str(val).strip()
    return s if s else None


def _color_hex(color: Color | None) -> str | None:
    """Extract hex color string from openpyxl Color object."""
    if color is None:
        return None
    try:
        rgb = color.rgb
        if rgb and isinstance(rgb, str) and rgb != "00000000":
            return rgb
    except (AttributeError, TypeError):
        pass
    try:
        theme = color.theme
        if theme is not None:
            return f"theme:{theme}"
    except (AttributeError, TypeError):
        pass
    return None


def _fill_hex(cell) -> str | None:
    """Extract fill color from cell."""
    try:
        fill = cell.fill
        if fill and fill.fgColor:
            return _color_hex(fill.fgColor)
    except (AttributeError, TypeError):
        pass
    return None


def _font_hex(cell) -> str | None:
    """Extract font color from cell."""
    try:
        font = cell.font
        if font and font.color:
            return _color_hex(font.color)
    except (AttributeError, TypeError):
        pass
    return None


def _last_name(name: str) -> str:
    name = name.strip().rstrip("*").strip()
    if "," in name:
        return name.split(",")[0].strip().lower()
    parts = name.split()
    return parts[-1].lower() if parts else name.lower()


def _parse_pgy(role_str: str | None) -> int | None:
    if not role_str:
        return None
    role = str(role_str).strip().upper()
    if "PGY" in role or "PGY " in role:
        for part in role.replace("PGY", "").strip().split():
            try:
                return int(part)
            except ValueError:
                continue
    if role in ("1", "2", "3"):
        return int(role)
    return None


def _parse_block_number(tab_name: str) -> int | None:
    try:
        return int(tab_name.replace("Block ", "").replace("Deprecated ", ""))
    except ValueError:
        return None


def extract_block(
    wb_path: Path,
    tab_name: str,
    db_data: dict | None = None,
    templates: dict[str, dict] | None = None,
) -> list[dict]:
    """Extract visual features from one block tab.

    Returns list of feature dicts, one per half-day cell.
    """
    wb = load_workbook(wb_path, data_only=True)
    ws = wb[tab_name]
    block_num = _parse_block_number(tab_name)

    # Extract dates from row 3
    dates: list[date | None] = []
    for i in range(MAX_DAYS):
        col = COL_SCHEDULE_START + i * COLS_PER_DAY
        v = ws.cell(row=ROW_DATES, column=col).value
        if isinstance(v, datetime):
            dates.append(v.date())
        elif isinstance(v, date):
            dates.append(v)
        else:
            dates.append(None)

    # Build DB lookup by last name
    db_people: dict[str, dict] = {}
    if db_data:
        for person in db_data.get("residents", []) + db_data.get("faculty", []):
            ln = _last_name(person["name"])
            db_people[ln] = person

    features: list[dict] = []

    for row_start, row_end, person_type in [
        (ROW_RESIDENTS_START, ROW_RESIDENTS_END, "resident"),
        (ROW_FACULTY_START, ROW_FACULTY_END, "faculty"),
    ]:
        for row in range(row_start, row_end + 1):
            name_raw = _clean(ws.cell(row=row, column=COL_NAME).value)
            if not name_raw:
                continue
            # Skip schedule codes leaked into name column
            if len(name_raw) <= 4 and "," not in name_raw and name_raw.upper() == name_raw:
                continue

            rot1 = _clean(ws.cell(row=row, column=COL_ROTATION1).value)
            rot2 = _clean(ws.cell(row=row, column=COL_ROTATION2).value)
            template = _clean(ws.cell(row=row, column=COL_TEMPLATE).value)
            role = _clean(ws.cell(row=row, column=COL_ROLE).value)

            # Fix mis-detected columns: if col1 has a template code, shift
            if rot1 and rot1.upper() in _TEMPLATE_CODES:
                if not template:
                    template = rot1
                if not role and rot2:
                    role = rot2
                rot1 = None
                rot2 = None

            pgy = _parse_pgy(role)

            ln = _last_name(name_raw)
            db_person = db_people.get(ln, {})
            db_days = {i: d for i, d in enumerate(db_person.get("days", []))}

            # Use DB rotation if available (more reliable)
            db_rot1 = db_person.get("rotation1") or rot1
            db_rot2 = db_person.get("rotation2") or rot2

            for i in range(MAX_DAYS):
                d = dates[i] if i < len(dates) else None
                am_col = COL_SCHEDULE_START + i * COLS_PER_DAY
                pm_col = am_col + 1

                for half, col in [("am", am_col), ("pm", pm_col)]:
                    cell = ws.cell(row=row, column=col)
                    truth_code = _clean(cell.value)

                    # Get DB code for this cell
                    db_day = db_days.get(i, {})
                    db_code = db_day.get(half)

                    # Skip if both empty
                    if not truth_code and not db_code:
                        continue

                    # Get previous/next codes for context
                    prev_code = None
                    next_code = None
                    other_half_code = None

                    if half == "am":
                        other_cell = ws.cell(row=row, column=pm_col)
                        other_half_code = _clean(other_cell.value)
                    else:
                        other_cell = ws.cell(row=row, column=am_col)
                        other_half_code = _clean(other_cell.value)

                    if i > 0:
                        prev_col = COL_SCHEDULE_START + (i - 1) * COLS_PER_DAY + (0 if half == "am" else 1)
                        prev_code = _clean(ws.cell(row=row, column=prev_col).value)

                    if i < MAX_DAYS - 1:
                        next_col = COL_SCHEDULE_START + (i + 1) * COLS_PER_DAY + (0 if half == "am" else 1)
                        next_code = _clean(ws.cell(row=row, column=next_col).value)

                    # Template expected code lookup
                    dow = d.weekday() if d else None
                    tpl_expected = None
                    if templates:
                        tpl_expected = lookup_template_code(
                            db_rot1, pgy, block_num, dow, half, templates
                        )

                    # DB-derived context features (available at prediction time)
                    db_other_half = None
                    if db_day:
                        db_other_half = db_day.get("pm" if half == "am" else "am")
                    db_matches_tpl = (
                        db_code == tpl_expected
                        if db_code and tpl_expected
                        else None
                    )
                    week_in_block = i // 7 if i is not None else None

                    feat = {
                        "person": name_raw,
                        "person_type": person_type,
                        "row": row,
                        "rotation1": db_rot1,
                        "rotation2": db_rot2,
                        "template": template,
                        "role": role,
                        "pgy_level": pgy,
                        "block_number": block_num,
                        "day_index": i,
                        "date": d.isoformat() if d else None,
                        "day_of_week": dow,
                        "is_weekend": d.weekday() in (5, 6) if d else False,
                        "half": half,
                        "truth_code": truth_code,
                        "db_code": db_code,
                        "template_expected": tpl_expected,
                        "db_other_half": db_other_half,
                        "db_matches_template": db_matches_tpl,
                        "week_in_block": week_in_block,
                        "fill_rgb": _fill_hex(cell),
                        "font_rgb": _font_hex(cell),
                        "font_bold": cell.font.bold if cell.font else None,
                        "prev_code": prev_code,
                        "next_code": next_code,
                        "other_half_code": other_half_code,
                    }
                    features.append(feat)

    wb.close()
    return features


def _print_stats(features: list[dict], label: str = "") -> None:
    """Print extraction statistics."""
    total = len(features)
    with_truth = sum(1 for f in features if f["truth_code"])
    with_db = sum(1 for f in features if f["db_code"])
    with_both = sum(1 for f in features if f["truth_code"] and f["db_code"])
    matches = sum(1 for f in features
                  if f["truth_code"] and f["db_code"]
                  and f["truth_code"] == f["db_code"])
    with_color = sum(1 for f in features if f["fill_rgb"])

    if label:
        print(f"\n  [{label}]")
    print(f"  Total cells: {total}")
    print(f"  With truth code: {with_truth}")
    print(f"  With DB code: {with_db}")
    print(f"  With both: {with_both}")
    if with_both:
        print(f"  Exact matches: {matches} ({100*matches/with_both:.1f}%)")
    print(f"  With fill color: {with_color}")

    codes = sorted(set(f["truth_code"] for f in features if f["truth_code"]))
    print(f"  Unique truth codes: {len(codes)}")

    # Template lookup stats
    with_tpl = sum(1 for f in features if f.get("template_expected"))
    tpl_match = sum(1 for f in features
                    if f.get("template_expected") and f.get("truth_code")
                    and f["template_expected"] == f["truth_code"])
    print(f"  With template lookup: {with_tpl}")
    if with_tpl:
        print(f"  Template matches truth: {tpl_match}/{with_tpl} ({100*tpl_match/with_tpl:.1f}%)")

    colors = sorted(set(f["fill_rgb"] for f in features if f["fill_rgb"]))
    print(f"  Unique fill colors: {len(colors)}")


def main():
    parser = argparse.ArgumentParser(description="Extract visual features for ML")
    parser.add_argument("--workbook", required=True, help="Path to handjam .xlsx")
    parser.add_argument("--db-json", default=None, help="DB export JSON (single block)")
    parser.add_argument("--block", default=None, help="Single block tab name")
    parser.add_argument("--blocks", nargs="+", default=None,
                        help="Multiple block tab names")
    parser.add_argument("--db-dir", default=None,
                        help="Dir with blockN_data.json files")
    parser.add_argument("--output", default="data/features.json", help="Output JSON")
    args = parser.parse_args()

    wb_path = Path(args.workbook).expanduser()

    # Parse Templates sheet for lookup
    print("Parsing Templates sheet...")
    templates = parse_templates(wb_path)
    print(f"  Loaded {len(templates)} rotation templates")

    # Determine which blocks to extract
    if args.blocks:
        block_tabs = args.blocks
    elif args.block:
        block_tabs = [args.block]
    else:
        parser.error("Must specify --block or --blocks")

    all_features: list[dict] = []

    for tab in block_tabs:
        # Find matching DB JSON
        db_data = None
        block_num = _parse_block_number(tab)

        if args.db_json and len(block_tabs) == 1:
            db_data = json.loads(Path(args.db_json).read_text())
        elif args.db_dir and block_num is not None:
            db_path = Path(args.db_dir) / f"block{block_num}_data.json"
            if db_path.exists():
                db_data = json.loads(db_path.read_text())
                print(f"  Found DB data: {db_path.name}")
            else:
                print(f"  No DB data for {tab} (looked for {db_path})")
        elif args.db_json is None and block_num is not None:
            # Auto-discover in /tmp/
            db_path = Path(f"/tmp/block{block_num}_data.json")
            if db_path.exists():
                db_data = json.loads(db_path.read_text())
                print(f"  Auto-found DB data: {db_path}")

        print(f"\nExtracting features from {wb_path.name} / {tab}...")
        features = extract_block(wb_path, tab, db_data, templates)
        _print_stats(features, tab)
        all_features.extend(features)

    if len(block_tabs) > 1:
        print(f"\n{'='*50}")
        _print_stats(all_features, "COMBINED")

    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(all_features, indent=2))
    print(f"\n  Wrote {len(all_features)} feature vectors to {out}")


if __name__ == "__main__":
    main()
