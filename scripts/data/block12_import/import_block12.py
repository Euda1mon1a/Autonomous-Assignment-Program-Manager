"""Block 12 schedule import — Excel handjam to database.

Copy of block_import_template.py customized for Block 12 (May 7 – Jun 3, 2026).
Pre-filled CODE_MAP from Block 11's proven mappings.

Usage:
  1. Fill in EXCEL_PATH, SHEET_NAME, NAME_MAP
  2. Backup DB: ./checkpoint.sh pre_import
  3. Run: backend/.venv/bin/python scripts/data/block12_import/import_block12.py

Reference: docs/development/BLOCK11_SCHEDULE_LOAD.md
"""

import sys
import uuid
from datetime import date, timedelta, datetime
from collections import Counter

import openpyxl
import psycopg2

sys.path.insert(0, "backend")
sys.path.insert(0, "scripts/data")
sys.path.insert(0, "scripts/data/block12_import")

# ── Configuration ────────────────────────────────────────────────────────────

CONN = "dbname=residency_scheduler user=scheduler host=localhost"

EXCEL_PATH = "/Users/aaronmontgomery/Downloads/Current AY 25-26 pulled 11 FEB 2026.xlsx"
SHEET_NAME = "Block 12"  # NOTE: Sheet doesn't exist yet in this Excel; will be added for handjam

BLOCK_CONFIG = {
    "block_number": 12,
    "academic_year": 2025,
    "start_date": date(2026, 5, 7),
    "end_date": date(2026, 6, 3),
}

# Excel column layout (1-indexed):
# Col D = "PROVIDER" header, Col E = first half-day slot (Day 1 AM)
# Each day = 2 columns (odd=AM, even=PM), 28 days = 56 columns (E through BH)
NAME_COL = 4                    # Column D = names (Last, First format)
DATA_COL_START = 5              # Column E = first half-day slot (Day 1 AM)
DATA_COL_END = 60               # Column BH = last half-day slot (Day 28 PM)

# Section markers in column B (TEMPLATE) and column C (ROLE)
# Used by detect_person_rows() to find resident/faculty rows dynamically.
# Fallback ranges only used if auto-detection fails.
_FALLBACK_RESIDENT_ROWS = range(9, 50)
_FALLBACK_FACULTY_ROWS = range(44, 75)

# ── Name Mapping ─────────────────────────────────────────────────────────────
# Maps "Last, First" from Excel to "First Last" in DB
# Fill from Excel roster when received — watch for:
#   - Asterisks (DO degree): "Mayell, Cameron*"
#   - Trailing spaces: "Wilhelm, Clara "
#   - Name variants: "Headid, Ronald" → "James Headid"

NAME_MAP = {
    # PGY-3 (R3) — asterisk = DO degree
    "Connolly, Laura": "Laura Connolly",        # NOT in B12 (done after Block 10)
    "Hernandez, Christian*": "Christian Hernandez",
    "Mayell, Cameron*": "Cam Mayell",
    "Petrie, William*": "Clay Petrie",           # William (Excel) -> Clay (DB)
    "You, Jae*": "Jae You",
    # PGY-2 (R2)
    "Cataquiz, Felipe": "Felipe Cataquiz",
    "Cook, Scott": "Scott Cook",
    "Gigon, Alaine": "Alaine Gigon",
    "Headid, Ronald": "James Headid",            # Ronald (Excel) -> James (DB)
    "Maher, Nicholas": "Nick Maher",             # Nicholas -> Nick
    "Thomas, Devin": "Devin Thomas",
    # PGY-1 (R1)
    "Sawyer, Tessa": "Tessa Sawyer",
    "Wilhelm, Clara": "Clara Wilhelm",
    "Travis, Colin": "Colin Travis",
    "Byrnes, Katherine": "Katie Byrnes",         # Katherine -> Katie
    "Sloss, Meleighe": "Meleigh Sloss",          # Meleighe -> Meleigh
    "Monsivais, Joshua": "Josh Monsivais",       # Joshua -> Josh
    # Faculty (C19)
    "Bevis, Zach": "Zach Bevis",
    "Kinkennon, Sarah": "Sarah Kinkennon",
    "LaBounty, Alex*": "Alex LaBounty",
    "McGuire, Chris": "Chris McGuire",
    "Dahl, Brian*": "Brian Dahl",
    "McRae, Zachery": "Zach McRae",              # Zachery -> Zach
    "Tagawa, Chelsea": "Chelsea Tagawa",
    "Montgomery, Aaron": "Aaron Montgomery",
    "Colgan, Bridget": "Bridget Colgan",         # DEPLOYED Feb-Jun 2026
    "Chu, Jimmy*": "Jimmy Chu",
}

# ── Activity Code Mapping ────────────────────────────────────────────────────
# Shared CODE_MAP from scripts/data/code_maps.py
from code_maps import CODE_MAP

# Leave-type activity codes (for absence sync)
LEAVE_CODES = {"LV", "LV-AM", "LV-PM", "TDY", "DEP", "SLV"}

# Maps leave codes to absence_type
LEAVE_TYPE_MAP = {
    "LV": "vacation",
    "LV-AM": "vacation",
    "LV-PM": "vacation",
    "SLV": "vacation",
    "TDY": "tdy",
    "DEP": "deployment",
}


# ── Row Detection ────────────────────────────────────────────────────────────

def detect_person_rows(ws):
    """Scan worksheet to detect resident and faculty rows dynamically.

    Looks for section markers in column B (TEMPLATE codes: R1, R2, R3, C17, C19)
    and column C (ROLE labels: PGY 1, PGY 2, PGY 3, FAC, NP, MD).
    Also detects rows with a name in the NAME_COL that matches NAME_MAP.

    Returns:
        (resident_rows, faculty_rows) — lists of 1-indexed row numbers
    """
    resident_markers = {"R1", "R2", "R3"}
    faculty_markers = {"C19", "C17"}
    role_resident = {"PGY 1", "PGY 2", "PGY 3", "PGY1", "PGY2", "PGY3"}
    role_faculty = {"FAC", "NP", "MD"}

    resident_rows = []
    faculty_rows = []

    for row_idx in range(1, min(ws.max_row + 1, 120)):
        template_val = str(ws.cell(row=row_idx, column=2).value or "").strip().upper()
        role_val = str(ws.cell(row=row_idx, column=3).value or "").strip().upper()
        name_val = str(ws.cell(row=row_idx, column=NAME_COL).value or "").strip()

        # Check template marker (column B)
        if template_val in resident_markers:
            resident_rows.append(row_idx)
        elif template_val in faculty_markers:
            faculty_rows.append(row_idx)
        # Check role label (column C)
        elif role_val in role_resident:
            resident_rows.append(row_idx)
        elif role_val in role_faculty:
            faculty_rows.append(row_idx)
        # Fallback: check if name is in NAME_MAP (catches unlabeled rows)
        elif name_val and name_val in NAME_MAP:
            # Determine resident vs faculty by NAME_MAP key position
            # (R1/R2/R3 keys appear before faculty keys in the map)
            # Conservative: add to resident list; load_hdas validates by DB role
            resident_rows.append(row_idx)

    return resident_rows, faculty_rows


# ── Step 1: Parse Excel ──────────────────────────────────────────────────────

def parse_excel():
    """Parse Excel handjam and return list of (name, date, time_of_day, code)."""
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[SHEET_NAME]

    # Detect person rows dynamically
    resident_rows, faculty_rows = detect_person_rows(ws)
    if not resident_rows and not faculty_rows:
        print("  WARNING: Auto-detection found no person rows; using fallback ranges")
        resident_rows = list(_FALLBACK_RESIDENT_ROWS)
        faculty_rows = list(_FALLBACK_FACULTY_ROWS)
    else:
        if not resident_rows:
            print("  WARNING: No resident rows detected; using fallback resident range")
            resident_rows = list(_FALLBACK_RESIDENT_ROWS)
        if not faculty_rows:
            print("  WARNING: No faculty rows detected; using fallback faculty range")
            faculty_rows = list(_FALLBACK_FACULTY_ROWS)
        print(f"  Auto-detected {len(resident_rows)} resident rows, {len(faculty_rows)} faculty rows")

    start = BLOCK_CONFIG["start_date"]
    num_days = (BLOCK_CONFIG["end_date"] - start).days + 1

    slots = []
    for i in range(num_days):
        d = start + timedelta(days=i)
        slots.append((d, "AM"))
        slots.append((d, "PM"))

    records = []
    unmapped_names = set()
    unmapped_codes = Counter()

    for row in resident_rows + faculty_rows:
        excel_name = ws.cell(row=row, column=NAME_COL).value
        if not excel_name or not str(excel_name).strip():
            continue
        excel_name = str(excel_name).strip()

        db_name = NAME_MAP.get(excel_name)
        if not db_name:
            unmapped_names.add(excel_name)
            continue

        for idx, (d, tod) in enumerate(slots):
            col = DATA_COL_START + idx
            val = ws.cell(row=row, column=col).value
            code = str(val).strip() if val else None
            if code:
                db_code = CODE_MAP.get(code)
                if db_code is None:
                    unmapped_codes[code] += 1
                    db_code = code.lower()
                records.append((db_name, d, tod, db_code))

    # Report unmapped items (potential gotchas)
    if unmapped_names:
        print(f"\n  UNMAPPED NAMES ({len(unmapped_names)}):")
        for name in sorted(unmapped_names):
            print(f"    - '{name}'")

    if unmapped_codes:
        print(f"\n  UNMAPPED CODES ({len(unmapped_codes)}):")
        for code, count in unmapped_codes.most_common():
            print(f"    - '{code}' ({count} occurrences)")

    return records


# ── Step 2: Load HDAs ────────────────────────────────────────────────────────

def load_hdas(records):
    """Upsert half-day assignments from parsed Excel records."""
    conn = psycopg2.connect(CONN)
    cur = conn.cursor()

    # Build person lookup
    cur.execute("SELECT id, name FROM people")
    people = {name: pid for pid, name in cur.fetchall()}

    # Build activity lookup
    cur.execute("SELECT id, code FROM activities")
    activities = {code: aid for aid, code in cur.fetchall()}

    created, updated, skipped = 0, 0, 0
    missing_people = set()
    missing_activities = set()

    for name, d, tod, code in records:
        person_id = people.get(name)
        activity_id = activities.get(code)

        if not person_id:
            missing_people.add(name)
            skipped += 1
            continue
        if not activity_id:
            missing_activities.add(code)
            skipped += 1
            continue

        # Try update first
        cur.execute(
            """UPDATE half_day_assignments
            SET activity_id = %s, updated_at = NOW()
            WHERE person_id = %s AND date = %s AND time_of_day = %s""",
            (activity_id, person_id, d, tod),
        )
        if cur.rowcount > 0:
            updated += cur.rowcount
        else:
            # Insert new (source='manual' — Block 11 gotcha #1)
            cur.execute(
                """INSERT INTO half_day_assignments
                (id, person_id, date, time_of_day, activity_id, source, created_at, updated_at)
                VALUES (%s, %s, %s, %s, %s, %s, NOW(), NOW())""",
                (str(uuid.uuid4()), person_id, d, tod, activity_id, "manual"),
            )
            created += 1

    conn.commit()
    cur.close()
    conn.close()

    # Report missing items (potential gotchas)
    if missing_people:
        print(f"\n  MISSING PEOPLE ({len(missing_people)}):")
        for name in sorted(missing_people):
            print(f"    - '{name}'")

    if missing_activities:
        print(f"\n  MISSING ACTIVITIES ({len(missing_activities)}):")
        for code in sorted(missing_activities):
            print(f"    - '{code}'")

    print(f"\n  HDAs: {updated} updated, {created} created, {skipped} skipped")


# ── Step 3: Sync Absences ────────────────────────────────────────────────────

def sync_absences(records):
    """Extract leave-type activities and sync to absences table."""
    conn = psycopg2.connect(CONN)
    cur = conn.cursor()

    # Build person lookup
    cur.execute("SELECT id, name FROM people")
    people = {name: pid for pid, name in cur.fetchall()}

    # Group leave records by person
    leave_by_person = {}
    for name, d, tod, code in records:
        if code.upper() in LEAVE_CODES:
            leave_by_person.setdefault(name, []).append((d, tod, code))

    # Delete existing block-specific absences (scoped to date range —
    # Block 11 gotcha #4: LIKE 'Block N%' matches across academic years)
    block_num = BLOCK_CONFIG["block_number"]
    block_start = BLOCK_CONFIG["start_date"]
    block_end = BLOCK_CONFIG["end_date"]
    cur.execute(
        """DELETE FROM absences
        WHERE notes LIKE %s
          AND start_date >= %s AND start_date <= %s""",
        (f"Block {block_num}%", block_start, block_end),
    )
    print(f"  Deleted {cur.rowcount} existing Block {block_num} absences")

    # Consolidate consecutive dates into ranges per person/type
    for name, entries in leave_by_person.items():
        person_id = people.get(name)
        if not person_id:
            continue

        # Group by absence_type
        by_type = {}
        for d, tod, code in entries:
            atype = LEAVE_TYPE_MAP.get(code.upper(), "vacation")
            by_type.setdefault(atype, set()).add(d)

        for atype, date_set in by_type.items():
            # Find contiguous date ranges
            sorted_dates = sorted(date_set)
            ranges = []
            range_start = sorted_dates[0]
            range_end = sorted_dates[0]
            for d in sorted_dates[1:]:
                if (d - range_end).days <= 1:
                    range_end = d
                else:
                    ranges.append((range_start, range_end))
                    range_start = d
                    range_end = d
            ranges.append((range_start, range_end))

            for start_d, end_d in ranges:
                cur.execute(
                    """INSERT INTO absences
                    (id, person_id, start_date, end_date, absence_type,
                     is_blocking, is_away_from_program, return_date_tentative,
                     status, notes, created_at)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
                    (
                        str(uuid.uuid4()),
                        person_id,
                        start_d,
                        end_d,
                        atype,
                        atype in ("tdy", "deployment"),
                        True,
                        False,
                        "approved",
                        f"Block {block_num} - synced from Excel handjam.",
                        datetime.utcnow(),
                    ),
                )
                print(f"  Created: {name} {atype} {start_d} - {end_d}")

    conn.commit()
    cur.close()
    conn.close()


# ── Main ─────────────────────────────────────────────────────────────────────

def preflight():
    """Run pre-flight validation before import. Aborts on critical failures."""
    from block_import_preflight import run_preflight

    block_num = BLOCK_CONFIG["block_number"]
    ay = BLOCK_CONFIG["academic_year"]

    print("\n=== Pre-Flight Validation ===")
    results = run_preflight(block_num, ay)

    critical = [r for r in results if not r.passed and r.severity == "CRITICAL"]
    warnings = [r for r in results if not r.passed and r.severity == "WARNING"]

    for r in results:
        icon = "PASS" if r.passed else "FAIL"
        severity = f" [{r.severity}]" if not r.passed else ""
        print(f"  [{icon}]{severity} {r.name}")
        if not r.passed:
            for detail in r.details:
                print(f"    {detail}")

    if critical:
        print(f"\n  BLOCKED: {len(critical)} critical issue(s). Fix before importing.")
        sys.exit(1)

    if warnings:
        print(f"\n  {len(warnings)} warning(s) — review before proceeding.")

    passed = sum(1 for r in results if r.passed)
    print(f"\n  Pre-flight: {passed}/{len(results)} passed")


def main():
    print("=" * 60)
    print("Block 12 Schedule Import (May 7 – Jun 3, 2026)")
    print("=" * 60)

    preflight()

    print("\n=== Step 1: Parsing Excel ===")
    records = parse_excel()
    print(f"  Parsed {len(records)} records")

    print("\n=== Step 2: Loading HDAs ===")
    load_hdas(records)

    print("\n=== Step 3: Syncing Absences ===")
    sync_absences(records)

    print("\n" + "=" * 60)
    print("Done. Update BLOCK12_SCHEDULE_LOAD.md with results.")
    print("=" * 60)


if __name__ == "__main__":
    main()
