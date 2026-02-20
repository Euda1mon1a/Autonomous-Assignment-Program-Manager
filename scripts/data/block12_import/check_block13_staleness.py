"""Check Block 13 for stale block_assignments (same problem as Block 12).

Block 12 had 100% stale data — all 16 block_assignments contained Block 11
rotation mappings. This script checks if Block 13 has the same problem by
comparing DB block_assignments against the Excel Block Schedule.

Read-only — does NOT modify the database.

Usage:
  backend/.venv/bin/python scripts/data/block12_import/check_block13_staleness.py
"""

from datetime import date

import openpyxl
import psycopg2

CONN = "dbname=residency_scheduler user=scheduler host=localhost"

EXCEL_PATH = "/Users/aaronmontgomery/Downloads/Current AY 25-26 pulled 11 FEB 2026.xlsx"
BLOCK_SCHEDULE_SHEET = "Block Schedule"

# Block 13: Jun 4 – Jul 1, 2026 (AY 2025)
BLOCK_NUM = 13
ACADEMIC_YEAR = 2025

# Excel Block Schedule column indices for Block 13 (0-indexed from openpyxl)
# Columns 26-27 in the rotation grid (adjust if needed)
EXCEL_COL_PRIMARY = 26    # Primary rotation column for B13
EXCEL_COL_NF = 27         # NF assignment column for B13

# Name mapping (same as Block 12 — Excel "Last, First" to DB "First Last")
NAME_MAP = {
    "Connolly, Laura": "Laura Connolly",
    "Hernandez, Christian*": "Christian Hernandez",
    "Mayell, Cameron*": "Cam Mayell",
    "Petrie, William*": "Clay Petrie",
    "You, Jae*": "Jae You",
    "Cataquiz, Felipe": "Felipe Cataquiz",
    "Cook, Scott": "Scott Cook",
    "Gigon, Alaine": "Alaine Gigon",
    "Headid, Ronald": "James Headid",
    "Maher, Nicholas": "Nick Maher",
    "Thomas, Devin": "Devin Thomas",
    "Sawyer, Tessa": "Tessa Sawyer",
    "Wilhelm, Clara": "Clara Wilhelm",
    "Travis, Colin": "Colin Travis",
    "Byrnes, Katherine": "Katie Byrnes",
    "Sloss, Meleighe": "Meleigh Sloss",
    "Monsivais, Joshua": "Josh Monsivais",
}


def main():
    print("=" * 70)
    print(f"Block {BLOCK_NUM} Staleness Check (AY {ACADEMIC_YEAR})")
    print("=" * 70)

    # ── Read Excel Block Schedule ──────────────────────────────────────
    print(f"\nReading Excel: {EXCEL_PATH}")
    print(f"Sheet: {BLOCK_SCHEDULE_SHEET}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[BLOCK_SCHEDULE_SHEET]

    # Find resident rows — scan column 1 for names matching "Last, First" pattern
    excel_assignments = {}
    for row in range(1, ws.max_row + 1):
        name_cell = ws.cell(row=row, column=1).value
        if not name_cell or not isinstance(name_cell, str):
            continue
        name_cell = name_cell.strip()

        db_name = NAME_MAP.get(name_cell)
        if not db_name:
            continue

        primary = ws.cell(row=row, column=EXCEL_COL_PRIMARY).value
        nf = ws.cell(row=row, column=EXCEL_COL_NF).value

        primary_str = str(primary).strip() if primary else "(blank)"
        nf_str = str(nf).strip() if nf else ""

        excel_assignments[db_name] = (primary_str, nf_str)

    print(f"\nExcel Block Schedule entries for Block {BLOCK_NUM}: {len(excel_assignments)}")
    for name, (primary, nf) in sorted(excel_assignments.items()):
        nf_str = f" + {nf}" if nf else ""
        print(f"  {name:30s} -> {primary}{nf_str}")

    # ── Read DB block_assignments ──────────────────────────────────────
    conn = psycopg2.connect(CONN)
    cur = conn.cursor()

    cur.execute("""
        SELECT p.name, rt.abbreviation, rt2.abbreviation
        FROM block_assignments ba
        JOIN people p ON ba.resident_id = p.id
        JOIN rotation_templates rt ON ba.rotation_template_id = rt.id
        LEFT JOIN rotation_templates rt2 ON ba.secondary_rotation_template_id = rt2.id
        WHERE ba.block_number = %s AND ba.academic_year = %s
        ORDER BY p.name
    """, (BLOCK_NUM, ACADEMIC_YEAR))
    db_assignments = {name: (rot, sec or "") for name, rot, sec in cur.fetchall()}

    print(f"\nDB block_assignments for Block {BLOCK_NUM}: {len(db_assignments)}")
    for name, (rot, sec) in sorted(db_assignments.items()):
        sec_str = f" / {sec}" if sec else ""
        print(f"  {name:30s} -> {rot}{sec_str}")

    # ── Compare ────────────────────────────────────────────────────────
    print("\n" + "=" * 70)
    print("COMPARISON")
    print("=" * 70)

    all_names = sorted(set(list(excel_assignments.keys()) + list(db_assignments.keys())))
    mismatches = 0
    matches = 0
    missing_db = 0
    missing_excel = 0

    print(f"\n{'Resident':30s} | {'DB Rotation':20s} | {'Excel Rotation':20s} | Status")
    print("-" * 100)

    for name in all_names:
        db_rot = db_assignments.get(name, (None, None))
        excel_rot = excel_assignments.get(name, (None, None))

        if db_rot[0] is None:
            print(f"  {name:30s} | {'(not in DB)':20s} | {excel_rot[0]:20s} | MISSING IN DB")
            missing_db += 1
        elif excel_rot[0] is None or excel_rot[0] == "(blank)":
            print(f"  {name:30s} | {db_rot[0]:20s} | {'(blank/missing)':20s} | NOT IN EXCEL")
            missing_excel += 1
        else:
            # Simple string match (imperfect — Excel names differ from DB abbreviations)
            status = "OK?" if db_rot[0].upper() in excel_rot[0].upper() or excel_rot[0].upper() in db_rot[0].upper() else "MISMATCH?"
            if status == "MISMATCH?":
                mismatches += 1
            else:
                matches += 1
            print(f"  {name:30s} | {db_rot[0]:20s} | {excel_rot[0]:20s} | {status}")

    print(f"\n{'=' * 70}")
    print(f"SUMMARY: {matches} likely match, {mismatches} likely mismatch, "
          f"{missing_db} missing in DB, {missing_excel} not in Excel")
    print(f"{'=' * 70}")

    if mismatches > 0 or missing_db > 0:
        print("\nWARNING: Block 13 may have stale data. Manual review needed.")
        print("NOTE: String matching is approximate — Excel uses display names,")
        print("DB uses template abbreviations. Verify mismatches manually.")
    else:
        print("\nBlock 13 appears consistent (approximate match).")

    cur.close()
    conn.close()


if __name__ == "__main__":
    main()
