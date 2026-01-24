#!/usr/bin/env python3
"""
Direct CSV → Excel Template Filler (Mac Compatible)

Bypasses VBA entirely. Reads SCHEDULE_OUTPUT.csv and writes directly
to Block10_TEMPLATE.xlsx using openpyxl.

Usage:
    python csv_to_excel_direct.py SCHEDULE_OUTPUT.csv Block10_TEMPLATE.xlsx Block10_FILLED.xlsx
"""

import argparse
import csv
import sys
from pathlib import Path
from typing import Optional

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


# Block 10 configuration
DATA_START_COL = 6  # Column F = first schedule data column
RESIDENT_START_ROW = 9
RESIDENT_END_ROW = 25
FACULTY_START_ROW = 31
FACULTY_END_ROW = 42
NAME_COL = 5  # Column E = names


def normalize_name(name: str) -> str:
    """Normalize name for matching (lowercase, strip whitespace, remove asterisks)."""
    return name.strip().lower().replace("  ", " ").replace("*", "").strip()


# Name aliases: CSV name → Template name patterns
NAME_ALIASES = {
    "mayell, cam": "mayell, cameron",
    "petrie, clay": "petrie, william",
    "headid, james": "headid, ronald",
    "maher, nick": "maher, nicholas",
    "byrnes, katie": "byrnes, katherine",
    "monsivais, josh": "monsivais, joshua",
    "sloss, meleigh": "sloss, meleighe",
    "labounty, alex": "labounty, alex",
    "mcrae, zach": "mcrae, zachery",
    "you, jae": "you, jae",
}


def resolve_name(csv_name: str, template_names: dict) -> Optional[str]:
    """Try to match CSV name to template name."""
    normalized = normalize_name(csv_name)

    # Direct match
    if normalized in template_names:
        return normalized

    # Try alias
    if normalized in NAME_ALIASES:
        alias = NAME_ALIASES[normalized]
        if alias in template_names:
            return alias

    # Try partial match (last name match)
    csv_last = normalized.split(",")[0] if "," in normalized else normalized
    for template_name in template_names:
        template_last = template_name.split(",")[0] if "," in template_name else template_name
        if csv_last == template_last:
            return template_name

    return None


def fill_template(
    csv_path: str,
    template_path: str,
    output_path: str,
    sheet_name: Optional[str] = None,
    clear_first: bool = False,
) -> dict:
    """
    Fill Excel template with schedule data from CSV.

    Returns stats dict.
    """
    # Load template
    print(f"Loading template: {template_path}")
    wb = load_workbook(template_path)

    # Select sheet - use specified name or default to Block Template2
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    elif "Block Template2" in wb.sheetnames:
        ws = wb["Block Template2"]
    else:
        ws = wb.active
    print(f"Using sheet: {ws.title}")

    # Optionally clear existing schedule data first
    if clear_first:
        cleared = 0
        for row in range(RESIDENT_START_ROW, RESIDENT_END_ROW + 1):
            for col in range(DATA_START_COL, 62):  # Cols F-BJ (56 slots)
                if ws.cell(row=row, column=col).value:
                    ws.cell(row=row, column=col).value = None
                    cleared += 1
        for row in range(FACULTY_START_ROW, FACULTY_END_ROW + 1):
            for col in range(DATA_START_COL, 62):
                if ws.cell(row=row, column=col).value:
                    ws.cell(row=row, column=col).value = None
                    cleared += 1
        print(f"Cleared {cleared} existing cells")

    # Build name → row mapping from template
    name_to_row = {}

    # Residents (rows 9-25)
    for row in range(RESIDENT_START_ROW, RESIDENT_END_ROW + 1):
        cell_val = ws.cell(row=row, column=NAME_COL).value
        if cell_val:
            name_to_row[normalize_name(str(cell_val))] = row

    # Faculty (rows 31-42)
    for row in range(FACULTY_START_ROW, FACULTY_END_ROW + 1):
        cell_val = ws.cell(row=row, column=NAME_COL).value
        if cell_val:
            name_to_row[normalize_name(str(cell_val))] = row

    print(f"Found {len(name_to_row)} names in template")

    # Read CSV and apply to template
    updates = 0
    matched = 0
    not_found = []
    preserved_cells = []  # Track (name, count) of preserved manual entries

    with open(csv_path, "r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames or []

        for csv_row in reader:
            csv_name = csv_row.get("name", "")

            # Use resolve_name for fuzzy matching
            resolved = resolve_name(csv_name, name_to_row)

            if resolved is None:
                if csv_name and csv_name not in not_found:
                    not_found.append(csv_name)
                continue

            target_row = name_to_row[resolved]
            matched += 1

            # Write schedule data (columns 5+ in CSV → columns 6+ in Excel)
            # PRESERVE existing values - only write to empty cells
            col_idx = DATA_START_COL
            skipped = 0
            for key in fieldnames:
                if "_AM" in key or "_PM" in key:
                    value = csv_row.get(key, "")
                    cell = ws.cell(row=target_row, column=col_idx)
                    existing = cell.value
                    # Preserve only real values (not formulas, not empty)
                    # Formulas start with "=", skip those
                    is_formula = isinstance(existing, str) and existing.startswith("=")
                    is_real_value = (
                        existing is not None
                        and existing != ""
                        and not is_formula
                        and str(existing).strip() != ""
                    )
                    if is_real_value:
                        # Cell has actual manual data - preserve it
                        skipped += 1
                    elif value:
                        ws.cell(row=target_row, column=col_idx).value = value
                        updates += 1
                    col_idx += 1
            if skipped > 0:
                preserved_cells.append((csv_name, skipped))

    # Save output
    print(f"Saving to: {output_path}")
    wb.save(output_path)

    return {
        "updates": updates,
        "matched": matched,
        "names_in_template": len(name_to_row),
        "names_not_found": not_found,
        "preserved": preserved_cells,
        "output": output_path,
    }


def main():
    parser = argparse.ArgumentParser(
        description="Fill Excel template directly from CSV (Mac compatible)"
    )
    parser.add_argument("csv", help="Input CSV file (VBA format)")
    parser.add_argument("template", help="Excel template file")
    parser.add_argument("output", help="Output Excel file")
    parser.add_argument(
        "--sheet", "-s",
        help="Sheet name to use (default: Block Template2)"
    )
    parser.add_argument(
        "--clear", "-c",
        action="store_true",
        help="Clear existing schedule data before importing"
    )

    args = parser.parse_args()

    # Validate inputs
    if not Path(args.csv).exists():
        print(f"ERROR: CSV not found: {args.csv}")
        sys.exit(1)
    if not Path(args.template).exists():
        print(f"ERROR: Template not found: {args.template}")
        sys.exit(1)

    try:
        stats = fill_template(
            args.csv, args.template, args.output,
            sheet_name=args.sheet, clear_first=args.clear
        )

        print("\n✅ Complete!")
        print(f"   Cells updated: {stats['updates']}")
        print(f"   People matched: {stats['matched']}/{stats['names_in_template']}")

        if stats["preserved"]:
            total_preserved = sum(c for _, c in stats["preserved"])
            print(f"\n🔒 Preserved {total_preserved} existing cells:")
            for name, count in stats["preserved"][:10]:
                print(f"      - {name}: {count} cells kept")
            if len(stats["preserved"]) > 10:
                remaining = len(stats["preserved"]) - 10
                print(f"      ... and {remaining} more people")

        if stats["names_not_found"]:
            print("\n⚠️  Names in CSV not found in template:")
            for name in stats["names_not_found"][:10]:
                print(f"      - {name}")
            if len(stats["names_not_found"]) > 10:
                more = len(stats["names_not_found"]) - 10
                print(f"      ... and {more} more")

        print(f"\n📁 Output: {stats['output']}")
        print("   Open this file in Excel to review")

    except Exception as e:
        print(f"ERROR: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
