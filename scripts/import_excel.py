#!/usr/bin/env python3
"""
Excel import script for scheduling data.

Imports leave/absences and pre-assigned rotations from Excel files.
Designed to work with exports from xlsx_export.py format.

Usage:
    python scripts/import_excel.py <excel_file> [--dry-run] [--verbose]

Examples:
    python scripts/import_excel.py schedule.xlsx --dry-run
    python scripts/import_excel.py absences.xlsx --verbose

Prerequisites:
    - Run from project root or backend directory
    - Backend virtual environment activated with all dependencies
    - Database accessible (or use --database-url)
"""

from __future__ import annotations

import argparse
import logging
import sys
import uuid
from datetime import date, datetime
from pathlib import Path
from typing import TYPE_CHECKING, Any

# Type checking imports (not loaded at runtime)
if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet
    from sqlalchemy.orm import Session as SessionType
    from app.models.absence import Absence as AbsenceType
    from app.models.person import Person as PersonType

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)


# =============================================================================
# Absence Type Mapping (reverse of xlsx_export.py)
# =============================================================================

ABBREVIATION_TO_ABSENCE_TYPE = {
    # Standard types from xlsx_export.py
    "VAC": "vacation",
    "SICK": "sick",
    "MED": "medical",
    "CONF": "conference",
    "DEP": "deployment",
    "TDY": "tdy",
    "FEM": "family_emergency",
    "PER": "personal",
    # Additional types from absence.py
    "BER": "bereavement",
    "BERV": "bereavement",
    "BEREAVEMENT": "bereavement",
    "EMRG": "emergency_leave",
    "EMERGENCY": "emergency_leave",
    "CONV": "convalescent",
    "MAT": "maternity_paternity",
    "PAT": "maternity_paternity",
    "MATPAT": "maternity_paternity",
    # Common variations
    "VACATION": "vacation",
    "LEAVE": "vacation",
    "LV": "vacation",
    "V": "vacation",
    "S": "sick",
    "M": "medical",
    "C": "conference",
    "D": "deployment",
    "T": "tdy",
    "P": "personal",
    "ABS": "vacation",  # Generic absence defaults to vacation
}

# Valid absence types (from app/models/absence.py)
VALID_ABSENCE_TYPES = {
    "vacation", "deployment", "tdy", "medical", "family_emergency",
    "conference", "bereavement", "emergency_leave", "sick",
    "convalescent", "maternity_paternity"
}


def get_absence_type(abbreviation: str) -> str | None:
    """
    Convert abbreviation to absence type.

    Args:
        abbreviation: The abbreviation from Excel (e.g., "VAC", "SICK")

    Returns:
        The absence type string or None if not recognized
    """
    if not abbreviation:
        return None

    abbrev_upper = abbreviation.strip().upper()
    return ABBREVIATION_TO_ABSENCE_TYPE.get(abbrev_upper)


# =============================================================================
# Name Normalization and Matching
# =============================================================================


def normalize_name(name: str) -> str:
    """
    Normalize name for case-insensitive matching.

    Handles:
    - Case normalization
    - "Dr." prefix removal
    - Whitespace trimming
    - "Last, First" vs "First Last" format detection

    Args:
        name: The raw name string

    Returns:
        Normalized name for comparison
    """
    if not name:
        return ""

    name = name.lower().strip()

    # Remove common prefixes
    prefixes = [
        "dr.", "dr ", "doctor ", "dr ",
        "mr.", "mr ", "mrs.", "mrs ", "ms.", "ms ",
        "prof.", "prof ", "professor ",
    ]
    for prefix in prefixes:
        if name.startswith(prefix):
            name = name[len(prefix):].strip()

    # Remove common suffixes
    suffixes = [", md", " md", ", do", " do", ", phd", " phd"]
    for suffix in suffixes:
        if name.endswith(suffix):
            name = name[: -len(suffix)].strip()

    # Handle "Last, First" format - convert to "First Last"
    if ", " in name:
        parts = name.split(", ", 1)
        if len(parts) == 2:
            name = f"{parts[1]} {parts[0]}"

    # Normalize multiple spaces
    name = " ".join(name.split())

    return name


def find_person_by_name(
    session: "SessionType",
    name: str,
    persons_cache: dict[str, "PersonType"] | None = None,
) -> "PersonType | None":
    """
    Find a person by name with fuzzy matching.

    Args:
        session: Database session
        name: Name to search for
        persons_cache: Optional cache of normalized_name -> Person

    Returns:
        Person object or None if not found
    """
    if not name:
        return None

    normalized = normalize_name(name)

    # Use cache if provided
    if persons_cache is not None:
        return persons_cache.get(normalized)

    # Lazy imports
    from sqlalchemy import select
    from app.models.person import Person

    # Query database
    persons = session.execute(select(Person)).scalars().all()

    for person in persons:
        if normalize_name(person.name) == normalized:
            return person

    return None


def build_persons_cache(session: "SessionType") -> dict[str, "PersonType"]:
    """
    Build a cache of normalized names to Person objects.

    Args:
        session: Database session

    Returns:
        Dictionary mapping normalized names to Person objects
    """
    # Lazy imports
    from sqlalchemy import select
    from app.models.person import Person

    persons = session.execute(select(Person)).scalars().all()
    cache = {}

    for person in persons:
        normalized = normalize_name(person.name)
        cache[normalized] = person
        logger.debug(f"Cached person: '{person.name}' -> '{normalized}'")

    return cache


# =============================================================================
# Date Parsing
# =============================================================================


def parse_date(value: Any) -> date | None:
    """
    Parse a date from various formats.

    Handles:
    - datetime objects
    - date objects
    - Strings in various formats (YYYY-MM-DD, MM/DD/YYYY, DD-MMM-YYYY, etc.)
    - Excel serial numbers

    Args:
        value: The value to parse

    Returns:
        date object or None if parsing fails
    """
    if value is None:
        return None

    # Already a date
    if isinstance(value, date) and not isinstance(value, datetime):
        return value

    # datetime object
    if isinstance(value, datetime):
        return value.date()

    # Excel serial number (int or float)
    if isinstance(value, (int, float)):
        try:
            # Excel dates are days since 1899-12-30 (with a bug for 1900 leap year)
            from datetime import timedelta
            excel_epoch = datetime(1899, 12, 30)
            return (excel_epoch + timedelta(days=int(value))).date()
        except (ValueError, OverflowError):
            return None

    # String parsing
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None

        # Try various date formats
        formats = [
            "%Y-%m-%d",      # ISO format: 2024-01-15
            "%m/%d/%Y",      # US format: 01/15/2024
            "%m/%d/%y",      # Short year: 01/15/24
            "%d-%b-%Y",      # Day-Month-Year: 15-Jan-2024
            "%d-%B-%Y",      # Full month: 15-January-2024
            "%b %d, %Y",     # Month Day, Year: Jan 15, 2024
            "%B %d, %Y",     # Full month: January 15, 2024
            "%d/%m/%Y",      # European format: 15/01/2024
            "%Y/%m/%d",      # Alternative ISO: 2024/01/15
            "%d.%m.%Y",      # German format: 15.01.2024
        ]

        for fmt in formats:
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue

        logger.warning(f"Could not parse date: '{value}'")
        return None

    logger.warning(f"Unexpected date type: {type(value)}")
    return None


# =============================================================================
# Excel Sheet Processing
# =============================================================================


class ImportResult:
    """Tracks results of import operations."""

    def __init__(self):
        self.created = 0
        self.skipped = 0
        self.errors = 0
        self.error_messages: list[str] = []

    def add_error(self, message: str) -> None:
        """Add an error message."""
        self.errors += 1
        self.error_messages.append(message)
        logger.error(message)

    def summary(self) -> str:
        """Get summary string."""
        return (
            f"Created: {self.created}, Skipped: {self.skipped}, Errors: {self.errors}"
        )


def detect_sheet_type(ws: Worksheet) -> str:
    """
    Detect the type of data in a worksheet.

    Returns:
        'absences': Sheet contains absence/leave data
        'schedule': Sheet contains schedule/assignment data
        'unknown': Could not determine sheet type
    """
    # Check first few rows for header patterns
    headers = []
    for row in range(1, 6):
        for col in range(1, 15):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value and isinstance(cell_value, str):
                headers.append(cell_value.lower())

    headers_text = " ".join(headers)

    # Absence sheet patterns
    absence_keywords = ["absence", "leave", "vacation", "start date", "end date", "type"]
    if any(kw in headers_text for kw in absence_keywords):
        return "absences"

    # Schedule sheet patterns (Block N, dates with AM/PM)
    schedule_keywords = ["block", "provider", "template", "rotation", "am", "pm"]
    if any(kw in headers_text for kw in schedule_keywords):
        return "schedule"

    return "unknown"


def find_header_row(ws: Worksheet, keywords: list[str]) -> int | None:
    """
    Find the row containing headers matching keywords.

    Args:
        ws: Worksheet to search
        keywords: Keywords to look for in headers

    Returns:
        Row number (1-indexed) or None if not found
    """
    for row in range(1, 20):  # Check first 20 rows
        row_text = ""
        for col in range(1, 20):
            cell = ws.cell(row=row, column=col).value
            if cell:
                row_text += str(cell).lower() + " "

        if any(kw in row_text for kw in keywords):
            return row

    return None


def import_absences_from_sheet(
    ws: "Worksheet",
    session: "SessionType",
    persons_cache: dict[str, "PersonType"],
    dry_run: bool = False,
) -> ImportResult:
    """
    Import absences from a worksheet.

    Expected columns (flexible order):
    - Person/Name/Provider: Person's name
    - Type: Absence type (or abbreviation)
    - Start Date/From: Start date
    - End Date/To: End date
    - Notes (optional): Additional notes

    Args:
        ws: Worksheet to import from
        session: Database session
        persons_cache: Cache of normalized names to Person objects
        dry_run: If True, don't actually create records

    Returns:
        ImportResult with counts and errors
    """
    # Lazy imports
    from sqlalchemy import select
    from app.models.absence import Absence

    result = ImportResult()

    # Find header row
    header_row = find_header_row(
        ws, ["person", "name", "provider", "type", "start", "end", "from", "to"]
    )

    if header_row is None:
        result.add_error(f"Sheet '{ws.title}': Could not find header row")
        return result

    # Map column indices
    col_map = {}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=header_row, column=col).value
        if header:
            header_lower = str(header).lower().strip()

            if header_lower in ("person", "name", "provider", "staff", "resident", "faculty"):
                col_map["person"] = col
            elif header_lower in ("type", "absence type", "leave type", "absence_type"):
                col_map["type"] = col
            elif header_lower in ("start date", "start", "from", "begin", "start_date"):
                col_map["start_date"] = col
            elif header_lower in ("end date", "end", "to", "through", "end_date"):
                col_map["end_date"] = col
            elif header_lower in ("notes", "note", "comments", "comment", "reason"):
                col_map["notes"] = col
            elif header_lower in ("tdy location", "tdy_location", "location"):
                col_map["tdy_location"] = col
            elif header_lower in ("deployment orders", "deployment_orders", "orders"):
                col_map["deployment_orders"] = col

    # Validate required columns
    required = ["person", "type", "start_date", "end_date"]
    missing = [col for col in required if col not in col_map]
    if missing:
        result.add_error(
            f"Sheet '{ws.title}': Missing required columns: {', '.join(missing)}"
        )
        return result

    logger.info(f"Sheet '{ws.title}': Found columns at {col_map}")

    # Process data rows
    for row in range(header_row + 1, ws.max_row + 1):
        # Get person name
        person_name = ws.cell(row=row, column=col_map["person"]).value
        if not person_name:
            continue  # Skip empty rows

        person_name = str(person_name).strip()

        # Find person
        person = find_person_by_name(session, person_name, persons_cache)
        if not person:
            result.add_error(
                f"Row {row}: Person not found: '{person_name}'"
            )
            continue

        # Get absence type
        type_value = ws.cell(row=row, column=col_map["type"]).value
        if not type_value:
            result.add_error(f"Row {row}: Missing absence type")
            continue

        absence_type = get_absence_type(str(type_value))
        if not absence_type:
            # Try using the raw value if it's a valid type
            raw_type = str(type_value).lower().strip().replace(" ", "_")
            if raw_type in VALID_ABSENCE_TYPES:
                absence_type = raw_type
            else:
                result.add_error(
                    f"Row {row}: Unknown absence type: '{type_value}'"
                )
                continue

        # Parse dates
        start_date = parse_date(ws.cell(row=row, column=col_map["start_date"]).value)
        end_date = parse_date(ws.cell(row=row, column=col_map["end_date"]).value)

        if not start_date:
            result.add_error(f"Row {row}: Invalid start date")
            continue

        if not end_date:
            result.add_error(f"Row {row}: Invalid end date")
            continue

        if end_date < start_date:
            result.add_error(
                f"Row {row}: End date ({end_date}) is before start date ({start_date})"
            )
            continue

        # Optional fields
        notes = None
        if "notes" in col_map:
            notes_value = ws.cell(row=row, column=col_map["notes"]).value
            if notes_value:
                notes = str(notes_value).strip()

        tdy_location = None
        if "tdy_location" in col_map:
            tdy_value = ws.cell(row=row, column=col_map["tdy_location"]).value
            if tdy_value:
                tdy_location = str(tdy_value).strip()

        deployment_orders = False
        if "deployment_orders" in col_map:
            orders_value = ws.cell(row=row, column=col_map["deployment_orders"]).value
            if orders_value:
                deployment_orders = str(orders_value).lower().strip() in (
                    "yes", "true", "1", "y", "x"
                )

        # Check for duplicates
        existing = session.execute(
            select(Absence).where(
                Absence.person_id == person.id,
                Absence.absence_type == absence_type,
                Absence.start_date == start_date,
                Absence.end_date == end_date,
            )
        ).scalar_one_or_none()

        if existing:
            logger.info(
                f"Row {row}: Skipping duplicate absence for {person.name} "
                f"({absence_type}: {start_date} to {end_date})"
            )
            result.skipped += 1
            continue

        # Create absence record
        if not dry_run:
            absence = Absence(
                id=uuid.uuid4(),
                person_id=person.id,
                absence_type=absence_type,
                start_date=start_date,
                end_date=end_date,
                notes=notes,
                tdy_location=tdy_location,
                deployment_orders=deployment_orders,
            )
            session.add(absence)
            logger.info(
                f"Row {row}: Created absence for {person.name}: "
                f"{absence_type} ({start_date} to {end_date})"
            )
        else:
            logger.info(
                f"Row {row}: [DRY-RUN] Would create absence for {person.name}: "
                f"{absence_type} ({start_date} to {end_date})"
            )

        result.created += 1

    return result


def import_schedule_from_sheet(
    ws: "Worksheet",
    session: "SessionType",
    persons_cache: dict[str, "PersonType"],
    dry_run: bool = False,
) -> ImportResult:
    """
    Import schedule assignments from a worksheet (legacy format).

    This handles the format from xlsx_export.py with:
    - AM/PM columns per day
    - Person names in column E
    - Date headers in row 3

    Args:
        ws: Worksheet to import from
        session: Database session
        persons_cache: Cache of normalized names to Person objects
        dry_run: If True, don't actually create records

    Returns:
        ImportResult with counts and errors
    """
    # Lazy imports
    from sqlalchemy import select
    from app.models.assignment import Assignment
    from app.models.block import Block
    from app.models.rotation_template import RotationTemplate

    result = ImportResult()

    # Get rotation templates cache
    templates = session.execute(select(RotationTemplate)).scalars().all()
    template_cache = {}
    for template in templates:
        if template.abbreviation:
            template_cache[template.abbreviation.upper()] = template
        template_cache[template.name.upper()] = template

    # Find the date header row (usually row 3)
    date_row = 3

    # Find person column (usually E)
    person_col = 5  # Column E

    # Build date mapping from headers
    date_columns = {}  # {(date, time_of_day): column}

    for col in range(6, ws.max_column + 1):
        # Row 3 should have dates
        date_cell = ws.cell(row=date_row, column=col).value
        # Row 2 should have AM/PM
        time_cell = ws.cell(row=2, column=col).value

        parsed_date = parse_date(date_cell)
        if parsed_date and time_cell:
            time_str = str(time_cell).strip().upper()
            if time_str in ("AM", "PM"):
                date_columns[col] = (parsed_date, time_str)

    if not date_columns:
        result.add_error(
            f"Sheet '{ws.title}': Could not find date columns in schedule format"
        )
        return result

    logger.info(f"Sheet '{ws.title}': Found {len(date_columns)} date/time columns")

    # Process person rows (usually starting at row 9)
    for row in range(9, ws.max_row + 1):
        person_name = ws.cell(row=row, column=person_col).value
        if not person_name:
            continue

        person_name = str(person_name).strip()
        person = find_person_by_name(session, person_name, persons_cache)

        if not person:
            # Skip gaps and empty names
            if person_name and not person_name.isspace():
                logger.debug(f"Row {row}: Person not found: '{person_name}'")
            continue

        # Process each date column
        for col, (col_date, time_of_day) in date_columns.items():
            cell_value = ws.cell(row=row, column=col).value
            if not cell_value:
                continue

            value = str(cell_value).strip().upper()

            # Skip working ("W") and empty cells
            if value in ("W", ""):
                continue

            # Check if it's an absence abbreviation
            absence_type = get_absence_type(value)
            if absence_type:
                # This is an absence - should be in absences, not assignments
                # Skip here, let absences sheet handle it
                continue

            # Look up rotation template
            template = template_cache.get(value)
            if not template:
                logger.debug(
                    f"Row {row}, Col {col}: Unknown abbreviation '{value}' "
                    f"for {person.name} on {col_date} {time_of_day}"
                )
                continue

            # Find or create block
            block = session.execute(
                select(Block).where(
                    Block.date == col_date,
                    Block.time_of_day == time_of_day,
                )
            ).scalar_one_or_none()

            if not block:
                logger.debug(
                    f"Row {row}: Block not found for {col_date} {time_of_day}"
                )
                continue

            # Check for existing assignment
            existing = session.execute(
                select(Assignment).where(
                    Assignment.block_id == block.id,
                    Assignment.person_id == person.id,
                )
            ).scalar_one_or_none()

            if existing:
                logger.debug(
                    f"Row {row}: Skipping duplicate assignment for {person.name} "
                    f"on {col_date} {time_of_day}"
                )
                result.skipped += 1
                continue

            # Create assignment
            if not dry_run:
                assignment = Assignment(
                    id=uuid.uuid4(),
                    block_id=block.id,
                    person_id=person.id,
                    rotation_template_id=template.id,
                    role="primary",  # Default role
                    created_by="excel_import",
                )
                session.add(assignment)
                logger.info(
                    f"Row {row}: Created assignment for {person.name}: "
                    f"{template.name} on {col_date} {time_of_day}"
                )
            else:
                logger.info(
                    f"Row {row}: [DRY-RUN] Would create assignment for {person.name}: "
                    f"{template.name} on {col_date} {time_of_day}"
                )

            result.created += 1

    return result


# =============================================================================
# Main Import Function
# =============================================================================


def import_excel(
    file_path: Path,
    session: "SessionType",
    dry_run: bool = False,
    verbose: bool = False,
) -> dict[str, ImportResult]:
    """
    Import data from an Excel file.

    Automatically detects sheet types and processes accordingly:
    - Absence/Leave sheets: Creates Absence records
    - Schedule sheets: Creates Assignment records

    Args:
        file_path: Path to the Excel file
        session: Database session
        dry_run: If True, don't commit changes
        verbose: If True, log detailed information

    Returns:
        Dictionary mapping sheet names to ImportResult objects
    """
    # Lazy import
    from openpyxl import load_workbook

    if verbose:
        logger.setLevel(logging.DEBUG)

    logger.info(f"Loading Excel file: {file_path}")

    if not file_path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    wb = load_workbook(file_path, read_only=False, data_only=True)
    logger.info(f"Found {len(wb.sheetnames)} sheet(s): {', '.join(wb.sheetnames)}")

    # Build persons cache once for efficiency
    persons_cache = build_persons_cache(session)
    logger.info(f"Loaded {len(persons_cache)} persons from database")

    results = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        logger.info(f"\nProcessing sheet: '{sheet_name}'")

        sheet_type = detect_sheet_type(ws)
        logger.info(f"Detected sheet type: {sheet_type}")

        if sheet_type == "absences":
            results[sheet_name] = import_absences_from_sheet(
                ws, session, persons_cache, dry_run
            )
        elif sheet_type == "schedule":
            results[sheet_name] = import_schedule_from_sheet(
                ws, session, persons_cache, dry_run
            )
        else:
            logger.warning(f"Sheet '{sheet_name}': Unknown format, skipping")
            result = ImportResult()
            result.add_error(f"Unknown sheet format")
            results[sheet_name] = result

        logger.info(f"Sheet '{sheet_name}' result: {results[sheet_name].summary()}")

    wb.close()

    return results


# =============================================================================
# CLI Entry Point
# =============================================================================


def main():
    """CLI entry point."""
    # Lazy imports - only load when actually running
    # Add backend to path for imports
    sys.path.insert(0, str(Path(__file__).parent.parent / "backend"))

    try:
        from openpyxl import load_workbook
        from sqlalchemy import create_engine
        from sqlalchemy.orm import sessionmaker
        from app.core.config import get_settings
    except ImportError as e:
        print(f"Error: Missing required dependency: {e}")
        print("\nPlease ensure you're running from the project root with the backend")
        print("virtual environment activated. Required packages:")
        print("  - openpyxl")
        print("  - sqlalchemy")
        print("  - pydantic-settings")
        print("\nInstall with: pip install -r backend/requirements.txt")
        sys.exit(1)

    parser = argparse.ArgumentParser(
        description="Import scheduling data from Excel files.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python scripts/import_excel.py schedule.xlsx
  python scripts/import_excel.py absences.xlsx --dry-run
  python scripts/import_excel.py data.xlsx --verbose

Supported sheet formats:
  - Absences: Columns for Person, Type, Start Date, End Date, Notes
  - Schedule: Legacy format with AM/PM columns per day

Absence type abbreviations:
  VAC=vacation, SICK=sick, MED=medical, CONF=conference,
  DEP=deployment, TDY=tdy, FEM=family_emergency, PER=personal
        """,
    )

    parser.add_argument(
        "file",
        type=Path,
        help="Path to the Excel file to import",
    )

    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse and validate without creating records",
    )

    parser.add_argument(
        "--verbose", "-v",
        action="store_true",
        help="Enable verbose logging",
    )

    parser.add_argument(
        "--database-url",
        type=str,
        default=None,
        help="Database URL (defaults to DATABASE_URL env var)",
    )

    args = parser.parse_args()

    # Get database URL
    settings = get_settings()
    database_url = args.database_url or str(settings.DATABASE_URL)

    # For async databases, convert to sync URL
    if database_url.startswith("postgresql+asyncpg://"):
        database_url = database_url.replace("postgresql+asyncpg://", "postgresql://")

    logger.info(f"Connecting to database...")

    # Create engine and session
    engine = create_engine(database_url, echo=args.verbose)
    SessionLocal = sessionmaker(bind=engine)
    session = SessionLocal()

    try:
        # Run import
        results = import_excel(
            file_path=args.file,
            session=session,
            dry_run=args.dry_run,
            verbose=args.verbose,
        )

        # Print summary
        print("\n" + "=" * 60)
        print("IMPORT SUMMARY")
        print("=" * 60)

        total_created = 0
        total_skipped = 0
        total_errors = 0

        for sheet_name, result in results.items():
            print(f"\n{sheet_name}:")
            print(f"  Created: {result.created}")
            print(f"  Skipped: {result.skipped}")
            print(f"  Errors:  {result.errors}")

            if result.error_messages and args.verbose:
                for msg in result.error_messages[:10]:  # Show first 10 errors
                    print(f"    - {msg}")
                if len(result.error_messages) > 10:
                    print(f"    ... and {len(result.error_messages) - 10} more errors")

            total_created += result.created
            total_skipped += result.skipped
            total_errors += result.errors

        print("\n" + "-" * 60)
        print(f"TOTAL: Created {total_created}, Skipped {total_skipped}, Errors {total_errors}")

        if args.dry_run:
            print("\n[DRY-RUN] No changes were made to the database")
            session.rollback()
        else:
            if total_errors == 0 or input("\nCommit changes despite errors? (y/N): ").lower() == "y":
                session.commit()
                print("\nChanges committed successfully!")
            else:
                session.rollback()
                print("\nChanges rolled back")

    except Exception as e:
        logger.exception(f"Import failed: {e}")
        session.rollback()
        sys.exit(1)

    finally:
        session.close()

    # Exit with error code if there were errors
    if total_errors > 0:
        sys.exit(1)


if __name__ == "__main__":
    main()
