#!/usr/bin/env python3
"""Fill Excel template via AppleScript -- zero openpyxl writes.

Preserves 100% of template conditional formatting by using
Microsoft Excel.app as the write engine via macOS osascript.

Phase 1 (Python): DB -> JSON -> cell mapping {(row, col): value}
Phase 2 (AppleScript): Open template -> fill cells -> save copy

Usage:
    python scripts/scheduling/fill_template_osa.py \
        --block 12 --academic-year 2025 \
        --template ~/path/to/Template.xlsx

    # With pre-exported JSON (skip DB):
    python scripts/scheduling/fill_template_osa.py \
        --block 12 --academic-year 2025 \
        --template ~/path/to/Template.xlsx \
        --json /tmp/block12.json

    # Dry run (show cell map without touching Excel):
    python scripts/scheduling/fill_template_osa.py \
        --block 12 --academic-year 2025 \
        --template ~/path/to/Template.xlsx --dry-run
"""

from __future__ import annotations

import argparse
import os
import subprocess
import sys
from datetime import date, timedelta
from pathlib import Path
from typing import Any

REPO_ROOT = Path(__file__).resolve().parents[2]
BACKEND_DIR = REPO_ROOT / "backend"
sys.path.insert(0, str(BACKEND_DIR))


# ---------------------------------------------------------------
# Environment setup (from fill_handjam.py)
# ---------------------------------------------------------------


def _load_env() -> None:
    """Load .env file for DATABASE_URL etc."""
    env_path = REPO_ROOT / ".env"
    if not env_path.exists():
        return
    for raw in env_path.read_text().splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        value = value.strip().strip('"').strip("'")
        key = key.strip()
        existing = os.environ.get(key)
        if key == "CORS_ORIGINS":
            if not existing or '"' not in existing:
                os.environ[key] = value
            continue
        if not existing:
            os.environ[key] = value
    if not os.environ.get("DATABASE_URL") and os.environ.get("DB_PASSWORD"):
        os.environ["DATABASE_URL"] = (
            f"postgresql://scheduler:{os.environ['DB_PASSWORD']}@localhost:5432/"
            "residency_scheduler"
        )


_load_env()

from openpyxl import load_workbook  # noqa: E402


# ---------------------------------------------------------------
# Block Template2 layout constants
# ---------------------------------------------------------------
COL_ROTATION1 = 1
COL_ROTATION2 = 2
COL_NAME = 5
COL_SCHEDULE_START = 6
COLS_PER_DAY = 2
MAX_DAYS = 28

ROW_STAFF_CALL = 4
ROW_RESIDENT_START = 9
ROW_RESIDENT_END = 25
ROW_FACULTY_START = 31
ROW_FACULTY_END = 45

# Activities that should only come from half_day_assignments (solver output),
# not default weekly templates.
SOLVER_ASSIGNED_CODES = {"FMIT", "NF", "C-I"}


# ---------------------------------------------------------------
# Name matching (from fill_handjam.py)
# ---------------------------------------------------------------


def _normalize_name(name: str) -> str:
    return name.strip().rstrip("*").strip().lower()


def _extract_last_name(name: str) -> str:
    name = _normalize_name(name)
    if "," in name:
        return name.split(",")[0].strip()
    parts = name.split()
    return parts[-1] if parts else name


def _discover_name_rows(ws: Any, row_start: int, row_end: int) -> dict[str, int]:
    """Read names from column E -> last_name -> row mapping."""
    mapping: dict[str, int] = {}
    for row in range(row_start, row_end + 1):
        val = ws.cell(row=row, column=COL_NAME).value
        if val and isinstance(val, str) and val.strip():
            last = _extract_last_name(val)
            if last in mapping:
                for existing_name, existing_row in list(mapping.items()):
                    if existing_name == last:
                        full = _normalize_name(
                            ws.cell(row=existing_row, column=COL_NAME).value or ""
                        )
                        mapping[full] = existing_row
                        del mapping[last]
                        break
                mapping[_normalize_name(val)] = row
            else:
                mapping[last] = row
    return mapping


def _discover_rotation_rows(
    ws: Any,
    row_start: int,
    row_end: int,
) -> dict[int, tuple[str, str]]:
    """Read rotation codes from cols A+B for rows without names in col E.

    Returns {row: (rot1, rot2)} for empty-name rows only.
    """
    mapping: dict[int, tuple[str, str]] = {}
    for row in range(row_start, row_end + 1):
        name_val = ws.cell(row=row, column=COL_NAME).value
        if name_val and isinstance(name_val, str) and name_val.strip():
            continue  # Has a name -- skip (handled by name matching)
        rot1 = ws.cell(row=row, column=COL_ROTATION1).value or ""
        rot2 = ws.cell(row=row, column=COL_ROTATION2).value or ""
        rot1 = rot1.strip() if isinstance(rot1, str) else ""
        rot2 = rot2.strip() if isinstance(rot2, str) else ""
        # Skip non-breaking space and empty rows
        if rot1 and rot1 != "\xa0":
            mapping[row] = (rot1, rot2)
    return mapping


def _match_name(db_name: str, wb_names: dict[str, int]) -> int | None:
    last = _extract_last_name(db_name)
    if last in wb_names:
        return wb_names[last]
    norm = _normalize_name(db_name)
    if norm in wb_names:
        return wb_names[norm]
    parts = _normalize_name(db_name).split()
    if len(parts) >= 2:
        converted = f"{parts[-1]}, {' '.join(parts[:-1])}"
        if converted in wb_names:
            return wb_names[converted]
    return None


def _match_rotation(
    db_rot1: str,
    db_rot2: str,
    rotation_rows: dict[int, tuple[str, str]],
    claimed: set[int],
) -> int | None:
    """Match a person to a template row by rotation code (fallback).

    Exact match on rot1+rot2. First unclaimed match wins (for duplicates
    like two ELEC rows).
    """
    for row in sorted(rotation_rows.keys()):
        if row in claimed:
            continue
        tmpl_rot1, tmpl_rot2 = rotation_rows[row]
        if tmpl_rot1 == db_rot1 and tmpl_rot2 == db_rot2:
            return row
    return None


# ---------------------------------------------------------------
# Data loading (from fill_handjam.py)
# ---------------------------------------------------------------


def _load_schedule_data(
    block_number: int,
    academic_year: int,
    json_path: Path | None = None,
) -> dict:
    import json as json_mod

    if json_path:
        print(f"\nLoading from JSON: {json_path}")
        with open(json_path) as f:
            return json_mod.load(f)

    from app.db.session import SessionLocal
    from app.services.half_day_json_exporter import HalfDayJSONExporter
    from app.utils.academic_blocks import get_block_dates

    block_dates = get_block_dates(block_number, academic_year)
    print(f"\nQuerying DB for Block {block_number} AY{academic_year}...")
    session = SessionLocal()
    try:
        exporter = HalfDayJSONExporter(session)
        data = exporter.export(
            block_start=block_dates.start_date,
            block_end=block_dates.end_date,
            include_faculty=True,
            include_call=True,
            include_overrides=True,
        )
        # Enrich faculty with weekly template data
        _enrich_faculty_from_templates(
            session, data, block_dates.start_date, block_dates.end_date
        )
        return data
    finally:
        session.close()


def _enrich_faculty_from_templates(
    session: Any,
    data: dict,
    block_start: date,
    block_end: date,
) -> None:
    """Fill empty faculty half-day slots from weekly templates + overrides.

    Faculty schedules come from 3 sources (highest priority first):
    1. half_day_assignments (solver output) - already in data["faculty"]
    2. faculty_weekly_overrides (date-specific exceptions)
    3. faculty_weekly_templates (base weekly pattern)

    This function fills gaps in source #1 with data from #2 and #3.
    """
    from sqlalchemy.orm import selectinload

    from app.models.faculty_weekly_override import FacultyWeeklyOverride
    from app.models.faculty_weekly_template import FacultyWeeklyTemplate
    from app.models.person import Person

    # Build lookup: faculty name -> existing day data from HDA
    hda_by_name: dict[str, dict[str, dict[str, str]]] = {}
    for fac in data.get("faculty", []):
        day_map: dict[str, dict[str, str]] = {}
        for day in fac.get("days", []):
            day_map[day["date"]] = {"am": day.get("am", ""), "pm": day.get("pm", "")}
        hda_by_name[fac["name"]] = day_map

    # Get all faculty people with their IDs
    all_faculty_rows = (
        session.query(Person)
        .filter(Person.type == "faculty")
        .order_by(Person.name)
        .all()
    )
    all_faculty = [(str(f.id), f.name) for f in all_faculty_rows]

    # Load ALL weekly templates for ALL faculty (one query)
    template_rows = (
        session.query(FacultyWeeklyTemplate)
        .options(
            selectinload(FacultyWeeklyTemplate.person),
            selectinload(FacultyWeeklyTemplate.activity),
        )
        .all()
    )
    # templates: {name: {(dow, tod, week_num): display_code}}
    templates: dict[str, dict[tuple[int, str, int | None], str]] = {}
    for t in template_rows:
        name = t.person.name
        code = t.activity.code if t.activity else ""
        display = (t.activity.display_abbreviation if t.activity else "") or code
        if display in SOLVER_ASSIGNED_CODES or code in SOLVER_ASSIGNED_CODES:
            continue
        templates.setdefault(name, {})[
            (t.day_of_week, t.time_of_day, t.week_number)
        ] = display

    # Load ALL overrides for the block date range (one query)
    override_rows = (
        session.query(FacultyWeeklyOverride)
        .options(
            selectinload(FacultyWeeklyOverride.person),
            selectinload(FacultyWeeklyOverride.activity),
        )
        .filter(FacultyWeeklyOverride.effective_date.between(block_start, block_end))
        .all()
    )
    # overrides: {name: {(effective_date, dow, tod): display_code}}
    overrides: dict[str, dict[tuple[str, int, str], str]] = {}
    for o in override_rows:
        name = o.person.name
        code = o.activity.code if o.activity else ""
        display = (o.activity.display_abbreviation if o.activity else "") or code
        overrides.setdefault(name, {})[
            (str(o.effective_date), o.day_of_week, o.time_of_day)
        ] = display

    # For each faculty, build complete schedule
    existing_names = {f["name"] for f in data.get("faculty", [])}
    num_days = (block_end - block_start).days + 1
    enriched_count = 0

    for fac_id, fac_name in all_faculty:
        tmpl = templates.get(fac_name, {})
        ovr = overrides.get(fac_name, {})
        hda = hda_by_name.get(fac_name, {})

        if not tmpl and not ovr and not hda:
            continue  # No schedule data at all

        days: list[dict[str, str]] = []
        for day_offset in range(num_days):
            d = block_start + timedelta(days=day_offset)
            date_str = d.isoformat()
            dow = d.weekday()
            # Python weekday: Mon=0..Sun=6; DB stores Sun=0..Sat=6
            db_dow = (dow + 1) % 7

            # Week number within block (1-based, 1-4)
            week_num = (day_offset // 7) + 1

            # Find Monday of this week for override lookup
            monday = d - timedelta(days=dow)
            monday_str = monday.isoformat()

            # Priority: HDA > override > template
            existing = hda.get(date_str, {})
            am = existing.get("am", "")
            pm = existing.get("pm", "")

            if not am:
                # Check override
                ovr_key = (monday_str, db_dow, "AM")
                if ovr_key in ovr:
                    am = ovr[ovr_key]
                else:
                    # Check week-specific template first, then all-weeks
                    tmpl_key_specific = (db_dow, "AM", week_num)
                    tmpl_key_all = (db_dow, "AM", None)
                    if tmpl_key_specific in tmpl:
                        am = tmpl[tmpl_key_specific]
                    elif tmpl_key_all in tmpl:
                        am = tmpl[tmpl_key_all]

            if not pm:
                ovr_key = (monday_str, db_dow, "PM")
                if ovr_key in ovr:
                    pm = ovr[ovr_key]
                else:
                    tmpl_key_specific = (db_dow, "PM", week_num)
                    tmpl_key_all = (db_dow, "PM", None)
                    if tmpl_key_specific in tmpl:
                        pm = tmpl[tmpl_key_specific]
                    elif tmpl_key_all in tmpl:
                        pm = tmpl[tmpl_key_all]

            days.append({"date": date_str, "am": am, "pm": pm})

        if fac_name in existing_names:
            # Update existing faculty entry
            for fac in data["faculty"]:
                if fac["name"] == fac_name:
                    fac["days"] = days
                    enriched_count += 1
                    break
        else:
            # Add new faculty entry
            data.setdefault("faculty", []).append(
                {
                    "name": fac_name,
                    "rotation1": "",
                    "rotation2": "",
                    "days": days,
                }
            )
            enriched_count += 1

    print(f"  Faculty enriched from templates: {enriched_count} people")


# ---------------------------------------------------------------
# Cell map computation
# ---------------------------------------------------------------


def _compute_cell_map(
    data: dict,
    all_wb_names: dict[str, int],
    rotation_rows: dict[int, tuple[str, str]] | None = None,
) -> tuple[dict[tuple[int, int], str], list[str]]:
    """Compute (row, col) -> value mapping from schedule JSON + name rows.

    Matches fill_handjam.py logic: rotation cols, daily AM/PM, weekend W fill,
    call row. No display rule transformation.

    Falls back to rotation-code matching for rows without names in col E.
    """
    cell_map: dict[tuple[int, int], str] = {}
    unmatched: list[str] = []
    block_start = date.fromisoformat(data["block_start"])
    rotation_claimed: set[int] = set()

    for person in data.get("residents", []) + data.get("faculty", []):
        db_name = person.get("name", "")
        row = _match_name(db_name, all_wb_names)

        # Fallback: match by rotation code for unnamed rows
        if row is None and rotation_rows:
            db_rot1 = person.get("rotation1") or ""
            db_rot2 = person.get("rotation2") or ""
            row = _match_rotation(db_rot1, db_rot2, rotation_rows, rotation_claimed)
            if row is not None:
                rotation_claimed.add(row)
                # Write name into col E (Last, First format)
                parts = db_name.split()
                if len(parts) >= 2:
                    display_name = f"{parts[-1]}, {' '.join(parts[:-1])}"
                else:
                    display_name = db_name
                cell_map[(row, COL_NAME)] = display_name

        if row is None:
            unmatched.append(db_name)
            continue

        rot1 = person.get("rotation1") or ""
        rot2 = person.get("rotation2") or ""

        if rot1:
            cell_map[(row, COL_ROTATION1)] = rot1
        if rot2:
            cell_map[(row, COL_ROTATION2)] = rot2

        for day_idx, day in enumerate(person.get("days", [])):
            if day_idx >= MAX_DAYS:
                break

            am_code = day.get("am", "")
            pm_code = day.get("pm", "")

            day_date = date.fromisoformat(day["date"])
            is_weekend = day_date.weekday() in (5, 6)
            if not am_code and is_weekend:
                am_code = "W"
            if not pm_code and is_weekend:
                pm_code = "W"

            am_col = COL_SCHEDULE_START + (day_idx * COLS_PER_DAY)
            pm_col = am_col + 1

            if am_code:
                cell_map[(row, am_col)] = am_code
            if pm_code:
                cell_map[(row, pm_col)] = pm_code

    # Call row (row 4)
    for night in data.get("call", {}).get("nights", []):
        night_date = date.fromisoformat(night["date"])
        day_offset = (night_date - block_start).days
        if 0 <= day_offset < MAX_DAYS:
            am_col = COL_SCHEDULE_START + (day_offset * COLS_PER_DAY)
            staff_name = night.get("staff", "")
            if staff_name:
                if "," in staff_name:
                    last_name = staff_name.split(",")[0].strip()
                else:
                    parts = staff_name.split()
                    last_name = parts[-1] if parts else staff_name
                cell_map[(ROW_STAFF_CALL, am_col)] = last_name

    return cell_map, unmatched


# ---------------------------------------------------------------
# AppleScript generation and execution
# ---------------------------------------------------------------


def _col_to_letter(col: int) -> str:
    """Convert 1-based column index to Excel letter(s). 1->A, 26->Z, 27->AA."""
    result = ""
    while col > 0:
        col, remainder = divmod(col - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _build_applescript(
    cell_map: dict[tuple[int, int], str],
    sheet_name: str,
    workbook_name: str | None = None,
    output_path: Path | None = None,
) -> str:
    """Generate AppleScript to fill cells in Excel.

    Two modes:
    - workbook_name set: write to already-open workbook (fast, no open/close)
    - output_path set: open a copy, fill, save, close

    Uses row-range batching: one Apple Event per row (~30 calls vs ~1,500).
    """
    # Group cells by row
    rows_data: dict[int, dict[int, str]] = {}
    for (row, col), value in cell_map.items():
        rows_data.setdefault(row, {})[col] = value

    def _esc(s: str) -> str:
        """Escape a string for safe interpolation into AppleScript."""
        return s.replace("\\", "\\\\").replace('"', '\\"')

    safe_sheet = _esc(sheet_name)

    lines: list[str] = []
    lines.append('tell application "Microsoft Excel"')

    if workbook_name:
        # Target already-open workbook
        safe_wb = _esc(workbook_name)
        lines.append(f'    tell worksheet "{safe_sheet}" of workbook "{safe_wb}"')
    elif output_path:
        safe_path = _esc(str(output_path))
        lines.append(
            f"    set wb to open workbook workbook file name "
            f'(POSIX file "{safe_path}" as text)'
        )
        lines.append(f'    tell worksheet "{safe_sheet}" of wb')
    else:
        lines.append("    tell active sheet of active workbook")

    for row in sorted(rows_data.keys()):
        cols = rows_data[row]
        min_col = min(cols.keys())
        max_col = max(cols.keys())

        values: list[str] = []
        for c in range(min_col, max_col + 1):
            v = cols.get(c, "")
            safe_v = v.replace("\\", "\\\\").replace('"', '\\"')
            values.append(f'"{safe_v}"')

        range_ref = f"{_col_to_letter(min_col)}{row}:{_col_to_letter(max_col)}{row}"
        value_list = "{" + ", ".join(values) + "}"
        lines.append(f'        set value of range "{range_ref}" to {{{value_list}}}')

    lines.append("    end tell")

    if output_path and not workbook_name:
        lines.append("    save wb")
        lines.append("    close wb saving no")

    lines.append("end tell")

    return "\n".join(lines)


def _execute_applescript(script: str, timeout: int = 120) -> tuple[bool, str]:
    """Execute AppleScript via osascript with stdin piping."""
    try:
        result = subprocess.run(
            ["osascript"],
            input=script,
            capture_output=True,
            text=True,
            timeout=timeout,
        )
        if result.returncode != 0:
            return False, result.stderr.strip()
        return True, result.stdout.strip()
    except subprocess.TimeoutExpired:
        return False, f"AppleScript timed out after {timeout}s"
    except FileNotFoundError:
        return False, "osascript not found -- this script requires macOS"


# ---------------------------------------------------------------
# Main orchestrator
# ---------------------------------------------------------------


def fill_template_osa(
    template_path: Path,
    block_number: int,
    academic_year: int,
    output_path: Path,
    sheet_name: str | None = None,
    json_path: Path | None = None,
    dry_run: bool = False,
    verbose: bool = False,
    timeout: int = 120,
) -> dict:
    """Fill Excel template via AppleScript. Returns stats dict."""
    # Phase 0: Pre-flight
    if sys.platform != "darwin":
        print("ERROR: This script requires macOS (AppleScript/osascript).")
        return {"error": "not_macos"}

    if not template_path.exists():
        print(f"ERROR: Template not found: {template_path}")
        return {"error": "template_not_found"}

    # Phase 1a: Read names from template (openpyxl read-only)
    wb = load_workbook(template_path, read_only=True, data_only=True)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            print(f"ERROR: Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
            wb.close()
            return {"error": "sheet_not_found"}
        ws = wb[sheet_name]
    else:
        target = f"Block {block_number}"
        if target in wb.sheetnames:
            ws = wb[target]
        else:
            ws = wb.active
            print(
                f"Sheet 'Block {block_number}' not found, using active sheet: {ws.title}"
            )

    resident_names = _discover_name_rows(ws, ROW_RESIDENT_START, ROW_RESIDENT_END)
    faculty_names = _discover_name_rows(ws, ROW_FACULTY_START, ROW_FACULTY_END)
    # Rotation-code fallback for rows without names
    rotation_rows = _discover_rotation_rows(ws, ROW_RESIDENT_START, ROW_RESIDENT_END)
    wb.close()

    if verbose:
        print(f"\nWorkbook: {template_path.name}")
        print(f"Sheet:    {ws.title}")
        print(f"\nResidents by name ({len(resident_names)}):")
        for name, row in sorted(resident_names.items(), key=lambda x: x[1]):
            print(f"  Row {row:2d}: {name}")
        print(f"\nResidents by rotation ({len(rotation_rows)} unnamed rows):")
        for row, (r1, r2) in sorted(rotation_rows.items()):
            label = f"{r1}/{r2}" if r2 else r1
            print(f"  Row {row:2d}: {label}")
        print(f"\nFaculty ({len(faculty_names)}):")
        for name, row in sorted(faculty_names.items(), key=lambda x: x[1]):
            print(f"  Row {row:2d}: {name}")

    all_wb_names = {**resident_names, **faculty_names}

    if not all_wb_names and not rotation_rows:
        print("ERROR: No names or rotation codes found in template.")
        return {"error": "no_names"}

    # Phase 1b: Load schedule data
    data = _load_schedule_data(block_number, academic_year, json_path)
    block_start = date.fromisoformat(data["block_start"])
    block_end = date.fromisoformat(data["block_end"])
    num_days = (block_end - block_start).days + 1
    print(f"\nBlock {block_number}: {block_start} to {block_end} ({num_days} days)")

    # Phase 1c: Compute cell map
    cell_map, unmatched = _compute_cell_map(data, all_wb_names, rotation_rows)

    matched_count = (
        len(data.get("residents", [])) + len(data.get("faculty", [])) - len(unmatched)
    )

    stats = {
        "matched": matched_count,
        "unmatched": len(unmatched),
        "cells": len(cell_map),
    }

    if dry_run:
        print(f"\n[DRY RUN] Cell map: {len(cell_map)} cells")
        print(f"  Matched:   {matched_count} people")
        print(f"  Unmatched: {len(unmatched)} people")
        if unmatched and verbose:
            for name in unmatched:
                print(f"    - {name}")

        if verbose:
            print("\nCell assignments (first 50):")
            for i, ((row, col), value) in enumerate(sorted(cell_map.items())):
                if i >= 50:
                    print(f"  ... and {len(cell_map) - 50} more")
                    break
                print(f"  {_col_to_letter(col)}{row}: {value}")

        return stats

    # Phase 2: Fill via AppleScript
    # Copy template -> open copy -> fill -> save -> close (no conflicts)
    import shutil

    shutil.copy2(template_path, output_path)
    print(f"  Copied template -> {output_path}")

    script = _build_applescript(
        cell_map,
        sheet_name=ws.title,
        output_path=output_path,
    )

    if verbose:
        script_kb = len(script.encode()) / 1024
        print(f"\nAppleScript: {script_kb:.1f} KB, {len(cell_map)} cell writes")

    print(f"\nFilling via Excel.app ({len(cell_map)} cells)...")
    success, message = _execute_applescript(script, timeout)

    if not success:
        print(f"\nERROR: AppleScript failed: {message}")
        return {"error": "applescript_failed", "detail": message, **stats}

    print(f"\n{'=' * 50}")
    print("RESULTS")
    print(f"{'=' * 50}")
    print(f"  Matched:    {matched_count} people")
    print(f"  Unmatched:  {len(unmatched)} people")
    if unmatched:
        for name in unmatched:
            print(f"    - {name}")
    print(f"  Cells:      {len(cell_map)}")
    print(f"  Output:     {output_path}")

    return stats


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Fill Excel template via AppleScript (preserves all formatting)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s --block 12 --academic-year 2025 --template ~/Template.xlsx
  %(prog)s --block 12 --academic-year 2025 --template ~/Template.xlsx --dry-run
  %(prog)s --block 12 --academic-year 2025 --template ~/Template.xlsx --json /tmp/b12.json
        """,
    )
    parser.add_argument("--block", type=int, required=True, help="Block number (1-13)")
    parser.add_argument(
        "--academic-year",
        type=int,
        required=True,
        help="Starting year (e.g., 2025 for AY25-26)",
    )
    parser.add_argument(
        "--template", type=str, required=True, help="Path to Excel template (.xlsx)"
    )
    parser.add_argument(
        "--output",
        type=str,
        default=None,
        help="Output path (default: Block{N}_OSA.xlsx next to template)",
    )
    parser.add_argument("--sheet", type=str, default=None, help="Sheet name")
    parser.add_argument(
        "--json", type=str, default=None, help="Pre-exported JSON (skip DB query)"
    )
    parser.add_argument(
        "--dry-run", action="store_true", help="Print cell map, no Excel writes"
    )
    parser.add_argument(
        "--timeout", type=int, default=120, help="AppleScript timeout in seconds"
    )
    parser.add_argument(
        "--verbose", "-v", action="store_true", help="Show detailed output"
    )

    args = parser.parse_args()

    template_path = Path(args.template).expanduser().resolve()
    if not template_path.exists():
        print(f"ERROR: Template not found: {template_path}")
        return 1

    output_path = (
        Path(args.output).expanduser().resolve()
        if args.output
        else template_path.parent / f"Block{args.block}_OSA.xlsx"
    )

    json_path = Path(args.json).expanduser().resolve() if args.json else None

    result = fill_template_osa(
        template_path=template_path,
        block_number=args.block,
        academic_year=args.academic_year,
        output_path=output_path,
        sheet_name=args.sheet,
        json_path=json_path,
        dry_run=args.dry_run,
        verbose=args.verbose,
        timeout=args.timeout,
    )

    if "error" in result:
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
