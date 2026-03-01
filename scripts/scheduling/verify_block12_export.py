#!/usr/bin/env python3
"""Block 12 Export Zeroing Verification.

Comprehensive cell-by-cell comparison of:
  - XLSX export (from canonical_schedule_export_service)
  - DB schedule_grid view (ground truth)

This verifies that every cell in the exported XLSX matches the DB,
accounting for known display code transformations.

Usage:
    python scripts/scheduling/verify_block12_export.py [xlsx_path]

Default xlsx_path: /tmp/Block12_Export_Test.xlsx
Requires: psycopg2, openpyxl
"""

from __future__ import annotations

import sys
from collections import defaultdict
from datetime import date, timedelta
from pathlib import Path

import psycopg2
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell


# ── Constants ─────────────────────────────────────────────────────────────
DB_DSN = "host=localhost port=5432 dbname=residency_scheduler user=scheduler"
BLOCK_START = date(2026, 5, 7)
BLOCK_END = date(2026, 6, 3)
TOTAL_DAYS = 28

# BT2 layout
COL_ROTATION1 = 1
COL_ROTATION2 = 2
COL_TEMPLATE = 3
COL_ROLE = 4
COL_NAME = 5
COL_SCHEDULE_START = 6
COLS_PER_DAY = 2

# Row bands
RESIDENT_ROW_START = 9
RESIDENT_ROW_END = 30
FACULTY_ROW_START = 31
FACULTY_ROW_END = 80


def cell_value(ws, row, col):
    """Get cell value, handling MergedCell."""
    c = ws.cell(row=row, column=col)
    if isinstance(c, MergedCell):
        return None
    v = c.value
    if v is None:
        return None
    return str(v).strip()


def load_xlsx_grid(xlsx_path: str) -> dict:
    """Load XLSX into structured grid.

    Returns: {
        "people": [
            {
                "row": int,
                "name": str,
                "rotation1": str,
                "rotation2": str,
                "template": str,
                "role": str,
                "type": "resident" | "faculty",
                "codes": {(date, "AM"|"PM"): str, ...}
            }
        ],
        "sheet_name": str,
    }
    """
    wb = load_workbook(xlsx_path, data_only=True)

    # Find the schedule sheet
    sheet_name = None
    for name in wb.sheetnames:
        if "Block Template2" in name or "Block 12" in name:
            sheet_name = name
            break
    if not sheet_name:
        sheet_name = wb.sheetnames[0]

    ws = wb[sheet_name]
    people = []

    # Scan resident and faculty rows
    for row_start, row_end, ptype in [
        (RESIDENT_ROW_START, RESIDENT_ROW_END, "resident"),
        (FACULTY_ROW_START, FACULTY_ROW_END, "faculty"),
    ]:
        for row in range(row_start, row_end + 1):
            name = cell_value(ws, row, COL_NAME)
            if not name:
                continue

            rot1 = cell_value(ws, row, COL_ROTATION1) or ""
            rot2 = cell_value(ws, row, COL_ROTATION2) or ""
            template = cell_value(ws, row, COL_TEMPLATE) or ""
            role = cell_value(ws, row, COL_ROLE) or ""

            codes = {}
            for day_offset in range(TOTAL_DAYS):
                day_date = BLOCK_START + timedelta(days=day_offset)
                am_col = COL_SCHEDULE_START + day_offset * COLS_PER_DAY
                pm_col = am_col + 1

                am = cell_value(ws, row, am_col)
                pm = cell_value(ws, row, pm_col)

                if am is not None:
                    codes[(day_date, "AM")] = am
                if pm is not None:
                    codes[(day_date, "PM")] = pm

            people.append({
                "row": row,
                "name": name,
                "rotation1": rot1,
                "rotation2": rot2,
                "template": template,
                "role": role,
                "type": ptype,
                "codes": codes,
            })

    wb.close()
    return {"people": people, "sheet_name": sheet_name}


def load_db_grid(conn) -> dict:
    """Load schedule_grid view into structured dict.

    Returns: {
        name: {
            "person_type": str,
            "codes": {(date, "AM"|"PM"): str, ...},
            "sources": {(date, "AM"|"PM"): str, ...},
        }
    }
    """
    cur = conn.cursor()
    cur.execute("""
        SELECT name, person_type, date, am_code, pm_code, am_source, pm_source
        FROM schedule_grid
        WHERE date >= %s AND date <= %s
        ORDER BY person_type, name, date
    """, (BLOCK_START, BLOCK_END))

    people = {}
    for name, ptype, dt, am_code, pm_code, am_src, pm_src in cur.fetchall():
        if name not in people:
            people[name] = {
                "person_type": ptype,
                "codes": {},
                "sources": {},
            }
        if am_code:
            people[name]["codes"][(dt, "AM")] = am_code
            people[name]["sources"][(dt, "AM")] = am_src or ""
        if pm_code:
            people[name]["codes"][(dt, "PM")] = pm_code
            people[name]["sources"][(dt, "PM")] = pm_src or ""

    cur.close()
    return people


def load_display_abbreviations(conn) -> dict:
    """Load activity code -> display_abbreviation mapping."""
    cur = conn.cursor()
    cur.execute("""
        SELECT code, display_abbreviation
        FROM activities
        WHERE display_abbreviation IS NOT NULL
          AND display_abbreviation != code
    """)
    result = {}
    for code, display in cur.fetchall():
        result[code] = display
    cur.close()
    return result


def normalize_name_for_match(xlsx_name: str) -> str:
    """Convert 'Last, First' to 'First Last' for DB matching."""
    if "," in xlsx_name:
        parts = xlsx_name.split(",", 1)
        last = parts[0].strip()
        first = parts[1].strip()
        return f"{first} {last}"
    return xlsx_name


def match_xlsx_to_db(xlsx_people, db_people):
    """Match XLSX names to DB names. Returns list of (xlsx_person, db_name)."""
    matches = []
    unmatched_xlsx = []

    db_names_by_last = defaultdict(list)
    for name in db_people:
        last = name.split()[-1].lower() if " " in name else name.lower()
        db_names_by_last[last].append(name)

    for xp in xlsx_people:
        normalized = normalize_name_for_match(xp["name"])

        # Exact match
        if normalized in db_people:
            matches.append((xp, normalized))
            continue

        # Last name match
        last = normalized.split()[-1].lower() if " " in normalized else normalized.lower()
        candidates = db_names_by_last.get(last, [])
        if len(candidates) == 1:
            matches.append((xp, candidates[0]))
            continue

        unmatched_xlsx.append(xp)

    return matches, unmatched_xlsx


def run_checks(xlsx_path: str):
    """Run all zeroing checks."""
    print("=" * 70)
    print("BLOCK 12 EXPORT ZEROING VERIFICATION")
    print("=" * 70)
    print(f"XLSX: {xlsx_path}")
    print(f"DB:   schedule_grid view ({BLOCK_START} to {BLOCK_END})")
    print()

    # Load data
    print("Loading XLSX...", end=" ")
    xlsx_data = load_xlsx_grid(xlsx_path)
    print(f"OK ({len(xlsx_data['people'])} people on sheet '{xlsx_data['sheet_name']}')")

    print("Connecting to DB...", end=" ")
    conn = psycopg2.connect(DB_DSN)
    print("OK")

    print("Loading schedule_grid...", end=" ")
    db_grid = load_db_grid(conn)
    print(f"OK ({len(db_grid)} people)")

    print("Loading display abbreviations...", end=" ")
    display_abbrevs = load_display_abbreviations(conn)
    print(f"OK ({len(display_abbrevs)} mappings)")
    print()

    checks_passed = 0
    checks_failed = 0
    checks_warned = 0

    # ── Check 1: Headcount match ──────────────────────────────────────
    print("─" * 70)
    print("CHECK 1: Headcount")
    xlsx_residents = [p for p in xlsx_data["people"] if p["type"] == "resident"]
    xlsx_faculty = [p for p in xlsx_data["people"] if p["type"] == "faculty"]
    db_residents = [n for n, d in db_grid.items() if d["person_type"] == "resident"]
    db_faculty = [n for n, d in db_grid.items() if d["person_type"] == "faculty"]

    res_match = len(xlsx_residents) == len(db_residents)
    fac_match = len(xlsx_faculty) == len(db_faculty)

    print(f"  Residents: XLSX={len(xlsx_residents)}, DB={len(db_residents)} {'✓' if res_match else '✗'}")
    print(f"  Faculty:   XLSX={len(xlsx_faculty)}, DB={len(db_faculty)} {'✓' if fac_match else '✗'}")

    if res_match and fac_match:
        print("  PASS")
        checks_passed += 1
    else:
        print("  FAIL")
        checks_failed += 1
    print()

    # ── Check 2: Name matching ────────────────────────────────────────
    print("─" * 70)
    print("CHECK 2: Name matching (XLSX → DB)")
    matches, unmatched = match_xlsx_to_db(xlsx_data["people"], db_grid)
    print(f"  Matched: {len(matches)}")
    if unmatched:
        print(f"  UNMATCHED ({len(unmatched)}):")
        for u in unmatched:
            print(f"    Row {u['row']}: {u['name']}")
        print("  FAIL")
        checks_failed += 1
    else:
        print("  PASS")
        checks_passed += 1
    print()

    # ── Check 3: Cell coverage ────────────────────────────────────────
    print("─" * 70)
    print("CHECK 3: Cell coverage (56 slots per person)")
    expected_slots = TOTAL_DAYS * 2  # 28 days × 2 (AM/PM)
    coverage_issues = []
    for xp, db_name in matches:
        xlsx_count = len(xp["codes"])
        db_count = len(db_grid[db_name]["codes"])
        if xlsx_count != expected_slots:
            coverage_issues.append(f"  XLSX {xp['name']}: {xlsx_count}/{expected_slots}")
        if db_count != expected_slots:
            coverage_issues.append(f"  DB   {db_name}: {db_count}/{expected_slots}")

    if coverage_issues:
        for issue in coverage_issues[:20]:
            print(issue)
        if len(coverage_issues) > 20:
            print(f"  ... and {len(coverage_issues) - 20} more")
        print(f"  FAIL ({len(coverage_issues)} issues)")
        checks_failed += 1
    else:
        print(f"  All {len(matches)} people have {expected_slots} slots each")
        print("  PASS")
        checks_passed += 1
    print()

    # ── Check 4: Cell-by-cell code comparison ─────────────────────────
    print("─" * 70)
    print("CHECK 4: Cell-by-cell code comparison (XLSX vs DB)")

    exact_matches = 0
    display_transforms = 0
    transform_reasons = defaultdict(int)
    mismatches = []
    mismatch_details = []

    # Build reverse display abbreviation map: display_abbr -> code
    reverse_display = {}
    for code, display in display_abbrevs.items():
        reverse_display[display] = code

    for xp, db_name in matches:
        db_codes = db_grid[db_name]["codes"]
        db_sources = db_grid[db_name]["sources"]

        for key, xlsx_code in xp["codes"].items():
            db_code = db_codes.get(key, "")

            if not xlsx_code and not db_code:
                exact_matches += 1
                continue

            if xlsx_code == db_code:
                exact_matches += 1
                continue

            # Check if xlsx_code is the display_abbreviation of db_code
            expected_display = display_abbrevs.get(db_code)
            if expected_display and xlsx_code == expected_display:
                display_transforms += 1
                transform_reasons[f"{db_code}→{xlsx_code} (display_abbrev)"] += 1
                continue

            # Check case-insensitive
            if xlsx_code.upper() == db_code.upper():
                display_transforms += 1
                transform_reasons[f"{db_code}→{xlsx_code} (case)"] += 1
                continue

            # Check if db_code is leave variant (LV-AM, LV-PM -> LV)
            if db_code in ("LV-AM", "LV-PM") and xlsx_code == "LV":
                display_transforms += 1
                transform_reasons[f"{db_code}→LV (leave consolidation)"] += 1
                continue

            # Check W -> OFF (weekend -> weekday mapping) or OFF -> W
            if {xlsx_code, db_code} == {"W", "OFF"}:
                display_transforms += 1
                transform_reasons[f"{db_code}→{xlsx_code} (W/OFF)"] += 1
                continue

            # Recovery code
            if db_code == "recovery" and xlsx_code in ("REC", "recovery"):
                display_transforms += 1
                transform_reasons[f"recovery→{xlsx_code}"] += 1
                continue

            mismatches.append(key)
            dt, tod = key
            src = db_sources.get(key, "?")
            mismatch_details.append({
                "person": xp["name"],
                "db_name": db_name,
                "date": dt,
                "tod": tod,
                "xlsx": xlsx_code,
                "db": db_code,
                "source": src,
            })

    total_cells = exact_matches + display_transforms + len(mismatches)
    print(f"  Total cells compared:    {total_cells}")
    print(f"  Exact matches:           {exact_matches}")
    print(f"  Display transforms:      {display_transforms}")
    print(f"  True mismatches:         {len(mismatches)}")

    if transform_reasons:
        print()
        print("  Display transform breakdown:")
        for reason, count in sorted(transform_reasons.items(), key=lambda x: -x[1]):
            print(f"    {count:4d}  {reason}")

    if mismatch_details:
        print()
        print("  MISMATCHES:")
        # Group by person for readability
        by_person = defaultdict(list)
        for m in mismatch_details:
            by_person[m["person"]].append(m)

        for person, details in sorted(by_person.items()):
            print(f"    {person} ({len(details)} cells):")
            for d in sorted(details, key=lambda x: (x["date"], x["tod"])):
                print(
                    f"      {d['date']} {d['tod']}: "
                    f"XLSX='{d['xlsx']}' DB='{d['db']}' "
                    f"(source={d['source']})"
                )

    if len(mismatches) == 0:
        print("  PASS")
        checks_passed += 1
    elif len(mismatches) <= 10:
        print(f"  WARN ({len(mismatches)} mismatches)")
        checks_warned += 1
    else:
        print(f"  FAIL ({len(mismatches)} mismatches)")
        checks_failed += 1
    print()

    # ── Check 5: Metadata sheets ──────────────────────────────────────
    print("─" * 70)
    print("CHECK 5: Metadata sheets")
    wb = load_workbook(xlsx_path)
    meta_checks = {
        "__SYS_META__": False,
        "__REF__": False,
    }
    for name in wb.sheetnames:
        for meta_key in meta_checks:
            if meta_key in name:
                meta_checks[meta_key] = True

    for name, found in meta_checks.items():
        print(f"  {name}: {'FOUND' if found else 'MISSING'}")

    if all(meta_checks.values()):
        print("  PASS")
        checks_passed += 1
    else:
        print("  FAIL")
        checks_failed += 1
    wb.close()
    print()

    # ── Check 6: No empty cells in schedule region ────────────────────
    print("─" * 70)
    print("CHECK 6: No empty/None cells in schedule region")
    empty_cells = []
    for xp, db_name in matches:
        for day_offset in range(TOTAL_DAYS):
            day_date = BLOCK_START + timedelta(days=day_offset)
            for tod in ("AM", "PM"):
                key = (day_date, tod)
                code = xp["codes"].get(key)
                if code is None or code == "":
                    empty_cells.append((xp["name"], day_date, tod))

    if empty_cells:
        print(f"  Empty cells: {len(empty_cells)}")
        for name, dt, tod in empty_cells[:20]:
            print(f"    {name}: {dt} {tod}")
        print("  FAIL")
        checks_failed += 1
    else:
        total_expected = len(matches) * TOTAL_DAYS * 2
        print(f"  All {total_expected} schedule cells populated")
        print("  PASS")
        checks_passed += 1
    print()

    # ── Check 7: Row ordering ─────────────────────────────────────────
    print("─" * 70)
    print("CHECK 7: Row ordering (PGY desc + alpha for residents, alpha for faculty)")
    order_issues = []

    # Residents: PGY 3 first, then 2, then 1, alphabetical within
    res_rows = sorted(xlsx_residents, key=lambda p: p["row"])
    prev_pgy = 99
    prev_name = ""
    for p in res_rows:
        role = p["role"]
        pgy = 0
        if role:
            try:
                pgy = int(role.replace("PGY", "").replace("pgy", "").strip())
            except ValueError:
                pass

        if pgy > prev_pgy:
            order_issues.append(f"  Row {p['row']}: PGY {pgy} after PGY {prev_pgy}")
        elif pgy == prev_pgy:
            normalized = normalize_name_for_match(p["name"])
            if normalized.lower() < prev_name.lower():
                order_issues.append(
                    f"  Row {p['row']}: {p['name']} before {prev_name} (same PGY)"
                )
        prev_pgy = pgy
        prev_name = normalize_name_for_match(p["name"])

    # Faculty: alphabetical
    fac_rows = sorted(xlsx_faculty, key=lambda p: p["row"])
    prev_name = ""
    for p in fac_rows:
        normalized = normalize_name_for_match(p["name"])
        if normalized.lower() < prev_name.lower():
            order_issues.append(f"  Row {p['row']}: {p['name']} before {prev_name}")
        prev_name = normalized

    if order_issues:
        for issue in order_issues:
            print(issue)
        print(f"  WARN ({len(order_issues)} issues)")
        checks_warned += 1
    else:
        print("  Residents: PGY desc + alpha ✓")
        print("  Faculty: alpha ✓")
        print("  PASS")
        checks_passed += 1
    print()

    # ── Check 8: Weekend codes ────────────────────────────────────────
    print("─" * 70)
    print("CHECK 8: Weekend codes (Sat/Sun should have W or rotation-appropriate code)")

    weekend_issues = []
    # Weekend rotation codes that are valid on weekends
    VALID_WEEKEND_CODES = {
        "W", "LV", "LV-AM", "LV-PM", "DEP",
        "FMIT", "NF", "PedsNF", "PedNF", "L&D", "LDNF", "PedW",
        "NBN", "HILO", "JAPAN", "TDY", "IMW", "KAP",
        "pcat", "do", "PCAT", "DO",
        "PEM",  # Peds EM — emergency rotations work weekends
        "OFF", "off",  # NF residents sleep during day on weekends
        "CALL", "call",  # Weekend overnight call
    }

    for xp, db_name in matches:
        for day_offset in range(TOTAL_DAYS):
            day_date = BLOCK_START + timedelta(days=day_offset)
            if day_date.weekday() < 5:  # Not weekend
                continue
            for tod in ("AM", "PM"):
                key = (day_date, tod)
                code = xp["codes"].get(key, "")
                if code and code not in VALID_WEEKEND_CODES:
                    # Check DB too
                    db_code = db_grid[db_name]["codes"].get(key, "")
                    if db_code not in VALID_WEEKEND_CODES:
                        weekend_issues.append(
                            f"  {xp['name']} {day_date} {tod}: "
                            f"XLSX='{code}' DB='{db_code}'"
                        )

    if weekend_issues:
        for issue in weekend_issues[:20]:
            print(issue)
        print(f"  FAIL ({len(weekend_issues)} issues)")
        checks_failed += 1
    else:
        weekends = sum(
            1 for d in range(TOTAL_DAYS)
            if (BLOCK_START + timedelta(days=d)).weekday() >= 5
        )
        print(f"  {weekends} weekend days × {len(matches)} people × 2 slots = "
              f"{weekends * len(matches) * 2} checks")
        print("  PASS")
        checks_passed += 1
    print()

    # ── Summary ───────────────────────────────────────────────────────
    print("=" * 70)
    print("SUMMARY")
    print(f"  Passed:  {checks_passed}")
    print(f"  Warned:  {checks_warned}")
    print(f"  Failed:  {checks_failed}")
    total = checks_passed + checks_warned + checks_failed
    print(f"  Total:   {total}")
    print()

    if checks_failed == 0 and checks_warned == 0:
        print("  ★ ALL CHECKS PASSED — ZEROING VERIFIED ★")
    elif checks_failed == 0:
        print("  ⚠ PASSED WITH WARNINGS")
    else:
        print("  ✗ ZEROING INCOMPLETE — FIXES NEEDED")

    conn.close()
    return checks_failed


if __name__ == "__main__":
    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else "/tmp/Block12_Export_Test.xlsx"
    if not Path(xlsx_path).exists():
        print(f"ERROR: XLSX file not found: {xlsx_path}")
        print("Run the export first or provide the correct path.")
        sys.exit(1)
    exit_code = run_checks(xlsx_path)
    sys.exit(exit_code)
